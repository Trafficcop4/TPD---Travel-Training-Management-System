"""Shared styling helpers for the TPD Travel & Training workbook builders.

Both build scripts import from here so the two workbooks look like one system.
Palette: TPD navy header bars, light-gray page ground, white input cells.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

FONT = "Arial"

NAVY = "1F3B5C"      # header bars / titles
STEEL = "44607E"     # sub-headers
PAGE = "EDEFF2"      # page background (non-input area)
INPUT = "FFFFFF"     # white input boxes
ACCENT = "C8A24B"    # gold accent (TPD badge gold)
WARN = "C0392B"      # overdue / owed
OK = "1E7145"        # good / reconciled
LIGHT_BAND = "DCE3EB"  # table banding / tile fill
BORDER_GRAY = "9AA5B1"

thin = Side(style="thin", color=BORDER_GRAY)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
UNDER = Border(bottom=thin)

F_TITLE = Font(name=FONT, size=16, bold=True, color="FFFFFF")
F_SUB = Font(name=FONT, size=10, italic=True, color="FFFFFF")
F_SECTION = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_LABEL = Font(name=FONT, size=10, bold=True, color="1F3B5C")
F_BODY = Font(name=FONT, size=10)
F_SMALL = Font(name=FONT, size=8, italic=True, color="555555")
F_INPUT = Font(name=FONT, size=10, color="0000FF")   # blue = user-entered value
F_FORMULA = Font(name=FONT, size=10)                  # black = formula
F_KPI = Font(name=FONT, size=14, bold=True, color="1F3B5C")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_STEEL = PatternFill("solid", fgColor=STEEL)
FILL_PAGE = PatternFill("solid", fgColor=PAGE)
FILL_INPUT = PatternFill("solid", fgColor=INPUT)
FILL_BAND = PatternFill("solid", fgColor=LIGHT_BAND)
FILL_GOLD = PatternFill("solid", fgColor=ACCENT)
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")  # cells the user must review/fill

A_LEFT = Alignment(horizontal="left", vertical="center")
A_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
A_RIGHT = Alignment(horizontal="right", vertical="center")
A_CENTER = Alignment(horizontal="center", vertical="center")

CUR = '$#,##0.00'
CUR0 = '$#,##0'
DATE = 'mm/dd/yyyy'
TIME = 'h:mm AM/PM'


def paint(ws, min_row, max_row, min_col, max_col, fill=FILL_PAGE):
    """Flood-fill a rectangle so the sheet reads as a designed page."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            if cell.font.name != FONT:
                cell.font = F_BODY


def title_bar(ws, row, first_col, last_col, text, subtitle=None):
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=first_col, value=text)
    c.font = F_TITLE
    c.fill = FILL_NAVY
    c.alignment = A_LEFT
    for col in range(first_col, last_col + 1):
        ws.cell(row=row, column=col).fill = FILL_NAVY
    ws.row_dimensions[row].height = 26
    if subtitle:
        r2 = row + 1
        ws.merge_cells(start_row=r2, start_column=first_col, end_row=r2, end_column=last_col)
        c2 = ws.cell(row=r2, column=first_col, value=subtitle)
        c2.font = F_SUB
        c2.fill = FILL_NAVY
        c2.alignment = A_LEFT
        for col in range(first_col, last_col + 1):
            ws.cell(row=r2, column=col).fill = FILL_NAVY
        ws.row_dimensions[r2].height = 14


def section_bar(ws, row, first_col, last_col, text):
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=first_col, value=text)
    c.font = F_SECTION
    c.fill = FILL_STEEL
    c.alignment = A_LEFT
    for col in range(first_col, last_col + 1):
        ws.cell(row=row, column=col).fill = FILL_STEEL
    ws.row_dimensions[row].height = 16


def label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_LABEL
    c.alignment = A_RIGHT
    return c


def input_cell(ws, row, col, fmt=None, name=None, wb=None, sheet=None, span=1, value=None):
    """White bordered input box; optionally register a workbook-level defined name."""
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    for i in range(span):
        cc = ws.cell(row=row, column=col + i)
        cc.fill = FILL_INPUT
        cc.border = BOX
    c = ws.cell(row=row, column=col)
    c.font = F_INPUT
    c.alignment = A_LEFT
    if fmt:
        c.number_format = fmt
    if value is not None:
        c.value = value
    if name and wb is not None and sheet is not None:
        ref = f"'{sheet}'!${get_column_letter(col)}${row}"
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    return c


def define(wb, name, sheet, coord):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{coord}")


def col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
