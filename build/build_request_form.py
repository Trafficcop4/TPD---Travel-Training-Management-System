"""Builds workbooks/TPD_Travel_Request_Form.xlsx — the officer-facing one-page request form.

Officers fill in the white boxes and click the Submit button (added by
tools/Install-VBA.ps1, which also imports src/vba/request_form/modSubmit.bas
and saves the file as .xlsm). Every input box carries a workbook-level defined
name (rq_*) so the master workbook's import macro reads fields by name, never
by cell address — the layout can change without breaking the import.
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(__file__))
from xlstyle import (  # noqa: E402
    A_CENTER, A_LEFT, A_LEFT_WRAP, BOX, CUR, DATE, F_BODY, F_INPUT, F_LABEL,
    F_SMALL, FILL_GOLD, FILL_INPUT, TIME, col_widths, define, input_cell,
    label, paint, section_bar, title_bar,
)

OUT = os.path.join(os.path.dirname(__file__), "..", "workbooks", "TPD_Travel_Request_Form.xlsx")
SHEET = "Request Form"


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.sheet_view.showGridLines = False

    col_widths(ws, {"A": 2, "B": 24, "C": 27, "D": 23, "E": 21, "F": 3})
    paint(ws, 1, 34, 1, 6)

    title_bar(ws, 1, 1, 6, "TYLER POLICE DEPARTMENT — TRAVEL / TRAINING REQUEST",
              "Fill in the WHITE boxes only, then click  “Submit to Training Unit”.  "
              "The email opens for your review — nothing sends automatically.")
    # Gold rule under the header
    for c in range(1, 7):
        ws.cell(row=3, column=c).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3

    def ic(row, col, fmt=None, name=None, span=1, example=None):
        cell = input_cell(ws, row, col, fmt=fmt, name=name, wb=wb, sheet=SHEET, span=span)
        if example:
            cell.comment = Comment(f"Example: {example}", "TPD Training Unit")
        return cell

    # ---- 1. Officer ----------------------------------------------------
    section_bar(ws, 5, 2, 5, "1.  OFFICER")
    label(ws, 6, 2, "Name")
    ic(6, 3, name="rq_Name", example="J. Q. Officer")
    label(ws, 6, 4, "Rank")
    ic(6, 5, name="rq_Rank", example="Officer")
    label(ws, 7, 2, "Employee ID")
    ic(7, 3, name="rq_EmpID", example="4127")
    label(ws, 7, 4, "Division / Unit")
    ic(7, 5, name="rq_Division", example="Patrol – B Shift")
    label(ws, 8, 2, "Email")
    ic(8, 3, name="rq_Email", example="jofficer@tylertexas.com")
    label(ws, 8, 4, "Phone")
    ic(8, 5, name="rq_Phone", example="903-555-0142")
    label(ws, 9, 2, "Supervisor Name")
    ic(9, 3, name="rq_SupName", example="Sgt. A. Smith")
    label(ws, 9, 4, "Supervisor Email")
    ic(9, 5, name="rq_SupEmail", example="asmith@tylertexas.com")

    # ---- 2. Course / Event --------------------------------------------
    section_bar(ws, 11, 2, 5, "2.  COURSE / EVENT")
    label(ws, 12, 2, "Course / Event Title")
    ic(12, 3, name="rq_Course", span=3, example="Advanced Crash Reconstruction")
    label(ws, 13, 2, "Provider / Host")
    ic(13, 3, name="rq_Provider", example="TEEX / TAMU")
    label(ws, 13, 4, "TCOLE Course #")
    ic(13, 5, name="rq_TCOLE", example="3341 (optional)")
    label(ws, 14, 2, "Location (City, ST)")
    ic(14, 3, name="rq_Location", example="College Station, TX")
    label(ws, 14, 4, "Website / Info")
    ic(14, 5, name="rq_Website", example="teex.org/... (optional)")
    label(ws, 15, 2, "Class Start Date")
    ic(15, 3, fmt=DATE, name="rq_Start", example="08/24/2026")
    label(ws, 15, 4, "Class End Date")
    ic(15, 5, fmt=DATE, name="rq_End", example="08/28/2026")
    label(ws, 16, 2, "Departure Date")
    ic(16, 3, fmt=DATE, name="rq_DepDate", example="08/23/2026")
    label(ws, 16, 4, "Departure Time")
    ic(16, 5, fmt=TIME, name="rq_DepTime", example="1:00 PM")
    label(ws, 17, 2, "Return Date")
    ic(17, 3, fmt=DATE, name="rq_RetDate", example="08/28/2026")
    label(ws, 17, 4, "Return Time")
    ic(17, 5, fmt=TIME, name="rq_RetTime", example="6:30 PM")

    # ---- 3. Estimated costs -------------------------------------------
    section_bar(ws, 19, 2, 5, "3.  ESTIMATED COSTS  (best available numbers — the Training Unit verifies)")
    label(ws, 20, 2, "Registration Fee")
    ic(20, 3, fmt=CUR, name="rq_Reg", example="450.00 (0 if free)")
    label(ws, 20, 4, "Travel Mode")
    ic(20, 5, name="rq_Mode", example="City Vehicle")
    label(ws, 21, 2, "Lodging: # of Nights")
    ic(21, 3, fmt="0", name="rq_Nights", example="5 (0 if none)")
    label(ws, 21, 4, "Lodging Rate / Night")
    ic(21, 5, fmt=CUR, name="rq_LodgeRate", example="129.00")
    label(ws, 22, 2, "Round-Trip Miles (POV only)")
    ic(22, 3, fmt="0", name="rq_Miles", example="240")
    label(ws, 22, 4, "Airfare / Other Travel")
    ic(22, 5, fmt=CUR, name="rq_TravelOther", example="0.00")
    label(ws, 23, 2, "Other Costs")
    ic(23, 3, fmt=CUR, name="rq_Other", example="35.00")
    label(ws, 23, 4, "Other — Description")
    ic(23, 5, name="rq_OtherDesc", example="parking")

    # ---- 4. Justification ---------------------------------------------
    section_bar(ws, 25, 2, 5, "4.  JUSTIFICATION / NOTES")
    cell = input_cell(ws, 26, 2, wb=wb, sheet=SHEET, name="rq_Notes", span=4)
    ws.merge_cells(start_row=26, start_column=2, end_row=28, end_column=5)
    for r in range(26, 29):
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = FILL_INPUT
            ws.cell(row=r, column=c).border = BOX
    cell.alignment = A_LEFT_WRAP
    cell.font = F_INPUT
    cell.comment = Comment(
        "Example: Required for accident-investigator certification; "
        "covers TCOLE intermediate requirement.", "TPD Training Unit")

    # ---- Submit / footer ----------------------------------------------
    ws.merge_cells("B30:E30")
    c = ws.cell(row=30, column=2,
                value="▶  SUBMIT:  click the “Submit to Training Unit” button "
                      "(or press Ctrl+Shift+S). Review the email that opens, then press Send.")
    c.font = F_LABEL
    c.alignment = A_CENTER
    ws.row_dimensions[30].height = 22

    ws.merge_cells("B32:E32")
    c = ws.cell(row=32, column=2,
                value="Hover any white box for an example entry. Questions: contact the "
                      "Training Unit. Form v1.0 — do not rearrange this sheet; the intake "
                      "system reads it by field name.")
    c.font = F_SMALL
    c.alignment = A_LEFT

    # ---- Data validation ----------------------------------------------
    dv_mode = DataValidation(type="list", formula1='"City Vehicle,POV,Air,Other"',
                             allow_blank=True, showDropDown=False)
    dv_mode.error = "Choose City Vehicle, POV, Air, or Other."
    dv_mode.add(ws["E20"])
    ws.add_data_validation(dv_mode)

    dv_date = DataValidation(type="date", operator="greaterThan",
                             formula1="DATE(2020,1,1)", allow_blank=True)
    dv_date.errorTitle = "Enter a date"
    dv_date.error = "Enter a real calendar date, e.g. 08/24/2026."
    for coord in ("C15", "E15", "C16", "C17"):
        dv_date.add(ws[coord])
    ws.add_data_validation(dv_date)

    dv_time = DataValidation(type="time", operator="between",
                             formula1="0", formula2="0.99999", allow_blank=True)
    dv_time.errorTitle = "Enter a time"
    dv_time.error = "Enter a clock time, e.g. 6:30 AM."
    for coord in ("E16", "E17"):
        dv_time.add(ws[coord])
    ws.add_data_validation(dv_time)

    dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual",
                            formula1="0", allow_blank=True)
    dv_pos.error = "Enter a number (0 or more)."
    for coord in ("C20", "C21", "E21", "C22", "E22", "C23"):
        dv_pos.add(ws[coord])
    ws.add_data_validation(dv_pos)

    ws.print_area = "A1:F33"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ---- Hidden settings sheet ----------------------------------------
    cfg = wb.create_sheet("Form Settings")
    cfg["A1"] = "Setting"
    cfg["B1"] = "Value"
    cfg["C1"] = "Notes"
    for coord in ("A1", "B1", "C1"):
        cfg[coord].font = F_LABEL
    rows = [
        ("Training Unit email", "training@tylertexas.com",
         "PLACEHOLDER — set to the real Training Unit mailbox before distributing this form."),
        ("Email subject prefix", "Travel/Training Request",
         "The submit macro appends the officer's name and course title."),
        ("Form version", "1.0", ""),
    ]
    for i, (a, b, note) in enumerate(rows, start=2):
        cfg.cell(row=i, column=1, value=a).font = F_BODY
        cfg.cell(row=i, column=2, value=b).font = F_INPUT
        cfg.cell(row=i, column=3, value=note).font = F_SMALL
    define(wb, "rq_TrainingEmail", "Form Settings", "$B$2")
    define(wb, "rq_SubjectPrefix", "Form Settings", "$B$3")
    cfg.column_dimensions["A"].width = 24
    cfg.column_dimensions["B"].width = 34
    cfg.column_dimensions["C"].width = 70
    cfg.sheet_state = "hidden"

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {os.path.abspath(OUT)}")


if __name__ == "__main__":
    build()
