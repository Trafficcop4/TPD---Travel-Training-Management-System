"""Builds workbooks/TPD_Training_Master.xlsx — the Training Coordinator's master workbook.

Design principle: every calculation is a spreadsheet formula (transparent,
stable, survives personnel changes). VBA (imported later from src/vba/master
by tools/Install-VBA.ps1) only does what formulas cannot: import request
files, draft Outlook emails, push calendar reminders, print packets, archive.

Sheets:
  Dashboard    action queues + budget tiles + active-record selector
  Requests     the single source of truth — one row per trip, cols A..BI
  Application  internal approval application  (protected — do not modify)
  Pre 2-90     city travel form, pre-trip estimate (protected — do not modify)
  Post 2-90    city travel form, post-trip reconciliation (protected — do not modify)
  Notification attendee's trip notification sheet
  Ledger       financial ledger, written entirely by formulas from Requests
  YTD Report   monthly + category budget report, written by formulas
  Settings     every value a future coordinator may need to change
  Instructions in-workbook quick reference incl. the 10-minute year rollover
  Lists        dropdown source lists (hidden)

All per diem thresholds/rates, the mileage rate, deadlines, budget, email
addresses and email wording live on Settings as named cells — formulas and
VBA read them by name only.
"""
import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(__file__))
from xlstyle import (  # noqa: E402
    A_CENTER, A_LEFT, A_LEFT_WRAP, A_RIGHT, BOX, CUR, CUR0, DATE, F_BODY,
    F_INPUT, F_KPI, F_LABEL, F_SECTION, F_SMALL, FILL_BAND, FILL_GOLD,
    FILL_INPUT, FILL_NAVY, FILL_STEEL, FILL_YELLOW, FONT, NAVY, TIME, WARN,
    col_widths, define, label, paint, section_bar, title_bar,
)

OUT = os.path.join(os.path.dirname(__file__), "..", "workbooks", "TPD_Training_Master.xlsx")

NROWS = 250                 # request capacity per fiscal year
R1, R2 = 2, NROWS + 1       # first/last data row on Requests

F_HDR = Font(name=FONT, size=9, bold=True, color="FFFFFF")
F_CELL = Font(name=FONT, size=9)
F_NOTE = Font(name=FONT, size=9, italic=True, color="555555")

# ---------------------------------------------------------------- Requests
# (letter, header, width, number_format, kind)  kind: in=typed by coordinator/
# import macro, f=formula, h=numeric helper for dashboard/report formulas
REQ_COLS = [
    ("A", "Tracking #", 10, None, "in"),
    ("B", "Status", 16, None, "in"),
    ("C", "Officer", 18, None, "in"),
    ("D", "Rank", 10, None, "in"),
    ("E", "Emp ID", 8, None, "in"),
    ("F", "Division", 14, None, "in"),
    ("G", "Officer Email", 22, None, "in"),
    ("H", "Phone", 12, None, "in"),
    ("I", "Supervisor", 16, None, "in"),
    ("J", "Supervisor Email", 22, None, "in"),
    ("K", "Course / Event", 28, None, "in"),
    ("L", "Provider", 16, None, "in"),
    ("M", "Location", 16, None, "in"),
    ("N", "Class Start", 11, DATE, "in"),
    ("O", "Class End", 11, DATE, "in"),
    ("P", "Depart Date", 11, DATE, "in"),
    ("Q", "Depart Time", 10, TIME, "in"),
    ("R", "Return Date", 11, DATE, "in"),
    ("S", "Return Time", 10, TIME, "in"),
    ("T", "Registration Est", 12, CUR, "in"),
    ("U", "Nights", 7, "0", "in"),
    ("V", "Lodging Rate", 11, CUR, "in"),
    ("W", "Lodging Est", 11, CUR, "f"),
    ("X", "Travel Mode", 11, None, "in"),
    ("Y", "Miles (RT)", 9, "0", "in"),
    ("Z", "Airfare/Other Trv", 12, CUR, "in"),
    ("AA", "Travel Est", 11, CUR, "f"),
    ("AB", "Per Diem Est", 11, CUR, "f"),
    ("AC", "Other Est", 10, CUR, "in"),
    ("AD", "TOTAL EST", 12, CUR, "f"),
    ("AE", "Registration Act", 12, CUR, "in"),
    ("AF", "Lodging Act", 11, CUR, "in"),
    ("AG", "Travel Act", 11, CUR, "in"),
    ("AH", "Per Diem Act", 11, CUR, "in"),
    ("AI", "Other Act", 10, CUR, "in"),
    ("AJ", "TOTAL ACT", 12, CUR, "f"),
    ("AK", "Paid Direct by City", 12, CUR, "in"),
    ("AL", "Advance to Traveler", 12, CUR, "in"),
    ("AM", "Balance Due Traveler", 13, CUR, "f"),
    ("AN", "Submitted", 11, DATE, "in"),
    ("AO", "Approved", 11, DATE, "in"),
    ("AP", "Registered?", 9, None, "in"),
    ("AQ", "Hotel Booked?", 9, None, "in"),
    ("AR", "Receipts Due", 11, DATE, "f"),
    ("AS", "Receipts Rec'd", 11, DATE, "in"),
    ("AT", "Finance Packet Due (Tue noon)", 13, DATE, "f"),
    ("AU", "Sent to Finance", 11, DATE, "in"),
    ("AV", "Cal Pushed?", 9, None, "in"),
    ("AW", "Coordinator Notes", 24, None, "in"),
    ("AX", "hEst", 10, CUR, "h"),
    ("AY", "hSpent", 10, CUR, "h"),
    ("AZ", "hOpen", 10, CUR, "h"),
    ("BA", "qApproval", 8, "0", "h"),
    ("BB", "qRegister", 8, "0", "h"),
    ("BC", "qHotel", 8, "0", "h"),
    ("BD", "qReceipts", 8, "0", "h"),
    ("BE", "qFinance", 8, "0", "h"),
    ("BF", "TCOLE #", 9, None, "in"),
    ("BG", "Website", 18, None, "in"),
    ("BH", "Other Desc", 14, None, "in"),
    ("BI", "Justification", 30, None, "in"),
]

REQ_FORMULAS = {
    "W": '=IF($A{r}="","",N($U{r})*N($V{r}))',
    "AA": '=IF($A{r}="","",IF($X{r}="POV",N($Y{r})*MileageRate,0)+N($Z{r}))',
    "AB": ('=IF(OR($A{r}="",$P{r}="",$R{r}="",$Q{r}="",$S{r}=""),"",'
           'IF($R{r}=$P{r},'
           '($Q{r}<=T_DepBfast)*($S{r}>=T_RetBfast)*PD_Bfast'
           '+($Q{r}<=T_DepLunch)*($S{r}>=T_RetLunch)*PD_Lunch'
           '+($Q{r}<=T_DepDinner)*($S{r}>=T_RetDinner)*PD_Dinner'
           '+PD_Incid,'
           '($Q{r}<=T_DepBfast)*PD_Bfast+($Q{r}<=T_DepLunch)*PD_Lunch'
           '+($Q{r}<=T_DepDinner)*PD_Dinner'
           '+($S{r}>=T_RetBfast)*PD_Bfast+($S{r}>=T_RetLunch)*PD_Lunch'
           '+($S{r}>=T_RetDinner)*PD_Dinner'
           '+MAX($R{r}-$P{r}-1,0)*(PD_Bfast+PD_Lunch+PD_Dinner)'
           '+($R{r}-$P{r}+1)*PD_Incid))'),
    "AD": '=IF($A{r}="","",N($T{r})+N($W{r})+N($AA{r})+N($AB{r})+N($AC{r}))',
    "AJ": '=IF($A{r}="","",IF(COUNT($AE{r}:$AI{r})=0,"",SUM($AE{r}:$AI{r})))',
    "AM": '=IF(OR($A{r}="",$AJ{r}=""),"",$AJ{r}-N($AK{r})-N($AL{r}))',
    "AR": '=IF(OR($A{r}="",$R{r}=""),"",$R{r}+ReceiptDays)',
    # Tuesday on/before (departure - lead days); DATE(2024,1,2) was a Tuesday
    "AT": ('=IF(OR($A{r}="",$P{r}=""),"",'
           '($P{r}-FinanceLeadDays)-MOD(($P{r}-FinanceLeadDays)-DATE(2024,1,2),7))'),
    "AX": '=IF(OR($A{r}="",$B{r}="Denied",$B{r}="Cancelled"),0,N($AD{r}))',
    "AY": '=IF($A{r}="",0,N($AJ{r}))',
    "AZ": ('=IF(OR($A{r}="",$B{r}="Denied",$B{r}="Cancelled",$B{r}="Closed"),0,'
           'IF(N($AJ{r})>0,0,N($AD{r})))'),
    # Queue columns hold a running RANK (0 = not in queue, else 1,2,3… in row
    # order) so the Dashboard can list the k-th member with a plain MATCH.
    "BA": ('=IF($A{r}="",0,IF($B{r}="Awaiting Approval",'
           'COUNTIF(BA$1:BA{rm1},">0")+1,0))'),
    "BB": ('=IF($A{r}="",0,IF(AND($B{r}="Approved",$AP{r}<>"Y"),'
           'COUNTIF(BB$1:BB{rm1},">0")+1,0))'),
    "BC": ('=IF($A{r}="",0,IF(AND($B{r}="Approved",N($U{r})>0,$AQ{r}<>"Y"),'
           'COUNTIF(BC$1:BC{rm1},">0")+1,0))'),
    "BD": ('=IF($A{r}="",0,IF(AND($AR{r}<>"",N($AR{r})<TODAY(),$AS{r}="",'
           '$B{r}<>"Closed",$B{r}<>"Cancelled",$B{r}<>"Denied"),'
           'COUNTIF(BD$1:BD{rm1},">0")+1,0))'),
    "BE": ('=IF($A{r}="",0,IF(AND($AT{r}<>"",N($AT{r})<=TODAY()+7,$AU{r}="",'
           '$B{r}<>"Denied",$B{r}<>"Cancelled",$B{r}<>"Closed"),'
           'COUNTIF(BE$1:BE{rm1},">0")+1,0))'),
}

STATUSES = ["Awaiting Approval", "Approved", "Denied", "Travel Complete",
            "Closed", "Cancelled"]

# Fictional sample records — delete before go-live (see Instructions sheet).
d = datetime.date
t = datetime.time
SAMPLES = [
    {  # fully reconciled and closed
        "A": "T26-001", "B": "Closed", "C": "SAMPLE — J. Doe", "D": "Officer",
        "E": "4101", "F": "Patrol", "G": "jdoe@example.com", "H": "903-555-0101",
        "I": "Sgt. R. Lee", "J": "rlee@example.com",
        "K": "Interview & Interrogation", "L": "TMPA", "M": "Plano, TX",
        "N": d(2026, 3, 9), "O": d(2026, 3, 11), "P": d(2026, 3, 8),
        "Q": t(15, 0), "R": d(2026, 3, 11), "S": t(18, 30),
        "T": 350.0, "U": 3, "V": 119.0, "X": "City Vehicle", "Y": 0, "Z": 0.0,
        "AC": 0.0, "AE": 350.0, "AF": 371.07, "AG": 0.0, "AH": 168.0, "AI": 0.0,
        "AK": 721.07, "AL": 168.0, "AN": d(2026, 1, 12), "AO": d(2026, 1, 15),
        "AP": "Y", "AQ": "Y", "AS": d(2026, 3, 16), "AU": d(2026, 3, 17),
        "AV": "Y", "BI": "TCOLE intermediate certificate requirement.",
    },
    {  # trip done, receipts overdue -> dashboard queue 4
        "A": "T26-002", "B": "Travel Complete", "C": "SAMPLE — M. Rivera",
        "D": "Detective", "E": "3877", "F": "CID", "G": "mrivera@example.com",
        "H": "903-555-0102", "I": "Lt. K. Ford", "J": "kford@example.com",
        "K": "Homicide Investigation", "L": "ILEETA", "M": "Dallas, TX",
        "N": d(2026, 7, 27), "O": d(2026, 7, 28), "P": d(2026, 7, 27),
        "Q": t(6, 30), "R": d(2026, 7, 28), "S": t(19, 0),
        "T": 425.0, "U": 1, "V": 139.0, "X": "POV", "Y": 190, "Z": 0.0,
        "AC": 0.0, "AN": d(2026, 6, 1), "AO": d(2026, 6, 3),
        "AP": "Y", "AQ": "Y", "AV": "Y",
        "BI": "Case-load requires homicide certification.",
    },
    {  # approved, still needs registration + hotel; finance packet due soon
        "A": "T26-003", "B": "Approved", "C": "SAMPLE — A. Chen", "D": "Corporal",
        "E": "4215", "F": "Traffic", "G": "achen@example.com", "H": "903-555-0103",
        "I": "Sgt. R. Lee", "J": "rlee@example.com",
        "K": "Advanced Crash Reconstruction", "L": "TEEX", "M": "College Station, TX",
        "N": d(2026, 8, 24), "O": d(2026, 8, 28), "P": d(2026, 8, 23),
        "Q": t(13, 0), "R": d(2026, 8, 28), "S": t(18, 30),
        "T": 450.0, "U": 5, "V": 129.0, "X": "City Vehicle", "Y": 0, "Z": 0.0,
        "AC": 35.0, "AN": d(2026, 7, 20), "AO": d(2026, 7, 24),
        "BI": "Required for accident-investigator certification.",
    },
    {  # new request awaiting chain-of-command approval
        "A": "T26-004", "B": "Awaiting Approval", "C": "SAMPLE — T. Okafor",
        "D": "Officer", "E": "4302", "F": "Patrol", "G": "tokafor@example.com",
        "H": "903-555-0104", "I": "Sgt. B. Hall", "J": "bhall@example.com",
        "K": "Crisis Intervention Training", "L": "TCOLE Reg. Academy",
        "M": "Longview, TX",
        "N": d(2026, 9, 15), "O": d(2026, 9, 17), "P": d(2026, 9, 15),
        "Q": t(6, 0), "R": d(2026, 9, 17), "S": t(17, 30),
        "T": 0.0, "U": 2, "V": 109.0, "X": "City Vehicle", "Y": 0, "Z": 0.0,
        "AC": 0.0, "AN": d(2026, 8, 6),
        "BI": "40-hr CIT — department mental-health response initiative.",
    },
]


def pull(col, blank='""'):
    """Formula: field <col> of the record picked by the Dashboard selector."""
    return (f'=IF(ActiveRow=0,{blank},'
            f'INDEX(Requests!${col}${R1}:${col}${R2},ActiveRow))')


def rng(col):
    return f"Requests!${col}${R1}:${col}${R2}"


# =================================================================== sheets
def sheet_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 2, "B": 30, "C": 34, "D": 78})
    paint(ws, 1, 78, 1, 4)
    title_bar(ws, 1, 1, 4, "SETTINGS — everything a coordinator may ever need to change",
              "Change VALUES in column C only. Yellow cells MUST be set before go-live. "
              "Never rename this sheet or move cells — formulas and macros read them by name.")

    row = [4]  # mutable cursor

    def sec(text):
        section_bar(ws, row[0], 2, 4, text)
        row[0] += 1

    def item(labeltext, value, name, note="", fmt=None, must_set=False, tall=False):
        r = row[0]
        label(ws, r, 2, labeltext)
        c = ws.cell(row=r, column=3, value=value)
        c.font = F_INPUT
        c.border = BOX
        c.fill = FILL_YELLOW if must_set else FILL_INPUT
        c.alignment = A_LEFT_WRAP if tall else A_LEFT
        if fmt:
            c.number_format = fmt
        n = ws.cell(row=r, column=4, value=note)
        n.font = F_NOTE
        n.alignment = A_LEFT_WRAP
        if tall:
            ws.row_dimensions[r].height = 88
        define(wb, name, "Settings", f"$C${r}")
        row[0] += 1

    sec("CONTACTS")
    item("Training Unit mailbox", "training@tylertexas.com", "TrainingEmail",
         "PLACEHOLDER — set to the real Training Unit email.", must_set=True)
    item("Coordinator name", "Training Coordinator", "CoordName",
         "Appears on notifications and emails.", must_set=True)
    item("Coordinator phone", "903-555-0100", "CoordPhone", "", must_set=True)
    item("Finance office email", "finance@tylertexas.com", "FinanceEmail",
         "PLACEHOLDER — where completed 2-90 packets go.", must_set=True)

    sec("PER DIEM — MEAL RATES  (set to current City of Tyler policy)")
    item("Breakfast", 13.00, "PD_Bfast", "Placeholder = GSA standard M&IE split. "
         "Verify against the current city travel policy.", fmt=CUR, must_set=True)
    item("Lunch", 15.00, "PD_Lunch", "", fmt=CUR, must_set=True)
    item("Dinner", 26.00, "PD_Dinner", "", fmt=CUR, must_set=True)
    item("Incidentals (per day)", 5.00, "PD_Incid",
         "Paid for every travel day, first through last. Set 0 if the city pays none.",
         fmt=CUR, must_set=True)

    sec("PER DIEM — ELIGIBILITY CUTOFFS  (which meals a travel day earns)")
    item("Departure day: breakfast if leaving at/before", t(6, 0), "T_DepBfast",
         "First-day meal is reimbursable when departure is at/before this time.", fmt=TIME)
    item("Departure day: lunch if leaving at/before", t(12, 0), "T_DepLunch", "", fmt=TIME)
    item("Departure day: dinner if leaving at/before", t(18, 0), "T_DepDinner", "", fmt=TIME)
    item("Return day: breakfast if back at/after", t(8, 0), "T_RetBfast",
         "Last-day meal is reimbursable when return is at/after this time.", fmt=TIME)
    item("Return day: lunch if back at/after", t(13, 0), "T_RetLunch", "", fmt=TIME)
    item("Return day: dinner if back at/after", t(19, 0), "T_RetDinner", "", fmt=TIME)

    sec("RATES & DEADLINES")
    item("Mileage rate (POV, per mile)", 0.70, "MileageRate",
         "PLACEHOLDER — set to the current IRS/city rate.", fmt="$0.000", must_set=True)
    item("Receipts due (days after return)", 7, "ReceiptDays",
         "Drives the Receipts Due column and the overdue queue.", fmt="0")
    item("Finance lead time (days before departure)", 7, "FinanceLeadDays",
         "Packet deadline = the TUESDAY on/before (departure − this many days). "
         "Packets reaching Finance by Tuesday NOON get a check cut that week.", fmt="0")

    sec("FISCAL YEAR & BUDGET")
    item("Fiscal year label", "FY26", "FY_Label", "City FY runs Oct 1 – Sep 30.")
    item("Fiscal year start", d(2025, 10, 1), "FY_Start",
         "First day of the FY this workbook covers. Drives the YTD report months.",
         fmt=DATE)
    item("Tracking # prefix", "T26-", "TrkPrefix",
         "Import assigns T26-001, T26-002, … Change at rollover (see Instructions).")
    item("Training budget (this FY)", 75000, "Budget_Total",
         "PLACEHOLDER — set to the adopted budget line.", fmt=CUR, must_set=True)

    sec("FOLDERS  (full paths; used by the macros)")
    item("Incoming request forms", r"H:\Training\Requests\Inbox", "ImportFolder",
         "The Import button reads every officer request form saved here.", must_set=True)
    item("Processed request forms", r"H:\Training\Requests\Processed", "ProcessedFolder",
         "Imported forms are moved here, renamed to their tracking number.", must_set=True)
    item("Printed packets (PDF)", r"H:\Training\Packets", "PacketFolder",
         "Print Packet saves the PDF here before opening it.", must_set=True)
    item("Year-end archives", r"H:\Training\Archive", "ArchiveFolder",
         "Archive Year saves the closing FY's copy here.", must_set=True)

    sec("EMAIL TEMPLATES  — placeholders in {braces} are filled automatically")
    item("Approval request — subject",
         "Training approval needed: {OFFICER} — {COURSE} ({TRK})", "ET_ApprSubj",
         "Placeholders: {TRK} {OFFICER} {RANK} {COURSE} {PROVIDER} {LOCATION} "
         "{DATES} {DEPART} {RETURN} {TOTAL} {RECEIPTS_DUE} {BALANCE} {COORD} {PHONE}")
    item("Approval request — body",
         "{SUPERVISOR},\n\nPlease review the attached training application for "
         "{RANK} {OFFICER}:\n\n    Course:  {COURSE} — {PROVIDER}\n    Where:   "
         "{LOCATION}\n    Dates:   {DATES}\n    Estimated cost:  {TOTAL}\n\n"
         "Reply to this email with your approval or denial.\n\n{COORD}\n"
         "Training Unit — {PHONE}", "ET_ApprBody", tall=True)
    item("Officer notification — subject",
         "Your training is booked: {COURSE} ({TRK})", "ET_NotifSubj")
    item("Officer notification — body",
         "{RANK} {OFFICER},\n\nYou are confirmed for {COURSE} in {LOCATION}, "
         "{DATES}.\nDeparture {DEPART} — return {RETURN}.\n\nYour notification "
         "sheet is attached; it lists what is prepaid and your per diem.\n"
         "Keep EVERY itemized receipt — they are due to the Training Unit by "
         "{RECEIPTS_DUE}.\n\n{COORD}\nTraining Unit — {PHONE}", "ET_NotifBody",
         tall=True)
    item("Post-trip settlement — subject",
         "Training trip {TRK} — expense settlement", "ET_ReimbSubj")
    item("Post-trip settlement — body",
         "{RANK} {OFFICER},\n\nYour expense reconciliation for {COURSE} is "
         "complete.\nBalance: {BALANCE}\n(A positive balance is owed to you; a "
         "negative balance is owed back to the city.)\n\nContact me with any "
         "questions before this goes to Finance.\n\n{COORD}\nTraining Unit — "
         "{PHONE}", "ET_ReimbBody", tall=True)
    return ws


def sheet_lists(wb):
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Statuses"
    ws["B1"] = "Travel Modes"
    ws["C1"] = "Y/N"
    for i, s in enumerate(STATUSES, start=2):
        ws.cell(row=i, column=1, value=s)
    for i, s in enumerate(["City Vehicle", "POV", "Air", "Other"], start=2):
        ws.cell(row=i, column=2, value=s)
    for i, s in enumerate(["Y", "N"], start=2):
        ws.cell(row=i, column=3, value=s)
    for c in ("A1", "B1", "C1"):
        ws[c].font = F_LABEL
    define(wb, "StatusList", "Lists", f"$A$2:$A${len(STATUSES) + 1}")
    define(wb, "ModeList", "Lists", "$B$2:$B$5")
    define(wb, "YNList", "Lists", "$C$2:$C$3")
    ws.sheet_state = "hidden"
    return ws


def sheet_requests(wb):
    ws = wb.create_sheet("Requests")
    ws.freeze_panes = "C2"
    for letter, header, width, fmt, kind in REQ_COLS:
        ws.column_dimensions[letter].width = width
        cidx = ws[f"{letter}1"]
        cidx.value = header
        cidx.font = F_HDR
        cidx.fill = FILL_STEEL if kind == "h" else FILL_NAVY
        cidx.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    for r in range(R1, R2 + 1):
        band = (r % 2 == 0)
        for letter, header, width, fmt, kind in REQ_COLS:
            cell = ws[f"{letter}{r}"]
            cell.font = F_CELL
            if fmt:
                cell.number_format = fmt
            if band:
                cell.fill = FILL_BAND
            if kind in ("f", "h"):
                cell.value = REQ_FORMULAS[letter].format(r=r, rm1=r - 1)

    # sample records (values into input columns only)
    for i, rec in enumerate(SAMPLES):
        r = R1 + i
        for letter, value in rec.items():
            ws[f"{letter}{r}"] = value

    # dropdowns
    for name, cols in (("=StatusList", ["B"]), ("=ModeList", ["X"]),
                       ("=YNList", ["AP", "AQ", "AV"])):
        dv = DataValidation(type="list", formula1=name, allow_blank=True)
        for cl in cols:
            dv.add(f"{cl}{R1}:{cl}{R2}")
        ws.add_data_validation(dv)

    define(wb, "TrkList", "Requests", f"$A${R1}:$A${R2}")
    ws.auto_filter.ref = f"A1:BI{R2}"
    return ws


def sheet_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 2, "B": 11, "C": 20, "D": 30, "E": 13, "F": 3,
                    "G": 11, "H": 20, "I": 30, "J": 13, "K": 3,
                    "L": 11, "M": 20, "N": 30, "O": 13, "P": 2})
    for hcol in ("R", "S", "T", "U", "V"):
        ws.column_dimensions[hcol].width = 6
        ws.column_dimensions[hcol].hidden = True
    paint(ws, 1, 46, 1, 16)

    title_bar(ws, 1, 1, 16, "TPD TRAINING UNIT — COMMAND DASHBOARD",
              "What needs doing today. Lists refresh themselves; buttons drive the "
              "paperwork for the record picked in ACTIVE RECORD.")
    for c in range(1, 17):
        ws.cell(row=3, column=c).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3

    # --- active record selector + today
    label(ws, 4, 2, "ACTIVE ▶")
    sel = ws.cell(row=4, column=3, value="T26-003")  # sample default; user picks
    sel.fill = FILL_YELLOW
    sel.border = BOX
    sel.font = Font(name=FONT, size=12, bold=True, color="0000FF")
    sel.alignment = A_CENTER
    define(wb, "ActiveTrk", "Dashboard", "$C$4")
    dv = DataValidation(type="list", formula1="=TrkList", allow_blank=True)
    dv.add("C4")
    ws.add_data_validation(dv)
    c = ws.cell(row=4, column=4,
                value='=IF(ActiveRow=0,"← pick a tracking # — all forms & emails use it",'
                      f'INDEX({rng("C")},ActiveRow)&"   •   "&INDEX({rng("K")},ActiveRow))')
    c.font = F_NOTE
    ws.cell(row=4, column=14, value="Today:").font = F_LABEL
    ws.cell(row=4, column=14).alignment = A_RIGHT
    td = ws.cell(row=4, column=15, value="=TODAY()")
    td.number_format = DATE
    td.font = F_LABEL
    # hidden helper: row of the active record
    hr = ws.cell(row=4, column=18,
                 value=f'=IFERROR(MATCH(ActiveTrk,{rng("A")},0),0)')
    define(wb, "ActiveRow", "Dashboard", "$R$4")

    # --- KPI tiles (row 6-8)
    tiles = [
        ("BUDGET (FY)", "=Budget_Total", CUR0, None),
        ("SPENT", f'=SUM({rng("AY")})', CUR0, None),
        ("ENCUMBERED (open est.)", f'=SUM({rng("AZ")})', CUR0, None),
        ("REMAINING", f'=Budget_Total-SUM({rng("AY")})-SUM({rng("AZ")})', CUR0, None),
        ("REQUESTS (FY)", f'=SUMPRODUCT(({rng("A")}<>"")*1)', "0", None),
        ("DEPARTING ≤ 30 DAYS",
         f'=SUMPRODUCT(({rng("P")}>=TODAY())*({rng("P")}<=TODAY()+30)*'
         f'({rng("A")}<>"")*({rng("B")}<>"Denied")*({rng("B")}<>"Cancelled"))',
         "0", None),
    ]
    start_cols = [(2, 3), (4, 5), (7, 8), (9, 10), (12, 13), (14, 15)]
    for (c1, c2), (name, formula, fmt, _) in zip(start_cols, tiles):
        ws.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
        h = ws.cell(row=6, column=c1, value=name)
        h.font = Font(name=FONT, size=8, bold=True, color="FFFFFF")
        h.alignment = A_CENTER
        ws.merge_cells(start_row=7, start_column=c1, end_row=7, end_column=c2)
        v = ws.cell(row=7, column=c1, value=formula)
        v.font = F_KPI
        v.alignment = A_CENTER
        v.number_format = fmt
        for cc in range(c1, c2 + 1):
            ws.cell(row=6, column=cc).fill = FILL_STEEL
            ws.cell(row=7, column=cc).fill = FILL_BAND
    ws.row_dimensions[7].height = 22

    # --- queues: 5 blocks. layout: three across top (rows 10-19),
    #     two across bottom (rows 22-31). Each block: header row, col heads, 8 rows.
    QSIZE = 8
    queues = [
        ("1 — AWAITING APPROVAL", "BA", [("Trk", "A"), ("Officer", "C"),
         ("Course", "K"), ("Submitted", "AN")], 10, 2, "S"),
        ("2 — NEEDS REGISTRATION", "BB", [("Trk", "A"), ("Officer", "C"),
         ("Course", "K"), ("Departs", "P")], 10, 7, "T"),
        ("3 — NEEDS HOTEL", "BC", [("Trk", "A"), ("Officer", "C"),
         ("Course", "K"), ("Departs", "P")], 10, 12, "U"),
        ("4 — RECEIPTS OVERDUE", "BD", [("Trk", "A"), ("Officer", "C"),
         ("Course", "K"), ("Was Due", "AR")], 22, 2, "V"),
        ("5 — FINANCE PACKET DUE ≤ 7 DAYS  (Tuesday NOON)", "BE",
         [("Trk", "A"), ("Officer", "C"), ("Course", "K"), ("Due (Tue)", "AT")],
         22, 7, "R"),
    ]
    date_heads = {"Submitted", "Departs", "Was Due", "Due (Tue)"}
    for title, flag, cols, r0, c0, hcol in queues:
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 3)
        h = ws.cell(row=r0, column=c0,
                    value=f'="{title}   ("&COUNTIF({rng(flag)},">0")&")"')
        h.font = F_SECTION
        h.alignment = A_LEFT
        for cc in range(c0, c0 + 4):
            ws.cell(row=r0, column=cc).fill = FILL_NAVY
        for j, (head, _) in enumerate(cols):
            hc = ws.cell(row=r0 + 1, column=c0 + j, value=head)
            hc.font = Font(name=FONT, size=8, bold=True, color=NAVY)
            hc.border = BOX
        for k in range(1, QSIZE + 1):
            r = r0 + 1 + k
            # hidden helper: position (within the data range) of the k-th queue member
            ws.cell(row=r, column=col_of(hcol),
                    value=f'=IFERROR(MATCH({k},{rng(flag)},0),"")')
            for j, (head, reqcol) in enumerate(cols):
                cell = ws.cell(
                    row=r, column=c0 + j,
                    value=f'=IF(${hcol}{r}="","",'
                          f'INDEX({rng(reqcol)},${hcol}{r}))')
                cell.font = F_CELL
                cell.border = BOX
                if head in date_heads:
                    cell.number_format = DATE
                if flag == "BD" and head == "Was Due":
                    cell.font = Font(name=FONT, size=9, bold=True, color=WARN)

    # --- button parking strip + reminder text
    section_bar(ws, 33, 2, 15,
                "BUTTONS (added by the one-time installer): Import Requests • Draft Approval Email • "
                "Draft Officer Notification • Draft Settlement Email • Print Packet (PDF) • "
                "Push Deadlines to Outlook • Archive Year")
    ws.merge_cells("B35:O35")
    c = ws.cell(row=35, column=2,
                value="No buttons? Press Alt+F8 and run the same macros by name. Emails always "
                      "OPEN FOR REVIEW — nothing sends itself. Calendar push only adds items "
                      "not already pushed (Cal Pushed? column).")
    c.font = F_NOTE
    ws.row_dimensions[33].height = 28
    ws.print_area = "A1:P35"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def col_of(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def doc_header(ws, last_col, title, subtitle):
    title_bar(ws, 1, 1, last_col, title, subtitle)
    for c in range(1, last_col + 1):
        ws.cell(row=3, column=c).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3


def field(ws, r, c, text, formula, fmt=None, span=1, bold=False):
    label(ws, r, c, text)
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c + 1, end_row=r, end_column=c + span)
    cell = ws.cell(row=r, column=c + 1, value=formula)
    cell.font = Font(name=FONT, size=10, bold=bold)
    cell.alignment = A_LEFT
    for i in range(span):
        ws.cell(row=r, column=c + 1 + i).border = BOX
        ws.cell(row=r, column=c + 1 + i).fill = FILL_INPUT
    if fmt:
        cell.number_format = fmt
    return cell


def sig_line(ws, r, c, who, span=2):
    from openpyxl.styles import Border, Side
    under = Border(bottom=Side(style="thin", color="000000"))
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    for cc in range(c, c + span):
        ws.cell(row=r, column=cc).border = under
    w = ws.cell(row=r + 1, column=c, value=who + "  /  Date")
    w.font = F_SMALL


def sheet_application(wb):
    ws = wb.create_sheet("Application")
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 2, "B": 20, "C": 22, "D": 20, "E": 22, "F": 2})
    paint(ws, 1, 46, 1, 6)
    doc_header(ws, 6, "TYLER POLICE DEPARTMENT — TRAINING / TRAVEL APPLICATION",
               "Internal chain-of-command approval. Auto-filled from the Dashboard's "
               "ACTIVE RECORD — never type on this sheet.")

    field(ws, 5, 2, "Tracking #", pull("A"), bold=True)
    field(ws, 5, 4, "Date Submitted", pull("AN"), fmt=DATE)
    section_bar(ws, 7, 2, 5, "APPLICANT")
    field(ws, 8, 2, "Name", pull("C"))
    field(ws, 8, 4, "Rank", pull("D"))
    field(ws, 9, 2, "Employee ID", pull("E"))
    field(ws, 9, 4, "Division", pull("F"))
    field(ws, 10, 2, "Supervisor", pull("I"))
    field(ws, 10, 4, "Phone", pull("H"))

    section_bar(ws, 12, 2, 5, "COURSE / EVENT")
    field(ws, 13, 2, "Course", pull("K"), span=3)
    field(ws, 14, 2, "Provider / Host", pull("L"))
    field(ws, 14, 4, "TCOLE #", pull("BF"))
    field(ws, 15, 2, "Location", pull("M"))
    field(ws, 15, 4, "Class Dates", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("N")},ActiveRow),"mm/dd/yy")&" – "&TEXT(INDEX({rng("O")},ActiveRow),"mm/dd/yy"))')
    field(ws, 16, 2, "Departure", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("P")},ActiveRow),"mm/dd/yy")&" "&TEXT(INDEX({rng("Q")},ActiveRow),"h:mm AM/PM"))')
    field(ws, 16, 4, "Return", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("R")},ActiveRow),"mm/dd/yy")&" "&TEXT(INDEX({rng("S")},ActiveRow),"h:mm AM/PM"))')

    section_bar(ws, 18, 2, 5, "ESTIMATED COST  (calculated under city travel policy)")
    est_rows = [("Registration", "T"), ("Lodging", "W"), ("Travel / Mileage", "AA"),
                ("Per Diem (M&IE)", "AB"), ("Other", "AC")]
    r = 19
    for text, cl in est_rows:
        field(ws, r, 2, text, pull(cl, blank='""'), fmt=CUR)
        r += 1
    tot = field(ws, r, 2, "TOTAL ESTIMATE", pull("AD"), fmt=CUR, bold=True)
    tot.fill = FILL_BAND

    section_bar(ws, 26, 2, 5, "JUSTIFICATION")
    ws.merge_cells("B27:E29")
    j = ws.cell(row=27, column=2, value=pull("BI"))
    j.alignment = A_LEFT_WRAP
    j.font = F_BODY
    for rr in range(27, 30):
        for cc in range(2, 6):
            ws.cell(row=rr, column=cc).border = BOX
            ws.cell(row=rr, column=cc).fill = FILL_INPUT

    section_bar(ws, 31, 2, 5, "APPROVALS")
    sig_line(ws, 33, 2, "Immediate Supervisor", span=2)
    sig_line(ws, 33, 4, "Division Commander", span=2)
    sig_line(ws, 36, 2, "Training Coordinator", span=2)
    sig_line(ws, 36, 4, "Chief of Police / Designee", span=2)

    ws.merge_cells("B39:E39")
    f = ws.cell(row=39, column=2,
                value='="Generated by TPD Training Master — record "&IF(ActiveRow=0,"(none selected)",ActiveTrk)&" — "&TEXT(TODAY(),"mm/dd/yyyy")')
    f.font = F_SMALL
    ws.print_area = "A1:F40"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.protection.sheet = True
    return ws


def form_290(wb, name, post):
    """Pre/Post 2-90 recreation. post=False -> estimate; True -> reconciliation."""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 2, "B": 24, "C": 16, "D": 16, "E": 16, "F": 16, "G": 2})
    paint(ws, 1, 48, 1, 7)
    kind = "POST-TRIP RECONCILIATION" if post else "PRE-TRIP ESTIMATE & AUTHORIZATION"
    doc_header(ws, 7, f"CITY OF TYLER — TRAVEL EXPENSE FORM 2-90 • {kind}",
               "Functional recreation of the Finance form — auto-filled from the ACTIVE "
               "RECORD. If Finance issues a new 2-90, replace this sheet per docs/.")

    field(ws, 5, 2, "Department", '="Police"')
    field(ws, 5, 5, "Tracking #", pull("A"), bold=True)
    field(ws, 6, 2, "Employee", pull("C"))
    field(ws, 6, 5, "Title / Rank", pull("D"))
    field(ws, 7, 2, "Destination", pull("M"))
    field(ws, 7, 5, "Employee ID", pull("E"))
    field(ws, 8, 2, "Purpose (course/event)", pull("K"), span=4)
    field(ws, 9, 2, "Departure", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("P")},ActiveRow),"mm/dd/yy")&" "&TEXT(INDEX({rng("Q")},ActiveRow),"h:mm AM/PM"))')
    field(ws, 9, 5, "Return", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("R")},ActiveRow),"mm/dd/yy")&" "&TEXT(INDEX({rng("S")},ActiveRow),"h:mm AM/PM"))')

    section_bar(ws, 11, 2, 6, "EXPENSES" + ("  (estimate vs. actual)" if post else "  (estimated)"))
    heads = ["Category", "Account Code", "Estimated"] + (["Actual", "Variance"] if post else [])
    for j, h in enumerate(heads):
        c = ws.cell(row=12, column=2 + j, value=h)
        c.font = F_LABEL
        c.border = BOX
        c.fill = FILL_BAND
    cats = [("Registration / Tuition", "T", "AE"), ("Lodging", "W", "AF"),
            ("Meals & Incidentals (per diem)", "AB", "AH"),
            ("Transportation", "AA", "AG"), ("Other", "AC", "AI")]
    r = 13
    for text, est_c, act_c in cats:
        ws.cell(row=r, column=2, value=text).font = F_BODY
        ws.cell(row=r, column=2).border = BOX
        acct = ws.cell(row=r, column=3)   # account code — typed by coordinator
        acct.border = BOX
        acct.fill = FILL_YELLOW
        acct.font = F_INPUT
        acct.protection = Protection(locked=False)
        e = ws.cell(row=r, column=4, value=pull(est_c))
        e.number_format = CUR
        e.border = BOX
        if post:
            a = ws.cell(row=r, column=5, value=pull(act_c))
            a.number_format = CUR
            a.border = BOX
            v = ws.cell(row=r, column=6,
                        value=f'=IF(OR($D{r}="",$E{r}=""),"",$D{r}-$E{r})')
            v.number_format = CUR
            v.border = BOX
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = F_LABEL
    ws.cell(row=r, column=2).border = BOX
    ws.cell(row=r, column=3).border = BOX
    te = ws.cell(row=r, column=4, value=f'=IF(ActiveRow=0,"",SUM(D13:D{r-1}))')
    te.number_format = CUR
    te.font = F_LABEL
    te.border = BOX
    te.fill = FILL_BAND
    if post:
        ta = ws.cell(row=r, column=5, value=f'=IF(ActiveRow=0,"",SUM(E13:E{r-1}))')
        ta.number_format = CUR
        ta.font = F_LABEL
        ta.border = BOX
        ta.fill = FILL_BAND
        ws.cell(row=r, column=6).border = BOX

    r += 2
    if post:
        field(ws, r, 2, "Total actual expenses", pull("AJ"), fmt=CUR)
        r += 1
        field(ws, r, 2, "Less: paid directly by city", pull("AK", blank='""'), fmt=CUR)
        r += 1
        field(ws, r, 2, "Less: advance to employee", pull("AL", blank='""'), fmt=CUR)
        r += 1
        bal = field(ws, r, 2, "BALANCE DUE EMPLOYEE / (CITY)", pull("AM"), fmt=CUR, bold=True)
        bal.fill = FILL_BAND
        r += 2
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cert = ws.cell(row=r, column=2,
                       value="I certify these expenses were incurred on official city "
                             "business and itemized receipts are attached.")
        cert.font = F_NOTE
        r += 2
    else:
        label(ws, r, 2, "Advance requested (if any)")
        adv = ws.cell(row=r, column=4)
        adv.border = BOX
        adv.fill = FILL_YELLOW
        adv.font = F_INPUT
        adv.number_format = CUR
        adv.protection = Protection(locked=False)
        r += 2

    sig_line(ws, r + 1, 2, "Employee", span=2)
    sig_line(ws, r + 1, 5, "Department Head / Designee", span=2)
    sig_line(ws, r + 4, 2, "Finance", span=2)
    ws.print_area = f"A1:G{r + 6}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.protection.sheet = True
    return ws


def sheet_notification(wb):
    ws = wb.create_sheet("Notification")
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 2, "B": 22, "C": 24, "D": 20, "E": 24, "F": 2})
    paint(ws, 1, 44, 1, 6)
    doc_header(ws, 6, "TPD TRAINING UNIT — ATTENDEE TRAVEL NOTIFICATION",
               "Give this sheet to the traveler. Auto-filled from the ACTIVE RECORD.")

    field(ws, 5, 2, "To", pull("C"), bold=True)
    field(ws, 5, 4, "Tracking #", pull("A"), bold=True)
    field(ws, 6, 2, "From", "=CoordName")
    field(ws, 6, 4, "Date", "=TODAY()", fmt=DATE)

    section_bar(ws, 8, 2, 5, "YOUR TRIP")
    field(ws, 9, 2, "Course", pull("K"), span=3)
    field(ws, 10, 2, "Provider", pull("L"))
    field(ws, 10, 4, "Location", pull("M"))
    field(ws, 11, 2, "Class Dates", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("N")},ActiveRow),"mm/dd/yy")&" – "&TEXT(INDEX({rng("O")},ActiveRow),"mm/dd/yy"))')
    field(ws, 11, 4, "Travel Mode", pull("X"))
    field(ws, 12, 2, "Depart", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("P")},ActiveRow),"dddd mm/dd")&" at "&TEXT(INDEX({rng("Q")},ActiveRow),"h:mm AM/PM"))')
    field(ws, 12, 4, "Return", f'=IF(ActiveRow=0,"",TEXT(INDEX({rng("R")},ActiveRow),"dddd mm/dd")&" at "&TEXT(INDEX({rng("S")},ActiveRow),"h:mm AM/PM"))')

    section_bar(ws, 14, 2, 5, "WHAT IS ARRANGED FOR YOU")
    field(ws, 15, 2, "Registration",
          f'=IF(ActiveRow=0,"",IF(INDEX({rng("AP")},ActiveRow)="Y","PAID — "&TEXT(INDEX({rng("T")},ActiveRow),"$#,##0.00"),"Pending"))')
    field(ws, 15, 4, "Hotel",
          f'=IF(ActiveRow=0,"",IF(N(INDEX({rng("U")},ActiveRow))=0,"None needed",IF(INDEX({rng("AQ")},ActiveRow)="Y","BOOKED — "&INDEX({rng("U")},ActiveRow)&" nights","Pending")))')

    section_bar(ws, 17, 2, 5, "YOUR MONEY")
    field(ws, 18, 2, "Per Diem (M&IE)", pull("AB"), fmt=CUR, bold=True)
    field(ws, 18, 4, "Mileage / travel", pull("AA"), fmt=CUR)
    ws.merge_cells("B20:E22")
    msg = ws.cell(
        row=20, column=2,
        value='=IF(ActiveRow=0,"","Per diem is calculated from your departure and '
              'return times under city travel policy. KEEP EVERY ITEMIZED RECEIPT — '
              'turn them in to the Training Unit by "'
              f'&TEXT(INDEX({rng("AR")},ActiveRow),"dddd, mmmm d")&'
              '". Questions: "&CoordName&", "&CoordPhone&".")')
    msg.alignment = A_LEFT_WRAP
    msg.font = F_BODY
    for rr in range(20, 23):
        for cc in range(2, 6):
            ws.cell(row=rr, column=cc).border = BOX
            ws.cell(row=rr, column=cc).fill = FILL_INPUT
    ws.print_area = "A1:F24"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def sheet_ledger(wb):
    ws = wb.create_sheet("Ledger")
    ws.freeze_panes = "A2"
    headers = [
        ("A", "Tracking #", 11, None),
        ("B", "Month", 10, None),
        ("C", "Officer", 20, None),
        ("D", "Course / Event", 30, None),
        ("E", "Status", 15, None),
        ("F", "Estimated", 12, CUR),
        ("G", "Actual", 12, CUR),
        ("H", "Variance (E−A)", 12, CUR),
        ("I", "Balance Due Traveler", 14, CUR),
        ("J", "Receipts Rec'd", 12, DATE),
        ("K", "Sent to Finance", 12, DATE),
        ("L", "Reconciled?", 10, None),
    ]
    for letter, h, w, fmt in headers:
        ws.column_dimensions[letter].width = w
        c = ws[f"{letter}1"]
        c.value = h
        c.font = F_HDR
        c.fill = FILL_NAVY
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    forms = {
        "A": '=IF(Requests!$A{rr}="","",Requests!$A{rr})',
        "B": '=IF(OR(Requests!$A{rr}="",Requests!$P{rr}=""),"",TEXT(Requests!$P{rr},"yyyy-mm"))',
        "C": '=IF(Requests!$A{rr}="","",Requests!$C{rr})',
        "D": '=IF(Requests!$A{rr}="","",Requests!$K{rr})',
        "E": '=IF(Requests!$A{rr}="","",Requests!$B{rr})',
        "F": '=IF(Requests!$A{rr}="","",Requests!$AD{rr})',
        "G": '=IF(Requests!$A{rr}="","",Requests!$AJ{rr})',
        "H": '=IF(OR(Requests!$A{rr}="",Requests!$AJ{rr}=""),"",Requests!$AD{rr}-Requests!$AJ{rr})',
        "I": '=IF(Requests!$A{rr}="","",Requests!$AM{rr})',
        "J": '=IF(OR(Requests!$A{rr}="",Requests!$AS{rr}=""),"",Requests!$AS{rr})',
        "K": '=IF(OR(Requests!$A{rr}="",Requests!$AU{rr}=""),"",Requests!$AU{rr})',
        "L": ('=IF(Requests!$A{rr}="","",IF(AND(Requests!$AJ{rr}<>"",'
              'Requests!$AS{rr}<>"",Requests!$AU{rr}<>""),"Yes","Open"))'),
    }
    for r in range(R1, R2 + 1):
        band = (r % 2 == 0)
        for letter, h, w, fmt in headers:
            cell = ws[f"{letter}{r}"]
            cell.value = forms[letter].format(rr=r)
            cell.font = F_CELL
            if fmt:
                cell.number_format = fmt
            if band:
                cell.fill = FILL_BAND
    # totals row
    tr = R2 + 2
    ws.cell(row=tr, column=4, value="FISCAL YEAR TOTALS").font = F_LABEL
    for cl in ("F", "G", "H", "I"):
        c = ws[f"{cl}{tr}"]
        c.value = f"=SUM({cl}{R1}:{cl}{R2})"
        c.number_format = CUR
        c.font = F_LABEL
        c.fill = FILL_BAND
    ws.auto_filter.ref = f"A1:L{R2}"
    return ws


def sheet_ytd(wb):
    ws = wb.create_sheet("YTD Report")
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 11, "B": 12, "C": 12, "D": 14, "E": 14, "F": 14, "G": 3})
    paint(ws, 1, 46, 1, 7)
    title_bar(ws, 1, 1, 7, "TRAINING BUDGET — YEAR-TO-DATE REPORT",
              '="Fiscal year "&FY_Label&" (from "&TEXT(FY_Start,"mm/dd/yyyy")&") — '
              'updates itself from the Requests sheet. Printed: "&TEXT(TODAY(),"mm/dd/yyyy")')
    # subtitle needs to be a formula: set directly
    ws.cell(row=2, column=1).value = ('="Fiscal year "&FY_Label&" (from "&TEXT(FY_Start,'
                                     '"mm/dd/yyyy")&") — updates itself from the Requests '
                                     'sheet. Printed: "&TEXT(TODAY(),"mm/dd/yyyy")')
    for c in range(1, 8):
        ws.cell(row=3, column=c).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3

    section_bar(ws, 5, 1, 6, "BUDGET POSITION")
    rows = [
        ("Adopted budget", "=Budget_Total"),
        ("Spent (actuals recorded)", f'=SUM({rng("AY")})'),
        ("Encumbered (approved, not yet reconciled)", f'=SUM({rng("AZ")})'),
        ("REMAINING", f'=Budget_Total-SUM({rng("AY")})-SUM({rng("AZ")})'),
    ]
    r = 6
    for text, formula in rows:
        c = ws.cell(row=r, column=1, value=text)
        c.font = F_LABEL if text == "REMAINING" else F_BODY
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        v = ws.cell(row=r, column=4, value=formula)
        v.number_format = CUR
        v.font = F_LABEL if text == "REMAINING" else F_BODY
        if text == "REMAINING":
            v.fill = FILL_BAND
        r += 1

    section_bar(ws, 11, 1, 6, "BY MONTH  (by departure date)")
    heads = ["", "Month", "Trips", "Estimated", "Actual", ""]
    for j, h in enumerate(("Month", "Trips", "Estimated", "Actual"), start=2):
        c = ws.cell(row=12, column=j, value=h)
        c.font = F_LABEL
        c.border = BOX
        c.fill = FILL_BAND
    for i in range(12):
        r = 13 + i
        a = ws.cell(row=r, column=1,
                    value="=FY_Start" if i == 0 else f"=EDATE($A{r - 1},1)")
        a.number_format = DATE
        a.font = F_SMALL
        ws.cell(row=r, column=2, value=f'=TEXT($A{r},"mmm yyyy")').font = F_BODY
        ws.cell(row=r, column=2).border = BOX
        tr = ws.cell(row=r, column=3,
                     value=f'=SUMPRODUCT(({rng("P")}>=$A{r})*({rng("P")}<EDATE($A{r},1))*({rng("A")}<>""))')
        tr.number_format = "0"
        tr.border = BOX
        e = ws.cell(row=r, column=4,
                    value=f'=SUMIFS({rng("AX")},{rng("P")},">="&$A{r},{rng("P")},"<"&EDATE($A{r},1))')
        e.number_format = CUR
        e.border = BOX
        ac = ws.cell(row=r, column=5,
                     value=f'=SUMIFS({rng("AY")},{rng("P")},">="&$A{r},{rng("P")},"<"&EDATE($A{r},1))')
        ac.number_format = CUR
        ac.border = BOX
    r = 25
    ws.cell(row=r, column=2, value="TOTAL").font = F_LABEL
    ws.cell(row=r, column=2).border = BOX
    for j, cl in ((3, "C"), (4, "D"), (5, "E")):
        c = ws.cell(row=r, column=j, value=f"=SUM({cl}13:{cl}24)")
        c.number_format = CUR if j > 3 else "0"
        c.font = F_LABEL
        c.border = BOX
        c.fill = FILL_BAND
    ws.column_dimensions["A"].hidden = True

    section_bar(ws, 27, 1, 6, "BY CATEGORY")
    for j, h in enumerate(("Category", "Estimated", "Actual"), start=2):
        c = ws.cell(row=28, column=j, value=h)
        c.font = F_LABEL
        c.border = BOX
        c.fill = FILL_BAND
    cats = [("Registration", "T", "AE"), ("Lodging", "W", "AF"),
            ("Travel / Mileage", "AA", "AG"), ("Per Diem", "AB", "AH"),
            ("Other", "AC", "AI")]
    r = 29
    for text, est_c, act_c in cats:
        ws.cell(row=r, column=2, value=text).border = BOX
        e = ws.cell(row=r, column=3, value=f'=SUM({rng(est_c)})')
        e.number_format = CUR
        e.border = BOX
        a = ws.cell(row=r, column=4, value=f'=SUM({rng(act_c)})')
        a.number_format = CUR
        a.border = BOX
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = F_LABEL
    ws.cell(row=r, column=2).border = BOX
    for j, cl in ((3, "C"), (4, "D")):
        c = ws.cell(row=r, column=j, value=f"=SUM({cl}29:{cl}{r - 1})")
        c.number_format = CUR
        c.font = F_LABEL
        c.border = BOX
        c.fill = FILL_BAND
    ws.print_area = "A1:G36"
    return ws


def sheet_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False
    col_widths(ws, {"A": 3, "B": 110})
    paint(ws, 1, 74, 1, 2)
    title_bar(ws, 1, 1, 2, "HOW THIS WORKBOOK RUNS THE TRAINING DESK",
              "The short version. Full manuals live in the docs folder of the project.")
    lines = [
        ("h", "THE DAILY LOOP"),
        ("t", "1.  Open the Dashboard. The five queues ARE your to-do list."),
        ("t", "2.  Click IMPORT REQUESTS — every form officers emailed in is pulled in and "
              "given the next tracking number automatically."),
        ("t", "3.  Pick a record in ACTIVE RECORD. Application, 2-90s and Notification fill "
              "themselves. Draft the approval email; when approval comes back, set Status = "
              "Approved and enter the date."),
        ("t", "4.  Register, book the hotel, mark the Y columns. Draft the officer "
              "notification. PUSH DEADLINES TO OUTLOOK adds any calendar items not already pushed."),
        ("t", "5.  After the trip: enter actuals, receipts date, draft the settlement email, "
              "print the packet, send to Finance by TUESDAY NOON, enter the sent date."),
        ("h", "RULES THE WORKBOOK LIVES BY"),
        ("t", "•  All math is spreadsheet formulas — per diem, deadlines, ledger, YTD. Code "
              "only touches what formulas can't: email, printing, calendar, importing."),
        ("t", "•  Emails always OPEN for your review. Nothing ever sends itself."),
        ("t", "•  Application, Pre 2-90 and Post 2-90 are PROTECTED — never type on them; "
              "they fill from the active record. (Review > Unprotect Sheet if you must fix "
              "layout; no password.)"),
        ("t", "•  Every number or address you might need to change lives on Settings."),
        ("t", "•  Rows 2–5 of Requests are SAMPLE records (officers named 'SAMPLE —'). "
              "Delete their rows' typed values before go-live; leave formula columns alone."),
        ("t", "•  Need more than 250 rows? Select the last data row's formula cells and "
              "fill down as far as needed."),
        ("h", "YEAR-END ROLLOVER — THE 10-MINUTE PROCEDURE"),
        ("t", "1.  (2 min) Finish reconciling — Ledger column L should read Yes on every "
              "row that had travel; chase anything Open."),
        ("t", "2.  (1 min) Dashboard > ARCHIVE YEAR. It saves a full copy to the archive "
              "folder named 'TPD_Training_Master_<FY>.xlsx' — that file IS the permanent record."),
        ("t", "3.  (2 min) In THIS workbook: Settings — advance FY label (FY26→FY27), FY "
              "start (+1 year), tracking prefix (T26-→T27-), new budget figure."),
        ("t", "4.  (3 min) Requests sheet: delete the VALUES in the input columns of all "
              "data rows (don't touch formula columns W, AA, AB, AD, AJ, AM, AR, AT or the "
              "helper columns AX–BE). Carry forward any trip that straddles the new year by "
              "re-importing or retyping it with a new-year tracking number."),
        ("t", "5.  (2 min) Verify: Dashboard tiles read zero; YTD months show the new year; "
              "Save. Done — the docs folder has the same list with screenshots."),
        ("h", "IF A MACRO BUTTON IS MISSING"),
        ("t", "Alt+F8 lists every macro by name (ImportRequests, DraftApprovalEmail, "
              "DraftOfficerNotification, DraftSettlementEmail, PrintPacket, "
              "PushDeadlinesToOutlook, ArchiveYear). The installer in the project's tools "
              "folder re-adds the buttons."),
    ]
    r = 3
    for kind, text in lines:
        c = ws.cell(row=r, column=2, value=text)
        if kind == "h":
            section_bar(ws, r, 1, 2, text)
        else:
            c.font = F_BODY
            c.alignment = A_LEFT_WRAP
            ws.row_dimensions[r].height = 28
        r += 1
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)
    # order matters for build (names referenced across sheets), then reorder
    sheet_settings(wb)
    sheet_lists(wb)
    sheet_requests(wb)
    sheet_dashboard(wb)
    sheet_application(wb)
    form_290(wb, "Pre 2-90", post=False)
    form_290(wb, "Post 2-90", post=True)
    sheet_notification(wb)
    sheet_ledger(wb)
    sheet_ytd(wb)
    sheet_instructions(wb)
    order = ["Dashboard", "Requests", "Application", "Pre 2-90", "Post 2-90",
             "Notification", "Ledger", "YTD Report", "Settings", "Instructions",
             "Lists"]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {os.path.abspath(OUT)}")


if __name__ == "__main__":
    build()
