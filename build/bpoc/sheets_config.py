"""Config & master sheets: StartHere, Settings, Lists, Agencies, Instructors,
ChapterMaster, ExamMaster, ExamPlan, SkillsMaster, SpellingMaster,
WritingMaster, Control (holidays + class-day calendar), Schedule.
"""
from openpyxl.utils import get_column_letter

from xlb import (
    HDR_ROW, DATA_ROW, CADET_LAST, ROWS_SCHEDULE, ROWS_CLASSDAYS,
    F_HDR, FILL_HDR, F_CALC, FILL_CALC, F_LABEL, F_BODY, F_SMALL, F_INPUT,
    FILL_INPUT, FILL_BAND, FILL_YELLOW, FILL_PAGE, FILL_STEEL, F_SECTION,
    A_LEFT, A_LEFT_WRAP, A_CENTER, BOX, DATE, TIME,
    header_row, fill_rows, dv_list, name_range, sheet_note, cf_yes_no,
    cf_formula, FILL_WARNBG, FILL_OKBG, FILL_AMBER, col_widths, section_bar,
    label, define, protect, unlock_range, page_setup_landscape,
    page_setup_portrait,
)
import data_chapters as DC
import data_lists as DL
import data_spelling as DS
import data_writing as DW


# --------------------------------------------------------------------------
def build_starthere(wb):
    ws = wb.create_sheet("StartHere")
    ws.sheet_view.showGridLines = False
    steps = [
        ("1. Settings", "Set academy class label, start/end dates, weights, "
                        "caps and flag thresholds. Blue cells are yours."),
        ("2. Lists / Agencies / Instructors", "Confirm dropdown lists, agency "
                        "contacts and the instructor roster (PID, certs, SME letters)."),
        ("3. ChapterMaster & Schedule", "Confirm chapter hours, then enter the "
                        "class schedule (date, times, chapter, instructor). "
                        "Delivered hours reconcile automatically."),
        ("4. ExamPlan / WritingMaster / SpellingMaster", "Pick this academy's "
                        "exams; writing due dates compute from the schedule."),
        ("5. Cadets", "Enter the roster (PID, names, agency). Everything keys "
                        "off PID."),
        ("6. Daily operation", "Log exam scores, spelling, attendance "
                        "exceptions, makeup, skills, writing, incidents, "
                        "counseling, PT and medical as they happen."),
        ("7. Watch the Dashboard / WatchList", "Flags, retest deadlines and "
                        "blocking issues surface automatically."),
        ("8. Print Center", "Sign-in sheets, evals, spelling tests, writing "
                        "handouts, transcripts and rosters — one place."),
        ("9. Agency emails", "EmailPreview + buttons draft per-agency Outlook "
                        "emails for review. Never auto-sent."),
        ("10. New academy", "Save-As, then use the New Academy reset button. "
                        "See the Coordinator Guide in docs/."),
    ]
    header_row(ws, ["Step", "What to do"])
    r = DATA_ROW
    for step, what in steps:
        ws.cell(row=r, column=2, value=step).font = F_LABEL
        c = ws.cell(row=r, column=3, value=what)
        c.font = F_BODY
        c.alignment = A_LEFT_WRAP
        ws.row_dimensions[r].height = 30
        r += 1
    col_widths(ws, {"A": 3, "B": 34, "C": 90})
    sheet_note(ws, "Workbook map: white boxes with a blue border are yours "
                   "to type (your entries show in blue), gray = calculated, "
                   "sys* sheets are the locked engine.")
    return ws


# --------------------------------------------------------------------------
SETTINGS = [
    # (label, value, note, name, fmt)
    ("Academy Class", "BPOC-2026-01", "Label shown on reports", "cfgAcademyClass", None),
    ("Start Date", "2026-05-11", "First day of academy", "cfgStartDate", DATE),
    ("End Date", "2026-11-06", "Scheduled graduation date", "cfgEndDate", DATE),
    ("Total Scheduled Minutes", 145800, "Academy length in minutes", "cfgTotalScheduledMinutes", "#,##0"),
    ("Passing Score", 70, "Minimum passing score for exams", "cfgPassingScore", None),
    ("Pass Threshold Score", 70, "Category-average threshold", "cfgThresholdScore", None),
    ("Retake Recorded Cap", 70, "Recorded score after a passed retest", "cfgRetakeRecordedCap", None),
    ("Threshold After Exam #", 4, "Category avg enforced once this many exams of that type are recorded (or all planned ones are)", "cfgThresholdAfterExam", None),
    ("Retest Within (class days)", 5, "Policy 300.5 retest window", "cfgRetestClassDays", None),
    ("Memo Due (class days)", 3, "Deficiency memo due this many class days after assignment", "cfgMemoDueClassDays", None),
    ("Weight: Major", 0.4, "Major exams weight", "cfgWeightMajor", "0%"),
    ("Weight: Minor", 0.3, "Minor exams weight", "cfgWeightMinor", "0%"),
    ("Weight: Spelling", 0.1, "Spelling weight", "cfgWeightSpelling", "0%"),
    ("Weight: Final", 0.2, "Final exam weight", "cfgWeightFinal", "0%"),
    ("Spelling Intervention Avg", 75, "Below this spelling avg = early intervention (300.4.B)", "cfgSpellInterventionAvg", None),
    # There is NO classroom attendance allowance. TCOLE's IRG is explicit -
    # "Learners are required to attend all classroom hours as listed in this
    # instructor resource guide, there is no 10% attendance rule" - and
    # Academy policy 400 sets no percentage either: it bars missing skills
    # training at all, caps PT at five sessions, and requires minute-for-
    # minute makeup of everything else. A 5% cap shipped here for several
    # rounds and was an invention of this build, not a rule. The two values
    # below are EARLY-WARNING thresholds for the coordinator only; they
    # forgive nothing and gate nothing.
    ("Makeup Advisory (minutes owed)", 120, "Watch a cadet once outstanding makeup reaches this many minutes (early warning only - NOT an allowance)", "cfgMakeupAdvisoryMin", "#,##0"),
    ("Makeup Critical (minutes owed)", 480, "Escalate once outstanding makeup reaches this many minutes (early warning only - NOT an allowance)", "cfgMakeupCriticalMin", "#,##0"),
    ("PT Cap (sessions)", 5, "Max PT sessions missed", "cfgPTCapSessions", None),
    ("PT Advisory (% of 5-session cap)", 0.6, "Tier 1 warning at this share of the PT session cap (PT is the only real cap in policy)", "cfgAttendanceAdvisoryPct", "0%"),
    ("PT Critical (% of 5-session cap)", 0.8, "Tier 2 warning at this share of the PT session cap", "cfgAttendanceCriticalPct", "0%"),
    ("Home Agency (gets all cadets)", "TPD", "Agency code whose email includes every cadet", "cfgHomeAgency", None),
    ("Current Exam # (for emails)", 1, "Sequence # of the exam to email", "cfgCurrentExamNum", None),
    ("Current Spelling # (for emails)", 1, "Highest spelling test # administered", "cfgCurrentSpellingNum", None),
    ("Writing Due Time", "1700", "Due time shown on handouts", "cfgWritingDueTime", None),
    ("PT Final Min Points", 70, "Points needed to pass the final PT assessment (approved chart: 60 = every event minimum met, 100 = Tier 5 in all five)", "cfgPTFinalMinPoints", None),
    # flag engine thresholds
    ("Flag: consecutive exam fails", 2, "Consecutive graded-exam scores below passing", "cfgFlagConsecFails", None),
    ("Flag: grade drop (points)", 10, "Drop in recorded score vs cadet's prior exam", "cfgFlagGradeDrop", None),
    ("Flag: category margin", 5, "Category avg within this many points of 70", "cfgFlagCategoryMargin", None),
    ("Flag: PT % of session cap", 0.6, "Flag once PT sessions missed reach this share of the five-session cap", "cfgFlagAttendancePct", "0%"),
    ("Flag: open negative incidents", 2, "Open negative incidents at/above this count", "cfgFlagOpenIncidents", None),
    ("Flag: overdue writing", 1, "Overdue writing assignments at/above this count", "cfgFlagOverdueWriting", None),
]

CALC_SETTINGS = [
    ("Weights total (must = 100%)", "cfgWeightMajor+cfgWeightMinor+cfgWeightSpelling+cfgWeightFinal",
     "CALCULATED", "cfgWeightsTotal", "0%"),
]


def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Setting", "Value", "Notes", "Internal name",
                    "Detected from data", "Check"])
    r = DATA_ROW
    for lab, val, note, nm, fmt in SETTINGS:
        ws.cell(row=r, column=2, value=lab).font = F_LABEL
        c = ws.cell(row=r, column=3, value=val)
        c.fill = FILL_INPUT
        c.font = F_INPUT
        c.border = BOX
        if fmt:
            c.number_format = fmt
        if nm in ("cfgStartDate", "cfgEndDate"):
            from datetime import datetime
            c.value = datetime.strptime(val, "%Y-%m-%d")
            c.number_format = DATE
        ws.cell(row=r, column=4, value=note).font = F_SMALL
        ws.cell(row=r, column=5, value=nm).font = F_SMALL
        define(wb, nm, "Settings", f"$C${r}")
        r += 1
    for lab, fx, note, nm, fmt in CALC_SETTINGS:
        ws.cell(row=r, column=2, value=lab).font = F_LABEL
        c = ws.cell(row=r, column=3, value="=" + fx)
        c.fill = FILL_CALC
        c.font = F_CALC
        c.border = BOX
        if fmt:
            c.number_format = fmt
        ws.cell(row=r, column=4, value=note).font = F_SMALL
        ws.cell(row=r, column=5, value=nm).font = F_SMALL
        define(wb, nm, "Settings", f"$C${r}")
        r += 1
    # detection helpers for the "Current #" settings and the academy length
    for i, (lab, _, _, nm, _) in enumerate(SETTINGS):
        row = DATA_ROW + i
        if nm == "cfgTotalScheduledMinutes":
            # the Schedule is the authority on academy length. This no longer
            # scales any cap (there is none); it is the reference figure for
            # the 736-hour reconciliation and for sanity-checking the
            # schedule, so a stale value is still worth objecting to.
            ws.cell(row=row, column=6, value=(
                # nrSCH_TimeCheck,"OK": a block whose End is before its
                # Start still produces MOD-derived hours (a 13:00->09:00
                # typo reads 20 hrs), and those bogus hours used to be
                # summed straight into the figure this cell tells the
                # coordinator to copy into cfgTotalScheduledMinutes.
                '=IFERROR(ROUND(SUMIFS(nrSCH_Hrs,nrSCH_TimeCheck,"OK")'
                "*60,0),0)"
            )).font = F_CALC
            # the empty-schedule branch must still object when C holds the
            # PREVIOUS academy's minutes: New Academy Reset empties the
            # Schedule but leaves this value.
            ws.cell(row=row, column=7, value=(
                f'=IF(AND($F${row}=0,N($C${row})=0),"No schedule entered yet",'
                f'IF($F${row}=0,"CHECK - no schedule entered yet, but Total '
                f'Scheduled Minutes still reads "&TEXT($C${row},"#,##0")'
                f'&" (the previous academy\'s length, or the shipped default) '
                f'- enter this academy\'s schedule, then set this to match",'
                f'IF(ROUND($C${row},0)=ROUND($F${row},0),"OK",'
                f'"CHECK - schedule totals "&TEXT($F${row},"#,##0")&" min ("'
                f'&TEXT($F${row}/60,"#,##0")&" hrs)")))'
            )).font = F_SMALL
        # A blank / non-date Start Date empties the entire class-day
        # calendar (Control), and with it every retest deadline, memo due
        # date, writing date and sign-in sheet. Nothing used to say so here.
        # cfgHomeAgency drives the "All Cadets" draft in modAgencyEmail: a
        # value matching no AgencyID silently skipped the consolidated draft,
        # the all-agency discipline digest and the orphan-cadet warning, and
        # the run still reported the same draft count. It was the only cell in
        # the C25/C26/C27 email-run group with neither a dropdown nor a check.
        if nm == "cfgHomeAgency":
            ws.cell(row=row, column=6, value=(
                f'=IF($C${row}="","(blank)",'
                f'IFERROR(INDEX(rngAgencyNames,MATCH($C${row},rngAgencyIDs,0)),'
                f'"(no such AgencyID)"))')).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF(COUNTIF(rngAgencyIDs,$C${row})=0,'
                f'"CHECK - not an AgencyID on the Agencies sheet; the '
                f'\'All Cadets\' draft, the all-agency discipline digest and '
                f'the orphan-cadet warning are all SKIPPED",'
                f'IF(INDEX(rngAgencyActive,MATCH($C${row},rngAgencyIDs,0))'
                f'<>"Yes","CHECK - that agency is not marked Active","OK"))'
            )).font = F_SMALL
            # same dropdown Cadets!G and EmailPreview!C5 already carry — the
            # cell that drives the PREVIEW was validated while the cell that
            # drives the actual emails was free text
            dv_list(ws, "=rngAgencyIDs", [f"C{row}"])
        if nm == "cfgStartDate":
            ws.cell(row=row, column=6, value=(
                f'=IF(ISNUMBER($C${row}),TEXT($C${row},"ddd mm/dd/yyyy"),'
                f'"(not a date)")')).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF(NOT(ISNUMBER($C${row})),'
                f'"FIX - Start Date is blank or is not a real date. The '
                f'Control class-day calendar is EMPTY until it is set, so '
                f'every Day #, retest deadline, memo due date, writing date '
                f'and sign-in sheet is blank too.",'
                f'IF(WEEKDAY($C${row},2)>5,'
                f'"CHECK - Start Date falls on a weekend; Day 1 will be the '
                f'next weekday","OK"))')).font = F_SMALL
        if nm == "cfgEndDate":
            ws.cell(row=row, column=6, value=(
                f'=IF(ISNUMBER($C${row}),TEXT($C${row},"ddd mm/dd/yyyy"),'
                f'"(not a date)")')).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF(NOT(ISNUMBER($C${row})),'
                f'"FIX - End Date is blank or is not a real date; the class-'
                f'day calendar cannot mark days In Session.",'
                f'IF(AND(ISNUMBER(cfgStartDate),$C${row}<=cfgStartDate),'
                f'"FIX - End Date is on or before Start Date","OK"))'
            )).font = F_SMALL
        if nm == "cfgWritingDueTime":
            ws.cell(row=row, column=6, value=(
                '=IFERROR(TEXT(TIMEVALUE(LEFT(TEXT($C$%d,"0000"),2)&":"'
                '&RIGHT(TEXT($C$%d,"0000"),2)),"h:mm AM/PM"),"(unreadable)")'
                % (row, row))).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF($F${row}="(unreadable)",'
                f'"CHECK - enter a 24-hour time as 4 digits (e.g. 1700). '
                f'This is the due time printed on the Writing handout.",'
                f'"OK - shown on the WritingHandout Due column header")'
            )).font = F_SMALL
        if nm == "cfgCurrentExamNum":
            ws.cell(row=row, column=6, value=(
                "=IFERROR(MAX(FILTER(nrES_Seq,(nrES_Rec<>\"\")*(nrES_Seq<>\"\"))),0)"
            )).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF($F${row}=0,"No exam scores posted yet",'
                f'IF($C${row}>$F${row},"CHECK - set to Test "&$C${row}&" but data only has Test "&$F${row},'
                f'IF($F${row}>$C${row},"CHECK - data shows Test "&$F${row},"OK")))'
            )).font = F_SMALL
        if nm == "cfgCurrentSpellingNum":
            # scan the per-test counts ROW (Spelling!D57:O57), not the
            # per-cadet '# Taken' column, for the highest administered test
            ws.cell(row=row, column=6, value=(
                "=MAX(SUMPRODUCT(MAX((nrSpellCounts>0)*COLUMN(nrSpellCounts)))-COLUMN(Spelling!$D$1)+1,0)"
            )).font = F_CALC
            ws.cell(row=row, column=7, value=(
                f'=IF($C${row}<cfgCurrentExamNum,'
                f'"CHECK - spelling will be OMITTED this run",'
                f'IF($F${row}>$C${row},"CHECK - data shows Spelling "&$F${row},"OK"))'
            )).font = F_SMALL
    weights_row = r
    ws.cell(row=r + 1, column=2, value="Weights check:").font = F_LABEL
    ws.cell(row=r + 1, column=3, value=(
        '=IF(ROUND(cfgWeightsTotal,4)=1,"OK","FIX: weights must total 100%")'
    )).font = F_CALC
    r += 3

    # ---- PT final-assessment points rubric (approved chart v1 09/03/2026) --
    section_bar(ws, r, 2, 9, "Final PT Assessment — Points Rubric "
                             "(approved PT Test Score Chart, the EXIT "
                             "standard — each cell is the value a cadet must "
                             "REACH for that tier; higher than the entry "
                             "standard by design)")
    r += 1
    rub_hdr = r
    header_row(ws, ["Event", "Measure", "Tier 1", "Tier 2", "Tier 3",
                    "Tier 4", "Tier 5", "Notes"], row=r)
    r += 1
    # the tier point values are themselves editable, so a revised chart that
    # keeps five tiers but changes what they are worth needs no rebuild
    rub_pts = r
    ws.cell(row=r, column=2, value="Points per tier").font = F_LABEL
    ws.cell(row=r, column=3, value="(editable)").font = F_SMALL
    for i, pts in enumerate(DL.PT_TIER_POINTS):
        cc = ws.cell(row=r, column=4 + i, value=pts)
        cc.font = F_INPUT
        cc.border = BOX
    ws.cell(row=r, column=9,
            value="5 events x 20 = 100 max; x 12 = 60 (every minimum met). "
                  "Passing = PT Final Min Points above.").font = F_SMALL
    r += 1
    rub_first = r
    for ev, _src, hib, measure, bands in DL.PT_FINAL_BANDS:
        ws.cell(row=r, column=2, value=ev).font = F_LABEL
        ws.cell(row=r, column=3, value=measure).font = F_SMALL
        for i, v in enumerate(bands):
            cc = ws.cell(row=r, column=4 + i, value=v)
            cc.font = F_INPUT
            cc.border = BOX
        ws.cell(row=r, column=9, value=(
            "best tier whose threshold is met"
            + ("; higher is better" if hib else "; lower is better")
            + "; below Tier 1 scores 0")).font = F_SMALL
        r += 1
    ws.cell(row=r, column=2, value=(
        "Bench Press and Vertical Jump are BASELINE standards only — they "
        "are not scored at the final assessment.")).font = F_SMALL
    r += 1
    define(wb, "nrPTTierPts", "Settings", f"$D${rub_pts}:$H${rub_pts}")
    define(wb, "nrPTBands", "Settings", f"$D${rub_first}:$H${rub_first + 4}")
    define(wb, "nrPTBandEvents", "Settings",
           f"$B${rub_first}:$B${rub_first + 4}")
    col_widths(ws, {"A": 3, "B": 34, "C": 22, "D": 12, "E": 12, "F": 12,
                    "G": 12, "H": 12, "I": 30})
    sheet_note(ws, "White boxes with a blue border = edit per academy "
                   "(entries show in blue). Gray = calculated. Internal "
                   "names must not change. The PT rubric block holds the "
                   "approved PT Test Score Chart — each cell is the value a "
                   "cadet must REACH for that tier, and the PT sheet scores "
                   "itself from it.")
    return ws


# --------------------------------------------------------------------------
def build_lists(wb):
    ws = wb.create_sheet("Lists")
    ws.sheet_view.showGridLines = False
    keys = list(DL.LISTS.keys())
    header_row(ws, keys)
    for i, k in enumerate(keys):
        colL = get_column_letter(2 + i)
        vals = DL.LISTS[k]
        for j, v in enumerate(vals):
            c = ws.cell(row=DATA_ROW + j, column=2 + i, value=v)
            c.fill = FILL_INPUT
            c.font = F_INPUT
            c.border = BOX
        nm = ("lst" + "".join(w.capitalize() for w in
              k.replace("/", " ").replace("?", "").replace("#", "Num")
              .replace("-", " ").split()))
        define(wb, nm, "Lists", f"${colL}${DATA_ROW}:${colL}${DATA_ROW+len(vals)-1}")
        ws.column_dimensions[colL].width = max(16, len(k) + 2)
    ws.column_dimensions["A"].width = 3
    sheet_note(ws, "These feed every dropdown. Add items at the bottom of a "
                   "column, then extend the named range (Formulas > Name "
                   "Manager) if you exceed the seeded rows.")
    return ws


# --------------------------------------------------------------------------
def build_agencies(wb):
    ws = wb.create_sheet("Agencies")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["AgencyID", "AgencyName", "Primary Contact", "Email",
                    "Phone", "Mailing Address", "ActiveYN", "Last Email Sent"])
    last = DATA_ROW + 44
    for i, (aid, aname, contact, email, phone, addr, act) in enumerate(DL.AGENCIES):
        r = DATA_ROW + i
        for col, v in ((2, aid), (3, aname), (4, contact), (5, email),
                       (6, phone), (7, addr), (8, act)):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = FILL_INPUT
            c.font = F_INPUT
            c.border = BOX
    fill_rows(ws, DATA_ROW, last, {
        "B": (None, "in"), "C": (None, "in"), "D": (None, "in"),
        "E": (None, "in"), "F": (None, "in"), "G": (None, "in"),
        "H": (None, "in"),
        # MAXIFS returns 0, not an error, so IFERROR never fired and every
        # agency that has never been emailed showed 12/30/1899 here (and made
        # EmailPreview's "never" fallback dead code).
        "I": ('IF($B{r}="","",IFERROR(IF(MAXIFS(nrELdate,nrELagency,$B{r})=0,'
              '"",MAXIFS(nrELdate,nrELagency,$B{r})),""))', "fx"),
    })
    for r in range(DATA_ROW, last + 1):
        ws[f"I{r}"].number_format = DATE
    dv_list(ws, "=lstYesNo", [f"H{DATA_ROW}:H{last}"])
    define(wb, "rngAgencyIDs", "Agencies", f"$B${DATA_ROW}:$B${last}")
    define(wb, "rngAgencyNames", "Agencies", f"$C${DATA_ROW}:$C${last}")
    define(wb, "rngAgencyEmails", "Agencies", f"$E${DATA_ROW}:$E${last}")
    define(wb, "rngAgencyActive", "Agencies", f"$H${DATA_ROW}:$H${last}")
    define(wb, "nrAgencyLastEmail", "Agencies", f"$I${DATA_ROW}:$I${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "D": 20, "E": 46, "F": 14,
                    "G": 32, "H": 9, "I": 14})
    sheet_note(ws, "AgencyID is the short code used on Cadets. 'Last Email "
                   "Sent' is calculated from EmailLog and drives the "
                   "since-last-email discipline digest.")
    return ws


# --------------------------------------------------------------------------
def build_instructors(wb):
    ws = wb.create_sheet("Instructors")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Instructor Name", "Agency / Unit", "TCOLE PID",
                    "Instructor Type", "Certificates (list)",
                    "Cert Expiration", "SME Letter", "Bio On File?",
                    "Notes", "Audit Ready?", "On Schedule?",
                    "Chapters Taught", "Last Class Taught",
                    "Cert Status"])
    last = DATA_ROW + 89
    for i, nm in enumerate(DL.INSTRUCTORS):
        ws.cell(row=DATA_ROW + i, column=2, value=nm)
    for j, nm in enumerate(DL.GUEST_ENTITIES):
        r = DATA_ROW + len(DL.INSTRUCTORS) + j
        ws.cell(row=r, column=2, value=nm)
        ws.cell(row=r, column=5, value="Guest/Outside")
    fill_rows(ws, DATA_ROW, last, {
        "B": (None, "in"), "C": (None, "in"), "D": (None, "in"),
        "E": (None, "in"), "F": (None, "in"), "G": (None, "in"),
        "H": (None, "in"), "I": (None, "in"), "J": (None, "in"),
        # IRG: every instructor/co-teacher teaching LOs needs a documented
        # qualification (TCOLE cert or written SME approval) AND a bio.
        # Cert Expiration (G) used to be a dead field: it could be keyed
        # and nothing read it, so an instructor whose TCOLE certificate
        # lapsed BEFORE the day they taught still reported Audit Ready. O
        # now grades the expiration against the last date they actually
        # taught, and Audit Ready refuses an expired or undocumented one.
        "K": ('IF($B{r}="","",IF($E{r}="Guest/Outside","N/A",'
              'IF(AND(OR($E{r}="TCOLE Instructor",'
              'AND(LEFT($E{r},3)="SME",$H{r}="On File")),'
              'OR($I{r}="On File",$I{r}="N/A"),'
              'LEFT($O{r},7)<>"EXPIRED",$O{r}<>"MISSING EXPIRATION"'
              '),"Yes","No")))', "fx"),
        "L": ('IF($B{r}="","",IF(SUMPRODUCT(--ISNUMBER(SEARCH($B{r},'
              'nrSCH_Instr)))>0,"Yes",""))', "fx"),
        # nrSCH_ChNum carries chapter numbers as TEXT (they are the
        # workbook's join key), so a plain SORT put an instructor's chapters
        # in dictionary order - "1, 10, 2, 20" - on this sheet and in the
        # ChapterPacket instructor panel. SORTBY on the numeric value sorts
        # them the way a human reads them while keeping the text keys.
        "M": ('IF($B{r}="","",IFERROR(LET(v,UNIQUE(FILTER(nrSCH_ChNum,'
              'ISNUMBER(SEARCH($B{r},nrSCH_Instr))*(nrSCH_ChNum<>""))),'
              'TEXTJOIN(", ",TRUE,SORTBY(v,IFERROR(v+0,99999),1))),""))', "fx"),
        # deliberately LET-free so the LibreOffice recalc sweep can still
        # exercise both columns against real schedule data
        # SUMPRODUCT(MAX(...)), not MAX(IF(...)): the IF form needs an
        # array context this cell does not have, and it silently returned a
        # blank for 60 of the 65 instructors actually on the schedule (only
        # the two whose own row happened to intersect resolved). SUMPRODUCT
        # forces the array evaluation, and boolean*date makes the IF
        # unnecessary - non-matching rows contribute 0.
        "N": ('IF($B{r}="","",'
              'IF(SUMPRODUCT(MAX(ISNUMBER(SEARCH($B{r},nrSCH_Instr))'
              '*nrSCH_Date))=0,"",'
              'SUMPRODUCT(MAX(ISNUMBER(SEARCH($B{r},nrSCH_Instr))'
              '*nrSCH_Date))))', "fx"),
        "O": ('IF($B{r}="","",'
              'IF($E{r}="Guest/Outside","N/A",'
              'IF($L{r}<>"Yes","(not on schedule)",'
              'IF(LEFT($E{r},3)="SME","N/A (SME letter)",'
              'IF($E{r}<>"TCOLE Instructor","CHECK TYPE",'
              'IF($G{r}="","MISSING EXPIRATION",'
              'IF(NOT(ISNUMBER($G{r})),"CHECK DATE",'
              'IF(AND(N($N{r})>0,$G{r}<$N{r}),'
              '"EXPIRED BEFORE LAST CLASS TAUGHT "&TEXT($G{r},"mm/dd/yy"),'
              'IF($G{r}<cfgEndDate,'
              '"RENEW - expires during academy "&TEXT($G{r},"mm/dd/yy"),'
              '"Current")))))))))', "fx"),
    })
    for r in range(DATA_ROW, last + 1):
        ws[f"G{r}"].number_format = DATE
        ws[f"N{r}"].number_format = DATE
    dv_list(ws, "=lstInstructorType", [f"E{DATA_ROW}:E{last}"])
    dv_list(ws, "=lstMaterialsStatus",
            [f"H{DATA_ROW}:H{last}", f"I{DATA_ROW}:I{last}"])
    cf_yes_no(ws, f"K{DATA_ROW}:K{last}")
    cf_formula(ws, f"L{DATA_ROW}:L{last}",
               f'AND($L{DATA_ROW}="Yes",$K{DATA_ROW}="No")', FILL_WARNBG)
    define(wb, "nrInstrNames", "Instructors", f"$B${DATA_ROW}:$B${last}")
    define(wb, "nrInstrReady", "Instructors", f"$K${DATA_ROW}:$K${last}")
    define(wb, "nrInstrOnSched", "Instructors", f"$L${DATA_ROW}:$L${last}")
    define(wb, "nrInstrChTaught", "Instructors", f"$M${DATA_ROW}:$M${last}")
    define(wb, "nrInstrLastTaught", "Instructors", f"$N${DATA_ROW}:$N${last}")
    define(wb, "nrInstrCertStat", "Instructors", f"$O${DATA_ROW}:$O${last}")
    cf_formula(ws, f"O{DATA_ROW}:O{last}",
               f'OR(LEFT($O{DATA_ROW},7)="EXPIRED",'
               f'LEFT($O{DATA_ROW},5)="RENEW",'
               f'$O{DATA_ROW}="MISSING EXPIRATION",'
               f'$O{DATA_ROW}="CHECK DATE",$O{DATA_ROW}="CHECK TYPE")',
               FILL_WARNBG)
    col_widths(ws, {"A": 3, "B": 24, "C": 18, "D": 12, "E": 20, "F": 30,
                    "G": 13, "H": 12, "I": 12, "J": 26, "K": 11, "L": 12,
                    "M": 30, "N": 15, "O": 34})
    sheet_note(ws, "IRG: EVERY instructor or co-teacher who teaches a "
                   "learning objective needs a bio + documented "
                   "qualification (TCOLE instructor cert, or SME letter "
                   "approved in writing). 'On Schedule?' scans every "
                   "schedule block's instructor text — co-teachers in "
                   "multi-name entries are matched individually. Red = "
                   "teaching without documentation. Guest/Outside entries "
                   "(proctors, venues) are exempt.")
    return ws


# --------------------------------------------------------------------------
BANK_SLOTS = 10      # certified instructors per topic  -> columns C..L
# must match BANK_SLOTS: with fewer slots than the bank, a topic taught by
# more instructors than there are slots lost the overflow from the dropdown
# that governs those very Schedule rows. Columns M..V; if this changes,
# update ClearRange "InstructorBanks" in src/vba/bpoc/modNewAcademy.bas.
SEL_SLOTS = 10       # picked to teach this academy


def build_instructorbanks(wb):
    """Per-topic instructor banks: the certified pool persists across
    academies; the 'Teaching this academy' picks are cleared on reset and
    drive the Schedule's instructor dropdown for that topic's rows."""
    from openpyxl.utils import get_column_letter as gcl
    ws = wb.create_sheet("InstructorBanks")
    ws.sheet_view.showGridLines = False
    hdrs = (["Topic / Class"]
            + [f"Bank {i+1}" for i in range(BANK_SLOTS)]
            + [f"Teach {i+1}" for i in range(SEL_SLOTS)]
            + ["# Selected"])
    header_row(ws, hdrs)
    first = DATA_ROW
    topics = ([name for _m, _c, name, _mi, _tp in DC.CHAPTERS]
              + [name for name, _p, _t in DC.SUBTOPICS]
              + [a for a in DC.ACTIVITIES if not a.startswith("Test ")
                 and a != "Lunch"])
    r = first
    for t in topics:
        ws.cell(row=r, column=2, value=t)
        r += 1
    last = r + 9        # spare rows
    # with BANK_SLOTS = SEL_SLOTS = 10 these are Bank C..L, Teach M..V and
    # "# Selected" W. (These comments said M..T / U while the constants
    # already produced M..V / W - the same stale-column-letter mistake this
    # workbook keeps regressing on. Derive, never hand-count.)
    bank_first_c, bank_last_c = 3, 2 + BANK_SLOTS               # C..L
    sel_first_c, sel_last_c = 3 + BANK_SLOTS, 2 + BANK_SLOTS + SEL_SLOTS  # M..V
    cnt_c = sel_last_c + 1                                       # W
    cols = {"B": (None, "in")}
    for c in range(bank_first_c, sel_last_c + 1):
        cols[gcl(c)] = (None, "in")
    cols[gcl(cnt_c)] = (
        'IF($B{r}="","",COUNTA(%s{r}:%s{r}))'
        % (gcl(sel_first_c), gcl(sel_last_c)), "fx")
    fill_rows(ws, first, last, cols)
    dv_list(ws, "=nrInstrNames",
            [f"{gcl(bank_first_c)}{first}:{gcl(bank_last_c)}{last}"])
    # academy picks must come from THAT ROW's certified bank
    dv_list(ws, f"=INDEX(${gcl(bank_first_c)}${first}:${gcl(bank_last_c)}"
                f"${last},MATCH($B{first},$B${first}:$B${last},0),0)",
            [f"{gcl(sel_first_c)}{first}:{gcl(sel_last_c)}{last}"],
            enforce=False)
    define(wb, "nrBankTopics", "InstructorBanks", f"$B${first}:$B${last}")
    define(wb, "nrBankGrid", "InstructorBanks",
           f"${gcl(bank_first_c)}${first}:${gcl(bank_last_c)}${last}")
    define(wb, "nrBankSel", "InstructorBanks",
           f"${gcl(sel_first_c)}${first}:${gcl(sel_last_c)}${last}")
    col_widths(ws, {"A": 3, "B": 46})
    for c in range(bank_first_c, cnt_c):
        ws.column_dimensions[gcl(c)].width = 16
    ws.column_dimensions[gcl(cnt_c)].width = 10
    sheet_note(ws, "Bank = everyone certified/approved to teach the topic "
                   "(kept between academies). Teach = who you picked for "
                   "THIS academy — the Schedule's instructor dropdown for a "
                   "topic offers exactly these names. Topics without picks "
                   "fall back to the full roster. New Academy Reset clears "
                   "Teach columns only.")
    return ws


# --------------------------------------------------------------------------
def build_chaptermaster(wb):
    ws = wb.create_sheet("ChapterMaster")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Module", "Ch #", "Chapter Name", "TCOLE Min Hrs",
                    "Planned Hrs", "Delivered Hrs", "vs TCOLE Min",
                    "First Taught", "Last Taught", "Default Instructor",
                    # per-chapter TRAINING FILE items (IRG mandatory list)
                    "Lesson Plan (SME)", "Instructor Bio", "Sign-In Sheets",
                    "Assessment", "Grade Sheet", "Evals Collected",
                    "Handouts/PPT (opt.)",
                    "Special TCOLE Requirement", "Special Req Met?",
                    "File Complete?"])
    first = DATA_ROW
    r = first
    for mod, ch, name, minh, tpdh in DC.CHAPTERS:
        ws.cell(row=r, column=2, value=mod)
        ws.cell(row=r, column=3, value=ch)
        ws.cell(row=r, column=4, value=name)
        ws.cell(row=r, column=5, value=minh)
        ws.cell(row=r, column=6, value=tpdh)
        if ch in DC.SPECIAL_REQS:
            ws.cell(row=r, column=19, value=DC.SPECIAL_REQS[ch])
        r += 1
    last = r - 1
    fill_rows(ws, first, last, {
        "B": (None, "in"), "C": (None, "in"), "D": (None, "in"),
        "E": (None, "in"), "F": (None, "in"),
        # rollup by chapter NUMBER: schedule rows resolve sub-classes
        # (Traffic Code, Crash, TIM, Crim-Inv subs) to their parent chapter
        # nrSCH_TimeCheck,"OK" quarantines a swapped start/end: those blocks
        # are flagged on Schedule and counted by the sysAudit "impossible
        # times" line, but their MOD-derived hours were still rolled up here,
        # into nrCHtotalDelivered and into the 736-hour audit figure.
        "G": ('IF($D{r}="","",ROUND(SUMIFS(nrSCH_Hrs,nrSCH_ChNum,$C{r},'
              'nrSCH_TimeCheck,"OK"),2))', "fx"),
        "H": ('IF($D{r}="","",IF($G{r}=0,"",$G{r}-$E{r}))', "fx"),
        "I": ('IF($D{r}="","",IFERROR(IF(MINIFS(nrSCH_Date,nrSCH_ChNum,$C{r})=0,'
              '"",MINIFS(nrSCH_Date,nrSCH_ChNum,$C{r})),""))', "fx"),
        "J": ('IF($D{r}="","",IFERROR(IF(MAXIFS(nrSCH_Date,nrSCH_ChNum,$C{r})=0,'
              '"",MAXIFS(nrSCH_Date,nrSCH_ChNum,$C{r})),""))', "fx"),
        "K": (None, "in"),
        "L": (None, "in"), "M": (None, "in"), "N": (None, "in"),
        "O": (None, "in"), "P": (None, "in"), "Q": (None, "in"),
        "R": (None, "in"), "S": (None, "in"), "T": (None, "in"),
        "U": ('IF($D{r}="","",IF(AND(COUNTIF(L{r}:P{r},"On File")'
              '+COUNTIF(L{r}:P{r},"N/A")=5,'
              'OR($Q{r}="On File",$Q{r}="N/A"),'
              'OR($S{r}="",$T{r}="Yes",$T{r}="N/A")),"Yes","No"))', "fx"),
    })
    for r2 in range(first, last + 1):
        # chapter numbers are the workbook's join key and are stored as TEXT;
        # '@' keeps a hand-typed edit from becoming a number that no
        # XLOOKUP/MATCH against nrCHnum can match
        ws[f"C{r2}"].number_format = "@"
        ws[f"I{r2}"].number_format = DATE
        ws[f"J{r2}"].number_format = DATE
        ws[f"S{r2}"].alignment = A_LEFT_WRAP
    dv_list(ws, "=nrInstrNames", [f"K{first}:K{last}"])
    dv_list(ws, "=lstMaterialsStatus",
            [f"L{first}:L{last}", f"M{first}:M{last}", f"N{first}:N{last}",
             f"O{first}:O{last}", f"P{first}:P{last}", f"Q{first}:Q{last}",
             f"R{first}:R{last}"])
    dv_list(ws, '"Yes,No,N/A"', [f"T{first}:T{last}"])
    cf_formula(ws, f"H{first}:H{last}",
               f'AND($H{first}<>"",$H{first}<0)', FILL_WARNBG)
    cf_yes_no(ws, f"U{first}:U{last}")
    define(wb, "nrCHmod", "ChapterMaster", f"$B${first}:$B${last}")
    define(wb, "nrCHnum", "ChapterMaster", f"$C${first}:$C${last}")
    define(wb, "nrCHname", "ChapterMaster", f"$D${first}:$D${last}")
    define(wb, "nrCHmin", "ChapterMaster", f"$E${first}:$E${last}")
    define(wb, "nrCHplan", "ChapterMaster", f"$F${first}:$F${last}")
    define(wb, "nrCHdeliv", "ChapterMaster", f"$G${first}:$G${last}")
    define(wb, "nrCHfirst", "ChapterMaster", f"$I${first}:$I${last}")
    define(wb, "nrCHinstr", "ChapterMaster", f"$K${first}:$K${last}")
    define(wb, "nrCHlesson", "ChapterMaster", f"$L${first}:$L${last}")
    define(wb, "nrCHbio", "ChapterMaster", f"$M${first}:$M${last}")
    define(wb, "nrCHsignin", "ChapterMaster", f"$N${first}:$N${last}")
    define(wb, "nrCHassess", "ChapterMaster", f"$O${first}:$O${last}")
    define(wb, "nrCHgrade", "ChapterMaster", f"$P${first}:$P${last}")
    define(wb, "nrCHevals", "ChapterMaster", f"$Q${first}:$Q${last}")
    define(wb, "nrCHhandoutOpt", "ChapterMaster", f"$R${first}:$R${last}")
    define(wb, "nrCHspecial", "ChapterMaster", f"$S${first}:$S${last}")
    define(wb, "nrCHspecialMet", "ChapterMaster", f"$T${first}:$T${last}")
    define(wb, "nrCHfileOK", "ChapterMaster", f"$U${first}:$U${last}")
    # totals row
    tr = last + 2
    ws.cell(row=tr, column=4, value="Totals:").font = F_LABEL
    ws.cell(row=tr, column=5, value=f"=SUM(E{first}:E{last})").font = F_CALC
    ws.cell(row=tr, column=6, value=f"=SUM(F{first}:F{last})").font = F_CALC
    ws.cell(row=tr, column=7, value=f"=SUM(G{first}:G{last})").font = F_CALC
    ws.cell(row=tr + 1, column=4, value="Required TCOLE hours (exact):").font = F_LABEL
    ws.cell(row=tr + 1, column=5, value=DC.REQUIRED_TCOLE_HOURS).font = F_CALC
    ws.cell(row=tr + 1, column=6, value=(
        "Report the BPOC at exactly 736; excess goes to Addendum course #101"
    )).font = F_SMALL
    define(wb, "nrCHtotalDelivered", "ChapterMaster", f"$G${tr}")
    # The per-chapter TCOLE minimums seeded above total 734, printed directly
    # above a cell asserting the course is exactly 736 - a 2-hour gap an
    # auditor sees on one screen. The gap is real and is NOT silently papered
    # over here: chapter A (Administrative/Departmental Overview) carries no
    # TCOLE minimum in this table, and the two unassigned hours are called
    # out in words so nobody has to reverse-engineer the subtraction.
    ws.merge_cells(start_row=tr + 2, start_column=4, end_row=tr + 2,
                   end_column=11)
    M = f"SUM(E{first}:E{last})"
    G = f"(N(cfgRequiredHours)-SUM(E{first}:E{last}))"
    gap = ws.cell(row=tr + 2, column=4, value=(
        f'=IF({G}=0,"Per-chapter TCOLE minimums total exactly "&'
        f'TEXT({M},"0.##")&" hrs - the full reported course length.",'
        f'"NOTE: the per-chapter TCOLE minimums above total "&'
        f'TEXT({M},"0.##")&" hrs, not the "&N(cfgRequiredHours)&" hrs the '
        f'BPOC is reported at. The "&TEXT({G},"0.##")&"-hr difference means a '
        f'line on the TCOLE hours table is missing from or miskeyed in this '
        f'column. Check it against the IRG hours table (Instructor Resource '
        f'Guide, \"0. Abstract\", final pages) - note that Module R, End of '
        f'Course Review (2 hrs), is listed there WITHOUT a chapter number and '
        f'is easy to drop. This is a gap in the MINIMUMS column only - it does '
        f'not change delivered hours or the exactly-"&N(cfgRequiredHours)&'
        f'"-hour report, but resolve it before submission.")'))
    gap.font = F_SMALL
    gap.alignment = A_LEFT_WRAP
    ws.row_dimensions[tr + 2].height = 44

    # ---- TPD sub-classes (scheduled separately, roll up to a chapter) ----
    sr = tr + 3
    section_bar(ws, sr, 2, 9, "TPD sub-classes — schedule by THESE names; "
                              "hours roll up to the parent TCOLE chapter")
    sr += 1
    header_row(ws, ["Sub-class", None, None, "Parent Ch #", "TPD Target Hrs",
                    "Delivered Hrs", "vs Target"], row=sr)
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=4)
    sr += 1
    sub_first = sr
    for name, parent, target in DC.SUBTOPICS:
        ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=4)
        c = ws.cell(row=sr, column=2, value=name)
        c.fill = FILL_INPUT
        c.font = F_INPUT
        c.border = BOX
        p = ws.cell(row=sr, column=5, value=parent)
        p.fill = FILL_INPUT
        p.font = F_INPUT
        p.border = BOX
        p.number_format = "@"       # parent chapter # — text, like nrCHnum
        t = ws.cell(row=sr, column=6, value=target)
        t.fill = FILL_INPUT
        t.font = F_INPUT
        t.border = BOX
        d = ws.cell(row=sr, column=7, value=(
            f'=IF($B{sr}="","",ROUND(SUMIFS(nrSCH_Hrs,nrSCH_Act,$B{sr},'
            f'nrSCH_TimeCheck,"OK"),2))'))
        d.font = F_CALC
        d.fill = FILL_CALC
        d.border = BOX
        v = ws.cell(row=sr, column=8, value=(
            f'=IF(OR($B{sr}="",$G{sr}=0),"",ROUND($G{sr}-$F{sr},2))'))
        v.font = F_CALC
        v.border = BOX
        sr += 1
    sub_last = sr + 4          # a few blank rows for future sub-classes
    for rr in range(sr, sub_last + 1):
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        for col in (2, 5, 6):
            cc = ws.cell(row=rr, column=col)
            cc.fill = FILL_INPUT
            cc.font = F_INPUT
            cc.border = BOX
            if col == 5:            # parent chapter # — text, like nrCHnum
                cc.number_format = "@"
        ws.cell(row=rr, column=7, value=(
            f'=IF($B{rr}="","",ROUND(SUMIFS(nrSCH_Hrs,nrSCH_Act,$B{rr},'
            f'nrSCH_TimeCheck,"OK"),2))'
        )).font = F_CALC
        ws.cell(row=rr, column=8, value=(
            f'=IF(OR($B{rr}="",$G{rr}=0),"",ROUND($G{rr}-$F{rr},2))'
        )).font = F_CALC
    dv_list(ws, "=nrCHnum", [f"E{sub_first}:E{sub_last}"])
    cf_formula(ws, f"H{sub_first}:H{sub_last}",
               f'AND($H{sub_first}<>"",$H{sub_first}<0)', FILL_WARNBG)
    define(wb, "nrSUBname", "ChapterMaster", f"$B${sub_first}:$B${sub_last}")
    define(wb, "nrSUBparent", "ChapterMaster", f"$E${sub_first}:$E${sub_last}")
    define(wb, "nrSUBtarget", "ChapterMaster", f"$F${sub_first}:$F${sub_last}")

    col_widths(ws, {"A": 3, "B": 8, "C": 6, "D": 44, "E": 10, "F": 10,
                    "G": 11, "H": 10, "I": 12, "J": 12, "K": 20, "L": 14,
                    "M": 13, "N": 13, "O": 12, "P": 12, "Q": 13, "R": 15,
                    "S": 46, "T": 13, "U": 12})
    sheet_note(ws, "Per the IRG, every chapter's training file needs: SME "
                   "lesson plan (a PowerPoint alone does NOT count), "
                   "instructor bio, original sign-in sheets w/ PID, "
                   "assessment + grade sheet, and course evaluation. "
                   "'File Complete?' turns Yes when all are On File/NA and "
                   "any special TCOLE requirement is met.")
    return ws


# --------------------------------------------------------------------------
def build_exammaster(wb):
    ws = wb.create_sheet("ExamMaster")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["ExamCode", "Exam Name", "Default Type", "Default Passing",
                    "Default Seq"])
    r = DATA_ROW
    for code, name, typ, passing, seq in DL.EXAMS:
        for col, v in ((2, code), (3, name), (4, typ), (5, passing), (6, seq)):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = FILL_INPUT
            c.font = F_INPUT
            c.border = BOX
        r += 1
    last = r + 6  # room to add exams
    fill_rows(ws, r, last, {c: (None, "in") for c in "BCDEF"})
    dv_list(ws, "=lstExamType", [f"D{DATA_ROW}:D{last}"])
    define(wb, "rngEMcode", "ExamMaster", f"$B${DATA_ROW}:$B${last}")
    define(wb, "rngEMname", "ExamMaster", f"$C${DATA_ROW}:$C${last}")
    define(wb, "rngEMtype", "ExamMaster", f"$D${DATA_ROW}:$D${last}")
    define(wb, "rngEMpass", "ExamMaster", f"$E${DATA_ROW}:$E${last}")
    define(wb, "rngEMseq", "ExamMaster", f"$F${DATA_ROW}:$F${last}")
    col_widths(ws, {"A": 3, "B": 11, "C": 52, "D": 13, "E": 14, "F": 12})
    sheet_note(ws, "Library of all possible exams. ExamPlan picks which ones "
                   "THIS academy uses.")
    return ws


def build_examplan(wb):
    ws = wb.create_sheet("ExamPlan")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Use?", "ExamCode", "Exam Name", "Type", "Passing", "Seq",
                    "Linked Ch #", "Exam Date"])
    first, last = DATA_ROW, DATA_ROW + 24
    exam_ch = {"E01": "1", "E02": "7", "E03": "8", "E04": "9", "E05": "10",
               "E06": "15", "E07": "17", "E08": "20", "E09": "22",
               "E10": "23", "E11": "25", "E12": "26", "E13": "28",
               "E14": "29", "E15": "32", "E16": "34", "E17": "40",
               "FIN": ""}
    r = first
    for code, name, typ, passing, seq in DL.EXAMS:
        ws.cell(row=r, column=2, value="Yes")
        ws.cell(row=r, column=3, value=code)
        ws.cell(row=r, column=8, value=exam_ch.get(code, ""))
        r += 1
    fill_rows(ws, first, last, {
        "B": (None, "in"), "C": (None, "in"),
        "D": ('IF($C{r}="","",IFERROR(INDEX(rngEMname,MATCH($C{r},rngEMcode,0)),"?"))', "fx"),
        "E": ('IF($C{r}="","",IFERROR(INDEX(rngEMtype,MATCH($C{r},rngEMcode,0)),"?"))', "fx"),
        "F": ('IF($C{r}="","",IFERROR(INDEX(rngEMpass,MATCH($C{r},rngEMcode,0)),cfgPassingScore))', "fx"),
        "G": ('IF($C{r}="","",IFERROR(INDEX(rngEMseq,MATCH($C{r},rngEMcode,0)),""))', "fx"),
        "H": (None, "in"),
        # "Exam Date" used to be the LAST day the linked chapter was taught,
        # which disagreed with the date the exam was actually given (for E09
        # by three days) on an auditor-facing page. Now: the date the exam
        # was actually taken (earliest dated first attempt) when scores
        # exist, else the scheduled "Test n" / "Final Test" block, else the
        # chapter's last taught day as the estimate it always was.
        "I": ('IF($C{r}="","",'
              'IF(IFERROR(MINIFS(nrES_Date,nrES_Code,$C{r},nrES_Att,1,'
              'nrES_Date,">0"),0)>0,'
              'MINIFS(nrES_Date,nrES_Code,$C{r},nrES_Att,1,nrES_Date,">0"),'
              'IF(AND($G{r}<>"",IFERROR(MAXIFS(nrSCH_Date,nrSCH_Act,'
              '"Test "&$G{r}),0)>0),'
              'MAXIFS(nrSCH_Date,nrSCH_Act,"Test "&$G{r}),'
              'IF(AND($H{r}<>"",IFERROR(MAXIFS(nrSCH_Date,nrSCH_ChNum,'
              '$H{r}),0)>0),MAXIFS(nrSCH_Date,nrSCH_ChNum,$H{r}),""))))', "fx"),
    })
    for r2 in range(first, last + 1):
        ws[f"I{r2}"].number_format = DATE
        ws[f"H{r2}"].number_format = "@"   # chapter key is TEXT (nrCHnum)
    dv_list(ws, "=lstYesNo", [f"B{first}:B{last}"])
    dv_list(ws, "=rngEMcode", [f"C{first}:C{last}"])
    dv_list(ws, "=nrCHnum", [f"H{first}:H{last}"])
    define(wb, "rngEPuse", "ExamPlan", f"$B${first}:$B${last}")
    define(wb, "rngEPcode", "ExamPlan", f"$C${first}:$C${last}")
    define(wb, "rngEPname", "ExamPlan", f"$D${first}:$D${last}")
    define(wb, "rngEPtype", "ExamPlan", f"$E${first}:$E${last}")
    define(wb, "rngEPpass", "ExamPlan", f"$F${first}:$F${last}")
    define(wb, "rngEPseq", "ExamPlan", f"$G${first}:$G${last}")
    define(wb, "rngEPch", "ExamPlan", f"$H${first}:$H${last}")
    define(wb, "rngEPdate", "ExamPlan", f"$I${first}:$I${last}")
    col_widths(ws, {"A": 3, "B": 7, "C": 11, "D": 52, "E": 10, "F": 10,
                    "G": 7, "H": 11, "I": 12})
    sheet_note(ws, "Linked Ch # ties each exam to its chapter. 'Exam Date' "
                   "is the date the exam was actually given (earliest dated "
                   "first attempt on ExamScores); before any score is "
                   "entered it falls back to the scheduled 'Test n' block, "
                   "and failing that to the last scheduled day of the linked "
                   "chapter. Retest deadlines always run off the actual "
                   "score dates, never off this column.")
    return ws


def build_skillsmaster(wb):
    ws = wb.create_sheet("SkillsMaster")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Category", "Max Attempts", "Scoring Mode", "Passing Score"])
    first = DATA_ROW
    r = first
    for cat, mx, mode, passing in DL.SKILLS:
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=mx)
        ws.cell(row=r, column=4, value=mode)
        if passing is not None:
            ws.cell(row=r, column=5, value=passing)
        r += 1
    last = r + 3
    fill_rows(ws, first, last, {c: (None, "in") for c in "BCDE"})
    dv_list(ws, "=lstScoringMode", [f"D{first}:D{last}"])
    define(wb, "rngSM_cat", "SkillsMaster", f"$B${first}:$B${last}")
    define(wb, "rngSM_max", "SkillsMaster", f"$C${first}:$C${last}")
    define(wb, "rngSM_mode", "SkillsMaster", f"$D${first}:$D${last}")
    define(wb, "rngSM_pass", "SkillsMaster", f"$E${first}:$E${last}")
    col_widths(ws, {"A": 3, "B": 18, "C": 14, "D": 14, "E": 14})
    sheet_note(ws, "Attempt limits per policy 300.7/600: exhausting attempts "
                   "= separation. Firearms scored; others pass/fail.")
    return ws


# --------------------------------------------------------------------------
def build_spellingmaster(wb):
    ws = wb.create_sheet("SpellingMaster")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Word #"] + [f"Test {n}" for n in range(1, 13)])
    first = DATA_ROW
    for w in range(25):
        ws.cell(row=first + w, column=2, value=w + 1).font = F_LABEL
        for t in range(1, 13):
            c = ws.cell(row=first + w, column=2 + t,
                        value=DS.TESTS[t][w])
            c.fill = FILL_INPUT
            c.font = F_INPUT
            c.border = BOX
    last = first + 24
    define(wb, "nrSpellWords", "SpellingMaster", f"$C${first}:$N${last}")
    r = last + 2
    ws.cell(row=r, column=2, value="Each word is worth 4 points "
            "(25 words x 4 = 100).").font = F_SMALL
    col_widths(ws, {"A": 3, "B": 8})
    for t in range(1, 13):
        ws.column_dimensions[get_column_letter(2 + t)].width = 16
    sheet_note(ws, "Word lists print from Print Center (test sheet or key). "
                   "Edit words here to change future academies.")
    return ws


# --------------------------------------------------------------------------
def build_writingmaster(wb):
    ws = wb.create_sheet("WritingMaster")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["#", "Assignment Title", "Linked Ch #", "Assign Delay (wks)",
                    "Due Offset (wks)", "Computed Assigned", "Computed Due",
                    "Override Assigned", "Override Due", "Assigned", "Due",
                    "Prompt"])
    first = DATA_ROW
    r = first
    for num, title, ch, delay, off, prompt in DW.ASSIGNMENTS:
        ws.cell(row=r, column=2, value=num)
        ws.cell(row=r, column=3, value=title)
        ws.cell(row=r, column=4, value=ch)
        ws.cell(row=r, column=5, value=delay)
        ws.cell(row=r, column=6, value=off)
        pc = ws.cell(row=r, column=13, value=prompt)
        pc.alignment = A_LEFT_WRAP
        r += 1
    last = r - 1
    # Computed Assigned: first class day of (chapter-first-taught week + delay)
    fill_rows(ws, first, last, {
        "B": (None, "in"), "C": (None, "in"), "D": (None, "in"),
        "E": (None, "in"), "F": (None, "in"),
        # MINIFS returns 0 (not an error) when the assign-week start falls
        # past the last generated class day, so IFERROR never fired and G -
        # and the Assigned column K that reads it - printed 12/30/1899 on the
        # writing handout. Zero-guarded exactly like its sibling H below.
        "G": ('IF($D{r}="","",LET(f,IFERROR(XLOOKUP($D{r},nrCHnum,nrCHfirst),""),'
              'IF(OR(f="",f=0),"",LET(wkstart,f-WEEKDAY(f,2)+1+$E{r}*7,'
              'd,IFERROR(MINIFS(nrCDdate,nrCDdate,">="&wkstart),0),'
              'IF(N(d)=0,"",d)))))', "fx"),
        "H": ('IF($G{r}="","",LET(a,$G{r},ws2,a-WEEKDAY(a,2)+1+$F{r}*7,'
              'd,MAXIFS(nrCDdate,nrCDdate,">="&ws2,nrCDdate,"<"&(ws2+7)),'
              'IF(d=0,"",d)))', "fx"),
        "I": (None, "in"), "J": (None, "in"),
        "K": ('IF($B{r}="","",IF($I{r}<>"",$I{r},$G{r}))', "fx"),
        "L": ('IF($B{r}="","",IF($J{r}<>"",$J{r},$H{r}))', "fx"),
        "M": (None, "in"),
    })
    for r2 in range(first, last + 1):
        for cl in "GHIJKL":
            ws[f"{cl}{r2}"].number_format = DATE
        ws[f"D{r2}"].number_format = "@"   # chapter key is TEXT (nrCHnum)
        ws[f"M{r2}"].alignment = A_LEFT_WRAP
        ws.row_dimensions[r2].height = 42
    dv_list(ws, "=nrCHnum", [f"D{first}:D{last}"])
    define(wb, "rngWMnum", "WritingMaster", f"$B${first}:$B${last}")
    define(wb, "rngWMtitle", "WritingMaster", f"$C${first}:$C${last}")
    define(wb, "rngWMassigned", "WritingMaster", f"$K${first}:$K${last}")
    define(wb, "rngWMdue", "WritingMaster", f"$L${first}:$L${last}")
    define(wb, "rngWMprompt", "WritingMaster", f"$M${first}:$M${last}")
    nr = last + 2
    ws.cell(row=nr, column=2, value="Requirements:").font = F_LABEL
    c = ws.cell(row=nr, column=3, value=DW.REQUIREMENTS_NOTE)
    c.font = F_SMALL
    c.alignment = A_LEFT_WRAP
    ws.merge_cells(start_row=nr, start_column=3, end_row=nr, end_column=12)
    col_widths(ws, {"A": 3, "B": 5, "C": 38, "D": 10, "E": 12, "F": 12,
                    "G": 14, "H": 14, "I": 14, "J": 14, "K": 12, "L": 12,
                    "M": 80})
    sheet_note(ws, "Dates COMPUTE from the Schedule: assigned = first class "
                   "day of the week the linked chapter is first taught (+ "
                   "delay); due = last class day of the due week, 1700. "
                   "Override columns always win.")
    return ws


# --------------------------------------------------------------------------
def build_control(wb):
    """Holidays + generated class-day calendar."""
    ws = wb.create_sheet("Control")
    ws.sheet_view.showGridLines = False
    # holidays B:D
    header_row(ws, ["Holiday", "Observed (Yr 1)", "Observed (Yr 2)", None,
                    "Extra Closure Dates", "Closure Check", "Day #",
                    "Class Date", "Week #", "In Session?"])
    first = DATA_ROW
    # ---- Easter / Good Friday computus, computed in plain arithmetic ----
    # Good Friday was the workbook's only LET-dependent holiday. A single
    # errored holiday poisons nrAllClosures and silently blanks ALL 170
    # class days (and therefore every retest/memo/writing deadline), so the
    # computus is done stepwise in helper cells that any spreadsheet engine
    # can evaluate. P..AF, rows 6-7 (year 1 / year 2).
    # the warning lives one row ABOVE the step labels: written into P5 it was
    # erased a moment later by the loop below, so the 34 live formula cells in
    # P6:AF7 shipped with no label and no warning on an unprotected sheet
    ws.cell(row=HDR_ROW - 1, column=16,
            value="Easter computus (helper - do not edit)").font = F_SMALL
    # a blank / non-date Start Date is the one input that can silently empty
    # this whole sheet; say so where the empty calendar is being looked at
    warn = ws.cell(row=HDR_ROW - 2, column=8, value=(
        '=IF(NOT(ISNUMBER(cfgStartDate)),'
        '"START DATE on Settings is blank or is not a real date - no class '
        'days can be generated, so every retest deadline, memo due date and '
        'writing date is blank until it is fixed.",'
        'IF(NOT(ISNUMBER(cfgEndDate)),'
        '"END DATE on Settings is blank or is not a real date - the '
        '\'In Session?\' column cannot be computed.",""))'))
    warn.font = F_LABEL
    steps = [
        ("Y", "{Y}"),
        ("a", "MOD($P{r},19)"),
        ("b", "INT($P{r}/100)"),
        ("c", "MOD($P{r},100)"),
        ("d", "INT($R{r}/4)"),
        ("e", "MOD($R{r},4)"),
        ("f", "INT(($R{r}+8)/25)"),
        ("g", "INT(($R{r}-$V{r}+1)/3)"),
        ("h", "MOD(19*$Q{r}+$R{r}-$T{r}-$W{r}+15,30)"),
        ("i", "INT($S{r}/4)"),
        ("k", "MOD($S{r},4)"),
        ("l", "MOD(32+2*$U{r}+2*$Y{r}-$X{r}-$Z{r},7)"),
        ("m", "INT(($Q{r}+11*$X{r}+22*$AA{r})/451)"),
        ("mth", "INT(($X{r}+$AA{r}-7*$AB{r}+114)/31)"),
        ("day", "MOD($X{r}+$AA{r}-7*$AB{r}+114,31)+1"),
        ("Easter", "DATE($P{r},$AC{r},$AD{r})"),
        ("GoodFri", "$AE{r}-2"),
    ]
    for yr_off, hr in ((0, first), (1, first + 1)):
        ybase = "YEAR(cfgStartDate)" + ("+1" if yr_off else "")
        for i, (lab, tmpl) in enumerate(steps):
            col = 16 + i
            lc = ws.cell(row=HDR_ROW, column=col, value=lab)
            lc.font = F_SMALL
            hc = ws.cell(row=hr, column=col)
            hc.value = "=" + tmpl.format(Y=ybase, r=hr)
            hc.font = F_SMALL
            hc.fill = FILL_CALC          # reads as an engine cell, not input
            if lab in ("Easter", "GoodFri"):
                hc.number_format = DATE
    GOODFRI = {first: "$AF$6", first + 1: "$AF$7"}

    r = first
    for name, tmpl in DL.HOLIDAYS:
        ws.cell(row=r, column=2, value=name).font = F_LABEL
        if name == "Good Friday":
            f1, f2 = "=" + GOODFRI[first], "=" + GOODFRI[first + 1]
        else:
            f1 = "=" + tmpl.format(Y="YEAR(cfgStartDate)")
            f2 = "=" + tmpl.format(Y="YEAR(cfgStartDate)+1")
        c1 = ws.cell(row=r, column=3, value=f1)
        c2 = ws.cell(row=r, column=4, value=f2)
        for c in (c1, c2):
            c.number_format = DATE
            c.font = F_CALC
            c.fill = FILL_CALC
            c.border = BOX
        r += 1
    hol_last = r - 1
    define(wb, "nrHolidays1", "Control", f"$C${first}:$C${hol_last}")
    define(wb, "nrHolidays2", "Control", f"$D${first}:$D${hol_last}")
    # extra closures (manual)
    extra_last = first + 14
    # column G is the Row Check the extra-closure block never had: a value
    # that is not a real date is DISCARDED by the calendar (see the ISNUMBER
    # guard on column M below), so it has to be visible here rather than
    # silently changing the class calendar. The dropdown-free date validation
    # blocks typing one; Row Check catches the pastes validation lets through.
    fill_rows(ws, first, extra_last, {
        "F": (None, "in"),
        "G": ('IF($F{r}="","",IF(NOT(ISNUMBER($F{r})),'
              '"NOT A DATE - this closure is IGNORED",'
              'IF(COUNTIF($F$%d:$F$%d,$F{r})>1,"DUPLICATE closure date",'
              'IF(NOT(ISNUMBER(cfgStartDate)),"OK",'
              'IF(OR($F{r}<cfgStartDate,$F{r}>N(cfgEndDate)+90),'
              '"OK (outside this academy - no effect)",'
              'IF(WEEKDAY($F{r},2)>5,"OK (weekend - already closed)",'
              '"OK"))))))' % (first, extra_last), "fx"),
    })
    for r2 in range(first, extra_last + 1):
        ws[f"F{r2}"].number_format = DATE
    from openpyxl.worksheet.datavalidation import DataValidation as _DV
    _dvd = _DV(type="date", operator="between",
               formula1="DATE(1900,1,1)", formula2="DATE(2199,12,31)",
               allow_blank=True, showErrorMessage=True)
    _dvd.errorTitle = "Extra Closure Date"
    _dvd.error = ("Enter a real date (e.g. 12/23/2026). Text is ignored by "
                  "the class-day calendar and the Closure Check column will "
                  "say so.")
    _dvd.add(f"F{first}:F{extra_last}")
    ws.add_data_validation(_dvd)
    cf_formula(ws, f"G{first}:G{extra_last}",
               f'AND($G{first}<>"",LEFT($G{first},2)<>"OK")', FILL_WARNBG)
    define(wb, "nrExtraClosures", "Control", f"$F${first}:$F${extra_last}")
    define(wb, "nrExtraClosureCheck", "Control", f"$G${first}:$G${extra_last}")
    # class-day calendar H:K
    cd_last = first + ROWS_CLASSDAYS - 1
    fill_rows(ws, first, cd_last, {
        "H": (('v', None), "fx"),
        # the START argument is guarded too, not just the holidays list: a
        # blank or non-date Start Date made cfgStartDate-1 evaluate to -1 (or
        # #VALUE!) and turned all 170 class days into errors, poisoning every
        # retest deadline, memo due date, writing date and sign-in sheet
        # downstream with nothing on Settings to say why. A bad Start Date now
        # leaves the calendar BLANK and says so in the banner on row 3 and in
        # the Settings "Check" column beside Start Date.
        "I": ('IF(OR($H{r}="",NOT(ISNUMBER(cfgStartDate))),"",'
              'IFERROR(WORKDAY.INTL(cfgStartDate-1,$H{r},"0000011",'
              'nrAllClosures),""))', "fx"),
        "J": ('IF(OR($H{r}="",$I{r}=""),"",'
              'INT(($I{r}-(cfgStartDate-WEEKDAY(cfgStartDate,2)+1))/7)+1)', "fx"),
        "K": ('IF(OR($H{r}="",$I{r}=""),"",IF($I{r}<=cfgEndDate,"Yes","No"))', "fx"),
    })
    for i, r2 in enumerate(range(first, cd_last + 1)):
        ws[f"H{r2}"].value = i + 1
        ws[f"I{r2}"].number_format = DATE
    # combined closures name (holidays yr1+yr2+extra) via helper column M
    ws.cell(row=HDR_ROW, column=13, value="AllClosures").font = F_SMALL
    n_h = hol_last - first + 1
    # every closure is IFERROR-guarded: WORKDAY.INTL fails outright if ANY
    # holiday cell is an error or text, which would silently blank all
    # class days (and every deadline computed from them). A bad holiday now
    # degrades to a harmless sentinel instead of taking the calendar down.
    SENT = "DATE(1900,1,1)"
    for i in range(n_h):
        ws.cell(row=first + i, column=13,
                value=f"=IFERROR(N(C{first+i})+0,{SENT})").number_format = DATE
        ws.cell(row=first + n_h + i, column=13,
                value=f"=IFERROR(N(D{first+i})+0,{SENT})").number_format = DATE
    # N() coerces TEXT to 0 without raising, so a pasted "12/23/2026" or a
    # label like "Thanksgiving week" used to be swallowed silently: the
    # closure simply never happened, IFERROR never fired, and the coordinator
    # got a wrong class calendar with no error, no flag and no audit line.
    # ISNUMBER keeps the sentinel for anything that is not a real date, and
    # the Check column beside the input says exactly what was rejected.
    for i in range(15):
        ws.cell(row=first + 2 * n_h + i, column=13,
                value=f'=IF(ISNUMBER(F{first+i}),F{first+i},{SENT})'
                ).number_format = DATE
    all_last = first + 2 * n_h + 14
    define(wb, "nrAllClosures", "Control", f"$M${first}:$M${all_last}")
    define(wb, "nrCDnum", "Control", f"$H${first}:$H${cd_last}")
    define(wb, "nrCDdate", "Control", f"$I${first}:$I${cd_last}")
    define(wb, "nrCDweek", "Control", f"$J${first}:$J${cd_last}")
    define(wb, "nrCDinsession", "Control", f"$K${first}:$K${cd_last}")
    col_widths(ws, {"A": 3, "B": 26, "C": 15, "D": 15, "E": 2, "F": 16,
                    "G": 34, "H": 8, "I": 13, "J": 8, "K": 11, "L": 2,
                    "M": 12})
    sheet_note(ws, "Class days = Mon-Fri from Start Date, skipping observed "
                   "holidays and any Extra Closure Dates you add. Retest "
                   "deadlines, writing dates and sign-in sheets all key off "
                   "this calendar. An Extra Closure Date must be a REAL date "
                   "- anything else is ignored by the calendar, and the "
                   "Closure Check column beside it says so. If Start Date on "
                   "Settings is blank or not a date, this whole table is "
                   "blank on purpose (see the red line above).")
    return ws


# --------------------------------------------------------------------------
def build_schedule(wb):
    ws = wb.create_sheet("Schedule")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Date", "Day", "Start", "End", "Hours",
                    "Chapter / Activity", "Ch #", "Instructor", "Location",
                    "Week #", "Day #", "Notes", "Instructor OK?",
                    "Time Check", "Bank Check"])
    first, last = DATA_ROW, DATA_ROW + ROWS_SCHEDULE - 1
    fill_rows(ws, first, last, {
        "B": (None, "in"),
        "C": ('IF($B{r}="","",TEXT($B{r},"ddd"))', "fx"),
        "D": (None, "in"), "E": (None, "in"),
        # MOD(...,1): a block that runs to or past midnight (22:00->02:00)
        # used to compute as NEGATIVE hours, which no error, no row check and
        # no audit line caught — and those negatives were SUBTRACTED from
        # ChapterMaster "Delivered Hrs" and from the Settings schedule-minutes
        # detector. Column O ("Time Check") catches the swapped-time typo
        # that MOD would otherwise turn into a plausible 22-hour block.
        "F": ('IF(OR($B{r}="",$D{r}="",$E{r}=""),"",'
              'ROUND(MOD($E{r}-$D{r},1)*24,2))', "fx"),
        "G": (None, "in"),
        "H": ('IF($G{r}="","",IFERROR(XLOOKUP($G{r},nrCHname,nrCHnum),'
              'IFERROR(XLOOKUP($G{r},nrSUBname,nrSUBparent),"")))', "fx"),
        "I": (None, "in"), "J": (None, "in"),
        "K": ('IF($B{r}="","",IFERROR(XLOOKUP($B{r},nrCDdate,nrCDweek),""))', "fx"),
        "L": ('IF($B{r}="","",IFERROR(XLOOKUP($B{r},nrCDdate,nrCDnum),""))', "fx"),
        "M": (None, "in"),
        # (nrInstrNames<>"") guard: SEARCH("",text) returns 1, so blank
        # roster rows would otherwise match everything and hide typos.
        # ONE roster match is not enough: the pick-to-append macro builds
        # comma-separated co-instructor lists, so an any-match let an
        # off-roster co-teacher ride along inside an otherwise valid cell —
        # invisible to this check AND to the "instructors lacking
        # documentation" audit line, which only walks the roster. Every
        # comma-separated token must resolve. Parenthetical annotations
        # ("Judson Moore (+ Cadre)", "(Ch 41, 42)") are stripped first so
        # a comma inside one cannot false-flag the row.
        # ...and the roster match must be counted against the SAME stripped
        # text the token count is taken from. Counting matches in the raw
        # cell while counting tokens in the stripped one let every roster
        # name sitting inside a parenthetical ("Cadre" in the house-style
        # "(+ Cadre)") buy one free off-roster name in the main list - and
        # "(+ Cadre), <name>" is exactly what the pick-to-append macro
        # writes. Separators other than the comma are normalised first: the
        # schedule's own blocks join co-instructors with " & " and "/", and
        # no roster name or guest entity contains either character, so a
        # pair joined that way needed only ONE roster match to read OK.
        "N": ('IF(OR($B{r}="",$I{r}=""),"",'
              'LET(a,$I{r},'
              'b0,IF(ISNUMBER(SEARCH("(",a)),TRIM(LEFT(a,SEARCH("(",a)-1)&'
              'IF(ISNUMBER(SEARCH(")",a)),MID(a,SEARCH(")",a)+1,255),"")),a),'
              'b,TRIM(SUBSTITUTE(SUBSTITUTE(b0,"&",","),"/",",")),'
              'n,LEN(b)-LEN(SUBSTITUTE(b,",",""))+1,'
              'm,SUMPRODUCT((nrInstrNames<>"")*ISNUMBER(SEARCH(nrInstrNames,b))),'
              'IF(m>=n,"OK","UNRECOGNIZED")))', "fx"),
        # a legitimate block that crosses midnight is short under MOD; a
        # swapped start/end reads as most of a day. 12 hours separates them.
        "O": ('IF(OR($B{r}="",$D{r}="",$E{r}=""),"",'
              'IF($E{r}=$D{r},"CHECK TIMES (start = end)",'
              'IF($F{r}>12,"CHECK TIMES (End is before Start?)","OK")))', "fx"),
        # The second instructor question, kept in its OWN column and
        # deliberately LET-free: N answers "does this text name anybody on
        # the roster", P answers "is every name it resolves to inside THAT
        # TOPIC's certified bank". InstructorBanks asserted who may teach
        # what and nothing enforced it - the Schedule dropdown only suggests
        # (warning-only), and a multi-name cell bypasses validation entirely.
        # Enforcement is skipped when the topic has no bank row or its bank
        # is empty: nothing has been asserted, so nothing can be violated.
        "P": ('IF(OR($B{r}="",$I{r}=""),"",'
              'IF(IFERROR(MATCH($G{r},nrBankTopics,0),0)=0,"OK",'
              'IF(COUNTA(INDEX(nrBankGrid,'
              'IFERROR(MATCH($G{r},nrBankTopics,0),1),0))=0,"OK",'
              'IF(SUMPRODUCT((nrInstrNames<>"")*'
              'ISNUMBER(SEARCH(nrInstrNames,$I{r}))*'
              '(COUNTIF(INDEX(nrBankGrid,'
              'IFERROR(MATCH($G{r},nrBankTopics,0),1),0),nrInstrNames)>0))'
              '<SUMPRODUCT((nrInstrNames<>"")*'
              'ISNUMBER(SEARCH(nrInstrNames,$I{r}))),'
              '"NOT IN BANK","OK"))))', "fx"),
    })
    for r2 in range(first, last + 1):
        ws[f"B{r2}"].number_format = DATE
        ws[f"D{r2}"].number_format = TIME
        ws[f"E{r2}"].number_format = TIME
    dv_list(ws, "=nrScheduleItems", [f"G{first}:G{last}"])
    # instructor dropdown depends on the row's topic: offers that topic's
    # "Teaching this academy" picks from InstructorBanks; topics without a
    # bank row fall back to the full roster. Warning-only so multi-name
    # combined entries (built by the pick-to-append macro) stay valid.
    # the roster fallback must also fire when the topic HAS a bank row but
    # nobody has been picked for it yet (the state of every topic in a fresh
    # template and after every New Academy Reset, which clears M6:T105) —
    # otherwise the dropdown is empty for all 65 real class topics.
    dv_list(ws, f"=IF(OR(IFERROR(MATCH($G{first},nrBankTopics,0),0)=0,"
                f"COUNTA(INDEX(nrBankSel,"
                f"IFERROR(MATCH($G{first},nrBankTopics,0),1),0))=0),"
                f"nrInstrNames,INDEX(nrBankSel,"
                f"IFERROR(MATCH($G{first},nrBankTopics,0),1),0))",
            [f"I{first}:I{last}"], enforce=False)
    define(wb, "nrSCH_Date", "Schedule", f"$B${first}:$B${last}")
    define(wb, "nrSCH_Start", "Schedule", f"$D${first}:$D${last}")
    define(wb, "nrSCH_End", "Schedule", f"$E${first}:$E${last}")
    define(wb, "nrSCH_Hrs", "Schedule", f"$F${first}:$F${last}")
    define(wb, "nrSCH_Act", "Schedule", f"$G${first}:$G${last}")
    define(wb, "nrSCH_ChNum", "Schedule", f"$H${first}:$H${last}")
    define(wb, "nrSCH_Instr", "Schedule", f"$I${first}:$I${last}")
    define(wb, "nrSCH_Loc", "Schedule", f"$J${first}:$J${last}")
    define(wb, "nrSCH_InstrOK", "Schedule", f"$N${first}:$N${last}")
    define(wb, "nrSCH_BankOK", "Schedule", f"$P${first}:$P${last}")
    define(wb, "nrSCH_TimeCheck", "Schedule", f"$O${first}:$O${last}")
    cf_formula(ws, f"N{first}:N{last}", f'$N{first}="UNRECOGNIZED"',
               FILL_WARNBG)
    cf_formula(ws, f"P{first}:P{last}", f'$P{first}="NOT IN BANK"',
               FILL_WARNBG)
    cf_formula(ws, f"O{first}:O{last}",
               f'AND($O{first}<>"",$O{first}<>"OK")', FILL_WARNBG)
    col_widths(ws, {"A": 3, "B": 12, "C": 6, "D": 9, "E": 9, "F": 8,
                    "G": 46, "H": 6, "I": 24, "J": 16, "K": 8, "L": 7,
                    "M": 30, "N": 15, "O": 30, "P": 15})
    # full grid, and out to O: the print area used to stop at M406, silently
    # dropping every block past row 406 and the whole "Instructor OK?" column
    # that was appended later. The trailing rows DO cost pages - fitToHeight
    # is 0, which means "as many pages tall as needed", not "one page" - and
    # all 900 rows carry formulas, borders and fill, so Excel paginates every
    # one of them. modPrint.btnPrintSchedule narrows the area to the rows
    # actually used before it prints and restores this one afterwards; this
    # static area stays full-height so a manual File > Print can never clip a
    # block.
    page_setup_landscape(ws, print_area=f"B{HDR_ROW}:P{last}",
                         repeat_rows=f"{HDR_ROW}:{HDR_ROW}")
    sheet_note(ws, "One row per time block (a day usually has several). "
                   "Hours, chapter reconciliation, sign-in sheets, writing "
                   "dates and eval headers all read from here. A block may "
                   "run past midnight (22:00-02:00 counts as 4.00 hrs); "
                   "Time Check flags a block whose End is at or before its "
                   "Start, because those hours move chapter totals and the "
                   "academy length. Printable as-is (File > Print).")
    return ws


def build_schedule_items_helper(wb):
    """Hidden helper: chapter names + activities as one dropdown list."""
    ws = wb.create_sheet("sysListsHelper")
    ws.sheet_view.showGridLines = False
    r = DATA_ROW
    ws.cell(row=HDR_ROW, column=2, value="Schedule dropdown items").font = F_LABEL
    n_ch = len(DC.CHAPTERS)
    for i in range(n_ch):
        ws.cell(row=r + i, column=2,
                value=f"=IF(ChapterMaster!D{DATA_ROW+i}=\"\",\"\","
                      f"ChapterMaster!D{DATA_ROW+i})")
    n_sub = len(DC.SUBTOPICS)
    for j, (name, _p, _t) in enumerate(DC.SUBTOPICS):
        ws.cell(row=r + n_ch + j, column=2, value=name)
    for j, act in enumerate(DC.ACTIVITIES):
        ws.cell(row=r + n_ch + n_sub + j, column=2, value=act)
    last = r + n_ch + n_sub + len(DC.ACTIVITIES) - 1
    define(wb, "nrScheduleItems", "sysListsHelper", f"$B${r}:$B${last}")
    # all linkable record IDs (Incidents I###, Attendance A###, Counseling
    # C###) stacked for the Memos "Linked Ref" dropdown
    ws.cell(row=HDR_ROW, column=4, value="Linkable record IDs").font = F_LABEL
    # formula-produced "" strings are NOT blank, so TOCOL(...,1) alone
    # would leave hundreds of empty entries in the dropdown — strip them
    ws.cell(row=DATA_ROW, column=4, value=(
        '=IFERROR(LET(ids,TOCOL(VSTACK(Incidents!$B$6:$B$405,'
        'Attendance!$B$6:$B$805,Counseling!$B$6:$B$405),1),'
        'FILTER(ids,ids<>"")),"")'))
    # 400 + 800 + 400 = 1600 rows maximum; the range used to be 16 rows
    # longer than the spill it wraps could ever be
    define(wb, "nrAllRefIDs", "sysListsHelper", f"$D${DATA_ROW}:$D${DATA_ROW+1599}")
    # Attendance EventIDs only, for the Makeup "Linked Event" dropdown. The
    # dropdown used to point straight at nrAT_ID (Attendance!$B$6:$B$805),
    # whose 800 cells are all FORMULAS: an unused row emits "" and a
    # formula-produced "" is NOT blank, so the picker listed ~795 empty
    # entries. Same strip-the-blanks pattern as nrAllRefIDs above.
    ws.cell(row=HDR_ROW, column=5, value="Attendance EventIDs").font = F_LABEL
    ws.cell(row=DATA_ROW, column=5, value=(
        '=IFERROR(FILTER(Attendance!$B$6:$B$805,'
        'Attendance!$B$6:$B$805<>""),"")'))
    define(wb, "nrAT_IDlist", "sysListsHelper", f"$E${DATA_ROW}:$E${DATA_ROW+799}")
    # locked and veryHidden like the rest of the engine: D6 is a dynamic
    # array (up to 1,600 record IDs) and column B generates the Schedule
    # dropdown, so one stray value typed in the spill zone kills both
    # dropdowns workbook-wide. "hidden" alone could be undone from the
    # right-click tab menu with no password.
    protect(ws)
    ws.sheet_state = "veryHidden"
    return ws


def build_all_config(wb):
    build_starthere(wb)
    build_settings(wb)
    build_lists(wb)
    build_agencies(wb)
    build_instructors(wb)
    build_instructorbanks(wb)
    build_chaptermaster(wb)
    build_exammaster(wb)
    build_examplan(wb)
    build_skillsmaster(wb)
    build_spellingmaster(wb)
    build_writingmaster(wb)
    build_control(wb)
    build_schedule_items_helper(wb)
    build_schedule(wb)
