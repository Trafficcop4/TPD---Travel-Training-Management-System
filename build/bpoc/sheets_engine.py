"""Engine sheets (locked, password-protected): sysGrades, sysAttendance,
sysSkills, sysIncidents, sysPT, sysFlags, sysChecks, sysAwards, sysAudit.

Every engine sheet is one row per cadet (rows 6..55), mirroring Cadets order.
"""
from openpyxl.utils import get_column_letter

from xlb import (
    HDR_ROW, DATA_ROW, CADET_LAST, F_CALC, FILL_CALC, F_LABEL, F_SMALL,
    A_LEFT_WRAP, BOX, DATE, header_row, fill_rows, sheet_note, cf_yes_no,
    cf_formula, FILL_WARNBG, FILL_OKBG, FILL_AMBER, col_widths, define,
    protect,
)

FIRST, LAST = DATA_ROW, CADET_LAST


def _mirror():
    return {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF($B{r}="","",Cadets!$F{r})', "fx"),
        "D": ('IF($B{r}="","",Cadets!$I{r})', "fx"),
    }


# --------------------------------------------------------------------------
def build_sysgrades(wb):
    ws = wb.create_sheet("sysGrades")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Status", "MajCnt", "MinCnt",
                    "SpCnt", "FinCnt", "MajorAvg", "MinorAvg", "SpellAvg",
                    "FinalScore", "Current Grade", "Final Grade", "Retakes",
                    "DismissRev", "MajBelow", "MinBelow", "GradeNum",
                    "RankElig", "Rank", "AcademicElig",
                    "Need Maj Avg Left", "Need Min Avg Left", "Need Final",
                    "LastScore", "PrevScore", "GradeDrop", "ConsecFails"])
    cols = _mirror()
    cols.update({
        # counts and averages must use the SAME criteria or the count=0 guard
        # on I/J/L is defeated: a pending retest row is the final attempt but
        # carries no Recorded score, so it must not be counted either.
        # The range is BOTH-ended: ExamScores has no upper-bound validation on
        # Raw Score (it deliberately accepts text such as "absent"), so a
        # mis-keyed 180 satisfied ">=0" and flowed into the transcript, the
        # class rank and sysAwards as a >100 category average. It is excluded
        # from the counts and the averages exactly as a non-numeric raw is;
        # the ExamScores Row Check ("RAW SCORE OUT OF RANGE") is what reports
        # the row itself.
        "E": ('IF($B{r}="","",COUNTIFS(nrES_PID,$B{r},nrES_Type,"Major",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"))', "fx"),
        "F": ('IF($B{r}="","",COUNTIFS(nrES_PID,$B{r},nrES_Type,"Minor",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"))', "fx"),
        "G": ('IF($B{r}="","",Spelling!Q{r})', "fx"),
        "H": ('IF($B{r}="","",COUNTIFS(nrES_PID,$B{r},nrES_Type,"Final",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"))', "fx"),
        "I": ('IF(OR($B{r}="",$E{r}=0),"",IFERROR(ROUND(AVERAGEIFS(nrES_Rec,nrES_PID,$B{r},nrES_Type,"Major",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"),2),""))', "fx"),
        "J": ('IF(OR($B{r}="",$F{r}=0),"",IFERROR(ROUND(AVERAGEIFS(nrES_Rec,nrES_PID,$B{r},nrES_Type,"Minor",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"),2),""))', "fx"),
        "K": ('IF($B{r}="","",Spelling!P{r})', "fx"),
        "L": ('IF(OR($B{r}="",$H{r}=0),"",IFERROR(ROUND(AVERAGEIFS(nrES_Rec,nrES_PID,$B{r},nrES_Type,"Final",nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"),2),""))', "fx"),
        "M": ('IF($B{r}="","",LET(w,($E{r}>0)*cfgWeightMajor+($F{r}>0)*cfgWeightMinor'
              '+(N($G{r})>0)*cfgWeightSpelling+($H{r}>0)*cfgWeightFinal,'
              'IF(w=0,"",ROUND((($E{r}>0)*cfgWeightMajor*N($I{r})'
              '+($F{r}>0)*cfgWeightMinor*N($J{r})'
              '+(N($G{r})>0)*cfgWeightSpelling*N($K{r})'
              '+($H{r}>0)*cfgWeightFinal*N($L{r}))/w,2))))', "fx"),
        "N": ('IF(OR($B{r}="",$H{r}=0),"",$M{r})', "fx"),
        "O": ('IF($B{r}="","",COUNTIFS(nrES_PID,$B{r},nrES_Att,2))', "fx"),
        "P": ('IF($B{r}="","",COUNTIFS(nrES_PID,$B{r},nrES_Dis,"Yes"))', "fx"),
        # "enforced yet?" for a category average. cfgThresholdAfterExam is
        # the documented grace period ("Category avg enforced after this many
        # exams"), and it used to live ONLY here - Q and R were read by
        # nothing at all, so the setting changed nothing and V below enforced
        # 70-in-each-category from the very first recorded exam. V now reads
        # these two cells, so the Settings row is live.
        # The grace can only ever apply MID-academy: it lapses as soon as the
        # cadet has every planned exam of that type on file, so a coordinator
        # who plans fewer than cfgThresholdAfterExam exams of a type cannot
        # turn the grace into a permanent waiver. ($I / $J are blank when the
        # count is 0, which is the "no records yet" waiver.)
        "Q": ('IF($B{r}="","",IF(AND($I{r}<>"",$I{r}<cfgThresholdScore,'
              'OR($E{r}>=cfgThresholdAfterExam,'
              '$E{r}>=COUNTIFS(rngEPuse,"Yes",rngEPtype,"Major"))),'
              '"Yes","No"))', "fx"),
        "R": ('IF($B{r}="","",IF(AND($J{r}<>"",$J{r}<cfgThresholdScore,'
              'OR($F{r}>=cfgThresholdAfterExam,'
              '$F{r}>=COUNTIFS(rngEPuse,"Yes",rngEPtype,"Minor"))),'
              '"Yes","No"))', "fx"),
        "S": ('IF($B{r}="","",IF($M{r}="",-1,$M{r}))', "fx"),
        # one source of truth for "is a dismissal review OPEN": sysChecks M.
        # Repeating the three raw triggers here meant a review CLOSED on the
        # DismissalLog still left the cadet permanently unranked.
        "T": ('IF($B{r}="","",IF(AND($D{r}="Active",sysChecks!$M{r}<>"Yes",'
              '$M{r}<>""),"Yes","No"))', "fx"),
        "U": ('IF($T{r}<>"Yes","",SUMPRODUCT(($T$%d:$T$%d="Yes")*'
              '(($S$%d:$S$%d>$S{r})+($S$%d:$S$%d=$S{r})*(ROW($S$%d:$S$%d)<ROW())))+1)'
              % (FIRST, LAST, FIRST, LAST, FIRST, LAST, FIRST, LAST), "fx"),
        # major/minor come from Q/R (which carry the cfgThresholdAfterExam
        # grace); spelling and the final exam have no grace period and are
        # tested here directly. Graduation stays safe because sysChecks R
        # requires every PLANNED exam of each type to be recorded, by which
        # point Q/R are enforcing.
        "V": ('IF($B{r}="","",IF(AND($Q{r}="No",$R{r}="No",'
              'OR(N($G{r})=0,$K{r}>=cfgThresholdScore),'
              'OR($H{r}=0,$L{r}>=cfgThresholdScore)),"Yes","No"))', "fx"),
        # projections: average needed on remaining exams to reach 70 in category
        "W": ('IF($B{r}="","",LET(tot,COUNTIFS(rngEPuse,"Yes",rngEPtype,"Major"),'
              'left,tot-$E{r},IF(left<=0,"done",'
              'MAX(0,ROUND((cfgThresholdScore*tot-N($I{r})*$E{r})/left,1)))))', "fx"),
        "X": ('IF($B{r}="","",LET(tot,COUNTIFS(rngEPuse,"Yes",rngEPtype,"Minor"),'
              'left,tot-$F{r},IF(left<=0,"done",'
              'MAX(0,ROUND((cfgThresholdScore*tot-N($J{r})*$F{r})/left,1)))))', "fx"),
        "Y": ('IF($B{r}="","",IF($H{r}>0,"done",cfgThresholdScore))', "fx"),
        # trend helpers: last & previous recorded first-attempt scores by seq
        "Z": ('IF($B{r}="","",LET(s,MAXIFS(nrES_Seq,nrES_PID,$B{r},nrES_Att,1,'
              'nrES_Raw,">=0"),IF(s=0,"",SUMIFS(nrES_Raw,nrES_PID,$B{r},'
              'nrES_Att,1,nrES_Seq,s))))', "fx"),
        # the PREVIOUS attempt must carry the same nrES_Raw,">=0" filter as
        # the latest one. Without it an unscored attempt-1 row (an EXCUSED
        # absence leaves Raw Score blank by policy) wins the MAXIFS, reads
        # as a recorded 0, and both fabricates a consecutive-fail flag (AC)
        # and erases the real grade-drop flag (AB, which goes negative).
        "AA": ('IF($B{r}="","",LET(s,MAXIFS(nrES_Seq,nrES_PID,$B{r},nrES_Att,1,'
               'nrES_Raw,">=0"),IF(s=0,"",LET(p,MAXIFS(nrES_Seq,nrES_PID,$B{r},'
               'nrES_Att,1,nrES_Seq,"<"&s,nrES_Raw,">=0"),IF(p=0,"",'
               'SUMIFS(nrES_Raw,nrES_PID,$B{r},nrES_Att,1,nrES_Seq,p))))))', "fx"),
        "AB": ('IF(OR($B{r}="",$Z{r}="",$AA{r}=""),"",ROUND($AA{r}-$Z{r},1))', "fx"),
        "AC": ('IF($B{r}="","",LET(s,MAXIFS(nrES_Seq,nrES_PID,$B{r},nrES_Att,1,'
               'nrES_Raw,">=0"),IF(s=0,0,LET(a,IF(SUMIFS(nrES_Raw,nrES_PID,$B{r},'
               'nrES_Att,1,nrES_Seq,s)<cfgPassingScore,1,0),'
               'p,MAXIFS(nrES_Seq,nrES_PID,$B{r},nrES_Att,1,nrES_Seq,"<"&s,'
               'nrES_Raw,">=0"),'
               'IF(OR(a=0,p=0),a,a+IF(SUMIFS(nrES_Raw,nrES_PID,$B{r},'
               'nrES_Att,1,nrES_Seq,p)<cfgPassingScore,1,0))))))', "fx"),
    })
    fill_rows(ws, FIRST, LAST, cols)
    define(wb, "nrGRrank", "sysGrades", f"$U${FIRST}:$U${LAST}")
    define(wb, "nrGRcurrent", "sysGrades", f"$M${FIRST}:$M${LAST}")
    define(wb, "nrGRfinal", "sysGrades", f"$N${FIRST}:$N${LAST}")
    # column L is the FINAL EXAM average; column N is the weighted composite.
    # Outputs that label a row "Final Exam" must read L, never N.
    define(wb, "nrGRfinalExam", "sysGrades", f"$L${FIRST}:$L${LAST}")
    define(wb, "nrGRacademic", "sysGrades", f"$V${FIRST}:$V${LAST}")
    define(wb, "nrGRmajavg", "sysGrades", f"$I${FIRST}:$I${LAST}")
    define(wb, "nrGRminavg", "sysGrades", f"$J${FIRST}:$J${LAST}")
    define(wb, "nrGRgradedrop", "sysGrades", f"$AB${FIRST}:$AB${LAST}")
    define(wb, "nrGRconsec", "sysGrades", f"$AC${FIRST}:$AC${LAST}")
    define(wb, "nrGRdisrev", "sysGrades", f"$P${FIRST}:$P${LAST}")
    define(wb, "nrGRretakes", "sysGrades", f"$O${FIRST}:$O${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    sheet_note(ws, "Weighted per policy 300.2 (Major 40 / Minor 30 / Spelling "
                   "10 / Final 20); pass requires 70 in EACH category. "
                   "Projection columns show the average needed on remaining "
                   "exams. Locked engine — do not edit.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_sysattendance(wb):
    ws = wb.create_sheet("sysAttendance")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Status", "Cl Missed", "Cl MadeUp",
                    "Cl Net", "Advisory At", "Cl Owed (min)", "Cl Tier",
                    "Cl Clear?",
                    "Cl Owed", "PT Missed", "PT MadeUp", "PT Net", "PT Cap",
                    "PT %", "PT Tier", "PT Elig", "Makeup Complete?",
                    "Tardy", "FullAbs", "PartAbs", "PTMissCnt", "Excused"])
    cols = _mirror()
    cols.update({
        "E": ('IF($B{r}="","",SUMIFS(nrAT_Min,nrAT_PID,$B{r},nrAT_Counts,"Yes",nrAT_IsPT,"No"))', "fx"),
        # made-up minutes are read from the PER-EVENT reconciliation on
        # Attendance (col P), which is capped at each event's own duration —
        # NOT from the raw Makeup rows. Summing nrMK_Min directly let a
        # 360-minute makeup booked against a 120-minute event pay off 240
        # minutes of a completely different absence, so the gate below read
        # "no makeup owed" while that event's own ledger still said OPEN.
        "F": ('IF($B{r}="","",SUMIFS(nrAT_MadeUpMin,nrAT_PID,$B{r},'
              'nrAT_Counts,"Yes",nrAT_IsPT,"No"))', "fx"),
        "G": ('IF($B{r}="","",MAX(0,$E{r}-$F{r}))', "fx"),
        # No classroom allowance exists. TCOLE: "there is no 10% attendance
        # rule"; policy 400 requires every missed minute to be made up. So
        # H is an early-warning threshold, not a cap; I is the minutes still
        # OWED (not a share of anything); J escalates on those minutes; and
        # K asks the only question policy actually asks - is anything still
        # outstanding?
        "H": ('IF($B{r}="","",cfgMakeupAdvisoryMin)', "fx"),
        "I": ('IF($B{r}="","",MAX(0,$E{r}-$F{r}))', "fx"),
        "J": ('IF($B{r}="","",IF($I{r}=0,"OK",'
              'IF($I{r}>=cfgMakeupCriticalMin,"CRITICAL",'
              'IF($I{r}>=cfgMakeupAdvisoryMin,"Advisory","Owes time"))))', "fx"),
        "K": ('IF($B{r}="","",IF($I{r}=0,"Yes","No"))', "fx"),
        "L": ('IF($B{r}="","",MAX(0,$E{r}-$F{r}))', "fx"),
        "M": ('IF($B{r}="","",SUMIFS(nrAT_Sess,nrAT_PID,$B{r},nrAT_Counts,"Yes",nrAT_IsPT,"Yes"))', "fx"),
        "N": ('IF($B{r}="","",SUMIFS(nrAT_MadeUpSess,nrAT_PID,$B{r},'
              'nrAT_Counts,"Yes",nrAT_IsPT,"Yes"))', "fx"),
        "O": ('IF($B{r}="","",MAX(0,$M{r}-$N{r}))', "fx"),
        "P": ('IF($B{r}="","",cfgPTCapSessions)', "fx"),
        "Q": ('IF($B{r}="","",IF($P{r}=0,0,$O{r}/$P{r}))', "fx"),
        "R": ('IF($B{r}="","",IF($Q{r}>=1,"AT CAP (5 sessions)",IF($Q{r}>=cfgAttendanceCriticalPct,'
              '"Critical",IF($Q{r}>=cfgAttendanceAdvisoryPct,"Advisory","OK"))))', "fx"),
        "S": ('IF($B{r}="","",IF($P{r}=0,"Yes",IF($O{r}<$P{r},"Yes","No")))', "fx"),
        "T": ('IF($B{r}="","",IF(AND($L{r}=0,$O{r}=0),"Yes","No"))', "fx"),
        "U": ('IF($B{r}="","",COUNTIFS(nrAT_PID,$B{r},nrAT_Type,"Tardy"))', "fx"),
        "V": ('IF($B{r}="","",COUNTIFS(nrAT_PID,$B{r},nrAT_Type,"Absent Full Day"))', "fx"),
        "W": ('IF($B{r}="","",COUNTIFS(nrAT_PID,$B{r},nrAT_Type,"Absent Partial Day"))', "fx"),
        "X": ('IF($B{r}="","",COUNTIFS(nrAT_PID,$B{r},nrAT_Type,"PT Missed"))', "fx"),
        "Y": ('IF($B{r}="","",COUNTIFS(nrAT_PID,$B{r},nrAT_Excused,"Excused"))', "fx"),
    })
    # fix column collision: header has 24 cols B..Y; L duplicated (Cl Owed)
    cols["L"] = ('IF($B{r}="","",MAX(0,$E{r}-$F{r}))', "fx")
    fill_rows(ws, FIRST, LAST, cols)
    # Q is still a fraction of a real cap (five PT sessions) and must show as
    # a percentage; shipped as General it displayed 0.6521739130434783. I is
    # no longer a fraction of anything - it is minutes still owed - so it is
    # formatted as a whole number.
    for rr in range(FIRST, LAST + 1):
        ws[f"I{rr}"].number_format = "#,##0"
        ws[f"Q{rr}"].number_format = "0.0%"
    define(wb, "nrATTclTier", "sysAttendance", f"$J${FIRST}:$J${LAST}")
    define(wb, "nrATTclElig", "sysAttendance", f"$K${FIRST}:$K${LAST}")
    define(wb, "nrATTclPct", "sysAttendance", f"$I${FIRST}:$I${LAST}")
    define(wb, "nrATTclOwed", "sysAttendance", f"$L${FIRST}:$L${LAST}")
    define(wb, "nrATTptTier", "sysAttendance", f"$R${FIRST}:$R${LAST}")
    define(wb, "nrATTptElig", "sysAttendance", f"$S${FIRST}:$S${LAST}")
    define(wb, "nrATTptPct", "sysAttendance", f"$Q${FIRST}:$Q${LAST}")
    define(wb, "nrATTmakeupOK", "sysAttendance", f"$T${FIRST}:$T${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    sheet_note(ws, "Classroom MINUTES STILL OWED (there is no attendance allowance - every missed minute is made up); PT sessions vs the "
                   "5-session cap (policy 400). 'Cl MadeUp'/'PT MadeUp' are "
                   "the sum of the PER-EVENT credits on Attendance (P/Q), "
                   "each capped at that event's own duration — so these can "
                   "never disagree with the OPEN/CLEARED banner. 'Cl Owed' = "
                   "missed minutes not yet made up; 'Makeup Complete?' "
                   "requires both owed classroom minutes AND owed PT sessions "
                   "at 0 before graduation. Locked.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_sysskills(wb):
    ws = wb.create_sheet("sysSkills")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Status", "Categories Attempted",
                    "Qualified", "Needs Remediation", "Pending",
                    "Failed Out Cats", "FailedOut?", "Skills Elig",
                    "Firearms Avg", "Firearms Best", "CoF1 Best",
                    "CoF2 Best",
                    '="Both CoF ≥"&IFERROR(INDEX(rngSM_pass,'
                    'MATCH("Firearms",rngSM_cat,0)),70)&"?"'])
    cols = _mirror()
    cols.update({
        "E": ('IF($B{r}="","",SUMPRODUCT((rngSM_cat<>"")*'
              '(COUNTIFS(nrSK_PID,$B{r},nrSK_Cat,rngSM_cat)>0)))', "fx"),
        "F": ('IF($B{r}="","",SUMPRODUCT((rngSM_cat<>"")*'
              '(COUNTIFS(nrSK_PID,$B{r},nrSK_Cat,rngSM_cat,nrSK_Status,"Qualified")>0)))', "fx"),
        "G": ('IF($B{r}="","",SUMPRODUCT((rngSM_cat<>"")*'
              '(COUNTIFS(nrSK_PID,$B{r},nrSK_Cat,rngSM_cat,nrSK_Status,"Needs Remediation")>0)))', "fx"),
        "H": ('IF($B{r}="","",SUMPRODUCT((rngSM_cat<>"")*'
              '(COUNTIFS(nrSK_PID,$B{r},nrSK_Cat,rngSM_cat,nrSK_Status,"Pending")>0)))', "fx"),
        # count failed-out CATEGORIES, not failed-out ROWS. Firearms records
        # ONE attempt as TWO rows (Course of Fire 1 and 2), so a firearms
        # fail-out with both courses failed produced 2 here while the board's
        # single DismissalLog closure counts 1 - sysChecks U's os>cs then
        # stayed TRUE forever and the cadet was permanently blocked and
        # unranked. Same SUMPRODUCT-over-rngSM_cat shape as E/F/G/H above.
        "I": ('IF($B{r}="","",SUMPRODUCT((rngSM_cat<>"")*'
              '(COUNTIFS(nrSK_PID,$B{r},nrSK_Cat,rngSM_cat,nrSK_Dis,"Yes")>0)))', "fx"),
        "J": ('IF($B{r}="","",IF($I{r}>0,"Yes","No"))', "fx"),
        "K": ('IF($B{r}="","",IF($J{r}="Yes","No",IF($G{r}>0,"No","Yes")))', "fx"),
        # every firearms aggregate carries the same 0-100 bound the exam
        # aggregates carry (nrES_Rec ">=0"/"<=100"). Without it a slipped
        # decimal (68 keyed as 680) printed as the transcript's firearms
        # average, won Top Gun, and satisfied the ch.41 both-courses-of-fire
        # gate that feeds sysChecks T and GraduationElig. Skills Row Check
        # (column S) names the offending row.
        "L": ('IF($B{r}="","",IFERROR(ROUND(AVERAGEIFS(nrSK_Score,nrSK_PID,$B{r},'
              'nrSK_Cat,"Firearms",nrSK_Score,">=0",nrSK_Score,"<=100"),2),""))', "fx"),
        "M": ('IF($B{r}="","",IFERROR(MAXIFS(nrSK_Score,nrSK_PID,$B{r},'
              'nrSK_Cat,"Firearms",nrSK_Score,">=0",nrSK_Score,"<=100"),""))', "fx"),
        "N": ('IF($B{r}="","",LET(v,MAXIFS(nrSK_Score,nrSK_PID,$B{r},'
              'nrSK_Cat,"Firearms",nrSK_CoF,1,nrSK_Score,">=0",'
              'nrSK_Score,"<=100"),IF(v=0,"",v)))', "fx"),
        "O": ('IF($B{r}="","",LET(v,MAXIFS(nrSK_Score,nrSK_PID,$B{r},'
              'nrSK_Cat,"Firearms",nrSK_CoF,2,nrSK_Score,">=0",'
              'nrSK_Score,"<=100"),IF(v=0,"",v)))', "fx"),
        # the firearms threshold is SkillsMaster's editable "Passing Score",
        # not a literal: with 70 hard-coded here, raising the standard on
        # SkillsMaster left this graduation gate (sysChecks T) and the
        # sysAudit firearms line passing cadets below the new mark.
        "P": ('IF($B{r}="","",IF(AND($N{r}="",$O{r}=""),"",'
              'LET(p,IFERROR(INDEX(rngSM_pass,MATCH("Firearms",rngSM_cat,0)),70),'
              'IF(AND($N{r}<>"",$O{r}<>"",N($N{r})>=p,N($O{r})>=p),"Yes",'
              'IF(OR(AND($N{r}<>"",N($N{r})<p),AND($O{r}<>"",N($O{r})<p)),'
              '"No","(one pending)")))))', "fx"),
    })
    fill_rows(ws, FIRST, LAST, cols)
    define(wb, "nrSKfailedout", "sysSkills", f"$J${FIRST}:$J${LAST}")
    define(wb, "nrSKelig", "sysSkills", f"$K${FIRST}:$K${LAST}")
    define(wb, "nrSKfirearmsAvg", "sysSkills", f"$L${FIRST}:$L${LAST}")
    define(wb, "nrSKbothCoF", "sysSkills", f"$P${FIRST}:$P${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    sheet_note(ws, "Column letters I/J intentionally match V5.4 cross-refs "
                   "(sysGrades reads sysSkills!$I / $J-style flags). Locked.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_sysincidents(wb):
    ws = wb.create_sheet("sysIncidents")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Status", "Total", "Negative",
                    "Positive", "Open Negative", "NegMajorCritical",
                    "OpenChainReview", "Incidents Elig",
                    "Counseling Cnt", "Open Counseling"])
    cols = _mirror()
    cols.update({
        "E": ('IF($B{r}="","",COUNTIF(nrIN_PID,$B{r}))', "fx"),
        "F": ('IF($B{r}="","",COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Negative"))', "fx"),
        "G": ('IF($B{r}="","",COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Positive"))', "fx"),
        "H": ('IF($B{r}="","",COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Negative",'
              'nrIN_Res,"Open")+COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Negative",'
              'nrIN_Res,"In Review"))', "fx"),
        "I": ('IF($B{r}="","",COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Negative",'
              'nrIN_Sev,"Major")+COUNTIFS(nrIN_PID,$B{r},nrIN_Dir,"Negative",'
              'nrIN_Sev,"Critical"))', "fx"),
        "J": ('IF($B{r}="","",COUNTIFS(nrIN_PID,$B{r},nrIN_Chain,"Yes",'
              'nrIN_Res,"Open")+COUNTIFS(nrIN_PID,$B{r},nrIN_Chain,"Yes",'
              'nrIN_Res,"In Review"))', "fx"),
        "K": ('IF($B{r}="","",IF($J{r}=0,"Yes","No"))', "fx"),
        "L": ('IF($B{r}="","",COUNTIF(nrCO_PID,$B{r}))', "fx"),
        "M": ('IF($B{r}="","",COUNTIFS(nrCO_PID,$B{r},nrCO_Status,"Open")'
              '+COUNTIFS(nrCO_PID,$B{r},nrCO_Status,"In Review"))', "fx"),
    })
    fill_rows(ws, FIRST, LAST, cols)
    define(wb, "nrINCopenNeg", "sysIncidents", f"$H${FIRST}:$H${LAST}")
    define(wb, "nrINCelig", "sysIncidents", f"$K${FIRST}:$K${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    sheet_note(ws, "Incident + counseling rollup per cadet. Locked.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_sysflags(wb):
    """Configurable flag engine: one column per flag, reason text, count."""
    ws = wb.create_sheet("sysFlags")
    ws.sheet_view.showGridLines = False
    # NOTE: the flag block is E..R and "Flag Count"/"Reasons" are S/T.
    # F:DismissReview was APPENDED at the end of the flag block (R); the
    # count and reason columns moved one right with it, and every consumer
    # (nrFLcount / nrFLreasons, EmailPreview's literal sysFlags!$T, and the
    # verify_build assertions) moved with them. Add new flags at the END of
    # the block, never in the middle.
    header_row(ws, ["PID", "Cadet Name", "Status", "F:ConsecFails",
                    "F:GradeDrop", "F:CategoryRisk", "F:Spelling",
                    "F:Attendance", "F:Incidents", "F:Writing", "F:Retest",
                    "F:PT", "F:Medical", "F:Certs", "F:Memos", "F:OpenTime",
                    "F:DismissReview",
                    "Flag Count", "Reasons", "F:MissedExam"])
    cols = _mirror()
    cols.update({
        "E": ('IF($B{r}="","",IF(N(nrGRconsec %s)>=cfgFlagConsecFails,1,0))'
              % "", "fx"),
        "F": ('IF($B{r}="","",IF(AND(sysGrades!$AB{r}<>"",'
              'sysGrades!$AB{r}>=cfgFlagGradeDrop),1,0))', "fx"),
        "G": ('IF($B{r}="","",IF(OR(AND(sysGrades!$I{r}<>"",'
              'sysGrades!$I{r}<cfgThresholdScore+cfgFlagCategoryMargin),'
              'AND(sysGrades!$J{r}<>"",sysGrades!$J{r}<cfgThresholdScore'
              '+cfgFlagCategoryMargin)),1,0))', "fx"),
        "H": ('IF($B{r}="","",IF(Spelling!$R{r}="INTERVENTION",1,0))', "fx"),
        # classroom: any outstanding makeup at/over the advisory threshold.
        # PT: still a real cap (five sessions, policy 400.1.D), so still a %.
        "I": ('IF($B{r}="","",IF(OR(N(sysAttendance!$I{r})>=cfgMakeupAdvisoryMin,'
              'sysAttendance!$Q{r}>=cfgFlagAttendancePct),1,0))', "fx"),
        "J": ('IF($B{r}="","",IF(sysIncidents!$H{r}>=cfgFlagOpenIncidents,1,0))', "fx"),
        "K": ('IF($B{r}="","",IF(N(Writing!$AS{r})>=cfgFlagOverdueWriting,1,0))', "fx"),
        # CHECK DATE counts too: a retest whose deadline could not be computed
        # is not "fine", it is a policy 300.5 clock that never started
        "L": ('IF($B{r}="","",IF(COUNTIFS(nrES_PID,$B{r},nrES_RetStat,"OVERDUE")'
              '+COUNTIFS(nrES_PID,$B{r},nrES_RetStat,"CHECK DATE")'
              '+COUNTIFS(nrES_PID,$B{r},nrES_RetStat,"LATE RETEST*")'
              '+COUNTIFS(nrES_PID,$B{r},nrES_RetStat,"RETEST UNDATED")'
              '+COUNTIFS(nrES_PID,$B{r},nrES_RetStat,"RETEST DATE UNCHECKED")'
              '>0,1,0))', "fx"),
        "M": ('IF($B{r}="","",IF(OR(PT!$K{r}="No",PT!$AB{r}="No",'
              'PT!$AB{r}="Incomplete"),1,0))', "fx"),
        "N": ('IF($B{r}="","",IF(COUNTIFS(nrMD_PID,$B{r},nrMD_Status,'
              '"RESTRICTION EXPIRED")>0,1,0))', "fx"),
        "O": ('IF($B{r}="","",IF(Certifications!$U{r}<>"",1,0))', "fx"),
        "P": ('IF($B{r}="","",IF(COUNTIFS(nrME_PID,$B{r},nrME_Status,'
              '"OVERDUE")+COUNTIFS(nrME_PID,$B{r},nrME_Status,'
              '"CHECK DATE")>0,1,0))', "fx"),
        "Q": ('IF($B{r}="","",IF(SUMPRODUCT((nrAT_PID=$B{r})*'
              '(nrAT_Cleared="OPEN"))>0,1,0))', "fx"),
        # an open dismissal review (failed retest, skills failed out, open
        # chain-of-command incident review) blocked graduation but raised no
        # flag at all: the cadet's Flag Count never moved, they never
        # appeared on WatchList and the Dashboard "Flagged cadets" tile
        # never counted them. Read the ONE gate (sysChecks M) rather than
        # re-deriving the triggers, so a review closed on the DismissalLog
        # clears the flag too.
        "R": ('IF($B{r}="","",IF(sysChecks!$M{r}="Yes",1,0))', "fx"),
        # policy: an unexcused missed exam is a recorded 0; a SECOND
        # occurrence is a removal trigger, so it must be visible well
        # before anyone reaches the graduation checklist.
        "U": ('IF($B{r}="","",IF(COUNTIFS(nrES_PID,$B{r},nrES_Absence,'
              '"Unexcused")>0,1,0))', "fx"),
        "S": ('IF($B{r}="","",SUM($E{r}:$R{r})+N($U{r}))', "fx"),
        "T": ('IF($B{r}="","",IF($S{r}=0,"",TEXTJOIN("; ",TRUE,'
              'IF($E{r}=1,"consecutive exam fails",""),'
              'IF($F{r}=1,"grade dropped "&sysGrades!$AB{r}&" pts",""),'
              'IF($G{r}=1,"category avg near 70",""),'
              'IF($H{r}=1,"spelling below "&cfgSpellInterventionAvg,""),'
              'IF($I{r}=1,"attendance at "&TEXT(MAX(sysAttendance!$I{r},'
              'sysAttendance!$Q{r}),"0%")&" of cap",""),'
              'IF($J{r}=1,"open negative incidents",""),'
              'IF($K{r}=1,N(Writing!$AS{r})&" overdue writing",""),'
              'IF($L{r}=1,"RETEST OVERDUE / LATE",""),'
              'IF($M{r}=1,"PT failure",""),'
              'IF($N{r}=1,"medical restriction expired",""),'
              'IF($O{r}=1,"cert copies outstanding: "&Certifications!$U{r},""),'
              'IF($P{r}=1,"OVERDUE MEMO",""),'
              'IF($Q{r}=1,SUMPRODUCT((nrAT_PID=$B{r})*(nrAT_Cleared="OPEN"))'
              '&" uncleared missed-time event(s)",""),'
              'IF($R{r}=1,"DISMISSAL REVIEW OPEN ("&'
              'sysChecks!$U{r}&")",""),'
              'IF($U{r}=1,LET(n,COUNTIFS(nrES_PID,$B{r},nrES_Absence,"Unexcused"),'
              'IF(n>=2,n&" unexcused missed exams (removal review)",'
              '"unexcused missed exam (0 recorded)")),""))))', "fx"),
    })
    # E needs the row-scoped reference, not the whole named range
    cols["E"] = ('IF($B{r}="","",IF(N(sysGrades!$AC{r})>=cfgFlagConsecFails,1,0))', "fx")
    fill_rows(ws, FIRST, LAST, cols)
    define(wb, "nrFLcount", "sysFlags", f"$S${FIRST}:$S${LAST}")
    define(wb, "nrFLmissedExam", "sysFlags", f"$U${FIRST}:$U${LAST}")
    define(wb, "nrFLreasons", "sysFlags", f"$T${FIRST}:$T${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "T": 70})
    sheet_note(ws, "Each flag threshold lives on Settings. WatchList sorts "
                   "by Flag Count and shows Reasons verbatim. Locked.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_syschecks(wb):
    ws = wb.create_sheet("sysChecks")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Status", "Academic", "Classroom",
                    "PT Sessions", "Skills", "Incidents", "Writing",
                    "Makeup Complete", "Final PT Pass", "DismissReview",
                    "GraduationElig", "Blocking Issues", "Final Exam Elig",
                    "Certs", "Exams Recorded", "Skills Assessed",
                    "Firearms CoF", "Open Reviews", "Exams Pending",
                    "Enroll Docs", "Skills P/F"])
    cols = _mirror()
    cols.update({
        "E": ('IF($B{r}="","",sysGrades!$V{r})', "fx"),
        "F": ('IF($B{r}="","",sysAttendance!$K{r})', "fx"),
        "G": ('IF($B{r}="","",sysAttendance!$S{r})', "fx"),
        "H": ('IF($B{r}="","",sysSkills!$K{r})', "fx"),
        "I": ('IF($B{r}="","",sysIncidents!$K{r})', "fx"),
        "J": ('IF($B{r}="","",IF(Writing!$AT{r}="Yes","Yes","No"))', "fx"),
        "K": ('IF($B{r}="","",sysAttendance!$T{r})', "fx"),
        # AFFIRMATIVE test: only an actual pass reads "Yes". A blank PT!AB
        # (final PT never assessed) and "(rubric pending)" (cfgPTFinalMinPoints
        # still 0) are distinct blocking states, not passes.
        "L": ('IF($B{r}="","",IF(PT!$AB{r}="Yes","Yes",IF(PT!$AB{r}="No","No",'
              'IF(PT!$AB{r}="(rubric pending)","Pending",'
              'IF(PT!$AB{r}="Incomplete","Incomplete","Not taken")))))', "fx"),
        # THE dismissal-review gate for the whole workbook (sysGrades T and
        # sysFlags R read this cell, they no longer re-derive it).
        # A raised review used to be unclosable: nothing consumed the
        # DismissalLog, so a failed retest blocked graduation and dropped the
        # cadet off the ranking permanently, even after a formal board
        # retained them. Closure is now counted PER TRIGGER against the
        # reviews actually recorded, approved and tagged on the DismissalLog,
        # so one closed review cannot pre-clear a future one.
        "M": ('IF($B{r}="","",IF($U{r}="","No","Yes"))', "fx"),
        # Coordinator-marked skills pass/fail (SkillsCheck). Deliberately a
        # manual call, not derived from the attempt log and not auto-raised
        # by missed skills time - parts of skills training can be made up in
        # practice, so the person who ran the evaluation decides. Any Fail
        # blocks graduation until it is cleared or changed to Pass.
        "X": ('IF($B{r}="","",IF(SkillsCheck!$P{r}="No","No","Yes"))', "fx"),
        "N": ('IF($B{r}="","",IF(AND($E{r}="Yes",$F{r}="Yes",$G{r}="Yes",'
              '$H{r}="Yes",$I{r}="Yes",$J{r}="Yes",$K{r}="Yes",$L{r}="Yes",'
              '$Q{r}="Yes",$R{r}="Yes",$S{r}="Yes",$T{r}="Yes",'
              '$M{r}="No",$V{r}="Yes",$W{r}="Yes",$X{r}="Yes"),"Yes","No"))',
              "fx"),
        "O": ('IF($B{r}="","",IF($N{r}="Yes","Eligible",TRIM('
              'IF($E{r}<>"Yes","Academic; ","")&'
              'IF($R{r}<>"Yes","Exams not all recorded; ","")&'
              'IF($F{r}<>"Yes","Classroom; ","")&'
              'IF($G{r}<>"Yes","PT sessions; ","")&'
              'IF($H{r}<>"Yes","Skills; ","")&'
              'IF($S{r}<>"Yes","Skills not all assessed; ","")&'
              'IF($T{r}<>"Yes","Firearms courses of fire; ","")&'
              'IF($I{r}<>"Yes","Incidents; ","")&'
              'IF($J{r}<>"Yes","Writing; ","")&'
              'IF($K{r}<>"Yes","Makeup owed; ","")&'
              'IF($L{r}="No","Final PT failed; ",'
              'IF($L{r}="Pending","Final PT rubric not set; ",'
              'IF($L{r}="Incomplete","Final PT partially scored; ",'
              'IF($L{r}<>"Yes","Final PT not assessed; ",""))))&'
              'IF($Q{r}<>"Yes","Certs; ","")&'
              'IF($V{r}<>"Yes","Exam pending (excused absence); ","")&'
              'IF($W{r}<>"Yes","Enrollment documents; ","")&'
              'IF($X{r}<>"Yes","Skills FAILED ("&SkillsCheck!$N{r}&"); ","")&'
              'IF($M{r}="Yes","Dismissal review; ",""))))', "fx"),
        "P": ('IF($B{r}="","",IF(PT!$AB{r}="Yes","Yes",IF(PT!$AB{r}="No","No",'
              'IF(PT!$AB{r}="(rubric pending)","Pending",'
              'IF(PT!$AB{r}="Incomplete","Incomplete","Not taken")))))', "fx"),
        "Q": ('IF($B{r}="","",IF(Certifications!$T{r}="Yes","Yes","No"))', "fx"),
        # completeness: sysGrades V waives a category whose count is 0 (and,
        # under cfgThresholdAfterExam, a category with only a handful of
        # exams on file) so mid-academy cadets don't read Academic="No".
        # Graduation must not inherit either waiver: "at least one record"
        # let a cadet with 1 of 10 major exams recorded read Exams
        # Recorded = "Yes", which is also what would have turned the
        # mid-academy grace period into a permanent academic waiver. Every
        # exam the ExamPlan marks Use? = Yes must actually be on file.
        "R": ('IF($B{r}="","",LET('
              'mj,COUNTIFS(rngEPuse,"Yes",rngEPtype,"Major"),'
              'mn,COUNTIFS(rngEPuse,"Yes",rngEPtype,"Minor"),'
              'fnl,COUNTIFS(rngEPuse,"Yes",rngEPtype,"Final"),'
              'IF(AND(N(sysGrades!$E{r})>0,N(sysGrades!$F{r})>0,'
              'N(sysGrades!$G{r})>0,N(sysGrades!$H{r})>0,'
              'N(sysGrades!$E{r})>=mj,N(sysGrades!$F{r})>=mn,'
              'N(sysGrades!$H{r})>=fnl),"Yes","No")))', "fx"),
        # same shape as R, for skills: sysSkills K only asks "not failed out
        # and nothing currently in remediation", so a cadet with ZERO skills
        # records scored as a pass. Graduation additionally requires every
        # SkillsMaster category to be Qualified (sysSkills F = categories
        # qualified). Deliberately NOT folded into sysSkills K / nrSKelig,
        # which stays the mid-academy "on track" label on CadetProfile.
        "S": ('IF($B{r}="","",IF(N(sysSkills!$F{r})>='
              'SUMPRODUCT((rngSM_cat<>"")*1),"Yes","No"))', "fx"),
        # ch.41 requires 70%+ on BOTH firearms courses of fire. sysSkills P
        # computed that and was consumed by NOTHING except one audit line
        # that only counted the outright "No", so a cadet who never fired
        # one course of fire (P = "(one pending)") — or whose firearms rows
        # carry no Course of Fire tag at all (P = "") — printed "Eligible"
        # on the graduation checklist. Only an affirmative "Yes" passes.
        "T": ('IF($B{r}="","",IF(sysSkills!$P{r}="Yes","Yes","No"))', "fx"),
        # WHICH engine-raised reviews are still open, counted per trigger
        # against the reviews recorded, approved and tagged on the
        # DismissalLog. This is the single source of truth: M is derived from
        # it, sysGrades T and sysFlags R read M, and sysFlags prints this
        # string verbatim on WatchList, so a trigger that has been closed can
        # never be named as the reason a cadet is still flagged.
        "U": ('IF($B{r}="","",LET('
              'oe,N(sysGrades!$P{r}),os,N(sysSkills!$I{r}),'
              'oi,N(sysIncidents!$J{r}),'
              'ce,COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Failed retest (exam)",nrDIS_Outcome,"Retained",nrDIS_Approval,"Yes")+COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Failed retest (exam)",nrDIS_Outcome,"Retained w/ Plan",nrDIS_Approval,"Yes"),'
              'cs,COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Skills failed out",nrDIS_Outcome,"Retained",nrDIS_Approval,"Yes")+COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Skills failed out",nrDIS_Outcome,"Retained w/ Plan",nrDIS_Approval,"Yes"),'
              'ci,COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Chain-of-command incident review",nrDIS_Outcome,"Retained",nrDIS_Approval,"Yes")+COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Chain-of-command incident review",nrDIS_Outcome,"Retained w/ Plan",nrDIS_Approval,"Yes"),'
              # policy: the SECOND unexcused missed exam is a removal
              # trigger. It used to raise a sysFlags line reading "REMOVAL
              # TRIGGER" that was wired to nothing: no review opened, no
              # graduation block, and no Closes Trigger value it could ever
              # be closed with. It is now one of the four engine triggers.
              # counted as "one review per unexcused absence AFTER the
              # first", so ONE board decision closes it - the same
              # count-what-the-board-decides lesson as sysSkills I, where an
              # open count of 2 against 1 closure blocked a cadet forever.
              'om,MAX(0,COUNTIFS(nrES_PID,$B{r},nrES_Absence,"Unexcused")-1),'
              'cm,COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Unexcused missed exams",nrDIS_Outcome,"Retained",nrDIS_Approval,"Yes")+COUNTIFS(nrDIS_PID,$B{r},nrDIS_Closes,"Unexcused missed exams",nrDIS_Outcome,"Retained w/ Plan",nrDIS_Approval,"Yes"),'
              'TRIM(IF(oe>ce,"failed retest; ","")'
              '&IF(os>cs,"skills failed out; ","")'
              '&IF(om>cm,"unexcused missed exams; ","")'
              '&IF(oi>ci,"chain-of-command incident review; ",""))))', "fx"),
        # policy: an EXCUSED absence DELAYS the first attempt - it is not a
        # zero and starts no clock, so nothing else on this sheet could see
        # that the cadet still owes the exam. Graduation waits for it.
        "V": ('IF($B{r}="","",IF(COUNTIFS(nrES_PID,$B{r},nrES_Absence,'
              '"Excused",nrES_Raw,"")=0,"Yes","No"))', "fx"),
        # the Audit sheet's per-cadet enrollment-documents grid (Rule 217.1
        # enrollment file: app, TCLEDDS L1, medical L2, psych L3,
        # background, photo ID/DL, Rules Ack) was consumed by NOTHING - no
        # named range, no audit line, no gate - so a cadet with an empty
        # enrollment file printed "Eligible". Matched by PID, never by a
        # hard-coded Audit row: the docs grid moves down every time an audit
        # check is added.
        "W": ('IF($B{r}="","",IF(IFERROR(INDEX(nrENRall,'
              'MATCH($B{r},rngCadetPIDs,0)),"")="Yes","Yes","No"))', "fx"),
    })
    fill_rows(ws, FIRST, LAST, cols)
    cf_yes_no(ws, f"N{FIRST}:N{LAST}")
    define(wb, "nrCKdismissRev", "sysChecks", f"$M${FIRST}:$M${LAST}")
    define(wb, "nrCKgradElig", "sysChecks", f"$N${FIRST}:$N${LAST}")
    define(wb, "nrCKblocking", "sysChecks", f"$O${FIRST}:$O${LAST}")
    define(wb, "nrCKfinalExamElig", "sysChecks", f"$P${FIRST}:$P${LAST}")
    define(wb, "nrCKexamsRecorded", "sysChecks", f"$R${FIRST}:$R${LAST}")
    define(wb, "nrCKskillsAssessed", "sysChecks", f"$S${FIRST}:$S${LAST}")
    define(wb, "nrCKfirearmsCoF", "sysChecks", f"$T${FIRST}:$T${LAST}")
    define(wb, "nrCKopenReviews", "sysChecks", f"$U${FIRST}:$U${LAST}")
    define(wb, "nrCKexamsPending", "sysChecks", f"$V${FIRST}:$V${LAST}")
    define(wb, "nrCKenrollDocs", "sysChecks", f"$W${FIRST}:$W${LAST}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "O": 50, "U": 44, "W": 12})
    sheet_note(ws, "Graduation gate per policy: 70 in each category, every "
                   "category actually recorded (Exams Recorded), under "
                   "attendance caps, makeup complete, EVERY skills category "
                   "actually qualified (Skills Assessed — no records is not "
                   "a pass), BOTH firearms courses of fire passed (Firearms "
                   "CoF — 'one pending' and an untagged Course of Fire are "
                   "blocks, not passes), "
                   "writing current, no open chain-of-command incident "
                   "review, final PT PASSED (a blank or '(rubric pending)' "
                   "final PT blocks — it is not a pass), all cert copies on "
                   "file, the Audit sheet's enrollment-document checklist "
                   "complete for the cadet (Enroll Docs), "
                   "no open dismissal review. 'Final Exam Elig' "
                   "enforces 500.1.H. Locked.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_sysawards(wb):
    ws = wb.create_sheet("sysAwards")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Award", "Computed Winner", "Basis", "Coordinator Override",
                    "FINAL Winner", "Notes"])
    r = FIRST
    rows = [
        ("Valedictorian",
         'IFERROR(INDEX(rngCadetNames,MATCH(1,nrGRrank,0)),"")',
         '"Rank #1 — final grade "&IFERROR(INDEX(nrGRfinal,MATCH(1,nrGRrank,0)),"n/a")'),
        ("Physical Fitness",
         'IFERROR(INDEX(rngCadetNames,MATCH(MAX(nrPT_FinalPts),nrPT_FinalPts,0)),"")',
         '"Top final PT points: "&IFERROR(MAX(nrPT_FinalPts),"n/a")'),
        ("Top Gun",
         'IFERROR(INDEX(rngCadetNames,MATCH(MAX(nrSKfirearmsAvg),nrSKfirearmsAvg,0)),"")',
         '"Top firearms average: "&IFERROR(ROUND(MAX(nrSKfirearmsAvg),2),"n/a")'),
        ("Grit (most improved)",
         'IFERROR(INDEX(rngCadetNames,MATCH(MAX(nrPT_Improve),nrPT_Improve,0)),"")',
         '"Best improvement index (PT deltas): "&IFERROR(MAX(nrPT_Improve),"n/a")'
         '&"% — verify with academics trend before finalizing"'),
    ]
    for name, winner_fx, basis_fx in rows:
        ws.cell(row=r, column=2, value=name).font = F_LABEL
        c = ws.cell(row=r, column=3, value="=" + winner_fx)
        c.font = F_CALC
        c.fill = FILL_CALC
        c.border = BOX
        c2 = ws.cell(row=r, column=4, value="=" + basis_fx)
        c2.font = F_SMALL
        c2.alignment = A_LEFT_WRAP
        from xlb import FILL_INPUT, F_INPUT
        o = ws.cell(row=r, column=5)
        o.fill = FILL_INPUT
        o.font = F_INPUT
        o.border = BOX
        f = ws.cell(row=r, column=6,
                    value=f'=IF($E{r}<>"",$E{r},$C{r})')
        f.font = F_LABEL
        f.border = BOX
        n = ws.cell(row=r, column=7)
        n.fill = FILL_INPUT
        n.font = F_INPUT
        n.border = BOX
        ws.row_dimensions[r].height = 30
        r += 1
    from xlb import dv_list as _dv
    _dv(ws, "=rngCadetNames", [f"E{FIRST}:E{r-1}"])
    define(wb, "nrAWfinal", "sysAwards", f"$F${FIRST}:$F${r-1}")
    define(wb, "nrAWnames", "sysAwards", f"$B${FIRST}:$B${r-1}")
    col_widths(ws, {"A": 3, "B": 22, "C": 24, "D": 52, "E": 24, "F": 24,
                    "G": 30})
    sheet_note(ws, "Data-driven picks with coordinator override (override "
                   "always wins). Grit is decision support — the computed "
                   "index only sees PT deltas; weigh academic improvement "
                   "and perseverance before finalizing. Only the override (E) "
                   "and Notes (G) columns are editable — the computed winner "
                   "and the FINAL winner feed the printed transcript.")
    # locked like every other engine sheet: C/F are formulas that reach the
    # OFFICIAL TRANSCRIPT's awards line. E and G are the reason this sheet
    # exists, so they are explicitly unlocked before protection goes on.
    protect(ws)
    from xlb import unlock_range as _unlock
    _unlock(ws, f"E{FIRST}:E{r-1}")
    _unlock(ws, f"G{FIRST}:G{r-1}")
    return ws


# --------------------------------------------------------------------------
# Module-level so the printable Audit sheet can size its mirror block from
# len(AUDIT_CHECKS) instead of a hard-coded literal — a check added here but
# not printed there is a red Dashboard tile with no visible cause.
AUDIT_CHECKS = [
        ("Delivered hours vs 736 (exact)",
         "nrCHtotalDelivered", "cfgRequiredHours",
         'IF(N(nrCHtotalDelivered)<N(cfgRequiredHours),"SHORT",'
         'IF(N(nrCHtotalDelivered)>N(cfgRequiredHours),"OVER - report excess as #101","OK"))',
         '"Report the BPOC at exactly 736 hrs; excess goes to the reporting course '
         'the Addendum sheet names - #2040 Arrest and Control, #2046 Professional '
         'Police Driving, #2055 Firearms, #101 Addendum to BPOC for the rest"'),
        # <>0 (not >0): a chapter whose delivered hours went NEGATIVE (a
        # swapped start/end time on a Schedule block) is the most broken
        # case there is and must not be skipped as "not taught yet".
        ("Chapters short of TCOLE minimum",
         'SUMPRODUCT((nrCHdeliv<>"")*(nrCHdeliv<>0)*(nrCHdeliv<nrCHmin))', "0",
         'IF(SUMPRODUCT((nrCHdeliv<>"")*(nrCHdeliv<>0)*(nrCHdeliv<nrCHmin))=0,"OK","CHECK")',
         '"Rule 218.1(C)(4): failure to meet minimum course length = denial of training"'),
        ("Chapter training files incomplete",
         'SUMPRODUCT((nrCHname<>"")*(nrCHfileOK="No"))', "0",
         'IF(SUMPRODUCT((nrCHname<>"")*(nrCHfileOK="No"))=0,"OK","CHECK")',
         '"SME lesson plan, instructor bio, sign-ins w/ PID, assessment, grade sheet, eval"'),
        ("Special TCOLE chapter requirements unmet",
         'SUMPRODUCT((nrCHspecial<>"")*(nrCHspecialMet<>"Yes")*(nrCHspecialMet<>"N/A"))', "0",
         'IF(SUMPRODUCT((nrCHspecial<>"")*(nrCHspecialMet<>"Yes")*(nrCHspecialMet<>"N/A"))=0,"OK","CHECK")',
         '"TIM, SFST, ALERRT, CPR/AED, TCIC #4800, Crime Stoppers, Seven Step, canine scenarios, CIT lead, SIT 3232"'),
        ("Chapters without evals collected",
         'SUMPRODUCT((nrCHname<>"")*(nrCHevals<>"On File")*(nrCHevals<>"N/A"))', "0",
         'IF(SUMPRODUCT((nrCHname<>"")*(nrCHevals<>"On File")*(nrCHevals<>"N/A"))=0,"OK","CHECK")',
         '"Course critiques per chapter (print from Print Center)"'),
        # N/A is an offered answer on the Audit sheet's own dropdown and its
        # conditional format paints it green — the engine must agree, or one
        # legitimate N/A pins this row red forever with nothing to fix.
        ("Program requirements unchecked (Audit sheet)",
         'SUMPRODUCT((nrPRGitems<>"")*(nrPRGmet<>"Yes")*(nrPRGmet<>"N/A"))', "0",
         'IF(SUMPRODUCT((nrPRGitems<>"")*(nrPRGmet<>"Yes")*(nrPRGmet<>"N/A"))=0,"OK","CHECK")',
         '"Rule 215.9 distribution, TCLEDDS roster, 736 exact reporting, facility, assessments"'),
        ("Teaching instructors lacking documentation",
         'SUMPRODUCT((nrInstrOnSched="Yes")*(nrInstrReady="No"))', "0",
         'IF(SUMPRODUCT((nrInstrOnSched="Yes")*(nrInstrReady="No"))=0,"OK","CHECK")',
         '"IRG: every instructor/co-teacher teaching an LO needs bio + TCOLE cert or SME letter"'),
        ("Schedule blocks with unrecognized instructor",
         'COUNTIF(nrSCH_InstrOK,"UNRECOGNIZED")', "0",
         'IF(COUNTIF(nrSCH_InstrOK,"UNRECOGNIZED")=0,"OK","CHECK")',
         '"Instructor text matches nobody on the Instructors roster - fix spelling or add them"'),
        ("Blocks taught outside the topic's certified bank",
         'COUNTIF(nrSCH_BankOK,"NOT IN BANK")', "0",
         'IF(COUNTIF(nrSCH_BankOK,"NOT IN BANK")=0,"OK","CHECK")',
         '"Someone is scheduled for a topic they are not in the InstructorBanks '
         'certified pool for - add them to that bank with the documentation that '
         'justifies it, or reassign the block"'),
        ("Instructor cert expired before a class they taught",
         'SUMPRODUCT((nrInstrOnSched="Yes")*(LEFT(nrInstrCertStat,7)="EXPIRED"))',
         "0",
         'IF(SUMPRODUCT((nrInstrOnSched="Yes")*(LEFT(nrInstrCertStat,7)="EXPIRED"))'
         '=0,"OK","CHECK")',
         '"Cert Expiration is earlier than the last date that instructor appears on '
         'the Schedule - TCOLE checks the instructor was licensed ON the day taught"'),
        ("Teaching instructor certs missing or expiring mid-academy",
         'SUMPRODUCT((nrInstrOnSched="Yes")*((nrInstrCertStat="MISSING EXPIRATION")'
         '+(LEFT(nrInstrCertStat,5)="RENEW")))', "0",
         'IF(SUMPRODUCT((nrInstrOnSched="Yes")*((nrInstrCertStat="MISSING EXPIRATION")'
         '+(LEFT(nrInstrCertStat,5)="RENEW")))=0,"OK","CHECK")',
         '"No expiration date recorded, or it falls inside the academy - collect the '
         'current certificate copy and update the Instructors sheet"'),
        # a swapped start/end time is arithmetic, not a typo Excel can see:
        # the block's hours land in ChapterMaster "Delivered Hrs" and in the
        # Settings schedule-minutes detector (academy-length reference)
        ("Schedule blocks with impossible times",
         'SUMPRODUCT((nrSCH_TimeCheck<>"")*(nrSCH_TimeCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrSCH_TimeCheck<>"")*(nrSCH_TimeCheck<>"OK"))=0,'
         '"OK","CHECK")',
         '"End time is at or before Start: the block\'s hours are wrong, '
         'which moves delivered chapter hours and the academy-length figure"'),
        # CHECK DATE rows are counted too: a retest dated on a weekend or a
        # holiday used to sit at "Pending" forever, so the 5-class-day clock
        # was unenforceable for exactly the rows that needed it.
        ("Overdue / late retests (incl. unusable dates)",
         'COUNTIF(nrES_RetStat,"OVERDUE")+COUNTIF(nrES_RetStat,"CHECK DATE")'
         '+COUNTIF(nrES_RetStat,"LATE RETEST*")'
         '+COUNTIF(nrES_RetStat,"RETEST UNDATED")'
         '+COUNTIF(nrES_RetStat,"RETEST DATE UNCHECKED")', "0",
         'IF(COUNTIF(nrES_RetStat,"OVERDUE")+COUNTIF(nrES_RetStat,"CHECK DATE")'
         '+COUNTIF(nrES_RetStat,"LATE RETEST*")'
         '+COUNTIF(nrES_RetStat,"RETEST UNDATED")'
         '+COUNTIF(nrES_RetStat,"RETEST DATE UNCHECKED")'
         '=0,"OK","ACT NOW")',
         '"Policy 300.5: retest within 5 class days. CHECK DATE = the exam date is '
         'missing or resolves past the last class day, so no deadline exists. '
         'LATE RETEST = the retest WAS taken, but after the 5-class-day '
         'deadline - it used to read a clean Retested and leave no trace"'),
        ("Cadets with makeup owed",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrATTmakeupOK="No"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrATTmakeupOK="No"))=0,"OK","CHECK")',
         '"Missed classroom minutes / PT sessions not yet made up (active cadets)"'),
        ("Active cadets failing a category",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrGRacademic="No"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrGRacademic="No"))=0,"OK","CHECK")',
         '"70 in each category (policy 300.2)"'),
        ("Spelling interventions open",
         'COUNTIF(nrSpellFlag,"INTERVENTION")', "0",
         'IF(COUNTIF(nrSpellFlag,"INTERVENTION")=0,"OK","CHECK")',
         '"Document intervention on Counseling log (300.4.B)"'),
        ("Cadets missing certification copies",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCERTmissing<>""))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrCERTmissing<>""))=0,"OK","CHECK")',
         '"TIM, SFST, TCIC, CPR/AED, ALERRT, ICS copies - see Certifications sheet"'),
        ("Certification requirements waived as N/A",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCERTwaived<>""))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrCERTwaived<>""))=0,"OK","CHECK")',
         '"Copy? = N/A removes that cert from the graduation gate entirely. '
         'Legitimate in places, but every waiver must be defensible - see the '
         'Certifications sheet\'s Waived (N/A) column for who and which"'),
        ("Cadets missing enrollment documents",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrENRall="No"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrENRall="No"))=0,'
         '"OK","CHECK")',
         '"Rule 217.1 enrollment file: application, TCLEDDS L1, medical L2, '
         'psych L3, background, photo ID/DL and the Rule 215.9 rules '
         'acknowledgment - all seven on the Audit sheet grid. Also blocks '
         'graduation (GradChecklist Enroll Docs)"'),
        # "(one pending)" — one course of fire never fired — is a fail of
        # this check too. Counting only the outright "No" let a half-fired
        # qualification read OK.
        ("Firearms: both courses of fire not passed",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrSKbothCoF<>"")*'
         '(nrSKbothCoF<>"Yes"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrSKbothCoF<>"")*'
         '(nrSKbothCoF<>"Yes"))=0,"OK","CHECK")',
         '"IRG requires 70%+ on BOTH firearms courses of fire (ch 41). '
         '\'(one pending)\' — only one course of fire on record — counts here"'),
        # the final-PT pass test used to compare the SUM of whatever rubric
        # points were entered against the minimum, so ONE event scored high
        # opened both the graduation gate and the final-exam gate. PT!AB now
        # reads "Incomplete" until all seven events are scored; this is the
        # line that says which cadets are sitting in that state.
        ("Cadets with a partially scored final PT",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrPT_FinalPass="Incomplete"))',
         "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*'
         '(nrPT_FinalPass="Incomplete"))=0,"OK","CHECK")',
         '"Fewer than 7 of the 7 final PT events have rubric points on the PT '
         'sheet. The partial total is NOT a pass - it blocks graduation and '
         'the Final Exam (500.1.H) until every event is scored"'),
        # the line used to fire on the FIRST unexcused absence, not the
        # second its title names, and the "removal review" it promised was
        # opened by nothing. It now counts the reviews sysChecks actually
        # holds open, so a board that retains the cadet clears it.
        ("Unexcused missed exams (2nd = removal review)",
         'SUMPRODUCT((nrCadetStatus="Active")*(LEN(nrCKopenReviews)-'
         'LEN(SUBSTITUTE(nrCKopenReviews,"unexcused missed exams",""))>0))',
         "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(LEN(nrCKopenReviews)-'
         'LEN(SUBSTITUTE(nrCKopenReviews,"unexcused missed exams",""))>0))=0,'
         '"OK","CHECK")',
         '"Policy: unexcused absence from an exam records a 0 and starts the retest clock; a SECOND occurrence opens a removal review that blocks graduation until the board closes it on the DismissalLog (Closes Trigger = Unexcused missed exams)"'),
        ("Exams still owed on an excused absence",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCKexamsPending="No"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrCKexamsPending="No"))=0,"OK","CHECK")',
         '"Excused absence delays the FIRST attempt - the exam is still owed and blocks graduation until it is taken"'),
        ("Advisory board met within last 12 months",
         'IF(COUNT(nrAB_Date)=0,"none",TEXT(MAX(nrAB_Date),"mm/dd/yyyy"))', "recent",
         'IF(COUNT(nrAB_Date)=0,"CHECK",IF(MAX(nrAB_Date)>=TODAY()-366,'
         '"OK","CHECK"))',
         '"Board meets 1-2x/year; running list + minutes folder on AdvisoryBoard sheet"'),
        ("Governance alignment done for this academy",
         'cfgBoardReviewed&" / "&cfgRulesAligned', "Yes / Yes",
         'IF(AND(cfgBoardReviewed="Yes",cfgRulesAligned="Yes"),"OK","CHECK")',
         '"Board minutes reviewed + rule changes applied to this workbook (Startup Review)"'),
        ("Policy manual version recorded",
         'IF(cfgPolicyVersion="","(blank)",cfgPolicyVersion)', "set",
         'IF(cfgPolicyVersion="","CHECK","OK")',
         '"Which policy manual this academy runs under (AdvisoryBoard sheet)"'),
        ("Log rows with unrecognized cadet name",
         'COUNTIF(nrES_PID,"?")+COUNTIF(nrAT_PID,"?")+COUNTIF(nrMK_PID,"?")'
         '+COUNTIF(nrSK_PID,"?")+COUNTIF(nrIN_PID,"?")+COUNTIF(nrCO_PID,"?")'
         '+COUNTIF(nrME_PID,"?")+COUNTIF(nrMD_PID,"?")', "0",
         'IF(COUNTIF(nrES_PID,"?")+COUNTIF(nrAT_PID,"?")+COUNTIF(nrMK_PID,"?")'
         '+COUNTIF(nrSK_PID,"?")+COUNTIF(nrIN_PID,"?")+COUNTIF(nrCO_PID,"?")'
         '+COUNTIF(nrME_PID,"?")+COUNTIF(nrMD_PID,"?")=0,"OK","CHECK")',
         '"A log row\'s cadet name matches nobody on the Cadets roster - '
         'that cadet\'s scores/events are orphaned; fix the spelling"'),
        # every SUMIFS/COUNTIFS in the engine is keyed on PID, and every log
        # sheet resolves PID by MATCH on the cadet NAME, so a repeat of
        # either silently MERGES two cadets' records in both directions.
        ("Duplicate cadet PID or name on the roster",
         'SUMPRODUCT((rngCadetPIDs<>"")*(COUNTIF(rngCadetPIDs,rngCadetPIDs)>1))'
         '+SUMPRODUCT((rngCadetNames<>"")*(COUNTIF(rngCadetNames,rngCadetNames)>1))',
         "0",
         'IF(SUMPRODUCT((rngCadetPIDs<>"")*(COUNTIF(rngCadetPIDs,rngCadetPIDs)>1))'
         '+SUMPRODUCT((rngCadetNames<>"")*(COUNTIF(rngCadetNames,rngCadetNames)>1))'
         '=0,"OK","CHECK")',
         '"Two Cadets rows sharing a PID (or a name) merge both cadets\' grades, '
         'attendance and gates - and swap them onto each other\'s transcript. '
         'The duplicate cells are highlighted red on Cadets"'),
        ("Exam rows failing Row Check",
         'SUMPRODUCT((nrES_RowCheck<>"")*(nrES_RowCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrES_RowCheck<>"")*(nrES_RowCheck<>"OK"))=0,"OK","CHECK")',
         '"ExamScores Row Check: a non-numeric or out-of-range Raw Score, a '
         'duplicate RecordID, an attempt 2 with no attempt 1 on file, or a '
         'retest row logged with no score - each silently distorts the '
         'category average or stops the 5-class-day retest clock"'),
        # A44: sysAudit rolled up ExamScores Row Check failures but had no
        # equivalent line for Makeup - "NO SUCH EVENT", "WRONG CADET",
        # "UNIT MISMATCH", a dateless clear or a negative credit were
        # visible only in Makeup column N, thirteen columns right of the
        # entry area, and never reached the Audit sheet or the Dashboard.
        ("Makeup rows failing Row Check",
         'SUMPRODUCT((nrMK_RowCheck<>"")*(nrMK_RowCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrMK_RowCheck<>"")*(nrMK_RowCheck<>"OK"))=0,'
         '"OK","CHECK")',
         '"Makeup Row Check: an unlinked row, another cadet\'s event, an '
         'EventID that no longer exists, a Type the caps do not credit, '
         'minutes booked against a session-counted PT event, a missing '
         'makeup DATE (which fabricates a CLEARED date on the attendance '
         'ledger) or a zero/NEGATIVE credit (which increases owed time). '
         'None of these earns credit, and each one moves the policy-400 '
         'ledger the wrong way"'),
        # A10 / A23: the same for Attendance, which had no Row Check at all
        ("Attendance rows failing Row Check",
         'SUMPRODUCT((nrAT_RowCheck<>"")*(nrAT_RowCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrAT_RowCheck<>"")*(nrAT_RowCheck<>"OK"))=0,'
         '"OK","CHECK")',
         '"Attendance Row Check: a counted absence with no minutes (or no '
         'sessions on a PT event) is dropped from BOTH caps and never shows '
         'OPEN; a row carrying the unit that does not match its own Is PT? '
         'flag has that value discarded from the balance, the makeup '
         'reconciliation and the caps"'),
        # A12: an open dismissal review blocked graduation but was counted
        # nowhere - not on the Audit sheet, not on the Dashboard tile
        ("Open dismissal reviews (active cadets)",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCKdismissRev="Yes"))', "0",
         'IF(SUMPRODUCT((nrCadetStatus="Active")*(nrCKdismissRev="Yes"))=0,'
         '"OK","CHECK")',
         '"Failed retest, skills failed out, or an open chain-of-command '
         'incident review. Policy 600.2.E: open a formal review on the '
         'DismissalLog, then close it there — Closes Trigger set to the '
         'trigger, Outcome Retained or Retained w/ Plan, Asst. Chief '
         'Approval Yes. Closure is per trigger, so a later one re-opens"'),
        # A29: a closure date that is not a real date is DISCARDED by the
        # class-day calendar, which moves every later class day silently
        ("Extra closure dates that are not dates",
         'SUMPRODUCT((nrExtraClosureCheck<>"")*'
         '(LEFT(nrExtraClosureCheck,2)<>"OK"))', "0",
         'IF(SUMPRODUCT((nrExtraClosureCheck<>"")*'
         '(LEFT(nrExtraClosureCheck,2)<>"OK"))=0,"OK","CHECK")',
         '"Control > Extra Closure Dates: a pasted text date or a label is '
         'ignored by the class-day calendar, so the academy silently runs a '
         'day long and every Day #, retest deadline, memo due date and '
         'writing date shifts"'),
        ("Skills rows failing Row Check",
         'SUMPRODUCT((nrSK_RowCheck<>"")*(nrSK_RowCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrSK_RowCheck<>"")*(nrSK_RowCheck<>"OK"))=0,'
         '"OK","CHECK")',
         '"Skills Row Check: a Score that is text or outside 0-100 counts '
         'toward nothing - not the transcript firearms average, not Top Gun '
         'and not the ch.41 both-courses-of-fire graduation gate. Also flags '
         'a decided row with no score, a score on a Pass/Fail category and a '
         'firearms result with no Course of Fire"'),
        ("Spelling rows failing Row Check",
         'SUMPRODUCT((nrSpellRowCheck<>"")*(nrSpellRowCheck<>"OK"))', "0",
         'IF(SUMPRODUCT((nrSpellRowCheck<>"")*(nrSpellRowCheck<>"OK"))=0,'
         '"OK","CHECK")',
         '"A spelling score that is text or outside 0-100 inflates the '
         'cadet average and can hide the policy 300.4.B intervention flag"'),
        ("Memos with an unusable due date",
         'COUNTIF(nrME_Status,"CHECK DATE")', "0",
         'IF(COUNTIF(nrME_Status,"CHECK DATE")=0,"OK","CHECK")',
         '"The Assigned date is missing, or resolves past the last class day, '
         'so no due date could be computed and the memo can never read '
         'OVERDUE"'),
        # NOTE: no audit row for "skills not all assessed". That condition is
        # true for every cadet for most of the academy, so an audit row would
        # sit red from day one and train the coordinator to ignore the block.
        # It is enforced per cadet instead — sysChecks S -> GradChecklist
        # ELIGIBLE + the named Blocking Issue.
    ]


def build_sysaudit(wb):
    """TCOLE audit-readiness rollups that need engine math."""
    ws = wb.create_sheet("sysAudit")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Check", "Value", "Target", "Status", "Detail"])
    r = FIRST
    checks = AUDIT_CHECKS
    for name, val_fx, target, stat_fx, detail_fx in checks:
        ws.cell(row=r, column=2, value=name).font = F_LABEL
        v = ws.cell(row=r, column=3, value="=" + val_fx)
        v.font = F_CALC
        v.border = BOX
        # cfg* targets are live formulas; digits are numbers; anything else
        # is a literal display string (never prefix those with "=")
        t = ws.cell(row=r, column=4, value="=" + target if target.startswith("cfg")
                    else (int(target) if target.isdigit() else target))
        t.font = F_CALC
        s = ws.cell(row=r, column=5, value="=" + stat_fx)
        s.font = F_LABEL
        s.border = BOX
        d = ws.cell(row=r, column=6, value="=" + detail_fx)
        d.font = F_SMALL
        d.alignment = A_LEFT_WRAP
        r += 1
    cf_formula(ws, f"E{FIRST}:E{r-1}", f'$E{FIRST}<>"OK"', FILL_WARNBG)
    cf_formula(ws, f"E{FIRST}:E{r-1}", f'$E{FIRST}="OK"', FILL_OKBG)
    define(wb, "cfgRequiredHours", "sysAudit", f"$H${FIRST}")
    ws.cell(row=FIRST, column=8, value=736).font = F_SMALL
    define(wb, "nrAUDstatus", "sysAudit", f"$E${FIRST}:$E${r-1}")
    define(wb, "nrAUDcheck", "sysAudit", f"$B${FIRST}:$B${r-1}")
    col_widths(ws, {"A": 3, "B": 36, "C": 10, "D": 9, "E": 11, "F": 58,
                    "G": 3, "H": 8})
    sheet_note(ws, "Engine half of the TCOLE Audit sheet — the printable "
                   "Audit sheet reads these. Locked.")
    protect(ws)
    return ws


def build_all_engine(wb):
    build_sysgrades(wb)
    build_sysattendance(wb)
    build_sysskills(wb)
    build_sysincidents(wb)
    build_sysflags(wb)
    build_syschecks(wb)
    build_sysawards(wb)
    build_sysaudit(wb)
