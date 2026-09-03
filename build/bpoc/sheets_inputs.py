"""Input sheets: Cadets, ExamScores, Spelling, Attendance, Makeup, Skills,
Writing, Incidents, Counseling, PT, Medical.

All keyed by PID; cadet grids run rows 6..55 (50 cadets); logs are long
tables. Blue cells = staff entry, gray = calculated.
"""
from openpyxl.utils import get_column_letter

from xlb import (
    HDR_ROW, DATA_ROW, CADETS, CADET_LAST, ROWS_EXAMSCORES, ROWS_ATTEND,
    ROWS_MAKEUP, ROWS_SKILLS, ROWS_INCIDENTS, ROWS_COUNSELING, ROWS_MEDICAL,
    F_HDR, F_CALC, FILL_CALC, F_LABEL, F_SMALL, F_INPUT, FILL_INPUT,
    FILL_YELLOW, A_LEFT_WRAP, A_CENTER, BOX, DATE, TIME,
    header_row, fill_rows, dv_list, dv_whole, sheet_note, cf_yes_no, cf_formula,
    FILL_WARNBG, FILL_OKBG, FILL_AMBER, col_widths, define,
)
import data_lists as DL
import data_writing as DW


# --------------------------------------------------------------------------
def build_cadets(wb):
    ws = wb.create_sheet("Cadets")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Last Name", "First Name", "MI", "Cadet Name",
                    "AgencyID", "Agency Name", "Status", "Sort",
                    "Enroll Date", "Separation Date", "Separation Reason"])
    first, last = DATA_ROW, CADET_LAST
    fill_rows(ws, first, last, {
        "B": (None, "in"), "C": (None, "in"), "D": (None, "in"),
        "E": (None, "in"),
        "F": ('IF($B{r}="","",$C{r}&", "&$D{r}&IF($E{r}="",""," "&$E{r}&"."))', "fx"),
        "G": (None, "in"),
        "H": ('IF($G{r}="","",IFERROR(INDEX(rngAgencyNames,MATCH($G{r},rngAgencyIDs,0)),"?"))', "fx"),
        "I": (None, "in"),
        "J": ('IF($B{r}="","",ROW()-{})'.replace("{}", str(HDR_ROW)), "fx"),
        "K": (None, "in"), "L": (None, "in"), "M": (None, "in"),
    })
    for r in range(first, last + 1):
        ws[f"K{r}"].number_format = DATE
        ws[f"L{r}"].number_format = DATE
    dv_list(ws, "=rngAgencyIDs", [f"G{first}:G{last}"])
    dv_list(ws, "=lstCadetStatus", [f"I{first}:I{last}"])
    # Two rows sharing a PID (a filled-down roster row, a mistyped digit)
    # merge both cadets' records in BOTH directions: every engine sheet is
    # mirrored by row off Cadets!B and then keyed on that PID, so one cadet
    # inherits the other's absences and gates while the other inherits their
    # grades and class rank — onto the printed transcript. A repeated
    # computed NAME does the same thing, because every log sheet resolves
    # PID by MATCH on the name and MATCH returns the first hit.
    # Added BEFORE the not-Active gray-out rule in gray_separated_rows() so
    # this warning takes CF priority over the gray fill.
    cf_formula(ws, f"B{first}:B{last}",
               f'AND($B{first}<>"",COUNTIF(rngCadetPIDs,$B{first})>1)',
               FILL_WARNBG)
    cf_formula(ws, f"F{first}:F{last}",
               f'AND($F{first}<>"",COUNTIF(rngCadetNames,$F{first})>1)',
               FILL_WARNBG)
    define(wb, "rngCadetPIDs", "Cadets", f"$B${first}:$B${last}")
    define(wb, "rngCadetNames", "Cadets", f"$F${first}:$F${last}")
    define(wb, "nrCadetAgency", "Cadets", f"$H${first}:$H${last}")
    define(wb, "nrCadetAgencyID", "Cadets", f"$G${first}:$G${last}")
    define(wb, "nrCadetStatus", "Cadets", f"$I${first}:$I${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 14, "D": 14, "E": 5, "F": 24,
                    "G": 10, "H": 22, "I": 11, "J": 6, "K": 12, "L": 12,
                    "M": 26})
    sheet_note(ws, "PID never changes once set — every other sheet keys off "
                   "it. Capacity 50 cadets. A red PID (or red Cadet Name) "
                   "means two rows share it: fix it immediately, both "
                   "cadets' records are being merged.")
    return ws


# --------------------------------------------------------------------------
def build_examscores(wb):
    ws = wb.create_sheet("ExamScores")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["RecordID", "Cadet Name", "PID", "Agency", "ExamCode",
                    "Exam Name", "Type", "Seq", "Passing", "Attempt #",
                    "Raw Score", "Recorded", "AttemptPass", "Pass?",
                    "FinalAttempt?", "RetakeReq?", "DismissReview?", "Date",
                    "Retest Due By", "Retest Status", "Entered By", "Notes",
                    "Row Check", "Absence"])
    first, last = DATA_ROW, DATA_ROW + ROWS_EXAMSCORES - 1
    fill_rows(ws, first, last, {
        "B": ('IF($C{r}="","",$D{r}&"-"&$F{r}&"-"&$K{r})', "fx"),
        "C": (None, "in"),
        "D": ('IF($C{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($C{r},rngCadetNames,0)),"?"))', "fx"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(nrCadetAgency,MATCH($D{r},rngCadetPIDs,0)),""))', "fx"),
        "F": (None, "in"),
        "G": ('IF($F{r}="","",IFERROR(INDEX(rngEPname,MATCH($F{r},rngEPcode,0)),"?"))', "fx"),
        "H": ('IF($F{r}="","",IFERROR(INDEX(rngEPtype,MATCH($F{r},rngEPcode,0)),"?"))', "fx"),
        "I": ('IF($F{r}="","",IFERROR(INDEX(rngEPseq,MATCH($F{r},rngEPcode,0)),""))', "fx"),
        "J": ('IF($F{r}="","",IFERROR(INDEX(rngEPpass,MATCH($F{r},rngEPcode,0)),cfgPassingScore))', "fx"),
        "K": (None, "in"), "L": (None, "in"),
        # a passed retake records at the cap — MAX(exam's own Passing, the
        # global 70) so an exam that passes above 70 is not capped below its
        # own passing mark and then reported as a fail. A raw score that is
        # not a number passes straight through (text sorts ABOVE any number,
        # so every >= test below would otherwise read as a pass); Row Check
        # flags the row instead.
        # every "a retest exists" test below requires a SCORED retest
        # (nrES_Raw,">=0"). An attempt-2 row logged when the retest is
        # SCHEDULED, before its score is entered, used to blank attempt 1's
        # Recorded score, drop the failed exam out of every category average
        # and flip AcademicElig from No to Yes while silently stopping the
        # policy-300.5 clock. Row Check reports the unscored row instead.
        # $K>2 records NOTHING: the else-branch below is written for the
        # single policy-300.5 retest, so an attempt-3 row used to record the
        # cadet's FAILED attempt-1 raw score, replacing the passed retest on
        # the grade sheet, the transcript and the agency email. Row Check
        # (column X) reports the row instead.
        "M": ('IF($C{r}="","",IF($L{r}="","",IF(NOT(ISNUMBER($L{r})),$L{r},'
              'IF(N($K{r})>2,"",'
              'IF($K{r}=1,'
              'IF(COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,2,'
              'nrES_Raw,">=0")>0,"",$L{r}),'
              'IF($L{r}>=$J{r},MAX($J{r},cfgRetakeRecordedCap),'
              'IF(COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,1)=0,$L{r},'
              'SUMIFS(nrES_Raw,nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,1))))))))', "fx"),
        # ISNUMBER guards: without them 'absent' (or a CSV-pasted text "85")
        # reported Pass? = Yes in green while the printed grade sheet, whose
        # SUMIFS coerces text to 0, printed the same record as 0 / FAIL.
        "N": ('IF($C{r}="","",IF($L{r}="","",IF(NOT(ISNUMBER($L{r})),'
              '"(not a score)",IF($L{r}>=$J{r},"Yes","No"))))', "fx"),
        "O": ('IF($C{r}="","",IF($M{r}="","",IF(NOT(ISNUMBER($M{r})),'
              '"(not a score)",IF($M{r}>=$J{r},"Yes","No"))))', "fx"),
        # a LATER attempt only supersedes this one once it has been scored,
        # otherwise the placeholder row became the "final" attempt and the
        # real (failed) score stopped counting anywhere
        # ...and it must not knock the real retest off either: the "is there
        # a later attempt" test is bounded at attempt 2, so a stray
        # attempt-3 row cannot flip the passed retest to FinalAttempt = No
        # and erase the exam from every average and printable.
        "P": ('IF($C{r}="","",IF($K{r}="","",IF(N($K{r})>2,"No",'
              'IF(COUNTIFS(nrES_PID,$D{r},'
              'nrES_Code,$F{r},nrES_Att,">"&$K{r},nrES_Att,"<=2",'
              'nrES_Raw,">=0")=0,'
              '"Yes","No"))))', "fx"),
        "Q": ('IF($C{r}="","",IF($L{r}="","",IF(AND($K{r}=1,ISNUMBER($L{r}),'
              '$L{r}<$J{r},'
              'COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,2,'
              'nrES_Raw,">=0")=0),"Yes","No")))', "fx"),
        # a FAILED second attempt always reaches dismissal review. Requiring
        # an attempt-1 row to exist meant a missing attempt-1 flipped the
        # academic-eligibility gate from No to Yes for a cadet who failed the
        # retest; the missing attempt-1 is reported by Row Check instead.
        # >=2, not =2: a failed THIRD attempt is still a failed retest, and
        # keying it as =2 meant the row opened no dismissal review at all.
        "R": ('IF($C{r}="","",IF($L{r}="","",IF(AND(N($K{r})>=2,'
              'ISNUMBER($L{r}),'
              '$L{r}<$J{r}),"Yes","No")))', "fx"),
        "S": (None, "in"),
        # match mode 1 = exact-or-next-larger, so an exam dated on a weekend,
        # a holiday or any non-class day starts the 5-class-day clock on the
        # next class day instead of giving up. Only a date past the LAST
        # class day has no answer, and that says so out loud.
        # ...and the answer must land on a day the academy is actually IN
        # SESSION. The class-day table is padded 44 rows past cfgEndDate, so
        # keying only on "did the lookup fall off the end" let a deadline
        # roll silently onto rows the same table marks In Session? = No.
        # keyed off the FAILED ATTEMPT 1 itself, not off Q ("RetakeReq?"):
        # Q flips to "No" the moment a scored attempt 2 exists, which erased
        # the deadline and made a late retest untraceable. The deadline that
        # applied stays on the row forever so U (and the audit) can compare
        # the retest's actual date against it.
        "T": ('IF(OR($C{r}="",$L{r}="",$K{r}<>1,NOT(ISNUMBER($L{r})),'
              '$L{r}>=$J{r}),"",IF($S{r}="","(enter date)",'
              'LET(dn,XLOOKUP($S{r},nrCDdate,nrCDnum,"",1),'
              'IF(dn="","(after last class day)",'
              'LET(dd,XLOOKUP(dn+cfgRetestClassDays,nrCDnum,nrCDdate,""),'
              'ok,XLOOKUP(dn+cfgRetestClassDays,nrCDnum,nrCDinsession,"No"),'
              'IF(OR(dd="",ok<>"Yes"),"(after last class day)",dd))))))', "fx"),
        # keyed off the failed attempt-1 itself (not Q, which flips to "No"
        # once attempt 2 exists) so completed retests display "Retested".
        # A due date that could not be computed reads CHECK DATE, never the
        # reassuring "Pending" that used to hide it from every flag and the
        # audit sheet forever.
        # ...and the completed case is CHECKED, not waved through. Before,
        # any scored attempt-2 row made this read "Retested" regardless of
        # when it was taken, so a retest months past the policy 300.5 window
        # left no trace anywhere in the workbook. LATE RETEST is counted by
        # the sysFlags retest flag and by the sysAudit overdue-retest line.
        # policy: ABSENT + EXCUSED = the first attempt is simply taken
        # later - no zero, no retest clock. The row would otherwise be
        # completely silent (blank score = blank everything), so the exam a
        # cadet still owes could never be seen. ABSENT + UNEXCUSED is a
        # recorded 0, which starts the ordinary 300.5 retest clock through
        # the branches below with no special casing needed.
        "U": ('IF($C{r}="","",'
              'IF(AND($Y{r}="Excused",$L{r}="",N($K{r})<=1),'
              '"EXCUSED - 1st attempt pending",'
              'IF(OR($L{r}="",$K{r}<>1,NOT(ISNUMBER($L{r})),'
              '$L{r}>=$J{r}),"",'
              'IF(COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},'
              'nrES_Att,2,nrES_Raw,">=0")>0,'
              'IF(IFERROR(MINIFS(nrES_Date,nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,2,nrES_Raw,">=0",nrES_Date,">0"),0)=0,"RETEST UNDATED",'
              'IF(NOT(ISNUMBER($T{r})),"RETEST DATE UNCHECKED",'
              'IF(MINIFS(nrES_Date,nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,2,nrES_Raw,">=0",nrES_Date,">0")>$T{r},"LATE RETEST "&TEXT('
              'MINIFS(nrES_Date,nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,2,nrES_Raw,">=0",nrES_Date,">0"),"mm/dd")&" (due "'
              '&TEXT($T{r},"mm/dd")&")","Retested"))),'
              'IF(NOT(ISNUMBER($T{r})),"CHECK DATE",'
              'IF(TODAY()>$T{r},"OVERDUE","Due "&TEXT($T{r},"mm/dd")))))))', "fx"),
        "V": (None, "in"), "W": (None, "in"), "Y": (None, "in"),
        # row integrity: every one of these silently distorts a category
        # average or a policy deadline, and none of them was visible before
        "X": ('IF($C{r}="","",'
              'IF(AND($L{r}<>"",NOT(ISNUMBER($L{r}))),"RAW SCORE NOT A NUMBER",'
              'IF(AND(ISNUMBER($L{r}),OR($L{r}<0,$L{r}>100)),'
              '"RAW SCORE OUT OF RANGE",'
              # policy: an UNEXCUSED absence is recorded as a zero, and that
              # zero is what starts the 300.5 retest clock. Marking the row
              # Unexcused without keying the 0 would leave the cadet with no
              # score, no clock and no trace.
              'IF(AND($Y{r}="Unexcused",OR($L{r}="",NOT(ISNUMBER($L{r})),'
              '$L{r}<>0)),"UNEXCUSED ABSENCE MUST RECORD 0",'
              'IF(COUNTIF(nrES_RecID,$B{r})>1,"DUPLICATE RECORD",'
              # policy 300.5 allows ONE retest. Every rule on this sheet is
              # written for attempts 1 and 2; a third attempt used to be a
              # legal dropdown pick that recorded the FAILED first attempt
              # over the passed retest with Row Check still reading OK.
              'IF(N($K{r})>2,"ATTEMPT 3+ NOT ALLOWED - policy 300.5 permits '
              'one retest per exam; this row records nothing",'
              'IF(AND($K{r}=2,COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},'
              'nrES_Att,1)=0),"ATTEMPT 2 WITHOUT ATTEMPT 1",'
              'IF(AND($K{r}=2,$L{r}=""),"RETEST ROW HAS NO SCORE",'
              # 500.1.H: a cadet whose FINAL PT was failed may not sit the
              # Final Exam. sysChecks column P / nrCKfinalExamElig computed
              # that rule and was then referenced by nothing at all - no
              # formula, no validation, no VBA, no printable. This is where
              # it lands: the moment a Final score is keyed for a cadet whose
              # final PT reads "No", the row says so and the sysAudit
              # "Exam rows failing Row Check" line turns red.
              'IF(AND($H{r}="Final",$L{r}<>"",'
              'IFERROR(INDEX(nrCKfinalExamElig,MATCH($D{r},rngCadetPIDs,0)),'
              '"")="No"),"FINAL EXAM AFTER FAILED FINAL PT (500.1.H)",'
              # a retest keyed against an attempt 1 that already PASSED is
              # silent corruption: column M blanks the passing attempt-1
              # score and records the retake cap (70) instead, so the
              # transcript, ScoresGrid, the category average, the rank and
              # the agency email all drop to 70 with nothing anywhere saying
              # why. Every other branch here exists for exactly this class of
              # mis-key. SUMIFS ignores a text raw score; a MISSING attempt 1
              # is already reported by the ATTEMPT 2 WITHOUT ATTEMPT 1 branch
              # above, and duplicate attempt-1 rows by DUPLICATE RECORD.
              'IF(AND($K{r}=2,ISNUMBER($L{r}),'
              'COUNTIFS(nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,1)>0,'
              'SUMIFS(nrES_Raw,nrES_PID,$D{r},nrES_Code,$F{r},nrES_Att,1)'
              '>=$J{r}),"RETEST AFTER A PASSING FIRST ATTEMPT",'
              '"OK"))))))))))', "fx"),
    })
    for r in range(first, last + 1):
        ws[f"S{r}"].number_format = DATE
        ws[f"T{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"C{first}:C{last}"])
    dv_list(ws, "=rngEPcode", [f"F{first}:F{last}"])
    # policy 300.5 = ONE retest per exam. lstAttemptNum (1-5) stays on
    # Skills, which really does allow up to 5 attempts.
    dv_list(ws, "=lstExamAttemptNum", [f"K{first}:K{last}"])
    dv_list(ws, '"Excused,Unexcused"', [f"Y{first}:Y{last}"])
    cf_formula(ws, f"Y{first}:Y{last}", f'$Y{first}="Unexcused"', FILL_WARNBG)
    cf_formula(ws, f"Y{first}:Y{last}", f'$Y{first}="Excused"', FILL_AMBER)
    cf_formula(ws, f"U{first}:U{last}",
               f'LEFT($U{first},7)="EXCUSED"', FILL_AMBER)
    cf_formula(ws, f"U{first}:U{last}", f'$U{first}="OVERDUE"', FILL_WARNBG)
    cf_formula(ws, f"U{first}:U{last}", f'$U{first}="CHECK DATE"', FILL_WARNBG)
    # LEFT(...,12) never equalled the 11-character literal, so the rule was
    # dead and a late retest carried no fill on the sheet it is entered on.
    cf_formula(ws, f"U{first}:U{last}",
               f'LEFT($U{first},11)="LATE RETEST"', FILL_WARNBG)
    cf_formula(ws, f"U{first}:U{last}",
               f'LEFT($U{first},7)="RETEST "', FILL_WARNBG)
    cf_formula(ws, f"X{first}:X{last}",
               f'AND($X{first}<>"",$X{first}<>"OK")', FILL_WARNBG)
    cf_yes_no(ws, f"O{first}:O{last}")
    define(wb, "nrES_RecID", "ExamScores", f"$B${first}:$B${last}")
    define(wb, "nrES_RowCheck", "ExamScores", f"$X${first}:$X${last}")
    define(wb, "nrES_PID", "ExamScores", f"$D${first}:$D${last}")
    define(wb, "nrES_Code", "ExamScores", f"$F${first}:$F${last}")
    define(wb, "nrES_Type", "ExamScores", f"$H${first}:$H${last}")
    define(wb, "nrES_Seq", "ExamScores", f"$I${first}:$I${last}")
    define(wb, "nrES_Att", "ExamScores", f"$K${first}:$K${last}")
    define(wb, "nrES_Raw", "ExamScores", f"$L${first}:$L${last}")
    define(wb, "nrES_Rec", "ExamScores", f"$M${first}:$M${last}")
    define(wb, "nrES_AttPass", "ExamScores", f"$N${first}:$N${last}")
    define(wb, "nrES_Final", "ExamScores", f"$P${first}:$P${last}")
    define(wb, "nrES_RetReq", "ExamScores", f"$Q${first}:$Q${last}")
    define(wb, "nrES_Dis", "ExamScores", f"$R${first}:$R${last}")
    define(wb, "nrES_Date", "ExamScores", f"$S${first}:$S${last}")
    define(wb, "nrES_RetDue", "ExamScores", f"$T${first}:$T${last}")
    define(wb, "nrES_RetStat", "ExamScores", f"$U${first}:$U${last}")
    define(wb, "nrES_Absence", "ExamScores", f"$Y${first}:$Y${last}")
    col_widths(ws, {"A": 3, "B": 16, "C": 22, "D": 9, "E": 18, "F": 10,
                    "G": 34, "H": 9, "I": 6, "J": 9, "K": 9, "L": 9, "M": 10,
                    "N": 11, "O": 8, "P": 12, "Q": 11, "R": 13, "S": 11,
                    "T": 12, "U": 12, "V": 14, "W": 26, "X": 26, "Y": 12})
    sheet_note(ws, "One row per attempt. Attempt 2 of a failed exam records "
                   "at the 70 cap when passed (policy 300.5). Retest Due By "
                   "= 5 class days after the failed attempt's date (a date "
                   "that is not a class day rolls to the next one). Watch "
                   "Row Check: a score that is not a number, a duplicate "
                   "RecordID, an attempt 2 with no attempt 1 on file, or an "
                   "attempt 2 logged against a first attempt that already "
                   "PASSED each distort the category average. ABSENCE (last "
                   "column): Unexcused = key Raw Score 0 as well - the zero "
                   "is the record and it starts the retest clock; Excused = "
                   "leave Raw Score blank, the first attempt is simply taken "
                   "later and Retest Status reads EXCUSED - 1st attempt "
                   "pending until the score is keyed on this same row.")
    return ws


# --------------------------------------------------------------------------
def build_spelling(wb):
    ws = wb.create_sheet("Spelling")
    ws.sheet_view.showGridLines = False
    tests = [f"S{n:02d}" for n in range(1, 13)]
    header_row(ws, ["PID", "Cadet Name"] + tests +
               ["Spelling Avg", "# Taken", "Intervention?", "Row Check"])
    first, last = DATA_ROW, CADET_LAST
    cols = {
        "B": (f'IF(Cadets!$B{{r}}="","",Cadets!$B{{r}})', "fx"),
        "C": (f'IF(Cadets!$B{{r}}="","",Cadets!$F{{r}})', "fx"),
        # the average and the count must EXCLUDE an out-of-range score the
        # same way every ExamScores aggregate excludes one (nrES_Rec
        # ">=0"/"<=100"). A pasted 950 used to inflate this average, the
        # cadet's Current Grade, his rank and the printed valedictorian
        # pick, and it masked the policy-300.4.B INTERVENTION flag. Row
        # Check (column S) reports the bad cell; the value no longer counts.
        "P": ('IF($C{r}="","",IF($Q{r}=0,"",ROUND(SUMIFS($D{r}:$O{r},'
              '$D{r}:$O{r},">=0",$D{r}:$O{r},"<=100")/$Q{r},1)))', "fx"),
        "Q": ('IF($C{r}="","",COUNTIFS($D{r}:$O{r},">=0",'
              '$D{r}:$O{r},"<=100"))', "fx"),
        "R": ('IF($C{r}="","",IF($P{r}="","",'
              'IF($P{r}<cfgSpellInterventionAvg,"INTERVENTION","OK")))', "fx"),
        # a mistyped score (950 for 95) used to sail through: it inflated the
        # average, MASKED the policy-300.4.B INTERVENTION flag this sheet
        # exists to raise, and poisoned the class-average row behind the
        # Dashboard chart. COUNT<>COUNTA catches text; the SUMPRODUCT catches
        # out-of-range numbers. Data validation alone would not: it does not
        # fire on paste.
        "S": ('IF($C{r}="","",'
              'IF(COUNT($D{r}:$O{r})<>COUNTA($D{r}:$O{r}),'
              '"SCORE NOT A NUMBER",'
              'IF(SUMPRODUCT(($D{r}:$O{r}<>"")*'
              '((N($D{r}:$O{r})<0)+(N($D{r}:$O{r})>100)))>0,'
              '"SCORE OUT OF RANGE","OK")))', "fx"),
    }
    for i in range(12):
        cols[get_column_letter(4 + i)] = (None, "in")
    fill_rows(ws, first, last, cols)
    cf_formula(ws, f"R{first}:R{last}", f'$R{first}="INTERVENTION"', FILL_WARNBG)
    cf_formula(ws, f"S{first}:S{last}",
               f'AND($S{first}<>"",$S{first}<>"OK")', FILL_WARNBG)
    dv_whole(ws, [f"D{first}:O{last}"], 0, 100,
             "Spelling scores are 0-100 (25 words x 4 points).")
    # class stats rows below the grid
    sr = last + 2
    ws.cell(row=sr, column=3, value="Scores entered:").font = F_LABEL
    ws.cell(row=sr + 1, column=3, value="Class average:").font = F_LABEL
    ws.cell(row=sr + 2, column=3, value="Test #:").font = F_SMALL
    for i in range(12):
        cl = get_column_letter(4 + i)
        # same 0-100 mask as the per-cadet average above: this row feeds the
        # printed class average and the Dashboard spelling chart.
        ws[f"{cl}{sr}"] = (f'=COUNTIFS({cl}{first}:{cl}{last},">=0",'
                           f'{cl}{first}:{cl}{last},"<=100")')
        ws[f"{cl}{sr}"].font = F_CALC
        ws[f"{cl}{sr+1}"] = (f"=IF({cl}{sr}=0,\"\","
                             f'ROUND(AVERAGEIFS({cl}{first}:{cl}{last},'
                             f'{cl}{first}:{cl}{last},">=0",'
                             f'{cl}{first}:{cl}{last},"<=100"),1))')
        ws[f"{cl}{sr+1}"].font = F_CALC
        ws[f"{cl}{sr+2}"] = i + 1
        ws[f"{cl}{sr+2}"].font = F_SMALL
    define(wb, "nrSpellCounts", "Spelling", f"$D${sr}:$O${sr}")
    define(wb, "nrSpellClassAvg", "Spelling", f"$D${sr+1}:$O${sr+1}")
    define(wb, "nrSpellTestNums", "Spelling", f"$D${sr+2}:$O${sr+2}")
    define(wb, "nrSpellAvg", "Spelling", f"$P${first}:$P${last}")
    define(wb, "nrSpellTaken", "Spelling", f"$Q${first}:$Q${last}")
    define(wb, "nrSpellFlag", "Spelling", f"$R${first}:$R${last}")
    define(wb, "nrSpellRowCheck", "Spelling", f"$S${first}:$S${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "P": 12, "Q": 9, "R": 14,
                    "S": 22})
    for i in range(12):
        ws.column_dimensions[get_column_letter(4 + i)].width = 7
    sheet_note(ws, "Scores 0-100 (25 words x 4 pts). Average below the "
                   "Settings intervention threshold (default 75) flags "
                   "INTERVENTION per policy 300.4.B — see Counseling log. "
                   "Row Check catches a score that is text or outside "
                   "0-100: one extra digit hides the intervention flag and "
                   "moves the class average.")
    return ws


# --------------------------------------------------------------------------
def build_attendance(wb):
    ws = wb.create_sheet("Attendance")
    ws.sheet_view.showGridLines = False
    # T ("Row Check") is APPENDED at the end - no existing column moves.
    header_row(ws, ["EventID", "Date", "Cadet Name", "PID", "Agency",
                    "Event Type", "Reason", "Minutes", "Sessions", "Is PT?",
                    "Doc Status", "Excused?", "Counts?", "Notes",
                    "Made-Up Min", "Made-Up Sess", "Balance", "Cleared?",
                    "Row Check"])
    first, last = DATA_ROW, DATA_ROW + ROWS_ATTEND - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","A"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": ('IF($E{r}="","",IFERROR(INDEX(nrCadetAgency,MATCH($E{r},rngCadetPIDs,0)),""))', "fx"),
        "G": (None, "in"), "H": (None, "in"), "I": (None, "in"),
        "J": (None, "in"),
        "K": ('IF($G{r}="","",IF(OR($G{r}="PT Missed",$G{r}="PT Modified",'
              '$G{r}="PT Refused"),"Yes","No"))', "fx"),
        "L": (None, "in"), "M": (None, "in"),
        "N": ('IF($D{r}="","",IF($M{r}="Excused","No",'
              'IF(AND($G{r}="PT Modified",$L{r}="Received"),"No","Yes")))', "fx"),
        "O": (None, "in"),
        # per-event makeup reconciliation: credited makeup rows linked to
        # this EventID clear the event when the balance reaches zero.
        # Same criteria as the sysAttendance roll-up (PID + credit + type) so
        # the CLEARED banner and 'Cl Owed'/'PT Net' can never disagree —
        # without the PID filter another cadet's makeup cleared this event.
        # MIN(...) caps the credit at THIS event's own duration: makeup is
        # minute-for-minute (policy 400.5), so surplus minutes booked against
        # one event must never spill over and pay off a different absence.
        # sysAttendance rolls these capped per-event figures up, so the cap
        # is enforced for the graduation gate too, not just for the banner.
        "P": ('IF(OR($B{r}="",N($I{r})=0),"",MIN(N($I{r}),'
              'SUMIFS(nrMK_Min,nrMK_Link,$B{r},'
              'nrMK_Credit,"Yes",nrMK_PID,$E{r},nrMK_Type,"Classroom")))', "fx"),
        "Q": ('IF(OR($B{r}="",N($J{r})=0),"",MIN(N($J{r}),'
              'SUMIFS(nrMK_Sess,nrMK_Link,$B{r},'
              'nrMK_Credit,"Yes",nrMK_PID,$E{r},nrMK_Type,"PT")))', "fx"),
        # balance unit follows the event's own Is PT? flag, not
        # whichever of minutes/sessions happens to be filled first
        "R": ('IF(OR($B{r}="",$N{r}<>"Yes"),"",'
              'IF($K{r}="Yes",IF(N($J{r})>0,N($J{r})-N($Q{r}),""),'
              'IF(N($I{r})>0,N($I{r})-N($P{r}),"")))', "fx"),
        # MAXIFS returns 0 when no credited makeup row carries a date, so a
        # credited Makeup row with a BLANK Date stamped a fabricated
        # "CLEARED 12/30" onto the policy-400 ledger. A dateless clear now
        # says so out loud; Makeup Row Check refuses credit for it as well,
        # so this branch should only ever be reachable on legacy data.
        "S": ('IF(OR($B{r}="",$R{r}=""),"",'
              'IF($R{r}<=0,'
              'IF(IFERROR(MAXIFS(nrMK_Date,nrMK_Link,$B{r},nrMK_Credit,"Yes",'
              'nrMK_PID,$E{r}),0)=0,"CLEARED (makeup date missing)",'
              '"CLEARED "&TEXT(MAXIFS(nrMK_Date,nrMK_Link,$B{r},'
              'nrMK_Credit,"Yes",nrMK_PID,$E{r}),"mm/dd")),'
              '"OPEN"))', "fx"),
        # A10 / A23: Attendance was the only major log with no Row Check.
        # Two silent losses lived here: a counted absence whose unit column
        # is BLANK (Sessions blank on a PT event, Minutes blank on a
        # classroom event) dropped out of BOTH caps and never showed OPEN;
        # and a row carrying BOTH minutes and sessions silently discarded
        # whichever one did not match the row's own "Is PT?" flag - in the
        # balance, in the makeup reconciliation and in the sysAttendance
        # caps. Neither was visible anywhere until now.
        "T": ('IF($D{r}="","",'
              'IF($C{r}="","NO DATE",'
              'IF($G{r}="","NO EVENT TYPE",'
              'IF(AND($I{r}<>"",NOT(ISNUMBER($I{r}))),"MINUTES NOT A NUMBER",'
              'IF(AND($J{r}<>"",NOT(ISNUMBER($J{r}))),"SESSIONS NOT A NUMBER",'
              'IF(OR(N($I{r})<0,N($J{r})<0),"NEGATIVE MINUTES/SESSIONS",'
              'IF(AND($K{r}="Yes",N($I{r})>0),'
              '"MINUTES ON A PT EVENT - not counted (PT counts sessions)",'
              'IF(AND($K{r}="No",N($J{r})>0),'
              '"SESSIONS ON A CLASSROOM EVENT - not counted (classroom '
              'counts minutes)",'
              'IF(AND($N{r}="Yes",$K{r}="Yes",N($J{r})=0),'
              '"COUNTED PT EVENT WITH NO SESSIONS - dropped from the PT cap",'
              'IF(AND($N{r}="Yes",$K{r}="No",N($I{r})=0),'
              '"COUNTED EVENT WITH NO MINUTES - dropped from the classroom '
              'cap","OK"))))))))))', "fx"),
    })
    for r in range(first, last + 1):
        ws[f"C{r}"].number_format = DATE
    cf_formula(ws, f"S{first}:S{last}", f'$S{first}="OPEN"', FILL_WARNBG)
    cf_formula(ws, f"S{first}:S{last}",
               f'$S{first}="CLEARED (makeup date missing)"', FILL_WARNBG)
    cf_formula(ws, f"S{first}:S{last}",
               f'LEFT($S{first},7)="CLEARED"', FILL_OKBG)
    cf_formula(ws, f"T{first}:T{last}",
               f'AND($T{first}<>"",$T{first}<>"OK")', FILL_WARNBG)
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    dv_list(ws, "=lstAttendanceEvent", [f"G{first}:G{last}"])
    dv_list(ws, "=lstReason", [f"H{first}:H{last}"])
    dv_list(ws, "=lstDocumentation", [f"L{first}:L{last}"])
    dv_list(ws, "=lstExcused", [f"M{first}:M{last}"])
    define(wb, "nrAT_ID", "Attendance", f"$B${first}:$B${last}")
    # capped per-event makeup credit — sysAttendance sums THESE (not the raw
    # Makeup rows) so an over-credited or unlinked makeup cannot inflate the
    # roll-up the graduation gate reads
    define(wb, "nrAT_MadeUpMin", "Attendance", f"$P${first}:$P${last}")
    define(wb, "nrAT_MadeUpSess", "Attendance", f"$Q${first}:$Q${last}")
    define(wb, "nrAT_Balance", "Attendance", f"$R${first}:$R${last}")
    define(wb, "nrAT_Cleared", "Attendance", f"$S${first}:$S${last}")
    define(wb, "nrAT_Date", "Attendance", f"$C${first}:$C${last}")
    define(wb, "nrAT_PID", "Attendance", f"$E${first}:$E${last}")
    define(wb, "nrAT_Type", "Attendance", f"$G${first}:$G${last}")
    define(wb, "nrAT_Min", "Attendance", f"$I${first}:$I${last}")
    define(wb, "nrAT_Sess", "Attendance", f"$J${first}:$J${last}")
    define(wb, "nrAT_IsPT", "Attendance", f"$K${first}:$K${last}")
    define(wb, "nrAT_Excused", "Attendance", f"$M${first}:$M${last}")
    define(wb, "nrAT_Counts", "Attendance", f"$N${first}:$N${last}")
    define(wb, "nrAT_RowCheck", "Attendance", f"$T${first}:$T${last}")
    col_widths(ws, {"A": 3, "B": 9, "C": 11, "D": 22, "E": 9, "F": 18,
                    "G": 16, "H": 12, "I": 9, "J": 9, "K": 7, "L": 12,
                    "M": 10, "N": 9, "O": 30, "T": 46})
    sheet_note(ws, "Exception log: only missed/modified time gets a row. "
                   "Minutes count toward the classroom cap; Sessions toward "
                   "the PT cap. Excused or documented-modified-PT rows don't "
                   "count (policy 400). Each counting event stays OPEN until "
                   "Makeup rows linked to its EventID cover the balance — "
                   "then it shows CLEARED with the makeup date. Watch Row "
                   "Check (last column): a counted row with no minutes (or "
                   "no sessions on a PT event) is dropped from BOTH caps, "
                   "and a row carrying the unit that does not match its own "
                   "'Is PT?' flag has that value discarded everywhere — "
                   "balance, makeup reconciliation and the caps.")
    return ws


def build_makeup(wb):
    ws = wb.create_sheet("Makeup")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["MakeupID", "Date", "Cadet Name", "PID", "Type",
                    "Minutes Credit", "Sessions Credit", "Linked Event",
                    "Doc Status", "Hold?", "Credit Applies?", "Notes",
                    "Row Check"])
    first, last = DATA_ROW, DATA_ROW + ROWS_MAKEUP - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","M"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": (None, "in"), "G": (None, "in"), "H": (None, "in"),
        "I": (None, "in"), "J": (None, "in"), "K": (None, "in"),
        # credit now also requires a SOUND row (N): the wrong cadet's event,
        # a dangling EventID, a type nothing credits, or minutes booked
        # against a session-denominated PT event all withhold credit instead
        # of silently clearing someone else's absence.
        "L": ('IF($D{r}="","",IF(AND($J{r}="Received",$K{r}<>"Yes",'
              'LEFT($N{r},2)="OK"),"Yes","No"))', "fx"),
        "M": (None, "in"),
        # a blank Linked Event is NOT ok: minute-for-minute reconciliation
        # and a TCOLE hours audit both require every credited minute to be
        # attached to the specific missed event it makes up.
        # A15: a credited row with a BLANK Date used to clear its event
        # with the fabricated banner "CLEARED 12/30" - a made-up make-up
        # date on the policy-400 ledger. A22: the sign and magnitude of the
        # credit were never checked, so a NEGATIVE credit passed as "OK" /
        # "Credit Applies = Yes" and INCREASED the cadet's owed time - a
        # silently wrong number rather than an error.
        "N": ('IF($D{r}="","",'
              'IF(AND($F{r}<>"Classroom",$F{r}<>"PT"),"TYPE NOT CREDITED",'
              'IF($C{r}="","NO MAKEUP DATE",'
              'IF(NOT(ISNUMBER($C{r})),"MAKEUP DATE NOT A DATE",'
              'IF(OR(AND($G{r}<>"",NOT(ISNUMBER($G{r}))),'
              'AND($H{r}<>"",NOT(ISNUMBER($H{r})))),"CREDIT NOT A NUMBER",'
              'IF(OR(N($G{r})<0,N($H{r})<0),"NEGATIVE CREDIT",'
              'IF(AND($F{r}="Classroom",N($G{r})<=0),"NO MINUTES CREDITED",'
              'IF(AND($F{r}="PT",N($H{r})<=0),"NO SESSIONS CREDITED",'
              'IF($I{r}="","NO LINKED EVENT",'
              'IF(COUNTIF(nrAT_ID,$I{r})=0,"NO SUCH EVENT",'
              'IF(INDEX(nrAT_PID,MATCH($I{r},nrAT_ID,0))<>$E{r},"WRONG CADET",'
              'IF(INDEX(nrAT_IsPT,MATCH($I{r},nrAT_ID,0))<>'
              'IF($F{r}="PT","Yes","No"),"UNIT MISMATCH","OK"))))))))))))', "fx"),
    })
    for r in range(first, last + 1):
        ws[f"C{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    # A28: the dropdown used to offer "Skills" and "Admin Approved", which
    # Row Check rejects unconditionally as TYPE NOT CREDITED - the sheet
    # invited two answers it then painted red. Only the two types the caps
    # actually credit are offered; a legacy row still carrying one of them
    # keeps reading TYPE NOT CREDITED, which is the truth.
    dv_list(ws, "=lstMakeupType", [f"F{first}:F{last}"])
    # nrAT_IDlist (a FILTERed helper), not nrAT_ID: the raw Attendance ID
    # column is 800 formula cells, so pointing the dropdown at it listed one
    # blank entry per unused row (~795 of 800 on the seeded workbook). Row
    # Check (column N) still validates the pick against the raw nrAT_ID.
    dv_list(ws, "=nrAT_IDlist", [f"I{first}:I{last}"], enforce=False)
    dv_list(ws, "=lstDocumentation", [f"J{first}:J{last}"])
    dv_list(ws, "=lstYesNo", [f"K{first}:K{last}"])
    define(wb, "nrMK_Date", "Makeup", f"$C${first}:$C${last}")
    define(wb, "nrMK_Link", "Makeup", f"$I${first}:$I${last}")
    define(wb, "nrMK_PID", "Makeup", f"$E${first}:$E${last}")
    define(wb, "nrMK_Type", "Makeup", f"$F${first}:$F${last}")
    define(wb, "nrMK_Min", "Makeup", f"$G${first}:$G${last}")
    define(wb, "nrMK_Sess", "Makeup", f"$H${first}:$H${last}")
    define(wb, "nrMK_Credit", "Makeup", f"$L${first}:$L${last}")
    define(wb, "nrMK_RowCheck", "Makeup", f"$N${first}:$N${last}")
    cf_formula(ws, f"N{first}:N{last}",
               f'AND($N{first}<>"",LEFT($N{first},2)<>"OK")', FILL_WARNBG)
    col_widths(ws, {"A": 3, "B": 10, "C": 11, "D": 22, "E": 9, "F": 14,
                    "G": 13, "H": 13, "I": 13, "J": 12, "K": 8, "L": 13,
                    "M": 30, "N": 22})
    sheet_note(ws, "Makeup is minute-for-minute (policy 400.5). Only "
                   "classroom minutes and PT sessions can be made up — "
                   "skills time cannot, and there is no administrative "
                   "waiver, so the Type dropdown offers only Classroom and "
                   "PT. Credit applies only when "
                   "documentation is Received, no Hold, and Row Check is OK. "
                   "Pick the missed event's ID in Linked Event — the "
                   "Attendance sheet clears that event automatically when its "
                   "balance hits zero. Every credited row must name the event "
                   "it makes up — Row Check refuses credit with no Linked "
                   "Event, for another cadet's event, an EventID that no "
                   "longer exists, a Type the caps don't credit, or minutes "
                   "booked against a PT (session-counted) event. It also "
                   "refuses a row with no makeup Date (which used to stamp a "
                   "fabricated CLEARED date on the attendance ledger) and a "
                   "zero or NEGATIVE credit (which used to pass as OK and "
                   "increase the cadet's owed time). Credit is "
                   "capped at the linked event's own minutes/sessions; split "
                   "a long session across rows to cover two events.")
    return ws


# --------------------------------------------------------------------------
def build_skills(wb):
    ws = wb.create_sheet("Skills")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["RecordID", "Cadet Name", "PID", "Category", "Max",
                    "Mode", "Passing", "Attempt #", "Result", "Score",
                    "Date", "Assessed By", "Attempts Used", "Status",
                    "DismissReview", "Notes", "Course of Fire", "Row Check"])
    first, last = DATA_ROW, DATA_ROW + ROWS_SKILLS - 1
    fill_rows(ws, first, last, {
        "B": ('IF($C{r}="","",$D{r}&"-"&$E{r}&"-"&$I{r})', "fx"),
        "C": (None, "in"),
        "D": ('IF($C{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($C{r},rngCadetNames,0)),"?"))', "fx"),
        "E": (None, "in"),
        "F": ('IF($E{r}="","",IFERROR(INDEX(rngSM_max,MATCH($E{r},rngSM_cat,0)),""))', "fx"),
        "G": ('IF($E{r}="","",IFERROR(INDEX(rngSM_mode,MATCH($E{r},rngSM_cat,0)),""))', "fx"),
        "H": ('IF($E{r}="","",IFERROR(INDEX(rngSM_pass,MATCH($E{r},rngSM_cat,0)),""))', "fx"),
        "I": (None, "in"), "J": (None, "in"), "K": (None, "in"),
        "L": (None, "in"), "M": (None, "in"),
        # "Attempts Used" counts SCORED ATTEMPT NUMBERS, not scored ROWS.
        # A row logged when the next attempt is merely SCHEDULED is not an
        # attempt used, and counting it pushed a cadet to FAILED OUT one
        # attempt early. Counting rows did the same thing to firearms, where
        # ONE attempt is recorded as TWO rows (Course of Fire 1 and 2, each
        # with its own Score) — a cadet with 3 permitted attempts read
        # "4 used / FAILED OUT" after his second. Column O computes the same
        # highest-scored-attempt expression for the supersede test.
        "N": ('IF($C{r}="","",MAX('
              'MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Pass"),'
              'MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Fail")'
              '))', "fx"),
        # same placeholder-row bug class as the exam retests: "latest" used
        # to be the highest attempt number of ANY row, so adding an unscored
        # next-attempt row made the real "Needs Remediation" / "FAILED OUT"
        # row read "(superseded)" and the new blank row read "Pending" -
        # which zeroed sysSkills "Needs Remediation" and flipped Skills Elig
        # (nrSKelig) from No back to Yes. Only a SCORED row supersedes.
        "O": ('IF($C{r}="","",'
              'IF(AND(MAX(MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Pass"),MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Fail"))>0,'
              'N($I{r})<MAX(MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Pass"),MAXIFS(nrSK_Att,nrSK_PID,$D{r},nrSK_Cat,$E{r},nrSK_Res,"Fail"))),"(superseded)",'
              'IF(OR($J{r}="",$J{r}="Pending"),"Pending",'
              'IF($J{r}="Pass","Qualified",'
              'IF(AND($J{r}="Fail",$N{r}>=$F{r}),"FAILED OUT",'
              '"Needs Remediation")))))', "fx"),
        "P": ('IF($C{r}="","",IF($O{r}="FAILED OUT","Yes","No"))', "fx"),
        "Q": (None, "in"),
        "R": (None, "in"),
        # Skills was the only graded number in the workbook with neither a
        # Row Check nor a range guard. 68 keyed as 680 satisfied the ch.41
        # "both courses of fire >= passing" gate (sysSkills P -> sysChecks T
        # -> GraduationElig) and won Top Gun, with nothing on the sheet
        # complaining. The aggregates in sysSkills now exclude an
        # out-of-range score; this column says why the row dropped out.
        "S": ('IF($C{r}="","",'
              'IF(AND($K{r}<>"",NOT(ISNUMBER($K{r}))),"SCORE NOT A NUMBER",'
              'IF(AND(ISNUMBER($K{r}),OR($K{r}<0,$K{r}>100)),'
              '"SCORE OUT OF RANGE - this score counts toward nothing",'
              'IF(AND($G{r}="Score",$K{r}="",$J{r}<>"",$J{r}<>"Pending"),'
              '"RESULT WITHOUT A SCORE",'
              'IF(AND($G{r}="Pass/Fail",$K{r}<>""),'
              '"SCORE ON A PASS-FAIL CATEGORY - it is ignored",'
              'IF(AND($E{r}="Firearms",$J{r}<>"",$J{r}<>"Pending",$R{r}=""),'
              '"FIREARMS ROW WITH NO COURSE OF FIRE",'
              '"OK"))))))', "fx"),
    })
    for r in range(first, last + 1):
        ws[f"L{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"C{first}:C{last}"])
    dv_list(ws, "=rngSM_cat", [f"E{first}:E{last}"])
    dv_list(ws, "=lstAttemptNum", [f"I{first}:I{last}"])
    dv_list(ws, "=lstSkillResult", [f"J{first}:J{last}"])
    dv_list(ws, '"1,2"', [f"R{first}:R{last}"])
    cf_formula(ws, f"O{first}:O{last}", f'$O{first}="FAILED OUT"', FILL_WARNBG)
    cf_formula(ws, f"S{first}:S{last}",
               f'AND($S{first}<>"",$S{first}<>"OK")', FILL_WARNBG)
    define(wb, "nrSK_RowCheck", "Skills", f"$S${first}:$S${last}")
    define(wb, "nrSK_PID", "Skills", f"$D${first}:$D${last}")
    define(wb, "nrSK_Cat", "Skills", f"$E${first}:$E${last}")
    define(wb, "nrSK_Att", "Skills", f"$I${first}:$I${last}")
    define(wb, "nrSK_Res", "Skills", f"$J${first}:$J${last}")
    define(wb, "nrSK_Score", "Skills", f"$K${first}:$K${last}")
    define(wb, "nrSK_Status", "Skills", f"$O${first}:$O${last}")
    define(wb, "nrSK_Dis", "Skills", f"$P${first}:$P${last}")
    define(wb, "nrSK_CoF", "Skills", f"$R${first}:$R${last}")
    col_widths(ws, {"A": 3, "B": 14, "C": 22, "D": 9, "E": 12, "F": 6,
                    "G": 10, "H": 9, "I": 9, "J": 9, "K": 8, "L": 11,
                    "M": 16, "N": 12, "O": 16, "P": 12, "Q": 26, "R": 12,
                    "S": 30})
    sheet_note(ws, "One row per attempt (firearms scores recorded; Top Gun "
                   "uses the firearms Score column). Firearms rows: set "
                   "Course of Fire 1 or 2 — TCOLE requires 70%+ on BOTH. "
                   "Exhausting max attempts = separation review (300.7). "
                   "A next-attempt row logged before it is scored does NOT "
                   "supersede the previous result and does not consume an "
                   "attempt: Status stays 'Pending' on the new row and the "
                   "real result keeps driving Skills eligibility until a "
                   "Pass or Fail is entered. Watch Row Check (last column): "
                   "a Score outside 0-100 counts toward NOTHING - not the "
                   "firearms average, not Top Gun and not the both-courses-"
                   "of-fire graduation gate - until it is corrected.")
    return ws


# --------------------------------------------------------------------------
def build_writing(wb):
    ws = wb.create_sheet("Writing")
    ws.sheet_view.showGridLines = False
    n = len(DW.ASSIGNMENTS)  # 40
    hdrs = ["PID", "Cadet Name"] + [f"#{i}" for i in range(1, n + 1)] + \
           [f"Complete (of {n})", "Overdue Missing", "Writing Current?"]
    header_row(ws, hdrs)
    first, last = DATA_ROW, CADET_LAST
    first_ac = 4                       # column D
    last_ac = first_ac + n - 1         # column AQ (col 43)
    last_acL = get_column_letter(last_ac)
    cols = {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF(Cadets!$B{r}="","",Cadets!$F{r})', "fx"),
    }
    for i in range(n):
        cols[get_column_letter(first_ac + i)] = (None, "in")
    cA, cB, cC = (get_column_letter(last_ac + 1), get_column_letter(last_ac + 2),
                  get_column_letter(last_ac + 3))
    cols[cA] = ('IF($B{r}="","",COUNTIF(D{r}:%s{r},"X"))' % last_acL, "fx")
    cols[cB] = ('IF($B{r}="","",SUMPRODUCT((UPPER(D{r}:%s{r})<>"X")*'
                '(TRANSPOSE(rngWMdue)<>"")*(TRANSPOSE(rngWMdue)<TODAY())))'
                % last_acL, "fx")
    cols[cC] = ('IF($B{r}="","",IF(%s{r}=0,"Yes","No"))' % cB, "fx")
    fill_rows(ws, first, last, cols)
    # X marks: centered; a workbook event (installed with the VBA) uppercases
    # any lowercase x typed here
    for r in range(first, last + 1):
        for i in range(n):
            ws[f"{get_column_letter(first_ac+i)}{r}"].alignment = A_CENTER
    dv_list(ws, '"X,x"', [f"D{first}:{last_acL}{last}"])
    cf_yes_no(ws, f"{cC}{first}:{cC}{last}")
    # blank cells highlighted red once the assignment's due date has passed
    for i in range(n):
        cl = get_column_letter(first_ac + i)
        cf_formula(ws, f"{cl}{first}:{cl}{last}",
                   f'AND($B{first}<>"",UPPER({cl}{first})<>"X",'
                   f'INDEX(rngWMdue,{i+1})<>"",INDEX(rngWMdue,{i+1})<TODAY())',
                   FILL_WARNBG)
    define(wb, "nrWRcurrent", "Writing", f"${cC}${first}:${cC}${last}")
    define(wb, "nrWRoverdue", "Writing", f"${cB}${first}:${cB}${last}")
    define(wb, "nrWRcomplete", "Writing", f"${cA}${first}:${cA}${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    for i in range(n):
        ws.column_dimensions[get_column_letter(first_ac + i)].width = 5
    for cl in (cA, cB, cC):
        ws.column_dimensions[cl].width = 13
    sheet_note(ws, "Type X when an assignment is received (lowercase x "
                   "auto-capitalizes; blank = not done). Red cells are past "
                   "their computed due date (see WritingMaster).")
    return ws


# --------------------------------------------------------------------------
def build_incidents(wb):
    ws = wb.create_sheet("Incidents")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["IncidentID", "Date", "Cadet Name", "PID", "Direction",
                    "Severity", "Category", "Description", "Reported By",
                    "Doc Ref", "ChainReview", "Resolution", "Notes",
                    "Report to Agency?"])
    first, last = DATA_ROW, DATA_ROW + ROWS_INCIDENTS - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","I"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": (None, "in"), "G": (None, "in"), "H": (None, "in"),
        "I": (None, "in"), "J": (None, "in"), "K": (None, "in"),
        "L": ('IF($D{r}="","",IF(AND($F{r}="Negative",OR($G{r}="Critical",'
              '$G{r}="Major")),"Yes","No"))', "fx"),
        "M": (None, "in"), "N": (None, "in"), "O": (None, "in"),
    })
    for r in range(first, last + 1):
        ws[f"C{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    dv_list(ws, "=lstIncidentDirection", [f"F{first}:F{last}"])
    dv_list(ws, "=lstSeverity", [f"G{first}:G{last}"])
    dv_list(ws, "=lstIssuetype", [f"H{first}:H{last}"])
    dv_list(ws, "=lstResolution", [f"M{first}:M{last}"])
    dv_list(ws, "=lstYesNo", [f"O{first}:O{last}"])
    define(wb, "nrIN_Report", "Incidents", f"$O${first}:$O${last}")
    define(wb, "nrIN_Date", "Incidents", f"$C${first}:$C${last}")
    define(wb, "nrIN_PID", "Incidents", f"$E${first}:$E${last}")
    define(wb, "nrIN_Dir", "Incidents", f"$F${first}:$F${last}")
    define(wb, "nrIN_Sev", "Incidents", f"$G${first}:$G${last}")
    define(wb, "nrIN_Desc", "Incidents", f"$I${first}:$I${last}")
    define(wb, "nrIN_Chain", "Incidents", f"$L${first}:$L${last}")
    define(wb, "nrIN_Res", "Incidents", f"$M${first}:$M${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 11, "D": 22, "E": 9, "F": 13,
                    "G": 13, "H": 14, "I": 44, "J": 14, "K": 12, "L": 11,
                    "M": 11, "N": 26, "O": 15})
    sheet_note(ws, "Positive and negative incidents. Negative Major/Critical "
                   "trigger chain-of-command review (policy 600). 'Report to "
                   "Agency?' = Yes includes it in that agency's next email "
                   "digest; blank keeps it an academy teaching moment.")
    return ws


# --------------------------------------------------------------------------
def build_counseling(wb):
    ws = wb.create_sheet("Counseling")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["EntryID", "Date", "Cadet Name", "PID", "Type",
                    "Trigger / Related Flag", "Description / Plan",
                    "Conducted By", "Agency Notified?", "Notified Date",
                    "Follow-up Date", "Status", "Notes"])
    first, last = DATA_ROW, DATA_ROW + ROWS_COUNSELING - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","C"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": (None, "in"), "G": (None, "in"), "H": (None, "in"),
        "I": (None, "in"), "J": (None, "in"), "K": (None, "in"),
        "L": (None, "in"), "M": (None, "in"), "N": (None, "in"),
    })
    for r in range(first, last + 1):
        for cl in ("C", "K", "L"):
            ws[f"{cl}{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    dv_list(ws, "=lstCounselingType", [f"F{first}:F{last}"])
    dv_list(ws, "=lstYesNo", [f"J{first}:J{last}"])
    dv_list(ws, "=lstResolution", [f"M{first}:M{last}"])
    define(wb, "nrCO_Date", "Counseling", f"$C${first}:$C${last}")
    define(wb, "nrCO_PID", "Counseling", f"$E${first}:$E${last}")
    define(wb, "nrCO_Type", "Counseling", f"$F${first}:$F${last}")
    define(wb, "nrCO_Desc", "Counseling", f"$H${first}:$H${last}")
    define(wb, "nrCO_Status", "Counseling", f"$M${first}:$M${last}")
    col_widths(ws, {"A": 3, "B": 9, "C": 11, "D": 22, "E": 9, "F": 18,
                    "G": 24, "H": 40, "I": 16, "J": 13, "K": 12, "L": 12,
                    "M": 11, "N": 26})
    define(wb, "nrCO_Report", "Counseling", f"$J${first}:$J${last}")
    sheet_note(ws, "Documents every intervention (tutoring, counseling, "
                   "agency notification, performance plan) — the early-"
                   "intervention paper trail policy 300.4.B expects. "
                   "'Agency Notified?' = Yes includes the entry in that "
                   "agency's next email digest; blank keeps it academy-only.")
    return ws


# --------------------------------------------------------------------------
def build_pt(wb):
    ws = wb.create_sheet("PT")
    ws.sheet_view.showGridLines = False
    events = [ev for ev, _, _, _ in DL.PT_EVENTS]
    short = ["Bench", "VertJump", "Pushups", "Situps", "Agility", "1.5Mi",
             "300m"]
    hdrs = (["PID", "Cadet Name"]
            + [f"Base {s}" for s in short] + ["Baseline Pass?"]
            + ["Midpoint Done?"]
            + [f"Final {s}" for s in short]
            + [f"Pts {s}" for s in short]
            + ["Final Points", "Final PT Pass?", "Improvement Index"])
    header_row(ws, hdrs)
    first, last = DATA_ROW, CADET_LAST
    cols = {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF(Cadets!$B{r}="","",Cadets!$F{r})', "fx"),
    }
    # base raw D..J, base pass K, midpoint L, final raw M..S, pts T..Z
    for i in range(7):
        cols[get_column_letter(4 + i)] = (None, "in")
    cols["K"] = (None, "in")
    cols["L"] = (None, "in")
    for i in range(7):
        cols[get_column_letter(13 + i)] = (None, "in")
    # Pts columns T..Z. Bench (T) and VertJump (U) are BASELINE-only
    # standards - the approved chart scores FIVE events, not seven - so they
    # say so instead of sitting there as blank inputs that look unfinished.
    cols["T"] = ('IF($B{r}="","","n/a (baseline only)")', "fx")
    cols["U"] = ('IF($B{r}="","","n/a")', "fx")
    # V..Z score themselves from the approved chart on Settings: the points
    # of the BEST tier whose threshold the result reaches, and 0 when it is
    # below the Tier 1 minimum (coordinator's rule - 0 for that event, the
    # total still decides). SUMPRODUCT(MAX(...)) forces the array evaluation
    # this needs; a half-entered rubric row scores nothing rather than
    # under-scoring a cadet against thresholds that are not all there yet.
    for i, (_ev, src, hib, _measure, _bands) in enumerate(DL.PT_FINAL_BANDS):
        val = ('$%s{r}*60' % src) if src == "R" else ('$%s{r}' % src)
        cmp_ = ">=" if hib else "<="
        cols[get_column_letter(22 + i)] = (
            'IF($B{r}="","",IF($%s{r}="","",'
            'IF(COUNT(INDEX(nrPTBands,%d,0))<5,"",'
            'SUMPRODUCT(MAX((%s%sINDEX(nrPTBands,%d,0))*nrPTTierPts)))))'
            % (src, i + 1, val, cmp_, i + 1), "fx")
    cols["AA"] = ('IF($B{r}="","",IF(COUNT($V{r}:$Z{r})=0,"",'
                  'SUM($V{r}:$Z{r})))', "fx")
    # COUNT(V:Z)<5 -> "Incomplete": the pass test used to compare the SUM of
    # whatever points happened to be present against the minimum, so ONE
    # high event score with the rest blank read "Yes" and opened both the
    # graduation gate (sysChecks L) and the final-exam gate (sysChecks P).
    # Five, not seven: bench and vertical jump are baseline-only and are no
    # longer scored, so a seven-count could never be satisfied. AA stays
    # numeric so CadetProfile still prints "20 pts" and the sysAwards MAX
    # still works; the completeness rule lives here.
    cols["AB"] = ('IF($B{r}="","",IF($AA{r}="","",'
                  'IF(cfgPTFinalMinPoints=0,"(rubric pending)",'
                  'IF(COUNT($V{r}:$Z{r})<5,"Incomplete",'
                  'IF($AA{r}>=cfgPTFinalMinPoints,"Yes","No")))))', "fx")
    # improvement index: mean % gain on countable events (pushups, situps)
    # plus % time cut on runs (agility, 1.5mi, 300m); ignores blanks
    # each delta only counts when BOTH baseline and final are entered —
    # a blank final would otherwise read as a +/-100% swing
    cols["AC"] = ('IF($B{r}="","",LET('
                  'pu,IF(COUNT($F{r},$O{r})=2,IFERROR(($O{r}-$F{r})/$F{r},""),""),'
                  'su,IF(COUNT($G{r},$P{r})=2,IFERROR(($P{r}-$G{r})/$G{r},""),""),'
                  'ag,IF(COUNT($H{r},$Q{r})=2,IFERROR(($H{r}-$Q{r})/$H{r},""),""),'
                  'mi,IF(COUNT($I{r},$R{r})=2,IFERROR(($I{r}-$R{r})/$I{r},""),""),'
                  'tm,IF(COUNT($J{r},$S{r})=2,IFERROR(($J{r}-$S{r})/$J{r},""),""),'
                  'vals,IFERROR(FILTER(HSTACK(pu,su,ag,mi,tm),'
                  'ISNUMBER(HSTACK(pu,su,ag,mi,tm))),""),'
                  'IF(COUNT(vals)=0,"",ROUND(AVERAGE(vals)*100,1))))', "fx")
    fill_rows(ws, first, last, cols)
    dv_list(ws, "=lstYesNo", [f"K{first}:K{last}", f"L{first}:L{last}"])
    cf_yes_no(ws, f"AB{first}:AB{last}")
    cf_yes_no(ws, f"K{first}:K{last}")

    define(wb, "nrPT_BasePass", "PT", f"$K${first}:$K${last}")
    define(wb, "nrPT_FinalPts", "PT", f"$AA${first}:$AA${last}")
    define(wb, "nrPT_FinalPass", "PT", f"$AB${first}:$AB${last}")
    define(wb, "nrPT_Improve", "PT", f"$AC${first}:$AC${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "AA": 11, "AB": 13, "AC": 14})
    for i in range(4, 27):
        ws.column_dimensions[get_column_letter(i)].width = 9
    sheet_note(ws, "Raw values per event: bench lbs, jump inches, pushup/situp "
                   "reps, agility & 300m seconds, 1.5mi minutes DECIMAL "
                   "(12:35 = 12.5833 — the engine converts to seconds to "
                   "score it). Pts columns COMPUTE from the approved PT Test "
                   "Score Chart in the rubric block on Settings: the points "
                   "of the best tier the result reaches, 0 if it is below "
                   "the Tier 1 minimum. Bench and vertical jump are baseline "
                   "standards only and are not scored at the final. All five "
                   "scored events must be recorded before Final PT Pass? "
                   "resolves — until then it reads Incomplete, which blocks. "
                   "Final PT failure blocks the Final Exam (policy 500.1.H).")
    return ws


# --------------------------------------------------------------------------
def build_medical(wb):
    ws = wb.create_sheet("Medical")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["EntryID", "Date", "Cadet Name", "PID", "Type",
                    "Description", "Restriction?", "Restriction Start",
                    "Restriction End", "Cleared Date", "Doc Status",
                    "Current Status", "Notes"])
    first, last = DATA_ROW, DATA_ROW + ROWS_MEDICAL - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","MD"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": (None, "in"), "G": (None, "in"), "H": (None, "in"),
        "I": (None, "in"), "J": (None, "in"), "K": (None, "in"),
        "L": (None, "in"),
        "M": ('IF($D{r}="","",IF($K{r}<>"","Cleared "&TEXT($K{r},"mm/dd"),'
              'IF($H{r}="Yes",IF(AND($J{r}<>"",$J{r}<TODAY()),"RESTRICTION EXPIRED",'
              '"ACTIVE RESTRICTION"),"Documented")))', "fx"),
        "N": (None, "in"),
    })
    for r in range(first, last + 1):
        for cl in ("C", "I", "J", "K"):
            ws[f"{cl}{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    dv_list(ws, "=lstIssuetype", [f"F{first}:F{last}"])
    dv_list(ws, "=lstYesNo", [f"H{first}:H{last}"])
    dv_list(ws, "=lstDocumentation", [f"L{first}:L{last}"])
    cf_formula(ws, f"M{first}:M{last}",
               f'OR($M{first}="ACTIVE RESTRICTION",$M{first}="RESTRICTION EXPIRED")',
               FILL_AMBER)
    define(wb, "nrMD_PID", "Medical", f"$E${first}:$E${last}")
    define(wb, "nrMD_Restr", "Medical", f"$H${first}:$H${last}")
    define(wb, "nrMD_Status", "Medical", f"$M${first}:$M${last}")
    col_widths(ws, {"A": 3, "B": 9, "C": 11, "D": 22, "E": 9, "F": 13,
                    "G": 36, "H": 12, "I": 14, "J": 14, "K": 13, "L": 12,
                    "M": 18, "N": 26})
    sheet_note(ws, "Injuries, clearances and restrictions (policy 400.7/500). "
                   "Modified PT inside a documented restriction is not a "
                   "missed session when objectives are met — log the "
                   "restriction here and the PT Modified event on Attendance.")
    return ws


def build_certifications(wb):
    """Per-cadet grid of IRG-mandatory certifications: completion date +
    copy-on-file status per cert, with a 'to collect' rollup that feeds the
    Dashboard reminders."""
    ws = wb.create_sheet("Certifications")
    ws.sheet_view.showGridLines = False
    hdrs = ["PID", "Cadet Name"]
    for nm, _src in DL.CERTS:
        hdrs += [f"{nm} Date", f"{nm} Copy?"]
    hdrs += ["All Certs?", "To Collect", "Waived (N/A)"]
    header_row(ws, hdrs)
    first, last = DATA_ROW, CADET_LAST
    cols = {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF(Cadets!$B{r}="","",Cadets!$F{r})', "fx"),
    }
    n = len(DL.CERTS)
    for i in range(n):
        cols[get_column_letter(4 + 2 * i)] = (None, "in")      # date
        cols[get_column_letter(5 + 2 * i)] = (None, "in")      # copy?
    col_all = get_column_letter(4 + 2 * n)       # T
    col_miss = get_column_letter(5 + 2 * n)      # U
    col_waived = get_column_letter(6 + 2 * n)    # V
    # missing = cert has no date OR copy not Yes/N/A
    parts = []
    for i, (nm, _src) in enumerate(DL.CERTS):
        dcol = get_column_letter(4 + 2 * i)
        ccol = get_column_letter(5 + 2 * i)
        parts.append(
            f'IF(AND(${ccol}{{r}}<>"N/A",OR(${dcol}{{r}}="",'
            f'${ccol}{{r}}<>"Yes")),"{nm}","")')
    cols[col_miss] = ('IF($B{r}="","",TEXTJOIN("; ",TRUE,' +
                      ",".join(parts) + "))", "fx")
    cols[col_all] = ('IF($B{r}="","",IF($%s{r}="","Yes","No"))' % col_miss, "fx")
    # Marking a Copy? cell "N/A" excuses that TCOLE-mandatory certification
    # completely: it drops out of To Collect, out of All Certs?, out of the
    # graduation block and out of the Audit sheet's count - silently, with
    # nothing anywhere recording that a requirement was waived or by whom.
    # Certifications are a hard graduation gate by decision ("they have to do
    # everything"), so the waiver stays available for the legitimate cases
    # but is no longer invisible: it is named per cadet here and counted on
    # the Audit sheet, where each one has to be defensible to a field agent.
    wparts = []
    for i, (nm, _src) in enumerate(DL.CERTS):
        ccol = get_column_letter(5 + 2 * i)
        wparts.append(f'IF(${ccol}{{r}}="N/A","{nm}","")')
    cols[col_waived] = ('IF($B{r}="","",TEXTJOIN("; ",TRUE,' +
                        ",".join(wparts) + "))", "fx")
    fill_rows(ws, first, last, cols)
    for r in range(first, last + 1):
        for i in range(n):
            ws[f"{get_column_letter(4+2*i)}{r}"].number_format = DATE
    copy_ranges = [f"{get_column_letter(5+2*i)}{first}:"
                   f"{get_column_letter(5+2*i)}{last}" for i in range(n)]
    dv_list(ws, '"Yes,Requested,N/A"', copy_ranges)
    cf_yes_no(ws, f"{col_all}{first}:{col_all}{last}")
    define(wb, "nrCERTall", "Certifications", f"${col_all}${first}:${col_all}${last}")
    define(wb, "nrCERTmissing", "Certifications", f"${col_miss}${first}:${col_miss}${last}")
    define(wb, "nrCERTwaived", "Certifications",
           f"${col_waived}${first}:${col_waived}${last}")
    cf_formula(ws, f"{col_waived}{first}:{col_waived}{last}",
               f'${col_waived}{first}<>""', FILL_WARNBG)
    # legend row under the grid
    lr = last + 2
    ws.cell(row=lr, column=2, value="Cert sources:").font = F_LABEL
    ws.cell(row=lr, column=3, value="; ".join(
        f"{nm} = {src}" for nm, src in DL.CERTS)).font = F_SMALL
    col_widths(ws, {"A": 3, "B": 10, "C": 24, col_all: 10, col_miss: 40,
                    col_waived: 30})
    for i in range(n):
        ws.column_dimensions[get_column_letter(4 + 2 * i)].width = 11
        ws.column_dimensions[get_column_letter(5 + 2 * i)].width = 9
    sheet_note(ws, "TCOLE-mandatory per-student completions. Enter the "
                   "completion date and mark Copy? = Yes once the cadet's "
                   "certificate copy is in the file. 'To Collect' feeds the "
                   "Dashboard reminder list; All Certs GATES GRADUATION - "
                   "every cert must have a date AND Copy? = Yes. Copy? = "
                   "N/A waives that requirement entirely, so every waiver is "
                   "named in 'Waived (N/A)' and counted on the Audit sheet; "
                   "be ready to justify each one to a TCOLE field agent.")
    return ws


def build_stateexam(wb):
    """State licensing exam: 250 questions, 70% pass, 3 attempts max —
    a third failure requires enrolling in a new BPOC."""
    ws = wb.create_sheet("StateExam")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet Name", "Attempt 1 Date", "Result 1",
                    "Attempt 2 Date", "Result 2", "Attempt 3 Date",
                    "Result 3", "Attempts", "Status", "Notes"])
    first, last = DATA_ROW, CADET_LAST
    fill_rows(ws, first, last, {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF(Cadets!$B{r}="","",Cadets!$F{r})', "fx"),
        "D": (None, "in"), "E": (None, "in"), "F": (None, "in"),
        "G": (None, "in"), "H": (None, "in"), "I": (None, "in"),
        "J": ('IF($B{r}="","",COUNTIF($E{r}:$I{r},"Pass")'
              '+COUNTIF($E{r}:$I{r},"Fail"))', "fx"),
        "K": ('IF($B{r}="","",IF(COUNTIF($E{r}:$I{r},"Pass")>0,'
              '"PASSED "&TEXT(MAX($D{r},$F{r},$H{r}),"mm/dd/yyyy"),'
              'IF($J{r}>=3,"FAILED 3 ATTEMPTS - new BPOC required",'
              'IF($J{r}=0,"Not yet attempted",'
              '"In progress ("&$J{r}&" attempt(s))"))))', "fx"),
        "L": (None, "in"),
    })
    for r in range(first, last + 1):
        for cl in ("D", "F", "H"):
            ws[f"{cl}{r}"].number_format = DATE
    dv_list(ws, '"Pass,Fail"',
            [f"E{first}:E{last}", f"G{first}:G{last}", f"I{first}:I{last}"])
    cf_formula(ws, f"K{first}:K{last}",
               f'LEFT($K{first},6)="FAILED"', FILL_WARNBG)
    cf_formula(ws, f"K{first}:K{last}",
               f'LEFT($K{first},6)="PASSED"', FILL_OKBG)
    define(wb, "nrSEstatus", "StateExam", f"$K${first}:$K${last}")
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "D": 13, "E": 9, "F": 13,
                    "G": 9, "H": 13, "I": 9, "J": 9, "K": 32, "L": 28})
    sheet_note(ws, "TCOLE licensing exam: 250 multiple-choice, 70% to pass, "
                   "maximum 3 attempts; a third failure means the learner "
                   "must enroll in and complete a new BPOC.")
    return ws


ROWS_MEMOS = 300
ROWS_DAILYLOG = 170


def build_memos(wb):
    """Deficiency memos: assigned when a cadet has a deficiency; the cadet
    writes lessons-learned + plan of action. Linked to the triggering
    Incident (I###), Attendance event (A###) or Counseling entry (C###)."""
    ws = wb.create_sheet("Memos")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["MemoID", "Assigned", "Cadet Name", "PID", "Linked Ref",
                    "Deficiency / Subject", "Due (computed)", "Due Override",
                    "Due", "Received", "Status", "Report to Agency?",
                    "Notes"])
    first, last = DATA_ROW, DATA_ROW + ROWS_MEMOS - 1
    fill_rows(ws, first, last, {
        "B": ('IF($D{r}="","","ME"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"), "D": (None, "in"),
        "E": ('IF($D{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($D{r},rngCadetNames,0)),"?"))', "fx"),
        "F": (None, "in"), "G": (None, "in"),
        # same fix as ExamScores "Retest Due By": match mode 1 rolls a memo
        # assigned on a non-class day to the next class day instead of
        # returning text, which used to freeze Status at "Pending" forever
        # ...and, like the retest clock, the due day must be one the academy
        # is IN SESSION for: the padded tail of the class-day table would
        # otherwise hand back a date days past graduation
        # ...and a memo with a cadet but NO Assigned date must not read
        # "Pending" forever either: H returning "" sent L down the
        # IF($H="","Pending",...) arm, so the row could never become OVERDUE
        # or CHECK DATE and was invisible to sysFlags P, to the WatchList
        # reason text and to the sysAudit "unusable due date" line. Text
        # (never "") for a live row, exactly like ExamScores T.
        "H": ('IF($D{r}="","",IF($C{r}="","(enter date)",'
              'LET(dn,XLOOKUP($C{r},nrCDdate,nrCDnum,"",1),'
              'IF(dn="","(after last class day)",'
              'LET(dd,XLOOKUP(dn+cfgMemoDueClassDays,nrCDnum,nrCDdate,""),'
              'ok,XLOOKUP(dn+cfgMemoDueClassDays,nrCDnum,nrCDinsession,"No"),'
              'IF(OR(dd="",ok<>"Yes"),"(after last class day)",dd))))))', "fx"),
        "I": (None, "in"),
        "J": ('IF($D{r}="","",IF($I{r}<>"",$I{r},IF(ISNUMBER($H{r}),$H{r},"")))', "fx"),
        "K": (None, "in"),
        # an uncomputable due date reads CHECK DATE, not "Pending": a memo
        # that can never become OVERDUE is invisible to every flag
        "L": ('IF($D{r}="","",IF($K{r}<>"","Received "&TEXT($K{r},"mm/dd"),'
              'IF(ISNUMBER($J{r}),IF(TODAY()>$J{r},"OVERDUE","Pending"),'
              'IF($H{r}="","Pending","CHECK DATE"))))', "fx"),
        "M": (None, "in"), "N": (None, "in"),
    })
    for r in range(first, last + 1):
        for cl in ("C", "H", "I", "J", "K"):
            ws[f"{cl}{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"D{first}:D{last}"])
    dv_list(ws, "=nrAllRefIDs", [f"F{first}:F{last}"], enforce=False)
    dv_list(ws, "=lstYesNo", [f"M{first}:M{last}"])
    cf_formula(ws, f"L{first}:L{last}", f'$L{first}="OVERDUE"', FILL_WARNBG)
    cf_formula(ws, f"L{first}:L{last}", f'$L{first}="CHECK DATE"', FILL_WARNBG)
    cf_formula(ws, f"L{first}:L{last}",
               f'LEFT($L{first},8)="Received"', FILL_OKBG)
    define(wb, "nrME_Assigned", "Memos", f"$C${first}:$C${last}")
    define(wb, "nrME_Cadet", "Memos", f"$D${first}:$D${last}")
    define(wb, "nrME_PID", "Memos", f"$E${first}:$E${last}")
    define(wb, "nrME_Ref", "Memos", f"$F${first}:$F${last}")
    define(wb, "nrME_Subject", "Memos", f"$G${first}:$G${last}")
    define(wb, "nrME_Due", "Memos", f"$J${first}:$J${last}")
    define(wb, "nrME_Received", "Memos", f"$K${first}:$K${last}")
    define(wb, "nrME_Status", "Memos", f"$L${first}:$L${last}")
    define(wb, "nrME_Report", "Memos", f"$M${first}:$M${last}")
    col_widths(ws, {"A": 3, "B": 9, "C": 11, "D": 22, "E": 9, "F": 12,
                    "G": 40, "H": 13, "I": 13, "J": 11, "K": 11, "L": 14,
                    "M": 15, "N": 26})
    sheet_note(ws, "Assigned for deficiencies: lessons learned + plan of "
                   "action. Linked Ref dropdown offers Incident (I), "
                   "Attendance (A) and Counseling (C) IDs — or leave blank "
                   "for stand-alone. Due auto-computes from the class-day "
                   "calendar (Settings: Memo Due). 'Report to Agency?' = Yes "
                   "puts it in that agency's next email digest; blank keeps "
                   "it an academy teaching moment.")
    return ws


def build_dailylog(wb):
    """One coordinator row per training day — the digital daily report.
    Roll-call and changes noted in seconds; counters compute."""
    ws = wb.create_sheet("DailyLog")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Date", "Day #", "Class Type", "Present", "Absent/Late "
                    "(names & reason)", "AM Notes (roll call / PT)",
                    "PM Notes (changes / departures)", "Incidents",
                    "Early Departs", "Memos In", "Issues?",
                    "Leader Report Scanned?", "Notes"])
    first, last = DATA_ROW, DATA_ROW + ROWS_DAILYLOG - 1
    fill_rows(ws, first, last, {
        "B": (None, "in"),
        "C": ('IF($B{r}="","",IFERROR(XLOOKUP($B{r},nrCDdate,nrCDnum),""))', "fx"),
        "D": ('IF($B{r}="","",TRIM(IF(COUNTIFS(nrSCH_Date,$B{r},nrSCH_Act,'
              '"PT*")>0,"PT + ","")&IF(COUNTIFS(nrSCH_Date,$B{r},nrSCH_Act,'
              '"*Test*")>0,"Test + ","")&"Classroom"))', "fx"),
        "E": (None, "in"), "F": (None, "in"), "G": (None, "in"),
        "H": (None, "in"),
        "I": ('IF($B{r}="","",COUNTIFS(nrIN_Date,$B{r}))', "fx"),
        "J": ('IF($B{r}="","",COUNTIFS(nrAT_Date,$B{r},nrAT_Type,'
              '"Early Departure"))', "fx"),
        "K": ('IF($B{r}="","",COUNTIFS(nrME_Received,$B{r}))', "fx"),
        "L": (None, "in"), "M": (None, "in"), "N": (None, "in"),
    })
    for r in range(first, last + 1):
        ws[f"B{r}"].number_format = DATE
    dv_list(ws, "=lstYesNo", [f"L{first}:L{last}", f"M{first}:M{last}"])
    cf_formula(ws, f"L{first}:L{last}", f'$L{first}="Yes"', FILL_AMBER)
    define(wb, "nrDL_Date", "DailyLog", f"$B${first}:$B${last}")
    col_widths(ws, {"A": 3, "B": 11, "C": 7, "D": 16, "E": 9, "F": 34,
                    "G": 34, "H": 34, "I": 10, "J": 12, "K": 10, "L": 8,
                    "M": 18, "N": 26})
    sheet_note(ws, "The digital daily report — one row per training day, "
                   "30 seconds after the class leader briefs you. Counters "
                   "(incidents, early departures, memos received) compute "
                   "from the logs; Issues=Yes highlights for follow-up. Mark "
                   "when the signed leader report is scanned into the file.")
    return ws


def build_advisoryboard(wb):
    """Governance alignment: the board meets 1-2x/year and its minutes live
    in a server folder — this sheet holds the POINTER (running meeting list
    + minutes location) and this academy's ALIGNMENT record (policy version
    in effect, minutes reviewed, workbook updated to match)."""
    from xlb import F_LABEL as _FL
    ws = wb.create_sheet("AdvisoryBoard")
    ws.sheet_view.showGridLines = False
    # ---- this academy's alignment record (prompted at academy startup) ----
    ws.cell(row=HDR_ROW, column=2,
            value="THIS ACADEMY — governance alignment").font = _FL
    rows = [
        ("Policy manual version in effect", "May 2026", "cfgPolicyVersion"),
        ("Board minutes reviewed for this academy?", None, "cfgBoardReviewed"),
        ("Rules/procedures aligned in this workbook?", None, "cfgRulesAligned"),
        ("Reviewed by / date", None, "cfgAlignReviewer"),
    ]
    r = DATA_ROW
    for lab, seedv, nm in rows:
        ws.cell(row=r, column=2, value=lab).font = _FL
        c = ws.cell(row=r, column=4)
        c.fill = FILL_INPUT
        c.font = F_INPUT
        c.border = BOX
        if seedv:
            c.value = seedv
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        define(wb, nm, "AdvisoryBoard", f"$D${r}")
        r += 1
    dv_list(ws, "=lstYesNo", [f"D{DATA_ROW+1}:D{DATA_ROW+2}"])
    r += 1
    # ---- running board-meeting reference list (kept across academies) ----
    hdr = r
    header_row(ws, ["Meeting Date", "Minutes Location (server folder)",
                    None, "Changes Affecting Academy?",
                    "What Changed / Workbook Updates Made", None,
                    "Reviewed By", "Date Reviewed"], row=r)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    r += 1
    first, last = r, r + 29
    for rr in range(first, last + 1):
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=7)
        for col in (2, 3, 5, 6, 8, 9):
            cc = ws.cell(row=rr, column=col)
            cc.fill = FILL_INPUT
            cc.font = F_INPUT
            cc.border = BOX
        ws[f"B{rr}"].number_format = DATE
        ws[f"I{rr}"].number_format = DATE
        ws.cell(row=rr, column=6).alignment = A_LEFT_WRAP
    dv_list(ws, "=lstYesNo", [f"E{first}:E{last}"])
    define(wb, "nrAB_Date", "AdvisoryBoard", f"$B${first}:$B${last}")
    define(wb, "nrAB_Changes", "AdvisoryBoard", f"$E${first}:$E${last}")
    col_widths(ws, {"A": 3, "B": 13, "C": 24, "D": 20, "E": 13, "F": 30,
                    "G": 20, "H": 16, "I": 13})
    sheet_note(ws, "The board meets 1-2x/year; minutes stay in their server "
                   "folder — record the pointer here. The New Academy reset "
                   "(and the Startup Review button) prompts for the latest "
                   "meeting, whether rules changed, and confirms this "
                   "workbook was aligned before cadets were entered. This "
                   "list persists across academies.")
    return ws


def build_all_inputs(wb):
    build_cadets(wb)
    build_examscores(wb)
    build_spelling(wb)
    build_attendance(wb)
    build_makeup(wb)
    build_skills(wb)
    build_writing(wb)
    build_incidents(wb)
    build_counseling(wb)
    build_pt(wb)
    build_medical(wb)
    build_certifications(wb)
    build_stateexam(wb)
    build_memos(wb)
    build_dailylog(wb)
    build_advisoryboard(wb)
