"""LO-compat test harness.

STATUS: normally a NO-OP now. The Good Friday LET was decomposed into plain
helper columns on Control (P..AF: Easter by the anonymous Gregorian
algorithm), which LibreOffice evaluates natively, so "replaced 0" is the
EXPECTED result and does not mean the substitution silently missed. Kept as a
safety net in case a LET-based Easter is ever reintroduced.

Original note: LibreOffice cannot evaluate LET, and the Good
Friday holiday uses it, which poisons nrAllClosures -> WORKDAY.INTL -> all
170 class days -> every date-derived value downstream. That hides real bugs
from the recalc sweep. This substitutes Python-computed Easter dates into a
TEST COPY only; the shipped workbook keeps the coordinator's proven formula."""
import sys, datetime
from openpyxl import load_workbook

def easter(y):                      # anonymous Gregorian algorithm
    a=y%19; b,c=divmod(y,100); d,e=divmod(b,4); f=(b+8)//25; g=(b-f+1)//3
    h=(19*a+b-d-g+15)%30; i,k=divmod(c,4); l=(32+2*e+2*i-h-k)%7
    m=(a+11*h+22*l)//451; mo,da=divmod(h+l-7*m+114,31)
    return datetime.date(y,mo,da+1)

src,dst=sys.argv[1],sys.argv[2]
wb=load_workbook(src)
ws=wb["Control"]
y1=wb["Settings"]["C7"].value
y1=y1.year if hasattr(y1,"year") else 2026
n=0
for r in range(6,20):
    for col in (3,4):               # C = year1, D = year2
        v=ws.cell(row=r,column=col).value
        if isinstance(v,str) and "LET(" in v:
            yr = y1 if col==3 else y1+1
            ws.cell(row=r,column=col).value = easter(yr)-datetime.timedelta(days=2)
            n+=1
wb.save(dst)
print(f"LO-compat: replaced {n} LET-based Easter formulas "
      f"({'expected - Control now uses helper columns' if n == 0 else 'substituted'}); "
      f"Good Friday {y1} = {easter(y1)-datetime.timedelta(days=2)}")

# Usage:
#   python3 build/bpoc/lo_compat_for_test.py <seeded.xlsx> <test_copy.xlsx>
# then recalc test_copy.xlsx in LibreOffice and sweep for errors. Without
# this step the Good Friday LET formula fails in LO, poisons nrAllClosures,
# and blanks all 170 class days -- hiding every date-derived defect
# downstream (retest deadlines, writing due dates, sign-in, DailyLog).
