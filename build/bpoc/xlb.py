"""BPOC workbook builder helpers.

Layout contract (per coordinator):
  - Column A is empty on every sheet.
  - Rows 1-4 are reserved for the coordinator's own headers/admin notes.
  - Table headers live on row 5; data starts on row 6.

Grid sizing:
  - Cadet grids: rows 6..55  (50 cadets)
  - Long logs:   sized per sheet (see ROWS_* constants)
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from xlstyle import (  # noqa: E402
    FONT, NAVY, STEEL, PAGE, ACCENT, WARN, OK, LIGHT_BAND, BORDER_GRAY,
    BOX, UNDER, F_TITLE, F_SUB, F_SECTION, F_LABEL, F_BODY, F_SMALL,
    F_INPUT, F_FORMULA, F_KPI, FILL_NAVY, FILL_STEEL, FILL_PAGE, FILL_INPUT,
    FILL_BAND, FILL_GOLD, FILL_YELLOW, A_LEFT, A_LEFT_WRAP, A_RIGHT, A_CENTER,
    DATE, TIME, paint, title_bar, section_bar, label, input_cell, define,
    col_widths,
)
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.formatting.rule import CellIsRule, FormulaRule  # noqa: E402

HDR_ROW = 5          # table header row
DATA_ROW = 6         # first data row
CADETS = 50          # roster capacity
CADET_LAST = DATA_ROW + CADETS - 1          # row 55

ROWS_EXAMSCORES = 1500
ROWS_ATTEND = 800
ROWS_MAKEUP = 500
ROWS_SKILLS = 600
ROWS_INCIDENTS = 400
ROWS_COUNSELING = 400
ROWS_MEDICAL = 200
ROWS_SCHEDULE = 900   # schedule blocks (day x time-block rows)
ROWS_CLASSDAYS = 170  # generated class-day calendar

F_HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
FILL_HDR = FILL_NAVY
F_CALC = Font(name=FONT, size=10, color="404040")
FILL_CALC = PatternFill("solid", fgColor="F3F4F6")
FILL_WARNBG = PatternFill("solid", fgColor="F8D7DA")
FILL_OKBG = PatternFill("solid", fgColor="D6E9DC")
FILL_AMBER = PatternFill("solid", fgColor="FCE8B2")

PROTECT_PW = "TPDAcademy"


def sheet_note(ws, text, col=2, row=4, span=8):
    """Small gray usage note at the bottom of the reserved header area."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_SMALL
    c.alignment = A_LEFT
    return c


def header_row(ws, headers, first_col=2, row=HDR_ROW, fill=FILL_HDR):
    """Write the row-5 table header. headers may contain None to skip cols."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=first_col + i)
        c.fill = fill
        c.font = F_HDR
        c.alignment = A_CENTER
        c.border = BOX
        if h is not None:
            c.value = h
    ws.row_dimensions[row].height = 26
    for i, h in enumerate(headers):
        ws.cell(row=row, column=first_col + i).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)


def fill_rows(ws, first_row, last_row, cols):
    """cols: dict col_letter -> (formula_template | ('v', value) | None, style)
    formula templates use {r} for the row number.
    style: 'in' input, 'fx' formula (calc gray), None default body."""
    for r in range(first_row, last_row + 1):
        for cl, spec in cols.items():
            tmpl, style = spec
            c = ws[f"{cl}{r}"]
            if tmpl is not None:
                if isinstance(tmpl, tuple) and tmpl[0] == "v":
                    c.value = tmpl[1]
                else:
                    c.value = "=" + tmpl.format(r=r)
            if style == "in":
                c.fill = FILL_INPUT
                c.font = F_INPUT
                c.border = BOX
            elif style == "fx":
                c.fill = FILL_CALC
                c.font = F_CALC
                c.border = BOX
            else:
                c.border = BOX


def dv_list(ws, formula, ranges, allow_blank=True, enforce=True):
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank,
                        showErrorMessage=enforce)
    for rng in ranges:
        dv.add(rng)
    ws.add_data_validation(dv)
    return dv


def name_range(wb, name, sheet, ref):
    define(wb, name, sheet, ref)


def cf_yes_no(ws, rng, yes="Yes", no="No"):
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=[f'"{yes}"'], fill=FILL_OKBG))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=[f'"{no}"'], fill=FILL_WARNBG))


def cf_formula(ws, rng, formula, fill):
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill))


def protect(ws, locked=True):
    if locked:
        ws.protection.sheet = True
        ws.protection.password = PROTECT_PW
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True


def unlock_range(ws, rng):
    """Mark cells unlocked so protection leaves inputs editable."""
    from openpyxl.styles import Protection
    for row in ws[rng]:
        for c in row:
            c.protection = Protection(locked=False)


def page_setup_portrait(ws, print_area=None, repeat_rows=None, fit_w=1, fit_h=0):
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = fit_w
    ws.page_setup.fitToHeight = fit_h
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if print_area:
        ws.print_area = print_area
    if repeat_rows:
        ws.print_title_rows = repeat_rows


def page_setup_landscape(ws, print_area=None, repeat_rows=None, fit_w=1, fit_h=0):
    page_setup_portrait(ws, print_area, repeat_rows, fit_w, fit_h)
    ws.page_setup.orientation = "landscape"
