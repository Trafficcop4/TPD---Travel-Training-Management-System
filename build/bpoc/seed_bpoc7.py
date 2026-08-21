"""Load the real BPOC 2026-01 (class #7) data into a copy of the V6 workbook.

Sources:
  - BPOC_Academy_Management_Workbook__Hardened_V5.4.xlsm  (cadets, agencies,
    scores, spelling, attendance, makeup, skills, writing, incidents,
    email log, settings)
  - BPOC_7_Hourly_Calendar__Locked_20260507.xlsx           (schedule blocks)

Usage:
    python3 build/bpoc/seed_bpoc7.py <v54.xlsm> <calendar.xlsx>
Output:
    workbooks/BPOC_Academy_Management_V6_BPOC7.xlsx   (seeded copy;
    the clean template BPOC_Academy_Management_V6.xlsx is untouched)
"""
import os
import sys
from datetime import datetime, time

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(HERE))
V6_PATH = os.path.join(ROOT, "workbooks", "BPOC_Academy_Management_V6.xlsx")
OUT_PATH = os.path.join(ROOT, "workbooks", "BPOC_Academy_Management_V6_BPOC7.xlsx")

# calendar class names -> V6 names. Traffic Code / Crash / TIM and the
# Criminal Investigations sub-classes are KEPT under their own names —
# they're TPD sub-classes on ChapterMaster and roll up to chapters 22/32.
NAME_MAP = {
    "ALERRT Levl 1": "ALERRT Level 1",
    "Interacting with Deaf and Hard of Hearing ":
        "Interacting with Deaf and Hard of Hearing",
}
# calendar instructor-text typos -> roster spellings (substring replace)
INSTR_FIXES = {
    "DJ, Schick": "DJ Schick",
    # stray comma split one co-teacher into two fragments on 13 blocks; the
    # workbook's substring matcher then dropped her from the only chapter
    # she co-taught
    "Rebekah, Hutson": "Rebekah Hutson",
}

HOLIDAY_WORDS = ("Memorial Day", "Juneteenth", "Independence Day",
                 "Labor Day", "Holiday", "July 4th", "4th of July")


def sval(c):
    v = c.value
    if v is None:
        return ""
    return str(v).strip() if not isinstance(v, (datetime, time)) else v


def map_class(name):
    name = str(name).strip()
    return NAME_MAP.get(name, name)


def seed():
    v54_path, cal_path = sys.argv[1], sys.argv[2]
    v6 = load_workbook(V6_PATH)                 # keep formulas
    v54 = load_workbook(v54_path, data_only=True, keep_vba=False)
    cal = load_workbook(cal_path, data_only=True)

    # ---------------- Settings ----------------
    ws = v6["Settings"]
    wanted = {
        "cfgAcademyClass": "BPOC-2026-01",
        "cfgCurrentExamNum": None, "cfgCurrentSpellingNum": None,
    }
    s54 = v54["Settings"]
    for r54 in range(5, 30):
        nm = sval(s54.cell(row=r54, column=5))
        # cfgTotalScheduledMinutes is deliberately NOT copied — it is
        # recomputed from the seeded Schedule at the end of this script
        if nm in ("cfgAcademyClass", "cfgStartDate", "cfgEndDate",
                  "cfgCurrentExamNum",
                  "cfgCurrentSpellingNum", "cfgHomeAgency"):
            wanted[nm] = s54.cell(row=r54, column=3).value
    for r in range(6, 60):
        nm = sval(ws.cell(row=r, column=5))
        if nm in wanted and wanted[nm] is not None:
            ws.cell(row=r, column=3).value = wanted[nm]
    print("Settings:", {k: str(v)[:12] for k, v in wanted.items() if v is not None})

    # ---------------- Agencies ----------------
    ws = v6["Agencies"]
    a54 = v54["Agencies"]
    name_to_id = {}
    n = 0
    for r54 in range(5, 46):
        r6 = r54 + 1
        for c54, c6 in ((2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8)):
            v = a54.cell(row=r54, column=c54).value
            if v is not None:
                ws.cell(row=r6, column=c6).value = v
        aid = sval(a54.cell(row=r54, column=2))
        aname = sval(a54.cell(row=r54, column=3))
        if aname:
            name_to_id[aname] = aid or aname
            n += 1
    print(f"Agencies: {n}")

    # ---------------- Cadets ----------------
    ws = v6["Cadets"]
    c54 = v54["Cadets"]
    n = 0
    roster = []                      # "Last, First" as the logs reference them
    for r54 in range(5, 35):
        if c54.cell(row=r54, column=2).value in (None, ""):
            continue
        r6 = r54 + 1
        for col in (2, 3, 4, 5, 7, 9):   # PID, last, first, MI, agency, status
            v = c54.cell(row=r54, column=col).value
            if v is not None:
                ws.cell(row=r6, column=col).value = v
        lastn = sval(c54.cell(row=r54, column=3))
        firstn = sval(c54.cell(row=r54, column=4))
        mi = sval(c54.cell(row=r54, column=5))
        if lastn:
            # exactly what Cadets!F computes: Last, First[ M.]
            nm = f"{lastn}, {firstn}" if firstn else lastn
            if mi:
                nm += f" {mi}."
            roster.append(nm)
        n += 1
    print(f"Cadets: {n}")

    # every log row's cadet name MUST resolve to a roster PID — a typo in
    # the source ("Delgoado, Ryan") would otherwise silently orphan that
    # cadet's records (PID lookup -> "?"). Exact match, else close-match
    # with a printed correction, else fail loudly.
    import difflib
    name_fixes = {}

    def fix_name(nm, where):
        nm = str(nm).strip()
        if nm in roster:
            return nm
        close = difflib.get_close_matches(nm, roster, n=1, cutoff=0.8)
        if close:
            key = (nm, close[0])
            name_fixes[key] = name_fixes.get(key, 0) + 1
            return close[0]
        raise SystemExit(
            f"seed_bpoc7: {where} cadet name {nm!r} matches nobody on the "
            f"roster {roster} — fix the source data")

    # ---------------- ExamScores ----------------
    ws = v6["ExamScores"]
    e54 = v54["ExamScores"]
    out = 6
    for r54 in range(5, 605):
        if sval(e54.cell(row=r54, column=3)) == "":
            continue
        ws.cell(row=out, column=3).value = fix_name(
            e54.cell(row=r54, column=3).value, "ExamScores")                   # name
        ws.cell(row=out, column=6).value = e54.cell(row=r54, column=6).value   # code
        ws.cell(row=out, column=11).value = e54.cell(row=r54, column=11).value # attempt
        ws.cell(row=out, column=12).value = e54.cell(row=r54, column=12).value # raw
        ws.cell(row=out, column=19).value = e54.cell(row=r54, column=19).value # date
        ws.cell(row=out, column=22).value = e54.cell(row=r54, column=20).value # by
        ws.cell(row=out, column=23).value = e54.cell(row=r54, column=21).value # notes
        out += 1
    print(f"ExamScores: {out-6} attempts")

    # ---------------- Spelling grid ----------------
    ws = v6["Spelling"]
    sp = v54["Spelling"]
    n = 0
    for r54 in range(5, 35):
        for col in range(4, 16):   # D..O
            v = sp.cell(row=r54, column=col).value
            if isinstance(v, (int, float)):
                ws.cell(row=r54 + 1, column=col).value = v
                n += 1
    print(f"Spelling scores: {n}")

    # ---------------- Attendance ----------------
    ws = v6["Attendance"]
    at = v54["Attendance"]
    out = 6
    for r54 in range(5, 805):
        if sval(at.cell(row=r54, column=4)) == "":
            continue
        m = {3: 3, 4: 4, 7: 7, 8: 8, 9: 9, 10: 10, 12: 12, 13: 13}
        for c54, c6 in m.items():
            v = at.cell(row=r54, column=c54).value
            if c6 == 4:
                v = fix_name(v, "Attendance")
            if c6 == 3 and isinstance(v, datetime):
                v = v.replace(hour=0, minute=0, second=0, microsecond=0)
            ws.cell(row=out, column=c6).value = v
        ws.cell(row=out, column=15).value = at.cell(row=r54, column=16).value  # notes
        out += 1
    print(f"Attendance events: {out-6}")

    # ---------------- Makeup ----------------
    ws = v6["Makeup"]
    mk = v54["Makeup"]
    out = 6
    for r54 in range(5, 505):
        if sval(mk.cell(row=r54, column=4)) == "":
            continue
        for col in (3, 4, 6, 7, 8, 9, 10, 11, 13):
            v = mk.cell(row=r54, column=col).value
            if col == 4:
                v = fix_name(v, "Makeup")
            ws.cell(row=out, column=col).value = v
        out += 1
    print(f"Makeup entries: {out-6}")

    # ---------------- Skills (per-category rows -> attempt rows) ----------
    ws = v6["Skills"]
    sk = v54["Skills"]
    out = 6
    for r54 in range(5, 605):
        if sval(sk.cell(row=r54, column=3)) == "":
            continue
        ws.cell(row=out, column=3).value = fix_name(
            sk.cell(row=r54, column=3).value, "Skills")                        # name
        ws.cell(row=out, column=5).value = sk.cell(row=r54, column=5).value    # category
        att = sk.cell(row=r54, column=9).value                                 # attempts used
        ws.cell(row=out, column=9).value = att if isinstance(att, (int, float)) and att else 1
        ws.cell(row=out, column=10).value = sk.cell(row=r54, column=10).value  # last result
        ws.cell(row=out, column=11).value = sk.cell(row=r54, column=11).value  # last score
        ws.cell(row=out, column=12).value = sk.cell(row=r54, column=14).value  # date
        ws.cell(row=out, column=13).value = sk.cell(row=r54, column=15).value  # assessed by
        out += 1
    print(f"Skills rows: {out-6}")

    # ---------------- Writing grid (Yes -> X) ----------------
    ws = v6["Writing"]
    wr = v54["Writing"]
    n = 0
    for r54 in range(5, 35):
        for col in range(4, 44):   # D..AQ
            if sval(wr.cell(row=r54, column=col)) == "Yes":
                ws.cell(row=r54 + 1, column=col).value = "X"
                n += 1
    print(f"Writing X marks: {n}")

    # ------- WritingMaster override dates (the academy's real handouts) ----
    # V6 DERIVES Assigned/Due from when the linked chapter is first taught.
    # That is the right fallback for a fresh academy but it is not what BPOC
    # 7 actually handed out: ch35 (Patrol Skills) is first taught 10/19 while
    # its assignments went out 09/07, so 21 of 40 derived due dates land up
    # to 6 weeks late and every cadet's overdue-writing count (Writing!AS ->
    # AT -> sysChecks graduation gate) reads short. Copy the authoritative
    # dates from V5.4 into the override columns I/J; K/L already prefer them.
    wm54 = v54["WritingMaster"]
    real = {}
    for r54 in range(5, 60):
        num = wm54.cell(row=r54, column=2).value
        if not isinstance(num, (int, float)):
            continue
        real[int(num)] = (wm54.cell(row=r54, column=4).value,
                          wm54.cell(row=r54, column=5).value)
    wm = v6["WritingMaster"]
    n, missing = 0, []
    for r in range(6, 46):
        num = wm.cell(row=r, column=2).value
        if not isinstance(num, (int, float)):
            continue
        pair = real.get(int(num))
        if pair is None or pair[0] is None or pair[1] is None:
            missing.append(int(num))
            continue
        wm.cell(row=r, column=9).value = pair[0]     # I Override Assigned
        wm.cell(row=r, column=10).value = pair[1]    # J Override Due
        n += 1
    print(f"WritingMaster override dates seeded: {n}"
          + (f"; NO published dates for #{missing}" if missing else ""))

    # ---------------- Incidents ----------------
    ws = v6["Incidents"]
    inc = v54["Incidents"]
    out = 6
    for r54 in range(5, 405):
        if sval(inc.cell(row=r54, column=4)) == "":
            continue
        for col in (3, 4, 6, 7, 8, 9, 10, 11, 13, 14):
            v = inc.cell(row=r54, column=col).value
            if col == 4:
                v = fix_name(v, "Incidents")
            ws.cell(row=out, column=col).value = v
        out += 1
    print(f"Incidents: {out-6}")
    if name_fixes:
        for (bad, good), cnt in sorted(name_fixes.items()):
            print(f"NAME FIXED: {bad!r} -> {good!r} ({cnt} row(s))")

    # ---------------- EmailLog ----------------
    ws = v6["EmailLog"]
    el = v54["EmailLog"]
    out = 6
    for r54 in range(5, 45):
        d = el.cell(row=r54, column=2).value
        if not isinstance(d, datetime):
            continue
        agency_name = sval(el.cell(row=r54, column=8)).replace(" (All)", "")
        ws.cell(row=out, column=2).value = d
        ws.cell(row=out, column=3).value = name_to_id.get(agency_name, agency_name)
        ws.cell(row=out, column=4).value = el.cell(row=r54, column=5).value
        ws.cell(row=out, column=5).value = el.cell(row=r54, column=6).value
        # the email macro writes this as a NUMBER; seeding it as text made
        # the 36 historic rows a different type from every future row
        _n = sval(el.cell(row=r54, column=11)).split(" ")[0]
        ws.cell(row=out, column=6).value = (int(_n) if _n.isdigit()
                                            else (_n or None))
        ws.cell(row=out, column=8).value = el.cell(row=r54, column=3).value
        out += 1
    print(f"EmailLog runs: {out-6}")

    # ---------------- Schedule (hourly calendar -> blocks) ----------------
    ws = v6["Schedule"]
    grid = cal["BPOC #7"]
    out = 6
    cur_date = None
    skipped = set()
    sched_minutes = 0
    for r in range(7, 520):
        d = grid.cell(row=r, column=4).value          # D: date (day rows)
        if isinstance(d, datetime):
            cur_date = d
        start = grid.cell(row=r, column=6).value       # F
        end = grid.cell(row=r, column=7).value         # G
        cls = sval(grid.cell(row=r, column=8))         # H
        instr = sval(grid.cell(row=r, column=9))       # I
        if cur_date is None or cls == "" or start is None or end is None:
            continue
        if cls == "Lunch" or any(h.lower() in cls.lower() for h in HOLIDAY_WORDS):
            if cls != "Lunch":
                skipped.add(f"{cur_date.date()} {cls}")
            continue
        ws.cell(row=out, column=2).value = cur_date.replace(hour=0, minute=0)
        ws.cell(row=out, column=4).value = start
        ws.cell(row=out, column=5).value = end
        ws.cell(row=out, column=7).value = map_class(cls)
        for bad, good in INSTR_FIXES.items():
            instr = instr.replace(bad, good)
        ws.cell(row=out, column=9).value = instr
        if isinstance(start, time) and isinstance(end, time):
            sched_minutes += ((end.hour * 60 + end.minute)
                              - (start.hour * 60 + start.minute))
        out += 1
    import data_lists as DLx
    import re as _re
    ros = DLx.INSTRUCTORS + DLx.GUEST_ENTITIES
    ros_norm = {" ".join(nm.split()).casefold() for nm in ros}
    unrec = {}
    for r in range(6, out):
        it = sval(ws.cell(row=r, column=9))
        if not it:
            continue
        # segment-level: co-teachers hidden in multi-name strings must each
        # resolve to a roster name (the workbook's any-match can't see them)
        t2 = _re.sub(r"\(.*?\)", "", str(it))
        for seg in _re.split(r"[&,]| and ", t2):
            seg = " ".join(seg.split())
            # EXACT match after normalising: the old bidirectional
            # substring test let fragments of a mangled name ("Rebekah",
            # "Hutson") both pass, so the typo that produced them was
            # never reported
            if seg and seg.casefold() not in ros_norm:
                unrec[seg] = unrec.get(seg, 0) + 1
    if unrec:
        print(f"WARNING unrecognized schedule instructors: {unrec}")
    print(f"Schedule blocks: {out-6}  (skipped holiday rows: {len(skipped)})")

    # ---- academy length comes from the schedule, not from the old workbook --
    # the V5.4 value (145,800 min = 2,430 hrs) is 2.5x the real 972-hr
    # calendar, which inflates cfgClassroomCapMinutes (5%) by the same factor
    st = v6["Settings"]
    for r in range(6, 60):
        if sval(st.cell(row=r, column=5)) == "cfgTotalScheduledMinutes":
            old = st.cell(row=r, column=3).value
            st.cell(row=r, column=3).value = sched_minutes
            print(f"cfgTotalScheduledMinutes: {old} -> {sched_minutes} "
                  f"({sched_minutes/60:.0f} hrs from the schedule); "
                  f"5% classroom cap = {round(sched_minutes*0.05)} min")
            break

    # ---------- instructor banks from actual schedule usage ---------------
    import data_lists as DLmod
    roster = DLmod.INSTRUCTORS + DLmod.GUEST_ENTITIES
    topic_instr = {}
    for r in range(6, out):
        topic = sval(ws.cell(row=r, column=7))
        itext = sval(ws.cell(row=r, column=9))
        if not topic or not itext:
            continue
        for nm in roster:
            if nm in itext:
                topic_instr.setdefault(topic, []).append(nm)
    ib = v6["InstructorBanks"]
    topics = {sval(ib.cell(row=r, column=2)): r for r in range(6, 110)
              if sval(ib.cell(row=r, column=2))}
    seeded, overflow = 0, []
    for topic, names in topic_instr.items():
        row = topics.get(topic)
        if row is None:
            continue
        uniq = list(dict.fromkeys(names))
        from sheets_config import BANK_SLOTS, SEL_SLOTS
        for i, nm in enumerate(uniq[:BANK_SLOTS]):
            ib.cell(row=row, column=3 + i).value = nm          # bank C..
        # slice by the real capacity, not a hard-coded 8: two instructors
        # already on 18 seeded Schedule blocks used to be dropped here
        for i, nm in enumerate(uniq[:SEL_SLOTS]):
            ib.cell(row=row, column=3 + BANK_SLOTS + i).value = nm   # teach
        if len(uniq) > min(BANK_SLOTS, SEL_SLOTS):
            overflow.append(f"{topic} ({len(uniq)})")
        seeded += 1
    print(f"Instructor banks seeded: {seeded} topics"
          + (f"; over 10 instructors: {overflow}" if overflow else ""))

    # ---------- backfill exam dates from the schedule's Test-day blocks ----
    # V5.4 stored no exam dates; exam seq n was administered on "Test n" day.
    test_dates = {}
    for r in range(6, out):
        act = sval(ws.cell(row=r, column=7))
        if act.startswith("Test ") and act[5:].isdigit():
            test_dates[int(act[5:])] = ws.cell(row=r, column=2).value
    code_seq = {f"E{i:02d}": i for i in range(1, 18)}
    es = v6["ExamScores"]
    n = 0
    for r in range(6, 1506):
        code = sval(es.cell(row=r, column=6))
        att = es.cell(row=r, column=11).value
        if code in code_seq and att == 1 and es.cell(row=r, column=19).value is None:
            d = test_dates.get(code_seq[code])
            if d is not None:
                es.cell(row=r, column=19).value = d
                n += 1
    print(f"Exam dates backfilled from schedule: {n}")

    v6.save(OUT_PATH)
    print(f"Saved {OUT_PATH}")


if __name__ == "__main__":
    seed()
