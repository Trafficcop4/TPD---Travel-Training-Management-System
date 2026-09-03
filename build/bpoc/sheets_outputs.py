"""Output & printable sheets: Dashboard, ScoresGrid, Ranking, WatchList,
CadetProfile, Transcript, GradChecklist, DismissalLog, Audit, PrintCenter,
SignIn, EvalSheet, SpellingPrint, WritingHandout, EmailPreview, EmailLog,
NamedRanges registry.
"""
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

from xlb import (
    HDR_ROW, DATA_ROW, CADETS, CADET_LAST, F_HDR, FILL_HDR, F_CALC, FILL_CALC,
    F_LABEL, F_SMALL, F_BODY, F_INPUT, FILL_INPUT, FILL_BAND, FILL_YELLOW,
    F_KPI, F_TITLE, F_SECTION, FILL_NAVY, FILL_STEEL, A_LEFT, A_LEFT_WRAP,
    A_RIGHT, A_CENTER, BOX, UNDER, DATE, header_row, fill_rows, dv_list,
    dv_whole,
    sheet_note, cf_yes_no, cf_formula, FILL_WARNBG, FILL_OKBG, FILL_AMBER,
    col_widths, define, section_bar, label, protect, unlock_range,
    page_setup_portrait, page_setup_landscape,
)
import data_lists as DL
import data_writing as DW

FIRST, LAST = DATA_ROW, CADET_LAST


def _kpi(ws, row, col, title, formula, wide=2):
    ws.merge_cells(start_row=row, start_column=col, end_row=row,
                   end_column=col + wide - 1)
    t = ws.cell(row=row, column=col, value=title)
    t.font = F_SMALL
    t.alignment = A_LEFT
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1,
                   end_column=col + wide - 1)
    v = ws.cell(row=row + 1, column=col, value="=" + formula)
    v.font = F_KPI
    v.alignment = A_LEFT
    for c in range(col, col + wide):
        ws.cell(row=row, column=c).fill = FILL_BAND
        ws.cell(row=row + 1, column=c).fill = FILL_BAND
    return v


# --------------------------------------------------------------------------
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value='=cfgAcademyClass&" — Dashboard"').font = F_KPI
    ws.cell(row=r, column=8, value='="As of "&TEXT(TODAY(),"mm/dd/yyyy")'
            ).font = F_SMALL
    g = ws.cell(row=r, column=10, value=(
        '=HYPERLINK("#InputGuide!B5","Data Entry Guide →")'))
    g.font = F_LABEL
    r += 2
    # key dates (final/state/graduation pull from the Schedule when entered)
    _kpi(ws, r, 2, "Academy start",
         'TEXT(cfgStartDate,"mm/dd/yyyy")')
    _kpi(ws, r, 4, "Final Exam date",
         'LET(d,MAXIFS(nrSCH_Date,nrSCH_Act,"Final Test"),'
         'IF(d=0,"(not on schedule)",TEXT(d,"mm/dd/yyyy")))')
    _kpi(ws, r, 6, "State Exam date",
         'LET(d,MAXIFS(nrSCH_Date,nrSCH_Act,"State Test"),'
         'IF(d=0,"(not on schedule)",TEXT(d,"mm/dd/yyyy")))')
    _kpi(ws, r, 8, "Graduation / end",
         'LET(d,MAXIFS(nrSCH_Date,nrSCH_Act,"Graduation"),'
         'IF(d=0,TEXT(cfgEndDate,"mm/dd/yyyy"),TEXT(d,"mm/dd/yyyy")))')
    _kpi(ws, r, 10, "Active cadets",
         'SUMPRODUCT((nrCadetStatus="Active")*1)')
    r += 3
    section_bar(ws, r, 2, 11, "Today")
    r += 1
    ws.cell(row=r, column=8, value=(
        '=IFERROR("Training Day #"&XLOOKUP(TODAY(),nrCDdate,nrCDnum)&'
        '"  (Week "&XLOOKUP(TODAY(),nrCDdate,nrCDweek)&")",'
        '"(not a class day)")')).font = F_LABEL
    # ONE today-schedule panel, not two. This anchor used to carry a second
    # copy of the identical FILTER capped at 7 rows, ten rows above the
    # "TODAY —" bar's copy capped at 8, so the Dashboard printed the same
    # table twice and the two disagreed the moment a day had 8+ blocks.
    ws.cell(row=r, column=2, value=(
        '=IF(COUNTIF(nrDL_Date,TODAY())=0,'
        '"DailyLog: no entry for today yet","DailyLog: entered ✓")'
    )).font = F_SMALL
    r += 2
    c = ws.cell(row=r, column=2, value=(
        '="TODAY — "&UPPER(TEXT(TODAY(),"dddd, mmmm d"))&'
        'IFERROR(" — TRAINING DAY #"&XLOOKUP(TODAY(),nrCDdate,nrCDnum)&'
        '", WEEK "&XLOOKUP(TODAY(),nrCDdate,nrCDweek),"")'))
    c.font = F_SECTION
    for col in range(2, 12):
        ws.cell(row=r, column=col).fill = FILL_STEEL
    ws.row_dimensions[r].height = 16
    r += 1
    # same cap (9) and the same blank-date guard as SignIn!B10 - one
    # question must not have two answers
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(FILTER(HSTACK(TEXT(nrSCH_Start,"h:mm AM/PM"),'
        'TEXT(nrSCH_End,"h:mm AM/PM"),nrSCH_Act,nrSCH_Instr,nrSCH_Loc),'
        '(nrSCH_Date<>"")*(nrSCH_Date=TODAY())),9),'
        '"No class scheduled today")'))
    r += 9
    _kpi(ws, r, 2, "Separated / total enrolled",
         'SUMPRODUCT((nrCadetStatus="Separated")*1)&" / "&'
         'SUMPRODUCT((nrCadetStatus<>"")*1)')
    _kpi(ws, r, 4, "Class current avg",
         # AVERAGE(IF(...)) is a legacy CSE array formula; stored as an
         # ordinary formula it silently evaluates to 0. AVERAGEIF needs no
         # array entry and is what the rest of the sheet already uses.
         'IFERROR(ROUND(AVERAGEIF(nrCadetStatus,"Active",nrGRcurrent),1),"—")')
    _kpi(ws, r, 6, "Graduation eligible",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCKgradElig="Yes"))')
    _kpi(ws, r, 8, "Flagged cadets",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrFLcount>0))')
    _kpi(ws, r, 10, "Retests overdue / late / undated",
         'COUNTIF(nrES_RetStat,"OVERDUE")+COUNTIF(nrES_RetStat,"CHECK DATE")'
         '+COUNTIF(nrES_RetStat,"LATE RETEST*")'
         '+COUNTIF(nrES_RetStat,"RETEST UNDATED")'
         '+COUNTIF(nrES_RetStat,"RETEST DATE UNCHECKED")')
    r += 3
    _kpi(ws, r, 2, "Spelling interventions",
         'COUNTIF(nrSpellFlag,"INTERVENTION")')
    _kpi(ws, r, 4, "Makeup owed (cadets)",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrATTmakeupOK="No"))')
    _kpi(ws, r, 6, "Delivered hours",
         'ROUND(N(nrCHtotalDelivered),0)&" / "&cfgRequiredHours')
    _kpi(ws, r, 8, "Audit checks failing",
         'SUMPRODUCT((nrAUDstatus<>"OK")*(nrAUDstatus<>""))')
    _kpi(ws, r, 10, "Days to graduation",
         'MAX(0,cfgEndDate-TODAY())')
    _kpi(ws, r, 12, "Cert copies to collect",
         'SUMPRODUCT((nrCadetStatus="Active")*(nrCERTmissing<>""))')
    r += 3
    section_bar(ws, r, 2, 11, "Watch list — highest flag counts first "
                              "(full list on WatchList)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(rngCadetNames,nrCadetAgency,'
        'nrFLcount,nrFLreasons),(nrFLcount>0)*(nrCadetStatus="Active")),'
        'FILTER(nrFLcount,(nrFLcount>0)*(nrCadetStatus="Active")),-1),11),'
        '"No flags — clear")'))
    watch_top = r
    r += 11
    section_bar(ws, r, 2, 11, "Certification reminders — copies to collect "
                              "from cadets (Certifications sheet)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(FILTER(HSTACK(rngCadetNames,nrCadetAgency,'
        'nrCERTmissing),(nrCERTmissing<>"")*(nrCadetStatus="Active")),9),'
        '"All certification copies collected")'))
    r += 9
    section_bar(ws, r, 2, 11, "Open missed-time events — not yet made up "
                              "(Attendance ↔ Makeup by EventID)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(nrAT_ID,Attendance!$D$6:$D$805,'
        'TEXT(nrAT_Date,"mm/dd"),Attendance!$G$6:$G$805,'
        'Attendance!$H$6:$H$805,nrAT_Balance),(nrAT_Cleared="OPEN")),'
        'FILTER(nrAT_Date,(nrAT_Cleared="OPEN")),-1),9),'
        '"All missed time cleared")'))
    r += 9
    section_bar(ws, r, 2, 11, "Outstanding memos (pending / overdue)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(Memos!$B$6:$B$305,nrME_Cadet,'
        'TEXT(nrME_Assigned,"mm/dd"),nrME_Ref,nrME_Subject,'
        'TEXT(nrME_Due,"mm/dd"),nrME_Status),'
        '(nrME_Cadet<>"")*(nrME_Received="")),'
        'FILTER(nrME_Due,(nrME_Cadet<>"")*(nrME_Received="")),1),8),'
        '"No memos outstanding")'))
    r += 8
    section_bar(ws, r, 2, 11, "Class average by exam")
    chart_anchor_row = r + 1
    r += 16
    section_bar(ws, r, 2, 11, "Spelling class average by test")
    chart2_anchor = r + 1

    # charts read the ScoresGrid class-average row and Spelling stats
    ch = LineChart()
    ch.title = "Class average by exam (recorded, final attempts)"
    ch.height = 7
    ch.width = 22
    data = Reference(wb["ScoresGrid"], min_col=4, max_col=28,
                     min_row=CADET_LAST + 2, max_row=CADET_LAST + 2)
    cats = Reference(wb["ScoresGrid"], min_col=4, max_col=28, min_row=HDR_ROW)
    # the data is one ROW (the class-average row): without from_rows openpyxl
    # emits one single-point series per column (25 of them), so the line has
    # nothing to connect and every marker lands on category 1
    ch.add_data(data, titles_from_data=False, from_rows=True)
    ch.set_categories(cats)
    ch.legend = None
    ch.y_axis.scaling.min = 50
    ch.y_axis.scaling.max = 100
    ws.add_chart(ch, f"B{chart_anchor_row}")

    sp = wb["Spelling"]
    sr = CADET_LAST + 2
    ch2 = BarChart()
    ch2.title = "Spelling class average by test"
    ch2.height = 7
    ch2.width = 22
    data2 = Reference(sp, min_col=4, max_col=15, min_row=sr + 1, max_row=sr + 1)
    cats2 = Reference(sp, min_col=4, max_col=15, min_row=HDR_ROW)
    ch2.add_data(data2, titles_from_data=False, from_rows=True)
    ch2.set_categories(cats2)
    ch2.legend = None
    ch2.y_axis.scaling.min = 0
    ch2.y_axis.scaling.max = 100
    ws.add_chart(ch2, f"B{chart2_anchor}")

    col_widths(ws, {"A": 3, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
                    "G": 14, "H": 14, "I": 14, "J": 14, "K": 40})
    # zero input cells: every value here is a formula or a chart, and this is
    # the landing page, so a stray keystroke on a KPI tile or a navigation
    # hyperlink was silently destructive. Locked (still fully selectable, and
    # the macro buttons the installer adds keep working on a locked sheet).
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_scoresgrid(wb):
    ws = wb.create_sheet("ScoresGrid")
    ws.sheet_view.showGridLines = False
    n_exams = 25
    hdrs = ["PID", "Cadet Name"]
    for i in range(n_exams):
        hdrs.append(f'=IF(INDEX(rngEPcode,{i+1})="","",INDEX(rngEPcode,{i+1}))')
    hdrs += ["Current Grade", "Rank"]
    # header formulas need manual write (header_row writes plain values)
    header_row(ws, ["PID", "Cadet Name"] + [None] * n_exams +
               ["Current Grade", "Rank"])
    for i in range(n_exams):
        c = ws.cell(row=HDR_ROW, column=4 + i,
                    value=f'=IF(INDEX(rngEPcode,{i+1})="","",INDEX(rngEPcode,{i+1}))')
    cols = {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF($B{r}="","",Cadets!$F{r})', "fx"),
    }
    for i in range(n_exams):
        cl = get_column_letter(4 + i)
        # blank (not 0) when the cadet has no recorded final attempt:
        # SUMIFS returns 0 on no match, so guard with COUNTIFS first
        cols[cl] = (
            'IF(OR($B{r}="",%s$%d=""),"",IF(COUNTIFS(nrES_PID,$B{r},'
            'nrES_Code,%s$%d,nrES_Final,"Yes",nrES_Rec,">=0")=0,"",'
            'SUMIFS(nrES_Rec,nrES_PID,$B{r},nrES_Code,%s$%d,'
            'nrES_Final,"Yes")))'
            % (cl, HDR_ROW, cl, HDR_ROW, cl, HDR_ROW), "fx")
    gcol = get_column_letter(4 + n_exams)
    rcol = get_column_letter(5 + n_exams)
    cols[gcol] = ('IF($B{r}="","",sysGrades!$M{r})', "fx")
    cols[rcol] = ('IF($B{r}="","",IFERROR(sysGrades!$U{r},""))', "fx")
    fill_rows(ws, FIRST, LAST, cols)
    # class average row
    ar = LAST + 2
    ws.cell(row=ar, column=3, value="Class average:").font = F_LABEL
    for i in range(n_exams):
        cl = get_column_letter(4 + i)
        # average recorded final attempts straight from the log so untaken
        # exams / cadets with no attempt can never drag the average down
        # same population as sysGrades: an out-of-range Recorded score is
        # excluded from the cadet's own average, so it must not move the
        # class average printed here and mailed to the agencies either.
        ws[f"{cl}{ar}"] = (f'=IF({cl}{HDR_ROW}="","",IFERROR(ROUND('
                           f'AVERAGEIFS(nrES_Rec,nrES_Code,{cl}{HDR_ROW},'
                           f'nrES_Final,"Yes",nrES_Rec,">=0",'
                           f'nrES_Rec,"<=100"),1),""))')
        ws[f"{cl}{ar}"].font = F_CALC
    define(wb, "nrSGclassavg", "ScoresGrid",
           f"$D${ar}:${get_column_letter(3+n_exams)}${ar}")
    # sub-70 highlighting
    lastcl = get_column_letter(3 + n_exams)
    # each column is a DIFFERENT exam, and ExamPlan lets an exam carry its
    # own passing score - comparing every column against the global
    # cfgPassingScore made these red cells contradict the authoritative
    # ExamScores "Pass?" column whenever an exam overrode it.
    cf_formula(ws, f"D{FIRST}:{lastcl}{LAST}",
               f'AND(D{FIRST}<>"",ISNUMBER(D{FIRST}),D{FIRST}<'
               f'IFERROR(INDEX(rngEPpass,MATCH(D${HDR_ROW},rngEPcode,0)),'
               f'cfgPassingScore))',
               FILL_WARNBG)
    col_widths(ws, {"A": 3, "B": 10, "C": 24})
    for i in range(n_exams):
        ws.column_dimensions[get_column_letter(4 + i)].width = 7
    sheet_note(ws, "Recorded score of each exam's final attempt (retest cap "
                   "applied). Red = below THAT exam's own passing score on "
                   "ExamPlan (70 unless the exam overrides it).")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_ranking(wb):
    ws = wb.create_sheet("Ranking")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Rank", "Cadet", "Agency", "Current Grade", "Major Avg",
                    "Minor Avg", "Spelling Avg", "Final", "Grad Elig?"])
    ws.cell(row=FIRST, column=2, value=(
        '=IFERROR(SORT(FILTER(HSTACK(nrGRrank,rngCadetNames,nrCadetAgency,'
        'nrGRcurrent,nrGRmajavg,sysGrades!$J$%d:$J$%d,sysGrades!$K$%d:$K$%d,'
        'nrGRfinal,nrCKgradElig),ISNUMBER(nrGRrank)),1),"No eligible cadets yet")'
        % (FIRST, LAST, FIRST, LAST)))
    col_widths(ws, {"A": 3, "B": 7, "C": 24, "D": 20, "E": 13, "F": 11,
                    "G": 11, "H": 12, "I": 9, "J": 11})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:J{LAST+5}",
                        repeat_rows=f"{HDR_ROW}:{HDR_ROW}")
    sheet_note(ws, "Live class ranking (dismissal-review, failed-out and "
                   "separated cadets are unranked per policy).")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_watchlist(wb):
    ws = wb.create_sheet("WatchList")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Cadet", "Agency", "Flags", "Reasons", "Current Grade",
                    "Makeup Owed (min)", "Overdue Writing",
                    "Open Counseling"])
    ws.cell(row=FIRST, column=2, value=(
        '=IFERROR(SORTBY(FILTER(HSTACK(rngCadetNames,nrCadetAgency,nrFLcount,'
        'nrFLreasons,nrGRcurrent,TEXT(nrATTclOwed,"#,##0"),nrWRoverdue,'
        'sysIncidents!$M$%d:$M$%d),(nrFLcount>0)*(nrCadetStatus="Active")),'
        'FILTER(nrFLcount,(nrFLcount>0)*(nrCadetStatus="Active")),-1),'
        '"No flagged cadets")' % (FIRST, LAST)))
    col_widths(ws, {"A": 3, "B": 24, "C": 18, "D": 7, "E": 64, "F": 13,
                    "G": 13, "H": 14, "I": 15})
    page_setup_landscape(ws, print_area=f"B{HDR_ROW}:I{LAST+5}",
                         repeat_rows=f"{HDR_ROW}:{HDR_ROW}")
    sheet_note(ws, "Every threshold is adjustable on Settings. Document your "
                   "response on the Counseling log — that's the early-"
                   "intervention record policy 300.4.B expects.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def _profile_label(ws, r, c, text, val_formula, wide=2):
    ws.cell(row=r, column=c, value=text).font = F_LABEL
    ws.merge_cells(start_row=r, start_column=c + 1, end_row=r,
                   end_column=c + wide)
    v = ws.cell(row=r, column=c + 1, value="=" + val_formula)
    v.font = F_BODY
    return v


def build_cadetprofile(wb):
    ws = wb.create_sheet("CadetProfile")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Cadet:").font = F_LABEL
    sel = ws.cell(row=r, column=3)
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    dv_list(ws, "=rngCadetNames", [f"C{r}"])
    define(wb, "cfgProfileCadet", "CadetProfile", f"$C${r}")
    ws.cell(row=r, column=7, value='=cfgAcademyClass&" — Cadet Profile — "'
            '&TEXT(TODAY(),"mm/dd/yyyy")').font = F_SMALL
    P = 'XLOOKUP(cfgProfileCadet,rngCadetNames,'
    r += 2
    section_bar(ws, r, 2, 10, "Identity & standing")
    r += 1
    _profile_label(ws, r, 2, "PID", P + 'rngCadetPIDs,"")')
    _profile_label(ws, r, 5, "Agency", P + 'nrCadetAgency,"")')
    _profile_label(ws, r, 8, "Status", P + 'nrCadetStatus,"")')
    r += 1
    _profile_label(ws, r, 2, "Current grade", P + 'nrGRcurrent,"")')
    _profile_label(ws, r, 5, "Rank", 'IFERROR(' + P + 'nrGRrank),"unranked")')
    _profile_label(ws, r, 8, "Grad eligible", P + 'nrCKgradElig,"")')
    r += 1
    _profile_label(ws, r, 2, "Blocking issues", P + 'nrCKblocking,"")', wide=8)
    r += 2
    section_bar(ws, r, 2, 10, "Academics (pass = 70 in each category)")
    r += 1
    # NOTE: "Final exam" is the FINAL-EXAM average (sysGrades L = nrGRfinalExam),
    # not the weighted composite (sysGrades N = nrGRfinal, shown as "Current
    # grade" above) — this block is the 70-in-each-category test.
    for lab2, rng in (("Major avg", "nrGRmajavg"),
                      ("Minor avg", f"sysGrades!$J${FIRST}:$J${LAST}"),
                      ("Spelling avg", f"sysGrades!$K${FIRST}:$K${LAST}"),
                      ("Final exam avg", "nrGRfinalExam")):
        _profile_label(ws, r, 2, lab2, P + rng + ',"—")')
        r += 1
    r -= 4
    _profile_label(ws, r, 5, "Needs on remaining Major",
                   P + f'sysGrades!$W${FIRST}:$W${LAST},"—")')
    _profile_label(ws, r + 1, 5, "Needs on remaining Minor",
                   P + f'sysGrades!$X${FIRST}:$X${LAST},"—")')
    _profile_label(ws, r + 2, 5, "Retakes used",
                   P + 'nrGRretakes,"—")')
    _profile_label(ws, r + 3, 5, "Consecutive fails",
                   P + 'nrGRconsec,"—")')
    _profile_label(ws, r, 8, "Spelling flag",
                   P + 'nrSpellFlag,"—")')
    _profile_label(ws, r + 1, 8, "Last vs prev score",
                   P + f'sysGrades!$Z${FIRST}:$Z${LAST},"—")&" / "&'
                   + P + f'sysGrades!$AA${FIRST}:$AA${LAST},"—")')
    r += 5
    section_bar(ws, r, 2, 10, "Attendance / PT / Skills / Writing")
    r += 1
    _profile_label(ws, r, 2, "Classroom missed (net min)",
                   P + f'sysAttendance!$G${FIRST}:$G${LAST},"—")')
    _profile_label(ws, r, 5, "Tier", P + 'nrATTclTier,"—")')
    _profile_label(ws, r, 8, "Makeup owed (min)", P + 'nrATTclOwed,"—")')
    r += 1
    _profile_label(ws, r, 2, "PT sessions missed (net)",
                   P + f'sysAttendance!$O${FIRST}:$O${LAST},"—")')
    _profile_label(ws, r, 5, "PT tier", P + 'nrATTptTier,"—")')
    _profile_label(ws, r, 8, "Final PT", P + 'nrPT_FinalPass,"—")')
    r += 1
    _profile_label(ws, r, 2, "Skills status",
                   'IF(' + P + 'nrSKfailedout,"—")="Yes","FAILED OUT",'
                   'IF(' + P + 'nrSKelig,"—")="Yes","On track","Remediation"))')
    _profile_label(ws, r, 5, "Firearms avg", P + 'nrSKfirearmsAvg,"—")')
    _profile_label(ws, r, 8, "Writing current",
                   P + 'nrWRcurrent,"—")&" ("&' + P +
                   'nrWRoverdue,"0")&" overdue)"')
    r += 2
    section_bar(ws, r, 2, 10, "Flags")
    r += 1
    _profile_label(ws, r, 2, "Flag reasons", P + 'nrFLreasons,"none")', wide=8)
    r += 2
    section_bar(ws, r, 2, 10, "Recent incidents & counseling (10 latest)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(TEXT(nrIN_Date,"mm/dd/yyyy"),'
        'nrIN_Dir,nrIN_Sev,'
        'nrIN_Desc),(nrIN_PID<>"")*(nrIN_PID=XLOOKUP(cfgProfileCadet,'
        'rngCadetNames,rngCadetPIDs,""))),FILTER(nrIN_Date,(nrIN_PID<>"")*'
        '(nrIN_PID=XLOOKUP(cfgProfileCadet,'
        'rngCadetNames,rngCadetPIDs,""))),-1),10),"none")'))
    r += 11
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(TEXT(nrCO_Date,"mm/dd/yyyy"),'
        'nrCO_Type,nrCO_Desc),(nrCO_PID<>"")*'
        '(nrCO_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,rngCadetPIDs,""))),'
        'FILTER(nrCO_Date,(nrCO_PID<>"")*'
        '(nrCO_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,'
        'rngCadetPIDs,""))),-1),10),"none")'))
    r += 11
    section_bar(ws, r, 2, 10, "Attendance exceptions & makeup (10 latest)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(TEXT(nrAT_Date,"mm/dd/yyyy"),'
        'nrAT_Type,'
        'Attendance!$H$6:$H$805,nrAT_Min,nrAT_Sess,nrAT_Excused),'
        '(nrAT_PID<>"")*'
        '(nrAT_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,rngCadetPIDs,""))),'
        'FILTER(nrAT_Date,(nrAT_PID<>"")*'
        '(nrAT_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,'
        'rngCadetPIDs,""))),-1),10),"none")'))
    r += 11
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(TEXT(Makeup!$C$6:$C$505,'
        '"mm/dd/yyyy"),nrMK_Type,'
        'nrMK_Min,nrMK_Sess,nrMK_Credit),(nrMK_PID<>"")*'
        '(nrMK_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,rngCadetPIDs,""))),'
        'FILTER(Makeup!$C$6:$C$505,(nrMK_PID<>"")*'
        '(nrMK_PID=XLOOKUP(cfgProfileCadet,'
        'rngCadetNames,rngCadetPIDs,""))),-1),10),"no makeup entries")'))
    r += 11
    section_bar(ws, r, 2, 10, "OPEN missed time — 6 most recent "
                              "(event / date / type / reason / balance owed)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(SORTBY(FILTER(HSTACK(nrAT_ID,TEXT(nrAT_Date,"mm/dd"),'
        'Attendance!$G$6:$G$805,Attendance!$H$6:$H$805,nrAT_Balance),'
        '(nrAT_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,rngCadetPIDs,""))*'
        '(nrAT_Cleared="OPEN")),FILTER(nrAT_Date,'
        '(nrAT_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,rngCadetPIDs,""))*'
        '(nrAT_Cleared="OPEN")),-1),6),"all missed time cleared")'))
    r += 6
    section_bar(ws, r, 2, 10, "OPEN memos — 7 shown (id / assigned / subject "
                              "/ due / status)")
    r += 1
    # TAKE(...,7) matches the 7 rows reserved below and the B5:J86 print
    # area: uncapped, an 8th open memo spilled off the bottom of the page
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(FILTER(HSTACK(Memos!$B$6:$B$305,'
        'TEXT(nrME_Assigned,"mm/dd"),nrME_Subject,TEXT(nrME_Due,"mm/dd"),'
        'nrME_Status),(nrME_PID=XLOOKUP(cfgProfileCadet,rngCadetNames,'
        'rngCadetPIDs,""))*(nrME_Cadet<>"")*(nrME_Received="")),7),'
        '"none outstanding")'))
    r += 6
    col_widths(ws, {"A": 3, "B": 24, "C": 14, "D": 14, "E": 16, "F": 14,
                    "G": 14, "H": 18, "I": 14, "J": 14})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:J{r}")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")   # C5 is the only picker
    sheet_note(ws, "Pick a cadet, print (Print Center button or File > "
                   "Print). Everything on one page for agency calls and "
                   "review boards.")
    return ws


# --------------------------------------------------------------------------
def build_transcript(wb):
    ws = wb.create_sheet("Transcript")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Cadet:").font = F_LABEL
    sel = ws.cell(row=r, column=3)
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    dv_list(ws, "=rngCadetNames", [f"C{r}"])
    define(wb, "cfgTranscriptCadet", "Transcript", f"$C${r}")
    P = 'XLOOKUP(cfgTranscriptCadet,rngCadetNames,'
    ws.cell(row=r, column=7, value=(
        '="TYLER POLICE ACADEMY — "&cfgAcademyClass&" — OFFICIAL TRANSCRIPT"'
    )).font = F_LABEL
    r += 2
    _profile_label(ws, r, 2, "PID", P + 'rngCadetPIDs,"")')
    _profile_label(ws, r, 5, "Agency", P + 'nrCadetAgency,"")')
    _profile_label(ws, r, 8, "Status", P + 'nrCadetStatus,"")')
    r += 1
    _profile_label(ws, r, 2, "Academy dates",
                   'TEXT(cfgStartDate,"mm/dd/yyyy")&" – "&TEXT(cfgEndDate,"mm/dd/yyyy")')
    _profile_label(ws, r, 5, "Final grade", P + 'nrGRfinal,"—")')
    _profile_label(ws, r, 8, "Class rank", 'IFERROR(' + P + 'nrGRrank),"—")')
    r += 2
    section_bar(ws, r, 2, 10, "Category averages (70 required in each)")
    r += 1
    # the four CATEGORY averages the 70-in-each rule is tested against;
    # the Final Exam row reads sysGrades L (final-exam average), NOT the
    # weighted composite in N which is already printed as "Final grade"
    for lab2, rng, w in (("Major (40%)", "nrGRmajavg", None),
                         ("Minor (30%)", f"sysGrades!$J${FIRST}:$J${LAST}", None),
                         ("Spelling (10%)", f"sysGrades!$K${FIRST}:$K${LAST}", None),
                         ("Final Exam (20%)", "nrGRfinalExam", None)):
        _profile_label(ws, r, 2, lab2, P + rng + ',"—")')
        r += 1
    r += 1
    section_bar(ws, r, 2, 10, "Exam record (recorded scores, final attempts)")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IFERROR(FILTER(HSTACK(ExamScores!$G$%d:$G$%d,ExamScores!$H$%d:$H$%d,'
        'ExamScores!$K$%d:$K$%d,ExamScores!$M$%d:$M$%d,'
        'TEXT(ExamScores!$S$%d:$S$%d,"mm/dd/yyyy")),'
        '(nrES_PID=%srngCadetPIDs,""))*(nrES_Final="Yes")*(nrES_Rec<>"")),"—")'
        % (FIRST, FIRST + 1499, FIRST, FIRST + 1499, FIRST, FIRST + 1499,
           FIRST, FIRST + 1499, FIRST, FIRST + 1499, P)))
    # cap to the reserved window so a long log can never #SPILL! into the
    # section bar below (25 exams max on the plan)
    v17 = ws.cell(row=r, column=2).value
    assert v17.startswith('=IFERROR(FILTER(') and v17.endswith('),"—")')
    ws.cell(row=r, column=2).value = (
        '=IFERROR(TAKE(' + v17[len('=IFERROR('):-len(',"—")')] + ',26),"—")')
    r += 26
    section_bar(ws, r, 2, 10, "Attendance / Skills / PT / Writing / Conduct")
    r += 1
    # no "of N cap": there is no classroom allowance to be a fraction of
    _profile_label(ws, r, 2, "Classroom minutes still owed",
                   P + f'sysAttendance!$G${FIRST}:$G${LAST},"0")&'
                   '" min (all missed time must be made up)"', wide=4)
    r += 1
    _profile_label(ws, r, 2, "PT sessions missed (net)",
                   P + f'sysAttendance!$O${FIRST}:$O${LAST},"0")&" of "&'
                   'cfgPTCapSessions&" cap"', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Makeup complete", P + 'nrATTmakeupOK,"—")', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Skills", 'IF(' + P + 'nrSKfailedout,"—")="Yes",'
                   '"FAILED OUT","Qualified/on record — see Skills log")', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Firearms average", P + 'nrSKfirearmsAvg,"—")', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Final PT assessment",
                   P + 'nrPT_FinalPass,"—")&" ("&' + P +
                   'nrPT_FinalPts,"—")&" pts)"', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Writing assignments complete",
                   P + 'nrWRcomplete,"0")&" of 40"', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Incidents (neg/pos)",
                   P + f'sysIncidents!$F${FIRST}:$F${LAST},"0")&" / "&'
                   + P + f'sysIncidents!$G${FIRST}:$G${LAST},"0")', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Awards",
                   'IFERROR(TEXTJOIN(", ",TRUE,IF(nrAWfinal=cfgTranscriptCadet,'
                   'nrAWnames,"")),"")', wide=4)
    r += 1
    _profile_label(ws, r, 2, "Required certifications",
                   'LET(m,' + P + 'nrCERTmissing,""),'
                   'IF(m="","All on file (TIM, SFST, TCIC, CPR/AED, ALERRT, '
                   'ICS)","Outstanding: "&m))', wide=4)
    r += 1
    _profile_label(ws, r, 2, "State licensing exam",
                   P + 'nrSEstatus,"Not yet attempted")', wide=4)
    r += 2
    ws.cell(row=r, column=2, value="Training Coordinator:").font = F_LABEL
    ws.cell(row=r, column=5, value="_______________________").font = F_BODY
    ws.cell(row=r, column=7, value="Date:").font = F_LABEL
    ws.cell(row=r, column=8, value="____________").font = F_BODY
    r += 2
    ws.cell(row=r, column=2, value=DL.ACADEMY_ADDRESS).font = F_SMALL
    col_widths(ws, {"A": 3, "B": 26, "C": 14, "D": 14, "E": 16, "F": 14,
                    "G": 16, "H": 14, "I": 12, "J": 12})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:J{r}")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")   # C5 is the only picker
    sheet_note(ws, "The end-of-academy deliverable per cadet — print one per "
                   "cadet for the agency packet.")
    return ws


# --------------------------------------------------------------------------
def build_gradchecklist(wb):
    ws = wb.create_sheet("GradChecklist")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["PID", "Cadet", "Agency", "Academic", "Classroom",
                    "PT Sessions", "Skills", "Incidents", "Writing",
                    "Makeup", "Final PT", "No Dismiss Rev", "Certs",
                    "Enroll Docs", "ELIGIBLE", "Blocking Issues"])
    cols = {
        "B": ('IF(Cadets!$B{r}="","",Cadets!$B{r})', "fx"),
        "C": ('IF($B{r}="","",Cadets!$F{r})', "fx"),
        "D": ('IF($B{r}="","",Cadets!$H{r})', "fx"),
        "E": ('IF($B{r}="","",sysChecks!$E{r})', "fx"),
        "F": ('IF($B{r}="","",sysChecks!$F{r})', "fx"),
        "G": ('IF($B{r}="","",sysChecks!$G{r})', "fx"),
        "H": ('IF($B{r}="","",sysChecks!$H{r})', "fx"),
        "I": ('IF($B{r}="","",sysChecks!$I{r})', "fx"),
        "J": ('IF($B{r}="","",sysChecks!$J{r})', "fx"),
        "K": ('IF($B{r}="","",sysChecks!$K{r})', "fx"),
        "L": ('IF($B{r}="","",sysChecks!$L{r})', "fx"),
        "M": ('IF($B{r}="","",IF(sysChecks!$M{r}="Yes","No","Yes"))', "fx"),
        "N": ('IF($B{r}="","",sysChecks!$Q{r})', "fx"),
        # Enroll Docs (sysChecks W) sits BEFORE the verdict columns so
        # ELIGIBLE and Blocking Issues stay last on the printed page.
        "O": ('IF($B{r}="","",sysChecks!$W{r})', "fx"),
        "P": ('IF($B{r}="","",sysChecks!$N{r})', "fx"),
        "Q": ('IF($B{r}="","",sysChecks!$O{r})', "fx"),
    }
    fill_rows(ws, FIRST, LAST, cols)
    cf_yes_no(ws, f"E{FIRST}:P{LAST}")
    # anything that is not a plain "Yes" is a block — "Pending" / "Not taken"
    # (final PT never assessed, rubric never configured) must read as red,
    # not as an uncoloured neutral value on the graduation gate
    cf_formula(ws, f"E{FIRST}:P{LAST}",
               f'AND(E{FIRST}<>"",E{FIRST}<>"Yes")', FILL_WARNBG)
    col_widths(ws, {"A": 3, "B": 10, "C": 24, "D": 18, "Q": 54})
    for cl in "EFGHIJKLMNOP":
        ws.column_dimensions[cl].width = 10
    page_setup_landscape(ws, print_area=f"B{HDR_ROW}:Q{LAST}",
                         repeat_rows=f"{HDR_ROW}:{HDR_ROW}")
    sheet_note(ws, "The final gate before the ceremony — columns Academic "
                   "through Enroll Docs must ALL read Yes and Blocking Issues must "
                   "read 'Eligible' — it is never blank for a passing cadet. "
                   "Anything else blocks and is spelled out in "
                   "Blocking Issues: 'No', 'Pending' (Final PT points rubric "
                   "not entered on Settings) and 'Not taken' (final PT never "
                   "assessed) are all blocks, not neutral values.")
    protect(ws)
    return ws


# --------------------------------------------------------------------------
def build_dismissallog(wb):
    ws = wb.create_sheet("DismissalLog")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["ReviewID", "Cadet Name", "PID", "Review Type",
                    "Trigger", "Opened", "Board/Reviewer", "Decision Date",
                    "Outcome", "Asst. Chief Approval", "Approval Date",
                    "Agency Notified", "Docs Ref", "Notes",
                    "Closes Trigger"])
    first, last = DATA_ROW, DATA_ROW + 99
    fill_rows(ws, first, last, {
        "B": ('IF($C{r}="","","R"&TEXT(ROW()-%d,"000"))' % HDR_ROW, "fx"),
        "C": (None, "in"),
        "D": ('IF($C{r}="","",IFERROR(INDEX(rngCadetPIDs,MATCH($C{r},rngCadetNames,0)),"?"))', "fx"),
        "E": (None, "in"), "F": (None, "in"), "G": (None, "in"),
        "H": (None, "in"), "I": (None, "in"), "J": (None, "in"),
        "K": (None, "in"), "L": (None, "in"), "M": (None, "in"),
        "N": (None, "in"), "O": (None, "in"), "P": (None, "in"),
    })
    for r in range(first, last + 1):
        for cl in ("G", "I", "L"):
            ws[f"{cl}{r}"].number_format = DATE
    dv_list(ws, "=rngCadetNames", [f"C{first}:C{last}"])
    dv_list(ws, "=lstReviewType", [f"E{first}:E{last}"])
    dv_list(ws, "=lstReviewOutcome", [f"J{first}:J{last}"])
    dv_list(ws, "=lstYesNo", [f"K{first}:K{last}", f"M{first}:M{last}"])
    dv_list(ws, "=lstDismissalTrigger", [f"P{first}:P{last}"])
    # the engine reads these back: recording "Retained"/"Retained w/ Plan"
    # with the Assistant-Chief approval and the trigger this review closes
    # is what actually clears sysChecks DismissReview. Before, the sheet fed
    # nothing at all, so a failed retest blocked graduation and unranked the
    # cadet FOREVER with no non-destructive remedy.
    define(wb, "nrDIS_PID", "DismissalLog", f"$D${first}:$D${last}")
    define(wb, "nrDIS_Outcome", "DismissalLog", f"$J${first}:$J${last}")
    define(wb, "nrDIS_Approval", "DismissalLog", f"$K${first}:$K${last}")
    define(wb, "nrDIS_Closes", "DismissalLog", f"$P${first}:$P${last}")
    col_widths(ws, {"A": 3, "B": 9, "C": 22, "D": 9, "E": 16, "F": 28,
                    "G": 11, "H": 20, "I": 12, "J": 15, "K": 16, "L": 12,
                    "M": 13, "N": 14, "O": 30, "P": 30})
    sheet_note(ws, "Formal record of every dismissal/academic/conduct review: "
                   "trigger, decision, Assistant-Chief approval (policy "
                   "600.2.E) and agency notification. This is the file "
                   "TCOLE/legal asks for. It is also the ONLY way to close a "
                   "review the engine opened: set Closes Trigger to the "
                   "trigger being reviewed, Outcome to Retained or Retained "
                   "w/ Plan, and Asst. Chief Approval to Yes — sysChecks "
                   "DismissReview then clears for THAT trigger only, so a "
                   "later failed retest / skills fail-out / conduct review "
                   "re-opens the block. Outcome 'Separated' closes nothing: "
                   "set the cadet's Status on the Cadets sheet instead.")
    return ws


# --------------------------------------------------------------------------
def build_audit(wb):
    ws = wb.create_sheet("Audit")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value=(
        '="TCOLE AUDIT READINESS — "&cfgAcademyClass&" — "&TEXT(TODAY(),"mm/dd/yyyy")'
    )).font = F_KPI
    r += 2
    section_bar(ws, r, 2, 8, "Program checks (live from the engine)")
    r += 1
    hdr2 = r
    header_row(ws, ["Check", "Value", "Target", "Status", "Detail"], row=r)
    r += 1
    # driven off the engine's own list — a hard-coded count meant a check
    # added to sysAudit was counted by the Dashboard "Audit checks failing"
    # tile but never printed on the packet handed to TCOLE, so the tile went
    # red with no visible cause
    from sheets_engine import AUDIT_CHECKS
    n_checks = len(AUDIT_CHECKS)
    for i in range(n_checks):
        sr = FIRST + i
        ws.cell(row=r, column=2, value=f"=sysAudit!$B${sr}").font = F_BODY
        ws.cell(row=r, column=3, value=f"=sysAudit!$C${sr}").font = F_CALC
        ws.cell(row=r, column=4, value=f"=sysAudit!$D${sr}").font = F_CALC
        s = ws.cell(row=r, column=5, value=f"=sysAudit!$E${sr}")
        s.font = F_LABEL
        ws.cell(row=r, column=6, value=f"=sysAudit!$F${sr}").font = F_SMALL
        r += 1
    cf_formula(ws, f"E{hdr2+1}:E{r-1}", f'AND($E{hdr2+1}<>"OK",$E{hdr2+1}<>"")',
               FILL_WARNBG)
    cf_formula(ws, f"E{hdr2+1}:E{r-1}", f'$E{hdr2+1}="OK"', FILL_OKBG)
    r += 1
    section_bar(ws, r, 2, 8, "Requirement sources — what is TCOLE rule vs "
                             "Academy policy")
    r += 1
    for txt in (
        "TCOLE Rule (Ch. 215/218/219): course #1000736 minimum hours per "
        "chapter, licensed/SME instructors, course records retention, "
        "exam security, student enrollment standards (L2/L3, medical), "
        "attendance records.",
        "Academy Policy (May 2026 manual): weighted grading (40/30/10/20), "
        "70-in-each-category rule, 12 spelling tests across the academy + "
        "75% intervention, "
        "40 writing assignments, PT program and caps, awards.",
    ):
        c = ws.cell(row=r, column=2, value="• " + txt)
        c.font = F_SMALL
        c.alignment = A_LEFT_WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    section_bar(ws, r, 2, 8, "Program-level TCOLE requirements (mark Yes "
                             "when satisfied)")
    r += 1
    header_row(ws, ["Requirement", None, None, None, "Met?", "Notes"], row=r)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    prg_first = r
    import data_chapters as DC
    for item, detail in DC.PROGRAM_REQS:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=2, value=item + " — " + detail)
        c.font = F_BODY
        c.alignment = A_LEFT_WRAP
        m = ws.cell(row=r, column=6)
        m.fill = FILL_INPUT
        m.font = F_INPUT
        m.border = BOX
        n = ws.cell(row=r, column=7)
        n.fill = FILL_INPUT
        n.font = F_INPUT
        n.border = BOX
        ws.row_dimensions[r].height = 26
        r += 1
    prg_last = r - 1
    dv_list(ws, '"Yes,No,N/A"', [f"F{prg_first}:F{prg_last}"])
    define(wb, "nrPRGitems", "Audit", f"$B${prg_first}:$B${prg_last}")
    define(wb, "nrPRGmet", "Audit", f"$F${prg_first}:$F${prg_last}")
    cf_formula(ws, f"F{prg_first}:F{prg_last}",
               f'AND($B{prg_first}<>"",$F{prg_first}<>"Yes",$F{prg_first}<>"N/A")',
               FILL_WARNBG)
    r += 1
    section_bar(ws, r, 2, 8, "Per-cadet enrollment documents (mark Yes when "
                             "on file)")
    r += 1
    doc_hdr = r
    header_row(ws, ["Cadet", "Enroll App", "TCLEDDS L1", "Medical (L2)",
                    "Psych (L3)", "Background", "Photo ID/DL", "Rules Ack",
                    "All Docs?"], row=r)
    r += 1
    doc_first = r
    doc_last = r + CADETS - 1
    # (no `cols` dict here: the grid is written cell-by-cell below. A dead
    # copy used to sit at this spot still describing the PRE-shift layout —
    # AllDocs in column I counting C:H = 6 — which is exactly the stale
    # column-letter bug this workbook keeps regressing on.)
    for rr in range(doc_first, doc_last + 1):
        src = FIRST + (rr - doc_first)
        ws.cell(row=rr, column=2,
                value=f'=IF(Cadets!$B{src}="","",Cadets!$F{src})').font = F_CALC
        for c in range(3, 10):
            cc = ws.cell(row=rr, column=c)
            cc.fill = FILL_INPUT
            cc.font = F_INPUT
            cc.border = BOX
        ws.cell(row=rr, column=10, value=(
            f'=IF($B{rr}="","",IF(COUNTIF(C{rr}:I{rr},"Yes")=7,"Yes","No"))'
        )).font = F_CALC
    dv_list(ws, "=lstYesNo", [f"C{doc_first}:I{doc_last}"])
    cf_yes_no(ws, f"J{doc_first}:J{doc_last}")
    # the roll-up was read by NOTHING: no named range, no sysAudit line, no
    # graduation gate — so a cadet with an empty enrollment file still read
    # ELIGIBLE and the Audit sheet's own live-check block said OK. Never
    # hard-code 67/116 here: adding an audit check shifts the whole grid.
    define(wb, "nrENRall", "Audit", f"$J${doc_first}:$J${doc_last}")
    col_widths(ws, {"A": 3, "B": 30, "C": 11, "D": 11, "E": 12, "F": 11,
                    "G": 12, "H": 12, "I": 10})
    # NO repeat_rows: Excel repeats print titles on EVERY page including
    # page 1, so naming the enrollment-docs header (which sits ~60 rows into
    # the print area) printed that header above the report banner on page 1
    # and twice on the page it naturally falls on.
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:J{doc_last}")
    # ---- A16: the Audit sheet was the last mixed input/formula deliverable
    # left completely unprotected - 207 formula cells (the sysAudit mirror,
    # the cadet-name column and the AllDocs roll-up) plus the requirement
    # labels could be typed over. Its input cells were never unlocked, so
    # protecting it without this would have frozen the sheet instead.
    unlock_range(ws, f"F{prg_first}:G{prg_last}")
    unlock_range(ws, f"C{doc_first}:I{doc_last}")
    protect(ws)
    sheet_note(ws, "Top: live program checks. Bottom: per-cadet enrollment "
                   "file checklist. Chapter-level records live on "
                   "ChapterMaster; instructor credentials on Instructors.")
    return ws


# --------------------------------------------------------------------------
def build_addendum(wb):
    """Printable excess-hours report: per-chapter delivered vs TCOLE minimum,
    with the correct separate-reporting course number for each excess."""
    import data_chapters as DC
    ws = wb.create_sheet("Addendum")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value=(
        '="TYLER POLICE ACADEMY — "&cfgAcademyClass&'
        '" — EXCESS HOURS REPORTING (ADDENDUM) — "&TEXT(TODAY(),"mm/dd/yyyy")'
    )).font = F_KPI
    r += 1
    c = ws.cell(row=r, column=2, value=(
        "The BPOC is reported to TCOLE at exactly 736 hours. Hours delivered "
        "beyond a chapter's TCOLE minimum are reported separately: Arrest & "
        "Control, Driving and Firearms under their own course numbers; all "
        "other excess under the Addendum to BPOC (#101). Verify course "
        "numbers against current TCOLE reporting guidance before submission."))
    c.font = F_SMALL
    c.alignment = A_LEFT_WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 40
    r += 2
    hdr = r
    header_row(ws, ["Ch #", "Chapter / Class", "TCOLE Min Hrs",
                    "Delivered Hrs", "Excess Hrs", "Report Under"], row=r)
    r += 1
    first = r
    n = len(DC.CHAPTERS)
    for i in range(n):
        rr = first + i
        src = DATA_ROW + i          # ChapterMaster data row
        ch = DC.CHAPTERS[i][1]
        ws.cell(row=rr, column=2, value=f"=ChapterMaster!$C${src}").font = F_BODY
        ws.cell(row=rr, column=3, value=f"=ChapterMaster!$D${src}").font = F_BODY
        ws.cell(row=rr, column=4, value=f"=ChapterMaster!$E${src}").font = F_CALC
        ws.cell(row=rr, column=5, value=(
            f'=IF(ChapterMaster!$G{src}=0,"",ChapterMaster!$G{src})'
        )).font = F_CALC
        ws.cell(row=rr, column=6, value=(
            f'=IF(OR(ChapterMaster!$G{src}="",ChapterMaster!$G{src}=0),"",'
            f'IF(ChapterMaster!$G{src}>ChapterMaster!$E{src},'
            f'ROUND(ChapterMaster!$G{src}-ChapterMaster!$E{src},2),""))'
        )).font = F_LABEL
        report_as = DC.SEPARATE_REPORT.get(ch, DC.ADDENDUM_COURSE)
        ws.cell(row=rr, column=7, value=(
            f'=IF($F{rr}="","","{report_as}")')).font = F_BODY
        for ccol in range(2, 8):
            ws.cell(row=rr, column=ccol).border = BOX
    last = first + n - 1
    cf_formula(ws, f"F{first}:F{last}",
               f'AND($F{first}<>"",$F{first}>0)', FILL_AMBER)
    r = last + 2
    ws.cell(row=r, column=3, value="Total excess to report:").font = F_LABEL
    tot = ws.cell(row=r, column=6, value=f"=SUM(F{first}:F{last})")
    tot.font = F_KPI
    ws.cell(row=r + 1, column=3,
            value="— of which under #101 (Addendum):").font = F_LABEL
    ws.cell(row=r + 1, column=6, value=(
        f'=SUMPRODUCT(($F{first}:$F{last}<>"")*'
        f'($G{first}:$G{last}="{DC.ADDENDUM_COURSE}")*'
        f'IFERROR($F{first}:$F{last}+0,0))')).font = F_CALC
    ws.cell(row=r + 2, column=3,
            value="— of which under separate course #s:").font = F_LABEL
    ws.cell(row=r + 2, column=6, value=(
        f'=SUMPRODUCT(($F{first}:$F{last}<>"")*'
        f'($G{first}:$G{last}<>"{DC.ADDENDUM_COURSE}")*'
        f'($G{first}:$G{last}<>"")*IFERROR($F{first}:$F{last}+0,0))'
    )).font = F_CALC
    r += 4
    ws.cell(row=r, column=2, value="Training Coordinator:").font = F_LABEL
    ws.cell(row=r, column=4, value="_______________________").font = F_BODY
    ws.cell(row=r, column=6, value="Date:").font = F_LABEL
    ws.cell(row=r, column=7, value="____________").font = F_BODY
    r += 2
    ws.cell(row=r, column=2, value=DL.ACADEMY_ADDRESS).font = F_SMALL
    col_widths(ws, {"A": 3, "B": 7, "C": 46, "D": 12, "E": 12, "F": 11,
                    "G": 30, "H": 4})
    # print titles must start at the FIRST row of the print area: naming
    # only the table header (row hdr, three rows in) made Excel repeat it at
    # the top of page 1 above the banner and again where it naturally falls.
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:G{r}",
                        repeat_rows=f"{HDR_ROW}:{hdr}")
    protect(ws)
    sheet_note(ws, "Fills from ChapterMaster/Schedule automatically — only "
                   "chapters with logged hours show; excess rows highlight. "
                   "Print for the TCOLE reporting packet.")
    return ws


# --------------------------------------------------------------------------
def build_chapterpacket(wb):
    """The auditor's 'show me your file for chapter X' page: pick a chapter,
    everything about it appears — hours, blocks, instructors, file status,
    special requirements, linked exam stats."""
    ws = wb.create_sheet("ChapterPacket")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Chapter:").font = F_LABEL
    sel = ws.cell(row=r, column=3, value="1")
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    # nrCHnum holds chapter numbers as TEXT; a General-formatted picker turns
    # a dropdown pick into a NUMBER and every exact XLOOKUP/MATCH/= below
    # silently returns #N/A. Text format keeps the key type uniform.
    sel.number_format = "@"
    dv_list(ws, "=nrCHnum", [f"C{r}"])
    define(wb, "cfgPacketChapter", "ChapterPacket", f"$C${r}")
    C = "cfgPacketChapter"
    ws.cell(row=r, column=5, value=(
        f'="TCOLE TRAINING FILE — "&cfgAcademyClass&" — Ch. "&{C}&" "'
        f'&IFERROR(XLOOKUP({C},nrCHnum,nrCHname),"")')).font = F_KPI
    r += 2
    X = f'IFERROR(XLOOKUP({C},nrCHnum,'
    _profile_label(ws, r, 2, "Module", X + 'nrCHmod),"")')
    _profile_label(ws, r, 5, "TCOLE min hrs", X + 'nrCHmin),"")')
    _profile_label(ws, r, 8, "Delivered hrs", X + 'nrCHdeliv),"")')
    r += 1
    _profile_label(ws, r, 2, "First taught",
                   'IF(IFERROR(N(XLOOKUP(' + C + ',nrCHnum,nrCHfirst)),0)=0,'
                   '"(not yet taught)",TEXT(XLOOKUP(' + C +
                   ',nrCHnum,nrCHfirst),"mm/dd/yyyy"))')
    # MAXIFS returns 0 (not an error) when the chapter has no schedule
    # blocks, so IFERROR never fired and this printed 12/30/1899 right beside
    # the "no hours logged" line on the same page.
    _profile_label(ws, r, 5, "Last taught",
                   'IF(IFERROR(MAXIFS(nrSCH_Date,nrSCH_ChNum,' + C + '),0)=0,'
                   '"(not yet taught)",TEXT(MAXIFS(nrSCH_Date,nrSCH_ChNum,'
                   + C + '),"mm/dd/yyyy"))')
    _profile_label(ws, r, 8, "vs TCOLE min",
                   'LET(d,IFERROR(XLOOKUP(' + C + ',nrCHnum,nrCHdeliv),0),'
                   'm,IFERROR(XLOOKUP(' + C + ',nrCHnum,nrCHmin),0),'
                   'IF(d=0,"no hours logged",TEXT(d-m,"+0.##;-0.##;0")))')
    r += 2
    section_bar(ws, r, 2, 9, "Training-file contents (required by the IRG)")
    r += 1
    for lab2, rng in (("Lesson plan (SME)", "nrCHlesson"),
                      ("Instructor bio(s)", "nrCHbio"),
                      ("Sign-in sheets (w/ PID)", "nrCHsignin"),
                      ("Assessment", "nrCHassess"),
                      ("Grade sheet", "nrCHgrade"),
                      ("Course evaluations", "nrCHevals"),
                      ("Handouts/PPT (optional)", "nrCHhandoutOpt")):
        _profile_label(ws, r, 2, lab2, X + rng + '),"")')
        r += 1
    r -= 7
    _profile_label(ws, r, 6, "Special TCOLE requirement",
                   X + 'nrCHspecial),"")&""', wide=4)
    _profile_label(ws, r + 1, 6, "Special requirement met",
                   X + 'nrCHspecialMet),"—")')
    _profile_label(ws, r + 2, 6, "File complete?", X + 'nrCHfileOK),"No")')
    lex = ('IFERROR(INDEX(rngEPcode,MATCH(' + C + ',rngEPch,0)),"")')
    _profile_label(ws, r + 3, 6, "Linked exam",
                   'IF(' + lex + '="","(none)",' + lex +
                   '&" — "&IFERROR(INDEX(rngEPname,MATCH(' + C +
                   ',rngEPch,0)),""))', wide=4)
    _profile_label(ws, r + 4, 6, "Exam class avg / low / fails",
                   'LET(c,' + lex + ',IF(c="","—",'
                   'IFERROR(ROUND(AVERAGEIFS(nrES_Rec,nrES_Code,c,'
                   'nrES_Final,"Yes",nrES_Rec,">=0",nrES_Rec,"<=100"),1),"—")'
                   '&" / "&'
                   'IFERROR(MINIFS(nrES_Rec,nrES_Code,c,nrES_Final,"Yes",'
                   'nrES_Rec,">=0",nrES_Rec,"<=100"),"—")'
                   '&" / "&COUNTIFS(nrES_Code,c,nrES_Final,"Yes",'
                   'nrES_Rec,">=0",nrES_Rec,"<"&cfgPassingScore)))', wide=4)
    r += 8
    section_bar(ws, r, 2, 9, "Instructors who taught this chapter "
                             "(from the schedule)")
    r += 1
    # a blank picker degenerates the delimiter-wrapped SEARCH to looking for
    # ", ," inside ", ," - which matches every instructor who has taught
    # NOTHING, filling the 12-row instructor reservation with blank rows.
    ws.cell(row=r, column=2, value=(
        '=IF(' + C + '="","— pick a chapter in C5 —",'
        'IFERROR(TAKE(FILTER(HSTACK(nrInstrNames,nrInstrReady,'
        'nrInstrChTaught),'
        '(nrInstrNames<>"")*(nrInstrChTaught<>"")*'
        'ISNUMBER(SEARCH(", "&' + C + '&",",", "&nrInstrChTaught&","))),20),'
        '"— none on the schedule yet —"))'))
    # the cap used to be 12 against exactly 12 free rows (the section bar
    # below is a MERGED range, which always blocks a spill), and BPOC 7
    # already puts 10 instructors on chapter 35 — an 11th and 12th would fit,
    # a 13th vanished from a TCOLE chapter training file with no error, no
    # count and no marker. The reservation is raised WITH the cap, and the
    # row after it says out loud when the list was still truncated.
    r += 20
    _ni = ('SUMPRODUCT((nrInstrNames<>"")*(nrInstrChTaught<>"")*'
           'ISNUMBER(SEARCH(", "&' + C + '&",",", "&nrInstrChTaught&",")))')
    ws.cell(row=r, column=2, value=(
        '=IF(' + C + '="","",IF(' + _ni + '>20,"+ "&(' + _ni + '-20)&'
        '" more instructor(s) not shown — see the Instructors sheet",""))'
    )).font = F_SMALL
    r += 1
    section_bar(ws, r, 2, 9, "Schedule blocks delivered")
    r += 1
    ws.cell(row=r, column=2, value=(
        '=IF(' + C + '="","— pick a chapter in C5 —",'
        'IFERROR(TAKE(FILTER(HSTACK(TEXT(nrSCH_Date,"mm/dd/yyyy"),'
        'TEXT(nrSCH_Start,"h:mm AM/PM"),TEXT(nrSCH_End,"h:mm AM/PM"),'
        'nrSCH_Act,nrSCH_Instr),(nrSCH_ChNum<>"")*(nrSCH_ChNum='
        + C + ')),32),'
        '"— no blocks scheduled —"))'))
    # same trap one section down: TAKE(...,32) against exactly 32 free rows,
    # and chapter 22 already sits at 23 blocks.
    r += 32
    _nb = 'COUNTIF(nrSCH_ChNum,' + C + ')'
    ws.cell(row=r, column=2, value=(
        '=IF(' + C + '="","",IF(' + _nb + '>32,"+ "&(' + _nb + '-32)&'
        '" more block(s) not shown — see the Schedule sheet",""))'
    )).font = F_SMALL
    r += 1
    ws.cell(row=r, column=2, value="Training Coordinator:").font = F_LABEL
    ws.cell(row=r, column=4, value="_______________________").font = F_BODY
    ws.cell(row=r, column=6, value="Date:").font = F_LABEL
    ws.cell(row=r, column=7, value="____________").font = F_BODY
    col_widths(ws, {"A": 3, "B": 24, "C": 16, "D": 16, "E": 14, "F": 22,
                    "G": 16, "H": 16, "I": 22})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:I{r}")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    sheet_note(ws, "Pick a chapter — the whole training file's status on one "
                   "printable page. Pair with the ExamSheet grade sheet and "
                   "the EvalSheet critique for the paper folder.")
    return ws


# --------------------------------------------------------------------------
def build_examsheet(wb):
    """IRG-required grade sheet per assessment: pick an exam, print the
    class's scores with pass/fail, retests, stats and a proctor signature."""
    ws = wb.create_sheet("ExamSheet")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Exam seq #:").font = F_LABEL
    sel = ws.cell(row=r, column=3, value=1)
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    define(wb, "cfgGradeSheetExam", "ExamSheet", f"$C${r}")
    E = "cfgGradeSheetExam"
    # this exam's OWN passing score (ExamPlan lets an exam override the
    # global 70). The printed grade sheet used to test every score against
    # cfgPassingScore, so an overriding exam printed a Pass?/FAIL flag and a
    # fail count that contradicted the authoritative ExamScores "Pass?"
    # column and the red cells on ScoresGrid.
    passing = ('IFERROR(INDEX(rngEPpass,MATCH(' + E +
               ',IFERROR(rngEPseq+0,-1),0)),cfgPassingScore)')
    r += 2
    ws.cell(row=r, column=2, value=(
        '="TYLER POLICE ACADEMY — "&cfgAcademyClass&" — GRADE SHEET — "'
        '&IFERROR(INDEX(rngEPname,MATCH(' + E + ',IFERROR(rngEPseq+0,-1),0)),'
        '"(set exam #)")')).font = F_KPI
    r += 1
    ws.cell(row=r, column=2, value=(
        # MAXIFS returns 0 (not an error) when nothing matches, so the old
        # IFERROR fallback never fired and an unscheduled exam printed
        # "Administered: 01/00/1900". The schedule block is also named
        # "Final Test", not "Test 20", so the final exam could never match
        # the literal key — fall back to the date the exam was actually
        # taken (first attempt on record) before giving up.
        '="Exam code: "&IFERROR(INDEX(rngEPcode,MATCH(' + E +
        ',IFERROR(rngEPseq+0,-1),0)),"?")&"   Administered: "&'
        'LET(s,MAXIFS(nrSCH_Date,nrSCH_Act,"Test "&' + E + '),'
        'a,IFERROR(MINIFS(nrES_Date,nrES_Code,INDEX(rngEPcode,MATCH(' + E +
        ',IFERROR(rngEPseq+0,-1),0)),nrES_Att,1,nrES_Date,">0"),0),'
        'IF(s>0,TEXT(s,"mm/dd/yyyy"),IF(a>0,TEXT(a,"mm/dd/yyyy"),'
        '"(not scheduled)")))'
        '&"   Passing score: "&' + passing + '&'
        '"   (passed retests record at "&cfgRetakeRecordedCap&")"'
    )).font = F_SMALL
    r += 2
    hdr = r
    header_row(ws, ["#", "Cadet", "PID", "Agency", "Raw Score",
                    "Recorded", "Pass?", "Retest", "Absence"], row=r)
    r += 1
    first = r
    code = ('IFERROR(INDEX(rngEPcode,MATCH(' + E +
            ',IFERROR(rngEPseq+0,-1),0)),"")')
    for i in range(CADETS):
        rr = first + i
        src = FIRST + i
        ws.cell(row=rr, column=2, value=(
            f'=IF(Cadets!$B{src}="","",ROW()-{first-1})')).font = F_CALC
        ws.cell(row=rr, column=3, value=(
            f'=IF(Cadets!$B{src}="","",Cadets!$F{src})')).font = F_BODY
        ws.cell(row=rr, column=4, value=(
            f'=IF($C{rr}="","",Cadets!$B{src})')).font = F_CALC
        ws.cell(row=rr, column=5, value=(
            f'=IF($C{rr}="","",Cadets!$H{src})')).font = F_CALC
        # both score columns must ask whether a score was RECORDED, not
        # merely whether a row exists. An excused first attempt is logged
        # with a blank Raw Score (see the ExamScores sheet note), and the
        # bare existence test used to print Raw 0 / Recorded 0 / FAIL for a
        # cadet who was lawfully excused - on the grade sheet that is filed
        # for the IRG. nrES_Rec,"<>" counted the Recorded FORMULA cell,
        # which returns "" and is therefore not an empty CELL; every other
        # exam consumer in the workbook uses the numeric nrES_Rec,">=0".
        ws.cell(row=rr, column=6, value=(
            f'=IF($C{rr}="","",LET(c,{code},IF(COUNTIFS(nrES_PID,'
            f'Cadets!$B{src},nrES_Code,c,nrES_Att,1,nrES_Raw,">=0")=0,"",'
            f'SUMIFS(nrES_Raw,nrES_PID,Cadets!$B{src},nrES_Code,c,'
            f'nrES_Att,1))))')).font = F_CALC
        ws.cell(row=rr, column=7, value=(
            f'=IF($C{rr}="","",LET(c,{code},IF(COUNTIFS(nrES_PID,'
            f'Cadets!$B{src},nrES_Code,c,nrES_Final,"Yes",nrES_Rec,">=0")=0,'
            f'"",SUMIFS(nrES_Rec,nrES_PID,Cadets!$B{src},nrES_Code,c,'
            f'nrES_Final,"Yes"))))')).font = F_CALC
        ws.cell(row=rr, column=8, value=(
            f'=IF(OR($C{rr}="",$G{rr}=""),"",'
            f'IF($G{rr}>={passing},"Pass","FAIL"))')).font = F_CALC
        # only a SCORED attempt-2 row is a retest: a row logged when the
        # retest is scheduled printed "Retested" on the TCOLE grade sheet
        # next to the untouched first-attempt score.
        ws.cell(row=rr, column=9, value=(
            f'=IF($C{rr}="","",LET(c,{code},IF(COUNTIFS(nrES_PID,'
            f'Cadets!$B{src},nrES_Code,c,nrES_Att,2,nrES_Raw,">=0")>0,'
            f'"Retested","")))'
        )).font = F_CALC
        # without this column a blank Raw/Recorded pair is indistinguishable
        # from "never sat the exam" on the filed grade sheet.
        ws.cell(row=rr, column=10, value=(
            f'=IF($C{rr}="","",LET(c,{code},'
            f'n,COUNTIFS(nrES_PID,Cadets!$B{src},nrES_Code,c,nrES_Att,1),'
            f's,COUNTIFS(nrES_PID,Cadets!$B{src},nrES_Code,c,nrES_Att,1,'
            f'nrES_Raw,">=0"),'
            f'x,COUNTIFS(nrES_PID,Cadets!$B{src},nrES_Code,c,nrES_Att,1,'
            f'nrES_Absence,"Excused"),'
            f'u,COUNTIFS(nrES_PID,Cadets!$B{src},nrES_Code,c,nrES_Att,1,'
            f'nrES_Absence,"Unexcused"),'
            f'IF(u>0,"Unexcused",IF(x>0,IF(s>0,"Excused - taken later",'
            f'"Excused - pending"),IF(OR(n=0,s>0),"","No score recorded")))))'
        )).font = F_CALC
        for ccol in range(2, 11):
            ws.cell(row=rr, column=ccol).border = BOX
    last = first + CADETS - 1
    cf_formula(ws, f"H{first}:H{last}", f'$H{first}="FAIL"', FILL_WARNBG)
    cf_formula(ws, f"J{first}:J{last}",
               f'OR($J{first}="Excused - pending",$J{first}='
               f'"No score recorded")', FILL_AMBER)
    cf_formula(ws, f"J{first}:J{last}", f'$J{first}="Unexcused"', FILL_WARNBG)
    r = last + 1
    ws.cell(row=r, column=3, value="Class average / low / high / fails:"
            ).font = F_LABEL
    ws.cell(row=r, column=6, value=(
        '=LET(c,' + code + ',IF(c="","",'
        'IFERROR(ROUND(AVERAGEIFS(nrES_Rec,nrES_Code,c,nrES_Final,"Yes",'
        'nrES_Rec,">=0",nrES_Rec,"<=100"),1),"—")'
        '&" / "&IFERROR(MINIFS(nrES_Rec,nrES_Code,c,nrES_Final,"Yes",'
        'nrES_Rec,">=0",nrES_Rec,"<=100"),"—")'
        '&" / "&IFERROR(MAXIFS(nrES_Rec,nrES_Code,c,nrES_Final,"Yes",'
        'nrES_Rec,">=0",nrES_Rec,"<=100"),"—")'
        '&" / "&COUNTIFS(nrES_Code,c,nrES_Final,"Yes",nrES_Rec,">=0",'
        'nrES_Rec,"<"&' + passing + ')))')).font = F_CALC
    r += 2
    ws.cell(row=r, column=2, value="Proctor / Instructor: ____________________"
            "____     Training Coordinator: ________________________"
            ).font = F_BODY
    r += 1
    ws.cell(row=r, column=2, value=DL.ACADEMY_ADDRESS).font = F_SMALL
    col_widths(ws, {"A": 3, "B": 5, "C": 28, "D": 10, "E": 16, "F": 11,
                    "G": 11, "H": 9, "I": 10, "J": 20})
    # print titles must start at the FIRST row of the print area (row 7)
    page_setup_portrait(ws, print_area=f"B{HDR_ROW+2}:J{r}",
                        repeat_rows=f"{HDR_ROW+2}:{hdr}")
    # the picker is the only cell a user can type in on a protected sheet,
    # so it carries a validation like every other one
    dv_list(ws, "=rngEPseq", [f"C{HDR_ROW}"])
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    sheet_note(ws, "The IRG requires a grade sheet for each assessment in "
                   "the training file — print one per exam and file it with "
                   "the chapter packet.")
    return ws


# --------------------------------------------------------------------------
def build_signin(wb):
    ws = wb.create_sheet("SignIn")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Sign-in date:").font = F_LABEL
    d = ws.cell(row=r, column=3)
    d.fill = FILL_INPUT
    d.font = F_INPUT
    d.border = BOX
    d.number_format = DATE
    define(wb, "cfgSignInDate", "SignIn", f"$C${r}")
    ws.cell(row=r, column=5, value=(
        '=IFERROR("Training Day #"&XLOOKUP(cfgSignInDate,nrCDdate,nrCDnum)'
        '&"  (Week "&XLOOKUP(cfgSignInDate,nrCDdate,nrCDweek)&")","")'
    )).font = F_LABEL
    r += 2
    ws.cell(row=r, column=2, value=(
        '="TYLER POLICE ACADEMY — "&cfgAcademyClass&" — DAILY REPORT & '
        'ATTENDANCE ROSTER — "&IF(cfgSignInDate="","(date: ____________)",'
        'TEXT(cfgSignInDate,"dddd, mmmm d, yyyy"))'
    )).font = F_KPI
    r += 2
    section_bar(ws, r, 2, 9, "Instruction scheduled this date")
    r += 1
    # ONE spilled column, not a five-column HSTACK. The strip sits above the
    # roster grid and therefore inherits the ROSTER's column widths (B=6 for
    # '#', D=10 for PID, E=18 for Agency), which clipped an 8-character start
    # time, every activity name over 10 characters (210 of the 313 blocks in
    # the seeded academy) and every instructor list over 18. Nothing can be
    # widened without wrecking the roster below, and no cell inside a
    # five-column spill can overflow because its neighbour is also part of
    # the spill. A single column overflows freely across the empty C..H of
    # these rows - the full print width - so the block reads on paper.
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(FILTER(TEXT(nrSCH_Start,"h:mm AM/PM")&" – "&'
        'TEXT(nrSCH_End,"h:mm AM/PM")&"   "&nrSCH_Act&'
        'IF(nrSCH_Instr="",""," ("&nrSCH_Instr&")")&'
        'IF(nrSCH_Loc="",""," — "&nrSCH_Loc),'
        '(nrSCH_Date<>"")*(nrSCH_Date=cfgSignInDate)),9),'
        '"— no schedule entered for this date —")'))
    r += 9
    section_bar(ws, r, 2, 9, "AM roll call")
    r += 1
    ws.cell(row=r, column=2, value=(
        "Present: ______ of ______     Absent/Late (name & reason): "
        "____________________________________________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    ws.cell(row=r, column=2, value=(
        "PT: [ ] Full participation   [ ] Modified: ______________   "
        "[ ] Missed: ______________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    section_bar(ws, r, 2, 9, "Cadet sign-in")
    r += 1
    header_row(ws, ["#", "Cadet", "PID", "Agency", "AM Signature",
                    "PM Signature", "Remarks"], row=r)
    r += 1
    first = r
    for i in range(CADETS):
        rr = first + i
        src = FIRST + i
        # gate the row number on the NAME (C), not merely on enrolment:
        # a separated cadet used to get a numbered but BLANK signature line
        # on the printed roster. COUNTIF keeps the numbering contiguous.
        ws.cell(row=rr, column=2, value=(
            f'=IF(C{rr}="","",COUNTIF($C${first}:C{rr},"?*"))')).font = F_CALC
        ws.cell(row=rr, column=3, value=(
            f'=IF(Cadets!$B{src}="","",IF(Cadets!$I{src}<>"Active","",'
            f'Cadets!$F{src}))')).font = F_BODY
        ws.cell(row=rr, column=4, value=(
            f'=IF(C{rr}="","",Cadets!$B{src})')).font = F_BODY
        ws.cell(row=rr, column=5, value=(
            f'=IF(C{rr}="","",Cadets!$H{src})')).font = F_BODY
        for c in range(2, 9):
            ws.cell(row=rr, column=c).border = BOX
        ws.row_dimensions[rr].height = 20
    last = first + CADETS - 1
    r = last + 1
    section_bar(ws, r, 2, 9, "PM report — changes through the day")
    r += 1
    ws.cell(row=r, column=2, value=(
        "Early departures (name / time out / reason — e.g. agency recall): "
        "______________________________________________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    ws.cell(row=r, column=2, value=(
        "____________________________________________________________"
        "____________________________________________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    ws.cell(row=r, column=2, value=(
        "Incidents / injuries: _______________________________________"
        "____________________________________________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    ws.cell(row=r, column=2, value=(
        "Memos turned in: ____________________________   Remarks: "
        "_____________________________________________")).font = F_BODY
    ws.row_dimensions[r].height = 20
    r += 1
    ws.cell(row=r, column=2, value=(
        "Class Leader signature: ______________________________     "
        "Coordinator: ______________________________")).font = F_BODY
    ws.row_dimensions[r].height = 22
    r += 1
    ws.cell(row=r, column=2, value="Instructor verification: ____________________"
            "____________     Time: ____________").font = F_BODY
    r += 1
    ws.cell(row=r, column=2, value=DL.ACADEMY_ADDRESS).font = F_SMALL
    col_widths(ws, {"A": 3, "B": 6, "C": 28, "D": 10, "E": 18, "F": 21,
                    "G": 21, "H": 22, "I": 12})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW}:H{r}")
    # the date picker is the only cell a user can type in on this protected
    # sheet, so it carries a validation like every other printable picker
    _dvd = DataValidation(type="date", operator="between",
                          formula1="cfgStartDate", formula2="cfgEndDate",
                          allow_blank=True, showErrorMessage=True)
    _dvd.errorTitle = "Not an academy date"
    _dvd.error = ("Pick a date between the academy's Start Date and End "
                  "Date (Settings).")
    _dvd.add(f"C{HDR_ROW}")
    ws.add_data_validation(_dvd)
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    sheet_note(ws, "The one-page daily report & roster: AM roll call, "
                   "sign-in, PM changes, signatures. Print on demand; the "
                   "signed sheet is scanned into that day's folder and the "
                   "scan is the legal original (confirmed by TPD records "
                   "management - see the Audit sheet item 'Scanned documents "
                   "established as legal originals'). The day is logged as "
                   "one DailyLog row. Exceptions still go on Attendance.")
    return ws


# --------------------------------------------------------------------------
def build_evalsheet(wb):
    ws = wb.create_sheet("EvalSheet")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Chapter:").font = F_LABEL
    sel = ws.cell(row=r, column=3)
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    # see ChapterPacket: nrCHnum is TEXT, so the picker must be text-formatted
    sel.number_format = "@"
    dv_list(ws, "=nrCHnum", [f"C{r}"])
    define(wb, "cfgEvalChapter", "EvalSheet", f"$C${r}")
    ws.cell(row=r, column=5, value='="(prints for: "&IFERROR(XLOOKUP('
            'cfgEvalChapter,nrCHnum,nrCHname),"pick a chapter")&")"'
            ).font = F_SMALL
    r += 2
    ws.cell(row=r, column=2, value="TYLER POLICE ACADEMY").font = F_KPI
    r += 1
    ws.cell(row=r, column=2, value=DL.ACADEMY_ADDRESS).font = F_SMALL
    r += 2
    ws.cell(row=r, column=2, value="CLASS EVALUATION AND ASSESSMENT").font = F_LABEL
    r += 1
    ws.cell(row=r, column=2, value=DL.EVAL_SCALE).font = F_SMALL
    r += 2
    _profile_label(ws, r, 2, "Course",
                   'cfgAcademyClass&" — Ch. "&cfgEvalChapter&" "&'
                   'IFERROR(XLOOKUP(cfgEvalChapter,nrCHnum,nrCHname),"")', wide=5)
    r += 1
    _profile_label(ws, r, 2, "Instructor",
                   'IFERROR(XLOOKUP(cfgEvalChapter,nrCHnum,nrCHinstr),"")', wide=3)
    _profile_label(ws, r, 6, "Date(s)",
                   'IFERROR(TEXT(XLOOKUP(cfgEvalChapter,nrCHnum,nrCHfirst),'
                   '"mm/dd")&" – "&TEXT(MAXIFS(nrSCH_Date,nrSCH_ChNum,'
                   'cfgEvalChapter),"mm/dd/yyyy"),"")', wide=2)
    r += 2
    header_row(ws, ["#", "Question", "1", "2", "3", "4", "5"], row=r)
    r += 1
    for i, q in enumerate(DL.EVAL_QUESTIONS):
        ws.cell(row=r, column=2, value=i + 1).font = F_BODY
        qc = ws.cell(row=r, column=3, value=q)
        qc.font = F_BODY
        qc.alignment = A_LEFT_WRAP
        for c in range(4, 9):
            ws.cell(row=r, column=c).border = BOX
        ws.cell(row=r, column=2).border = BOX
        qc.border = BOX
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="ADDITIONAL COMMENTS").font = F_LABEL
    r += 1
    for i in range(6):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2).border = UNDER
        ws.row_dimensions[r].height = 20
        r += 1
    c = ws.cell(row=r, column=2, value=DL.EVAL_FOOTER)
    c.font = F_SMALL
    c.alignment = A_LEFT_WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    col_widths(ws, {"A": 3, "B": 5, "C": 62, "D": 6, "E": 6, "F": 6,
                    "G": 6, "H": 6})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW+2}:H{r}")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    sheet_note(ws, "Modernized replica of the existing critique form. Pick a "
                   "chapter — course, instructor and dates fill from the "
                   "schedule. Print one per cadet (Print Center does stacks).")
    return ws


# --------------------------------------------------------------------------
def build_spellingprint(wb):
    ws = wb.create_sheet("SpellingPrint")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Test #:").font = F_LABEL
    s1 = ws.cell(row=r, column=3, value=1)
    s1.fill = FILL_INPUT
    s1.font = F_INPUT
    s1.border = BOX
    # the picker is the only cell a user can type in on this protected sheet,
    # and nrSpellWords is only 12 columns wide: an out-of-range test number
    # printed 25 #REF! cells (or, blank/0, a scrambled cross-test key) under a
    # heading that read "Spelling Test #13 — KEY". Validate the pick against
    # the same 1..12 list the Spelling sheet publishes...
    dv_list(ws, "=nrSpellTestNums", [f"C{r}"])
    define(wb, "cfgSpellPrintNum", "SpellingPrint", f"$C${r}")
    ws.cell(row=r, column=4, value="Mode:").font = F_LABEL
    s2 = ws.cell(row=r, column=5, value="Test")
    s2.fill = FILL_INPUT
    s2.font = F_INPUT
    s2.border = BOX
    dvm = dv_list(ws, '"Test,Key"', [f"E{r}"])
    define(wb, "cfgSpellPrintMode", "SpellingPrint", f"$E${r}")
    r += 2
    ws.cell(row=r, column=2, value="TYLER POLICE DEPARTMENT ACADEMY").font = F_KPI
    r += 1
    # ...and guard the printed page itself, because a PASTE bypasses the
    # validation (see dv_whole's note in xlb.py). N() makes the text "13" and
    # a blank picker both read 0, so every bad state prints one instruction
    # instead of a page of #REF!.
    BAD = 'OR(N(cfgSpellPrintNum)<1,N(cfgSpellPrintNum)>12)'
    ws.cell(row=r, column=2, value=(
        f'=IF({BAD},"Pick a spelling test number 1-12 in C5",'
        '"Spelling Test #"&cfgSpellPrintNum&IF(cfgSpellPrintMode="Key",'
        '" — KEY",""))')).font = F_LABEL
    r += 2
    ws.cell(row=r, column=2, value="Name: ______________________    "
            "Date: ____________    Score: ________").font = F_BODY
    r += 2
    first = r
    for i in range(13):
        rr = first + i
        ws.cell(row=rr, column=2, value=f"{i+1}.").font = F_BODY
        w1 = ws.cell(row=rr, column=3, value=(
            f'=IF({BAD},"",IF(cfgSpellPrintMode="Key",'
            f'INDEX(nrSpellWords,{i+1},'
            f'cfgSpellPrintNum),"_______________________"))'))
        w1.font = F_BODY
        if i + 14 <= 25:
            ws.cell(row=rr, column=5, value=f"{i+14}.").font = F_BODY
            w2 = ws.cell(row=rr, column=6, value=(
                f'=IF({BAD},"",IF(cfgSpellPrintMode="Key",'
                f'INDEX(nrSpellWords,{i+14},'
                f'cfgSpellPrintNum),"_______________________"))'))
            w2.font = F_BODY
        ws.row_dimensions[rr].height = 24
    r = first + 14
    ws.cell(row=r, column=2, value="Each word is worth 4 points.").font = F_SMALL
    col_widths(ws, {"A": 3, "B": 5, "C": 30, "D": 4, "E": 5, "F": 30})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW+2}:F{r}")
    protect(ws)
    # two real pickers here (C5 = test #, E5 = Test/Key mode); D5 holds the
    # "Mode:" label and stays locked so it cannot be erased
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    unlock_range(ws, f"E{HDR_ROW}:E{HDR_ROW}")
    sheet_note(ws, "Pick test # and Test/Key mode; print. Words come from "
                   "SpellingMaster.")
    return ws


# --------------------------------------------------------------------------
def build_writinghandout(wb):
    ws = wb.create_sheet("WritingHandout")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Week #:").font = F_LABEL
    s1 = ws.cell(row=r, column=3, value=1)
    s1.fill = FILL_INPUT
    s1.font = F_INPUT
    s1.border = BOX
    define(wb, "cfgHandoutWeek", "WritingHandout", f"$C${r}")
    r += 2
    ws.cell(row=r, column=2, value=(
        '="TYLER POLICE ACADEMY — "&cfgAcademyClass&" — WRITING ASSIGNMENTS '
        '— WEEK "&cfgHandoutWeek')).font = F_KPI
    r += 1
    c = ws.cell(row=r, column=2, value=DW.REQUIREMENTS_NOTE)
    c.font = F_SMALL
    c.alignment = A_LEFT_WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    r += 2
    header_row(ws, ["#", "Title", "Assigned", None, "Prompt"], row=r)
    # A35: cfgWritingDueTime was exposed on Settings as "Due time shown on
    # handouts" and referenced by nothing at all - the header hard-coded
    # 1700. It is now the live source of this column heading.
    dh = ws.cell(row=r, column=5, value=(
        '="Due ("&IFERROR(TEXT(TIMEVALUE(LEFT(TEXT(cfgWritingDueTime,"0000"),2)'
        '&":"&RIGHT(TEXT(cfgWritingDueTime,"0000"),2)),"h:mm AM/PM"),'
        'cfgWritingDueTime)&")"'))
    dh.fill = FILL_HDR
    dh.font = F_HDR
    dh.alignment = A_CENTER
    dh.border = BOX
    r += 1
    # the week is resolved with match_mode 1 (next larger), NOT an exact
    # match. An assignment dated on a CLOSURE day - BPOC 7 ships two dated
    # Labor Day - found no calendar row, fell to 0 and appeared on no
    # weekly handout at all, while the week it belonged to printed the
    # affirmatively false "no assignments start this week". The same
    # closure-day bracketing idiom already guards retest deadlines and memo
    # due dates. ISNUMBER keeps blank/text rows out; a date past the last
    # class day lands on the last week instead of disappearing.
    # TAKE(...,8): the only printable digest left without a cap. The sheet
    # styles exactly 8 rows and the print area stops at row 19, so a longer
    # week silently lost its wrap formatting and everything past row 19
    # never printed at all. Capped at what the page can actually hold.
    ws.cell(row=r, column=2, value=(
        '=IFERROR(TAKE(FILTER(HSTACK(rngWMnum,rngWMtitle,TEXT(rngWMassigned,'
        '"mm/dd"),TEXT(rngWMdue,"mm/dd"),rngWMprompt),'
        'IF(ISNUMBER(rngWMassigned),IFERROR(XLOOKUP(rngWMassigned,nrCDdate,'
        'nrCDweek,MAX(nrCDweek),1),0),0)=cfgHandoutWeek),'
        '8),"— no assignments start this week —")'))
    for rr in range(r, r + 8):
        ws.row_dimensions[rr].height = 60
        for ccol in range(2, 9):
            ws.cell(row=rr, column=ccol).alignment = A_LEFT_WRAP
    col_widths(ws, {"A": 3, "B": 5, "C": 32, "D": 10, "E": 10, "F": 80})
    page_setup_portrait(ws, print_area=f"B{HDR_ROW+2}:F{r+8}")
    # an unlocked picker on a protected sheet is the only cell a user can
    # type in, so it carries a validation like every other one.
    dv_whole(ws, [f"C{HDR_ROW}"], 1, 52,
             "Pick an academy week number (1 = the first week of class).")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    sheet_note(ws, "Pick an academy week; assignments whose computed assigned "
                   "date falls in that week appear with their prompts (up to "
                   "8 — the page holds 8; a week with more is split across "
                   "two handouts by moving the extras' Override Assigned "
                   "dates on WritingMaster). The Due column header shows the "
                   "Writing Due Time from Settings.")
    return ws


# --------------------------------------------------------------------------
def build_emailpreview(wb):
    ws = wb.create_sheet("EmailPreview")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Agency:").font = F_LABEL
    sel = ws.cell(row=r, column=3, value="TPD")
    sel.fill = FILL_INPUT
    sel.font = F_INPUT
    sel.border = BOX
    dv_list(ws, "=rngAgencyIDs", [f"C{r}"])
    define(wb, "cfgPreviewAgency", "EmailPreview", f"$C${r}")
    ws.cell(row=r, column=5, value=(
        '="Exam #"&cfgCurrentExamNum&" — "&IFERROR(INDEX(rngEPname,'
        'MATCH(cfgCurrentExamNum,IFERROR(rngEPseq+0,-1),0)),"?")&'
        # MAXIFS returns 0, not an error, so the IFERROR "never" fallback
        # was unreachable and an agency with no EmailLog row read 12/30/1899
        '" | Spelling through #"&cfgCurrentSpellingNum&" | Last emailed: "&'
        'IF(IFERROR(MAXIFS(nrELdate,nrELagency,cfgPreviewAgency),0)=0,"never",'
        'TEXT(MAXIFS(nrELdate,nrELagency,cfgPreviewAgency),"mm/dd/yyyy"))'
        )).font = F_SMALL
    r += 2
    section_bar(ws, r, 2, 12, "Cadet results (as the email will report them)")
    r += 1
    header_row(ws, ["Cadet", "Score", "Class Avg", "Retake?",
                    "Spelling (test #)", "Spelling Flag", "Attendance",
                    "Writing Current", "Open Flags"], row=r)
    r += 1
    grid_first = r
    for i in range(CADETS):
        rr = grid_first + i
        src = FIRST + i
        pred = (f'IF(OR(Cadets!$B{src}="",Cadets!$I{src}<>"Active"),FALSE,'
                f'OR(cfgPreviewAgency=cfgHomeAgency,'
                f'Cadets!$G{src}=cfgPreviewAgency))')
        ws.cell(row=rr, column=2, value=(
            f'=IF({pred},Cadets!$F{src},"")')).font = F_CALC
        ws.cell(row=rr, column=3, value=(
            f'=IF($B{rr}="","",IF(COUNTIFS(nrES_PID,Cadets!$B{src},'
            f'nrES_Seq,cfgCurrentExamNum,nrES_Final,"Yes",nrES_Rec,">=0")=0,'
            f'"",SUMIFS(nrES_Rec,nrES_PID,Cadets!$B{src},'
            f'nrES_Seq,cfgCurrentExamNum,nrES_Final,"Yes")))')).font = F_CALC
        ws.cell(row=rr, column=4, value=(
            f'=IF($B{rr}="","",IFERROR(ROUND(AVERAGEIFS(nrES_Rec,'
            f'nrES_Seq,cfgCurrentExamNum,nrES_Final,"Yes",nrES_Rec,">=0",'
            f'nrES_Rec,"<=100"),1),""))')).font = F_CALC
        # "(cap 70)" is a claim about the RECORDED score, so it may only
        # appear for a retest that was actually scored AND passed: an
        # attempt-2 row logged when the retest is merely SCHEDULED left the
        # attempt-1 score in place, and a FAILED retest records the raw
        # attempt-1 score, not the cap. Both used to be badged as capped.
        # modAgencyEmail.RetakeNote carries the identical three states.
        ws.cell(row=rr, column=5, value=(
            f'=IF($B{rr}="","",IF(COUNTIFS(nrES_PID,Cadets!$B{src},'
            f'nrES_Seq,cfgCurrentExamNum,nrES_Att,2,nrES_Raw,">=0",'
            f'nrES_AttPass,"Yes")>0,"RETEST (cap 70)",'
            f'IF(COUNTIFS(nrES_PID,Cadets!$B{src},nrES_Seq,cfgCurrentExamNum,'
            f'nrES_Att,2,nrES_Raw,">=0",nrES_AttPass,"No")>0,'
            f'"RETEST FAILED (first-attempt score shown)","")))'
        )).font = F_CALC
        # the draft sends the score for THIS spelling test (modAgencyEmail
        # spellCol = 3 + spellNum, i.e. Spelling column D+spellNum), with the
        # running average appended only when it is below the intervention
        # threshold. The preview used to show Spelling!$P - the running
        # average - so the coordinator reviewed a different number from the
        # one the agency received on every single row. Same three omit tests
        # as modAgencyEmail's omitSpelling, including spellNum > 12, which
        # the preview had no counterpart for.
        omit = ('OR(cfgCurrentSpellingNum<cfgCurrentExamNum,'
                'cfgCurrentSpellingNum<=0,cfgCurrentSpellingNum>12)')
        ws.cell(row=rr, column=6, value=(
            f'=IF($B{rr}="","",IF({omit},"(omitted)",'
            f'LET(sc,IF(INDEX(Spelling!$D{src}:$O{src},cfgCurrentSpellingNum)'
            f'="","",INDEX(Spelling!$D{src}:$O{src},cfgCurrentSpellingNum)),'
            f'av,Spelling!$P{src},'
            f'IF(AND(ISNUMBER(av),av<cfgSpellInterventionAvg),'
            f'sc&" (avg "&av&")",sc))))')).font = F_CALC
        # gated on the same omit test: the preview must never carry an
        # intervention note for a column the draft drops entirely
        ws.cell(row=rr, column=7, value=(
            f'=IF($B{rr}="","",IF({omit},"",'
            f'IF(Spelling!$R{src}="INTERVENTION",'
            f'"BELOW "&cfgSpellInterventionAvg&" — intervention","")))'
        )).font = F_CALC
        ws.cell(row=rr, column=8, value=(
            f'=IF($B{rr}="","",sysAttendance!$J{src}&" / PT "&'
            f'sysAttendance!$R{src})')).font = F_CALC
        ws.cell(row=rr, column=9, value=(
            f'=IF($B{rr}="","",Writing!$AT{src}&" ("&Writing!$AS{src}&'
            f'" overdue)")')).font = F_CALC
        ws.cell(row=rr, column=10, value=(
            # sysFlags Reasons is column T (the flag block runs E..R and
            # Flag Count is S) — see the header note in build_sysflags
            f'=IF($B{rr}="","",sysFlags!$T{src})')).font = F_CALC
    grid_last = grid_first + CADETS - 1
    define(wb, "nrEPVgrid", "EmailPreview",
           f"$B${grid_first}:$J${grid_last}")
    r = grid_last + 2
    section_bar(ws, r, 2, 12,
                "Marked-for-reporting items since last email to this agency "
                "(incidents / counseling / memos you flagged Yes)")
    r += 1
    define(wb, "nrEPVsinceRow", "EmailPreview", f"$B${r+1}")
    header_row(ws, ["Date", "Cadet", "Type", "Severity/Kind", "Description"],
               row=r)
    r += 1
    since_first = r
    # INT() + >= below: EmailLog!B is stamped with Now (date AND time) while
    # every log dates its rows date-only, so a plain > against the raw
    # timestamp dropped anything dated on a previous run's calendar day —
    # permanently, from this preview and from the draft. The VBA sender
    # (modAgencyEmail.LastEmailDate) applies the identical Int()/>= rule;
    # these two must stay in lockstep or the preview stops matching the draft.
    since = ('LET(cutoff,INT(IFERROR(MAXIFS(nrELdate,nrELagency,'
             'cfgPreviewAgency),0)),')
    # only rows YOU marked for agency reporting are included (Incidents
    # "Report to Agency?", Counseling "Agency Notified?", Memos "Report to
    # Agency?") — everything else stays an academy teaching moment
    ws.cell(row=r, column=2, value=(
        '=IFERROR(' + since +
        'inc,FILTER(HSTACK(nrIN_Date,IFERROR(XLOOKUP(nrIN_PID,rngCadetPIDs,'
        'rngCadetNames),""),IF(SEQUENCE(ROWS(nrIN_Date)),"Incident"),'
        'nrIN_Sev,nrIN_Desc),(nrIN_Date>=cutoff)*(nrIN_Report="Yes")*'
        'IFERROR((XLOOKUP(nrIN_PID,rngCadetPIDs,nrCadetAgencyID)='
        'cfgPreviewAgency)+(cfgPreviewAgency=cfgHomeAgency),0),'
        '{"","","","",""}),'
        'cns,FILTER(HSTACK(nrCO_Date,IFERROR(XLOOKUP(nrCO_PID,rngCadetPIDs,'
        'rngCadetNames),""),IF(SEQUENCE(ROWS(nrCO_Date)),"Counseling"),'
        'nrCO_Type,nrCO_Desc),(nrCO_Date>=cutoff)*(nrCO_Report="Yes")*'
        'IFERROR((XLOOKUP(nrCO_PID,rngCadetPIDs,nrCadetAgencyID)='
        'cfgPreviewAgency)+(cfgPreviewAgency=cfgHomeAgency),0),'
        '{"","","","",""}),'
        'mem,FILTER(HSTACK(nrME_Assigned,IFERROR(XLOOKUP(nrME_PID,'
        'rngCadetPIDs,rngCadetNames),""),IF(SEQUENCE(ROWS(nrME_Assigned)),'
        '"Memo"),nrME_Status,nrME_Subject),(nrME_Assigned>=cutoff)*'
        '(nrME_Report="Yes")*IFERROR((XLOOKUP(nrME_PID,rngCadetPIDs,'
        'nrCadetAgencyID)=cfgPreviewAgency)+(cfgPreviewAgency='
        'cfgHomeAgency),0),{"","","","",""}),'
        'all,VSTACK(inc,cns,mem),'
        'SORT(FILTER(all,INDEX(all,0,1)<>"",'
        '"— nothing marked for this agency since last email —"),1)),'
        '"— nothing marked for this agency since last email —")'))
    # this is the workbook's only multi-row spill that carries RAW date
    # values (it must: the SORT below is on that column, and sorting
    # mm/dd/yyyy text puts December before February). The spill zone shipped
    # as General, so every date rendered as a five-digit serial in the
    # preview the coordinator reads before approving the draft. Formatting
    # the landing column as a date fixes the display without breaking the
    # sort; the fallback string still prints as text.
    # ...and the zone must cover the WHOLE spill, which is uncapped by
    # design (the VBA draft emits every matching row, so a TAKE() here would
    # make the preview under-report the email it previews). 16 formatted rows
    # meant item 17 onward re-rendered as five-digit serials. Nothing else
    # occupies EmailPreview below this anchor.
    for rr in range(since_first, since_first + 300):
        ws.cell(row=rr, column=2).number_format = DATE
    r += 16
    col_widths(ws, {"A": 3, "B": 12, "C": 24, "D": 11, "E": 11, "F": 18,
                    "G": 22, "H": 20, "I": 18, "J": 60})
    sheet_note(ws, "Preview of what the Outlook draft will contain — the "
                   "same day-granular 'since last email' cutoff the sender "
                   "uses. The buttons on THIS sheet and on the Dashboard "
                   "build the drafts — for review, never auto-sent.")
    protect(ws)
    unlock_range(ws, f"C{HDR_ROW}:C{HDR_ROW}")
    return ws


# --------------------------------------------------------------------------
def build_emaillog(wb):
    ws = wb.create_sheet("EmailLog")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Date", "AgencyID", "Exam #", "Spelling #", "Cadets",
                    "Included Discipline Since", "Logged By", "Notes"])
    first, last = DATA_ROW, DATA_ROW + 499
    fill_rows(ws, first, last, {c: (None, "in") for c in "BCDEFGHI"})
    for r in range(first, last + 1):
        ws[f"B{r}"].number_format = DATE
    define(wb, "nrELdate", "EmailLog", f"$B${first}:$B${last}")
    define(wb, "nrELagency", "EmailLog", f"$C${first}:$C${last}")
    col_widths(ws, {"A": 3, "B": 12, "C": 10, "D": 8, "E": 10, "F": 9,
                    "G": 22, "H": 16, "I": 30})
    sheet_note(ws, "Appended by the email macro (one row per agency) only "
                   "after you confirm the draft was actually SENT — drafts "
                   "are never auto-sent, and a row here advances that "
                   "agency's since-last-report cutoff, so anything older "
                   "drops out of every later digest. 'Last Email Sent' on "
                   "Agencies and the digest cutoff read from here; delete a "
                   "row logged in error.")
    return ws


# --------------------------------------------------------------------------
def build_printcenter(wb):
    ws = wb.create_sheet("PrintCenter")
    ws.sheet_view.showGridLines = False
    r = HDR_ROW
    ws.cell(row=r, column=2, value="Print Center").font = F_KPI
    r += 2
    rows = [
        ("Daily report & roster (one day)", "Set the date on SignIn; blank "
         "date prints a generic form the leader dates by hand", "SignIn",
         "btnPrintSignIn"),
        ("Daily report & roster (week)", "Prints the next 5 class days, "
         "each pre-filled with its own date, day # and schedule", "SignIn",
         "btnPrintSignInWeek"),
        ("Academy book (all days)", "Every class day's form in one run — "
         "the class leader's bound book for the whole academy (print to "
         "PDF for the print shop)", "SignIn", "btnPrintSignInAcademy"),
        ("Course critique stack", "Set chapter on EvalSheet; prints one per "
         "active cadet", "EvalSheet", "btnPrintEvals"),
        ("Spelling test / key", "Set test # and mode on SpellingPrint",
         "SpellingPrint", "btnPrintSpelling"),
        ("Writing handout (week)", "Set week # on WritingHandout",
         "WritingHandout", "btnPrintWriting"),
        ("Cadet profile", "Set cadet on CadetProfile", "CadetProfile",
         "btnPrintProfile"),
        ("Transcript", "Set cadet on Transcript (button can run all cadets)",
         "Transcript", "btnPrintTranscript"),
        ("Class ranking", "Prints as-is", "Ranking", "btnPrintRanking"),
        ("Graduation checklist", "Prints as-is", "GradChecklist",
         "btnPrintGradCheck"),
        ("Audit packet", "Program checks + enrollment docs", "Audit",
         "btnPrintAudit"),
        ("Chapter packet", "Set chapter on ChapterPacket — the full "
         "training-file page for the auditor", "ChapterPacket",
         "btnPrintChapterPacket"),
        ("Exam grade sheet", "Set exam # on ExamSheet — the IRG-required "
         "grade sheet per assessment", "ExamSheet", "btnPrintGradeSheet"),
        ("Addendum (excess hours)", "Per-class excess vs TCOLE minimum with "
         "reporting course #s (#101 / #2040 / #2046 / #2055)", "Addendum",
         "btnPrintAddendum"),
        ("Schedule", "Full schedule listing", "Schedule", "btnPrintSchedule"),
    ]
    header_row(ws, ["What", "How", "Sheet", "Macro (button)"], row=r)
    r += 1
    for what, how, sheet, macro in rows:
        ws.cell(row=r, column=2, value=what).font = F_LABEL
        ws.cell(row=r, column=3, value=how).font = F_BODY
        ws.cell(row=r, column=4, value=sheet).font = F_BODY
        ws.cell(row=r, column=5, value=macro).font = F_SMALL
        ws.row_dimensions[r].height = 20
        r += 1
    r += 1
    c = ws.cell(row=r, column=2, value=(
        "Buttons are added by the BPOC VBA install — run "
        "tools/Install-BPOC-VBA.ps1 (see docs/BPOC-Coordinator-Guide.md, "
        "'Rebuilding from source'). Each button prints its sheet's defined "
        "print area to the default printer; Ctrl+P on the sheet does the "
        "same thing manually."))
    c.font = F_SMALL
    c.alignment = A_LEFT_WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 30
    # F blank gutter; G+H carry the install script's button strip, one
    # button per table row (rows 8..22) — they need real width or the
    # captions are clipped to a few characters
    col_widths(ws, {"A": 3, "B": 26, "C": 52, "D": 16, "E": 20,
                    "F": 3, "G": 20, "H": 20})
    # the only green/OUTPUT tab that shipped unprotected. It has no input
    # cells at all, and both VBA modules already assume it is protected:
    # ClearButtons restores the protection state it found, and the installer
    # unprotects it (Install-BPOC-VBA.ps1) without ever re-protecting.
    protect(ws)
    return ws


# --------------------------------------------------------------------------
INPUT_GUIDE = [
    ("Getting set up", None),
    ("Settings", "Academy label, start/end dates, grading weights, caps, "
     "retest window, flag thresholds, PT points rubric (yellow block)."),
    ("Lists", "Dropdown choices used across the workbook — add items here."),
    ("Agencies", "Sending agencies: codes, contacts, email addresses."),
    ("Instructors", "Roster with TCOLE PID, certs, SME letters, bio-on-file "
     "status. 'On Schedule?' shows who actually teaches — every co-teacher "
     "needs documentation."),
    ("InstructorBanks", "Per-topic banks: the certified pool (kept between "
     "academies) and who you picked to teach THIS academy — those picks "
     "become the Schedule's instructor dropdown for that topic."),
    ("ChapterMaster", "Chapter hours + the per-chapter TCOLE training-file "
     "record (lesson plan, bio, sign-ins, assessment, grade sheet, eval, "
     "special requirements)."),
    ("Control", "Extra closure dates (holidays compute automatically)."),
    ("Schedule", "The class schedule: one row per time block — date, times, "
     "chapter/activity, instructor. Drives hours, due dates, sign-ins."),
    ("ExamPlan", "Which exams this academy uses, their sequence and linked "
     "chapters."),
    ("ExamMaster", "The exam library (edit names/types/defaults)."),
    ("SkillsMaster", "Attempt limits and scoring mode per skill category."),
    ("SpellingMaster", "The 12 spelling word lists (printable via Print "
     "Center)."),
    ("WritingMaster", "The 40 assignments: prompts, linked chapters, "
     "date overrides."),
    ("Daily operation", None),
    ("Cadets", "The roster: PID, names, agency, status, enroll/separation."),
    ("ExamScores", "One row per exam attempt: cadet, exam, attempt #, raw "
     "score, date. Retest deadlines compute."),
    ("Spelling", "Spelling test scores per cadet per test."),
    ("Attendance", "Exception log: missed/modified time in minutes (or PT "
     "sessions), reason, documentation, excused."),
    ("Makeup", "Makeup credit: minutes/sessions, documentation, holds."),
    ("Skills", "Skills attempts: category, result, score, firearms course "
     "of fire."),
    ("Writing", "Type X when an assignment is received (auto-capitalizes)."),
    ("Incidents", "Positive/negative incidents with severity and resolution."),
    ("Counseling", "Every intervention: tutoring, counseling, agency "
     "notification, performance plans. 'Agency Notified?' = Yes puts it in "
     "the next email digest."),
    ("Memos", "Deficiency memos: assign, link to the I/A/C record, due "
     "auto-computes, mark received. 'Report to Agency?' is your call."),
    ("DailyLog", "One row per training day — the digital daily report. "
     "Counters compute; mark when the leader's signed report is scanned."),
    ("PT", "Baseline and final raw values per event; final points once the "
     "rubric arrives."),
    ("Medical", "Injuries, restrictions, clearances and expirations."),
    ("Certifications", "Per-cadet TIM/SFST/TCIC/CPR/ALERRT/ICS completion "
     "dates and certificate copies collected."),
    ("StateExam", "TCOLE licensing exam attempts and results (3 max)."),
    ("DismissalLog", "Formal reviews: trigger, decision, approvals."),
    ("AdvisoryBoard", "Governance: this academy's alignment record "
     "(policy version, minutes reviewed, workbook aligned) + the running "
     "board-meeting list with minutes folder locations. The Startup Review "
     "button fills it."),
    ("End of academy / as needed", None),
    ("Audit", "Program-requirement checklist answers + per-cadet enrollment "
     "documents grid."),
    ("EmailPreview", "Pick an agency to preview its email; the button "
     "builds Outlook drafts."),
    ("sysAwards", "Award override cells — your pick always wins."),
    ("PrintCenter", "Every printable in one place (buttons after VBA "
     "install)."),
]


def build_inputguide(wb):
    ws = wb.create_sheet("InputGuide")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Go to…", "What you enter there"])
    r = DATA_ROW
    for name, desc in INPUT_GUIDE:
        if desc is None:
            section_bar(ws, r, 2, 8, name)
            r += 1
            continue
        c = ws.cell(row=r, column=2,
                    value=f'=HYPERLINK("#{name}!B6","{name}")')
        c.font = F_LABEL
        d = ws.cell(row=r, column=3, value=desc)
        d.font = F_BODY
        d.alignment = A_LEFT_WRAP
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1
    col_widths(ws, {"A": 3, "B": 18, "C": 90})
    sheet_note(ws, "Only pages that take input are listed — gray sys* tabs "
                   "are the locked calculation engine; green tabs are "
                   "outputs. White boxes with a blue border are yours to type "
                   "(what you type shows in blue); gray cells calculate.")
    # pure navigation: every cell is a hyperlink formula or a static label
    protect(ws)
    return ws


def add_home_links(wb):
    """'◄ Dashboard' hyperlink in B1 of every visible sheet."""
    from openpyxl.styles import Font as _Font
    link_font = _Font(name="Arial", size=9, bold=True, color="1F3B5C",
                      underline="single")
    for ws in wb.worksheets:
        if ws.title in ("Dashboard", "sysListsHelper"):
            continue
        c = ws["B1"]
        c.value = '=HYPERLINK("#Dashboard!B5","◄ Dashboard")'
        c.font = link_font
        # LOCKED. A HYPERLINK formula is still clickable on a protected
        # sheet (protect() leaves every cell selectable), so unlocking B1
        # bought nothing and cost a lot: it was the only writable cell on
        # each of the eight sys* engine sheets, which made one keystroke
        # enough to delete the only navigation off a locked sheet.
        from openpyxl.styles import Protection as _Prot
        c.protection = _Prot(locked=True)


def gray_separated_rows(wb):
    """Gray out (and strike) rows of non-active cadets on cadet grids."""
    from openpyxl.styles import Font as _Font, PatternFill as _Fill
    from openpyxl.formatting.rule import FormulaRule as _FR
    gray_font = _Font(color="9AA5B1", strike=True)
    gray_fill = _Fill("solid", fgColor="EDEFF2")
    targets = {
        "Writing": "B6:AT55", "Spelling": "B6:S55", "PT": "B6:AC55",
        # AD, not AC: ScoresGrid runs out to AD ("Rank"), so a separated
        # cadet's row was struck through for 28 columns and then printed in
        # normal black for the Rank cell.
        "Certifications": "B6:U55", "ScoresGrid": "B6:AD55",
        "GradChecklist": "B6:Q55", "StateExam": "B6:L55",
        "Cadets": "B6:M55",
    }
    for name, rng in targets.items():
        ws = wb[name]
        ws.conditional_formatting.add(rng, _FR(
            formula=['AND(Cadets!$B6<>"",Cadets!$I6<>"Active")'],
            font=gray_font, fill=gray_fill))


def build_namedranges(wb):
    ws = wb.create_sheet("NamedRanges")
    ws.sheet_view.showGridLines = False
    header_row(ws, ["Name", "Refers To"])
    r = DATA_ROW
    for name, dn in sorted(wb.defined_names.items()):
        ws.cell(row=r, column=2, value=name).font = F_BODY
        ws.cell(row=r, column=3, value=dn.attr_text).font = F_SMALL
        r += 1
    col_widths(ws, {"A": 3, "B": 30, "C": 46})
    sheet_note(ws, "Auto-generated registry of every named range at build "
                   "time. Do not rename ranges without updating formulas.")
    protect(ws)
    return ws


def build_all_outputs(wb):
    build_scoresgrid(wb)
    build_ranking(wb)
    build_watchlist(wb)
    build_cadetprofile(wb)
    build_transcript(wb)
    build_gradchecklist(wb)
    build_dismissallog(wb)
    build_audit(wb)
    build_addendum(wb)
    build_chapterpacket(wb)
    build_examsheet(wb)
    build_signin(wb)
    build_evalsheet(wb)
    build_spellingprint(wb)
    build_writinghandout(wb)
    build_emailpreview(wb)
    build_emaillog(wb)
    build_printcenter(wb)
    build_inputguide(wb)
    build_dashboard(wb)     # after ScoresGrid/Spelling exist (charts)
    gray_separated_rows(wb)
    # NamedRanges BEFORE add_home_links: created after, it was the only
    # visible sheet with no '◄ Dashboard' link — and it is protected with
    # nothing unlocked, so there was no way back off it at all.
    build_namedranges(wb)   # captures every name defined above
    add_home_links(wb)
