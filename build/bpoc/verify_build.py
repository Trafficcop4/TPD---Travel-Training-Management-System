"""Structural verification of the built BPOC workbook.

Run after build_bpoc.py:  python3 build/bpoc/verify_build.py
Exits non-zero on any failure. (Formula *values* need Excel; this verifies
structure: names, prefixes, seeds, protection, print areas, postprocessor.)
"""
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from openpyxl import load_workbook

import postprocess
import data_spelling as DS
import data_writing as DW
import data_chapters as DC

WB_PATH = os.path.join(os.path.dirname(os.path.dirname(HERE)),
                       "workbooks", "BPOC_Academy_Management_V6.xlsx")

failures = []


def check(cond, msg):
    if cond:
        print(f"  ok  {msg}")
    else:
        failures.append(msg)
        print(f"FAIL  {msg}")


def test_postprocess_units():
    f = postprocess.fix_formula('=XLOOKUP(A1,B:B,C:C)')
    check(f == '=_xlfn.XLOOKUP(A1,B:B,C:C)', "postprocess: XLOOKUP prefixed")
    f = postprocess.fix_formula('=FILTER(A:A,B:B=1)')
    check(f == '=_xlfn._xlws.FILTER(A:A,B:B=1)', "postprocess: FILTER namespaced")
    f = postprocess.fix_formula('=LET(x,1,x+1)')
    check(f == '=_xlfn.LET(_xlpm.x,1,_xlpm.x+1)', "postprocess: LET params")
    f = postprocess.fix_formula('=LET(s,MAX(A:A),IF(s=0,"",LET(p,s+1,p*2)))')
    check("_xlpm.s" in f and "_xlpm.p" in f and f.count("_xlfn.LET(") == 2,
          "postprocess: nested LET")
    f = postprocess.fix_formula('=IF(A1="FILTER(","x",LET(d,1,d))')
    check('"FILTER("' in f and "_xlpm.d,1,_xlpm.d" in f,
          "postprocess: strings untouched")
    # WORKDAY.INTL / NETWORKDAYS.INTL are native ECMA-376 built-ins: an
    # _xlfn. prefix on them opens as #NAME? in Excel and blanks the whole
    # class-day calendar (Control!I) the workbook is built on.
    f = postprocess.fix_formula('=WORKDAY.INTL(A1,5,"0000011",B:B)')
    check(f == '=WORKDAY.INTL(A1,5,"0000011",B:B)',
          "postprocess: WORKDAY.INTL left unprefixed (native function)")
    from openpyxl.utils import FORMULAE as _BUILTIN
    native = sorted(k for k in postprocess.XLFN if k in _BUILTIN)
    check(not native,
          f"postprocess: no XLFN key is an ECMA-376 built-in {native}")


def test_workbook():
    wb = load_workbook(WB_PATH)
    check(len(wb.sheetnames) == 60, f"60 sheets ({len(wb.sheetnames)} found)")

    # every referenced name is defined
    defined = set(wb.defined_names.keys())
    pat = re.compile(r"\b(nr|cfg|lst|rng)[A-Za-z_0-9]+")
    missing = set()
    unprefixed = []
    # WORKDAY.INTL is deliberately absent: it is a native built-in and must
    # stay unprefixed (see test_postprocess_units).
    fnpat = re.compile(
        r'(?<![A-Za-z0-9_.])(XLOOKUP|FILTER|SORTBY|SORT|TEXTJOIN|MAXIFS|'
        r'MINIFS|HSTACK|VSTACK|TAKE|SEQUENCE|LET)\(')
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    for m in pat.finditer(v):
                        if m.group(0) not in defined:
                            missing.add(m.group(0))
                    if v.count("(") != v.count(")"):
                        unprefixed.append(f"paren {ws.title}!{c.coordinate}")
                    for m in fnpat.finditer(v):
                        unprefixed.append(
                            f"unprefixed {m.group(1)} {ws.title}!{c.coordinate}")
    check(not missing, f"all formula name refs defined {sorted(missing)[:5]}")
    check(not unprefixed, f"all modern fns prefixed {unprefixed[:5]}")

    # seed data
    sm = wb["SpellingMaster"]
    check(sm["C6"].value == DS.TESTS[1][0] and sm["N30"].value == DS.TESTS[12][24],
          "SpellingMaster seeded (12 tests x 25 words)")
    wm = wb["WritingMaster"]
    check(wm["B6"].value == 1 and wm["B45"].value == 40 and
          isinstance(wm["M6"].value, str) and "Code of Ethics" in wm["C6"].value,
          "WritingMaster seeded (40 assignments with prompts)")
    cm = wb["ChapterMaster"]
    check(cm["D7"].value == "Professionalism and Ethics" and
          cm["E7"].value == 12, "ChapterMaster seeded with TCOLE minimums")
    check(cm["S28"].value is not None and "TIM" in cm["S28"].value,
          "ChapterMaster special TCOLE requirements seeded (ch 22)")

    # engine protection — EVERY sys* sheet, not a hand-maintained list.
    # sysAwards and sysListsHelper were both shipping unprotected because
    # they were simply missing from the literal below.
    for n in [s for s in wb.sheetnames if s.startswith("sys")] + ["ScoresGrid"]:
        check(wb[n].protection.sheet, f"{n} locked")
    # the awards overrides are the whole point of sysAwards, so they must
    # stay editable through that protection
    aw = wb["sysAwards"]
    check(not aw["E6"].protection.locked and not aw["G6"].protection.locked
          and aw["C6"].protection.locked and aw["F6"].protection.locked,
          "sysAwards: override/notes unlocked, computed + FINAL locked")
    check(wb["sysListsHelper"].sheet_state == "veryHidden",
          "sysListsHelper cannot be unhidden from the tab menu")
    for n in ("Cadets", "ExamScores", "Attendance", "Counseling", "PT"):
        check(not wb[n].protection.sheet, f"{n} open for entry")
    # protection must never disable SELECTION: the OOXML flags are inverted
    # (True = "may not select"), and setting both froze every picker cell and
    # home link on the protected sheets.
    frozen = [ws.title for ws in wb.worksheets if ws.protection.sheet and
              (ws.protection.selectLockedCells or
               ws.protection.selectUnlockedCells)]
    check(not frozen, f"protected sheets stay selectable {frozen[:5]}")

    # printables have print areas
    for n in ("SignIn", "EvalSheet", "SpellingPrint", "WritingHandout",
              "Transcript", "CadetProfile", "Ranking", "GradChecklist",
              "Audit", "Addendum", "Schedule"):
        check(bool(wb[n].print_area), f"{n} has a print area")
    ad = wb["Addendum"]
    check("#2055" in str(ad["G50"].value or "") or
          any("#2055" in str(ad.cell(row=rr, column=7).value or "")
              for rr in range(10, 60)),
          "Addendum maps Firearms excess to course #2055")

    # key policy formulas present
    es = wb["ExamScores"]
    check("cfgRetakeRecordedCap" in es["M6"].value,
          "retake cap rule in ExamScores")
    check("cfgRetestClassDays" in es["T6"].value,
          "5-class-day retest deadline in ExamScores")
    # a date that is not on the class-day calendar must roll forward, and a
    # deadline that still cannot be computed must SAY so — "Pending" forever
    # made the policy 300.5 clock unenforceable and invisible to the audit
    check('nrCDdate,nrCDnum,"",1)' in es["T6"].value and
          '"CHECK DATE"' in es["U6"].value and
          "(not a class day)" not in es["T6"].value,
          "retest deadline rolls to the next class day; bad dates say so")
    check('nrCDdate,nrCDnum,"",1)' in wb["Memos"]["H6"].value and
          '"CHECK DATE"' in wb["Memos"]["L6"].value,
          "memo due date rolls to the next class day; bad dates say so")
    # text sorts above every number, so an unguarded >= made 'absent' a pass
    check("ISNUMBER" in es["N6"].value and "ISNUMBER" in es["O6"].value and
          "ISNUMBER" in es["Q6"].value and "ISNUMBER" in es["R6"].value and
          "ISNUMBER" in es["M6"].value,
          "ExamScores pass/retake columns require a numeric score")
    # a failed attempt 2 always reaches dismissal review, even when the
    # attempt-1 row was never keyed in
    check("nrES_Att,1)>0" not in es["R6"].value,
          "dismissal review does not depend on attempt 1 being on file")
    # ...and an orphan attempt 2 records at the cap, not at its raw score
    check("MAX($J6,cfgRetakeRecordedCap)" in es["M6"].value,
          "passed retake records at the cap (>= the exam's own passing)")
    check(es["X5"].value == "Row Check" and
          "RAW SCORE NOT A NUMBER" in es["X6"].value and
          "DUPLICATE RECORD" in es["X6"].value and
          "ATTEMPT 2 WITHOUT ATTEMPT 1" in es["X6"].value,
          "ExamScores Row Check flags the silent data-entry faults")
    sg = wb["sysGrades"]
    check("cfgWeightMajor" in sg["M6"].value, "weighted grade in sysGrades")
    # counts and averages share criteria so an unscored (pending-retest) row
    # can never leave the count>0 guard true while the average has no data
    check(all('nrES_Rec,">=0"' in sg[f"{cl}6"].value for cl in "EFHIJL") and
          all("IFERROR" in sg[f"{cl}6"].value for cl in "IJL"),
          "sysGrades counts/averages both require a recorded score")
    # ...and a BOUNDED one: ExamScores has no upper-bound validation on Raw
    # Score, so a mis-keyed 180 used to produce a >100 transcript average
    check(all('nrES_Rec,"<=100"' in sg[f"{cl}6"].value for cl in "EFHIJL"),
          "sysGrades excludes an out-of-range score from counts and averages")
    # the same population must drive every class aggregate, or the value the
    # transcript now refuses merely moves to the printed/emailed class average
    sgrid = wb["ScoresGrid"]
    avg_row = next(r for r in range(56, 60)
                   if str(sgrid.cell(row=r, column=3).value or "")
                   .startswith("Class average"))
    check('nrES_Rec,">=0"' in str(sgrid.cell(row=avg_row, column=4).value)
          and 'nrES_Rec,"<=100"' in str(sgrid.cell(row=avg_row, column=4).value),
          "ScoresGrid class average reads the same population as sysGrades")
    check('nrES_Rec,">=0"' in str(wb["EmailPreview"]["D9"].value),
          "the emailed class average reads the same population as sysGrades")
    ck = wb["sysChecks"]
    check("PT!$AB" in ck["L6"].value and 'PT!$AB6="Yes"' in ck["L6"].value and
          '"Not taken"' in ck["L6"].value,
          "final-PT gate wired into sysChecks (affirmative pass test)")
    check(ck["R5"].value == "Exams Recorded" and '$R6="Yes"' in ck["N6"].value
          and "Exams not all recorded; " in ck["O6"].value,
          "graduation requires every exam category actually recorded")
    check("Final PT failed; " in ck["O6"].value and
          "Final PT rubric not set; " in ck["O6"].value and
          "Final PT not assessed; " in ck["O6"].value,
          "blocking issues name the reason the final PT gate is not met")
    # absence of assessment is not a pass: sysSkills K only asks "not failed
    # out, nothing in remediation", so a cadet with zero skills records used
    # to clear the skills gate outright
    check(ck["S5"].value == "Skills Assessed" and
          "sysSkills!$F6" in ck["S6"].value and "rngSM_cat" in ck["S6"].value
          and '$S6="Yes"' in ck["N6"].value and
          "Skills not all assessed; " in ck["O6"].value,
          "graduation requires every skills category actually qualified")
    # ch.41: nrSKbothCoF used to feed ONE audit line that counted only the
    # outright "No", so a cadet who never fired one course of fire read
    # "Eligible" on the graduation checklist
    check(ck["T5"].value == "Firearms CoF" and
          'sysSkills!$P6="Yes"' in ck["T6"].value and
          '$T6="Yes"' in ck["N6"].value and
          "Firearms courses of fire; " in ck["O6"].value,
          "graduation requires BOTH firearms courses of fire passed")
    # A45: an engine-raised dismissal review must be CLOSABLE, per trigger,
    # and there must be exactly one gate the rest of the workbook reads
    check(ck["U5"].value == "Open Reviews" and
          "nrDIS_PID" in ck["U6"].value and "nrDIS_Closes" in ck["U6"].value
          and "nrDIS_Outcome" in ck["U6"].value
          and "nrDIS_Approval" in ck["U6"].value
          and 'nrDIS_Closes,"Failed retest (exam)"' in ck["U6"].value
          and 'nrDIS_Closes,"Skills failed out"' in ck["U6"].value
          and 'nrDIS_Closes,"Chain-of-command incident review"' in ck["U6"].value
          and '"Retained w/ Plan"' in ck["U6"].value
          and "*" not in ck["U6"].value        # no wildcard criteria
          and '$U6=""' in ck["M6"].value,
          "sysChecks DismissReview clears from the DismissalLog, per trigger")
    # the WatchList reason must come from the same per-trigger arithmetic, or
    # it can name a trigger whose review has already been closed
    check("sysChecks!$U6" in wb["sysFlags"]["T6"].value and
          "sysGrades!$P6>0" not in wb["sysFlags"]["T6"].value,
          "the dismissal-review flag reason names only UNCLOSED triggers")
    check("sysChecks!$M6" in sg["T6"].value and
          "sysGrades!$P6" not in sg["T6"].value and
          "sysChecks!$M6" in wb["sysFlags"]["R6"].value,
          "rank eligibility and the flag read the ONE dismissal-review gate")
    dl = wb["DismissalLog"]
    check(dl["P5"].value == "Closes Trigger" and
          str(wb.defined_names["nrDIS_Closes"].value).endswith("$P$6:$P$105")
          and dl.auto_filter.ref == "B5:P5",
          "DismissalLog carries the trigger the review closes")
    import data_lists as _DL0
    check(_DL0.LISTS["Dismissal Trigger"][:3] ==
          ["Failed retest (exam)", "Skills failed out",
           "Chain-of-command incident review"],
          "the Closes Trigger list matches the strings the engine matches")
    check(wb["Control"]["I6"].value.startswith("=IF")
          and "WORKDAY.INTL(" in wb["Control"]["I6"].value
          and "_xlfn.WORKDAY" not in wb["Control"]["I6"].value,
          "class-day calendar generated on Control (native WORKDAY.INTL)")
    cm2 = wb["ChapterMaster"]
    check("nrSCH_ChNum" in cm2["G6"].value,
          "chapter delivered-hours roll up by chapter number")
    check("nrSUBname" in wb["Schedule"]["H6"].value,
          "schedule resolves sub-classes to parent chapters")
    check("nrSUBname" in wb.defined_names and
          any(cm2.cell(row=r, column=2).value == "Crash Investigation"
              for r in range(50, 75)),
          "TPD sub-class block on ChapterMaster")
    ins = wb["Instructors"]
    check("SEARCH($B6,nrSCH_Instr)" in ins["L6"].value and
          "nrSCH_ChNum" in ins["M6"].value,
          "per-instructor on-schedule scan + chapters taught")
    check('$I6="On File"' in ins["K6"].value and
          '"Guest/Outside","N/A"' in ins["K6"].value.replace("'", '"'),
          "audit-ready requires bio; guests exempt")
    check("UNRECOGNIZED" in wb["Schedule"]["N6"].value and
          '(nrInstrNames<>"")' in wb["Schedule"]["N6"].value,
          "schedule flags unrecognized instructor entries (blank-safe)")
    # one roster match is not enough — the multi-select macro builds
    # comma-separated cells and an off-roster co-teacher used to ride along
    check('SUBSTITUTE(_xlpm.b,",","")' in wb["Schedule"]["N6"].value and
          'SEARCH("(",_xlpm.a)' in wb["Schedule"]["N6"].value,
          "every comma-separated instructor token must resolve "
          "(parentheticals stripped)")
    sa = wb["sysAudit"]
    # located by check NAME, never by row: hard-coded rows broke every time a
    # check was inserted above them, which is exactly what column shifts do
    sa_target = {}
    for r in range(6, 80):
        nm = str(sa.cell(row=r, column=2).value or "")
        if nm:
            sa_target[nm] = sa.cell(row=r, column=4).value
    lits = [v for k, v in sa_target.items()
            if isinstance(v, str) and not str(v).startswith("=")]
    check("recent" in lits and "Yes / Yes" in lits and "set" in lits,
          "sysAudit literal targets stored as text, not formulas")
    check(any("unrecognized cadet" in k for k in sa_target),
          "sysAudit orphaned-PID check present")
    # the printed Audit sheet must mirror EVERY engine check: a check the
    # Dashboard tile counts but the packet never prints is a red tile with
    # no visible cause
    import sheets_engine
    n_eng = len(sheets_engine.AUDIT_CHECKS)
    aud_rows = [r for r in range(5, 80)
                if str(wb["Audit"].cell(row=r, column=2).value or
                       "").startswith("=sysAudit!$B$")]
    check(len(aud_rows) == n_eng,
          f"Audit sheet prints all {n_eng} engine checks ({len(aud_rows)})")
    checks_txt = " | ".join(str(sa.cell(row=r, column=2).value or "")
                            for r in range(6, 6 + n_eng))
    check("Duplicate cadet PID or name" in checks_txt and
          "Exam rows failing Row Check" in checks_txt and
          "unusable due date" in checks_txt,
          "sysAudit covers duplicate PIDs, exam row checks and dead due dates")
    # N/A is an offered answer on the Audit sheet and painted green there —
    # the engine check must exempt it too or the row is red forever
    prg = next(str(sa.cell(row=r, column=3).value) for r in range(6, 6 + n_eng)
               if "Program requirements" in str(sa.cell(row=r, column=2).value))
    check('nrPRGmet<>"N/A"' in prg, "program-requirement check exempts N/A")
    check("COUNTIFS" in wb["ScoresGrid"]["D6"].value and
          "AVERAGEIFS" in str(wb["ScoresGrid"]["D57"].value or ""),
          "ScoresGrid blanks untaken exams; class avg from the log")
    check("nrBankTopics" in wb.defined_names and
          "nrBankSel" in wb.defined_names and
          any(wb["InstructorBanks"].cell(row=r, column=2).value ==
              "Arrest and Control" for r in range(6, 110)),
          "InstructorBanks sheet with per-topic rows")
    sch_dvs = [dv.formula1 for dv in wb["Schedule"].data_validations.dataValidation]
    check(any("nrBankSel" in (f or "") for f in sch_dvs),
          "schedule instructor dropdown is bank-dependent")
    cp = wb["ChapterPacket"]
    check("cfgPacketChapter" in wb.defined_names and
          any("nrInstrChTaught" in str(c.value) for row in
              cp.iter_rows(min_row=5, max_row=40) for c in row
              if isinstance(c.value, str)) and bool(cp.print_area),
          "ChapterPacket one-page training-file view")
    ex = wb["ExamSheet"]
    check("cfgGradeSheetExam" in wb.defined_names and bool(ex.print_area) and
          any("nrES_Raw" in str(c.value) for row in
              ex.iter_rows(min_row=10, max_row=12) for c in row
              if isinstance(c.value, str)),
          "ExamSheet per-assessment grade sheet")
    check(wb["ExamScores"].freeze_panes == "D6" and
          wb["Schedule"].freeze_panes == "B6" and
          wb["ExamScores"].auto_filter.ref == "B5:Y5" and
          wb["Makeup"].auto_filter.ref == "B5:N5",
          "freeze panes + log filters (filters cover Row Check)")

    # policy: absent+unexcused = a recorded 0 that starts the retest clock;
    # absent+excused = the FIRST attempt is merely delayed (no zero, no clock)
    es2 = wb["ExamScores"]
    check(es2["Y5"].value == "Absence" and
          "UNEXCUSED ABSENCE MUST RECORD 0" in str(es2["X6"].value),
          "unexcused absence must record a 0 (Row Check)")
    check("EXCUSED - 1st attempt pending" in str(es2["U6"].value),
          "excused absence shows a pending first attempt, no retest clock")
    check(str(es2["U6"].value).count("(") == str(es2["U6"].value).count(")"),
          "ExamScores Retest Status parens balanced")
    ck2 = wb["sysChecks"]
    check(ck2["V5"].value == "Exams Pending" and
          'nrES_Absence' in str(ck2["V6"].value) and
          '$V{}="Yes"'.format(6) in str(ck2["N6"].value),
          "a pending excused exam blocks graduation")
    fl2 = wb["sysFlags"]
    # the flag used to SAY "REMOVAL TRIGGER" and be wired to nothing: no
    # review opened, no graduation block, and no Closes Trigger value it
    # could be closed with. It is now one of the engine triggers, so the
    # assertion is that it reaches sysChecks and the DismissalLog.
    import data_lists as _DLm
    check(fl2["U5"].value == "F:MissedExam" and
          "unexcused missed exams (removal review)" in str(fl2["T6"].value),
          "second unexcused missed exam is named on the flag reasons")
    check('nrES_Absence' in str(ck2["U6"].value) and
          '"unexcused missed exams; "' in str(ck2["U6"].value) and
          "Unexcused missed exams" in _DLm.LISTS["Dismissal Trigger"],
          "second unexcused missed exam opens a closable removal review")
    check(any("nrSCH_Date=TODAY()" in str(c.value) for row in
              wb["Dashboard"].iter_rows(min_row=5, max_row=25) for c in row
              if isinstance(c.value, str)),
          "Dashboard Today panel")
    # a single-ROW Reference without from_rows becomes one series per column
    dash_charts = wb["Dashboard"]._charts
    check(len(dash_charts) == 2 and
          all(len(c.series) == 1 for c in dash_charts),
          "Dashboard charts plot one series across the categories "
          f"{[len(c.series) for c in dash_charts]}")
    wmG = wm["G6"].value
    check("nrCHfirst" in wmG and "nrCDdate" in wmG,
          "writing due dates computed from schedule")

    cert = wb["Certifications"]
    check("TEXTJOIN" in str(cert["U6"].value) and
          cert["D5"].value == "TIM Date",
          "Certifications grid with to-collect rollup")
    # decision on record: certifications are a HARD graduation block, not a
    # warning ("they have to do everything"). The gate must also fail CLOSED
    # on absent data - a cadet row with nothing entered blocks, it does not
    # sail through - and the one escape hatch (Copy? = N/A) must be visible.
    cf2 = wb["Certifications"]
    check(cf2["V5"].value == "Waived (N/A)" and
          '="N/A"' in str(cf2["V6"].value) and
          "nrCERTwaived" in wb.defined_names,
          "every N/A waiver is named per cadet, not silently dropped")
    check(str(cf2["T6"].value).endswith('IF($U6="","Yes","No"))'),
          "All Certs? fails closed: blank row is No, never Yes")
    check("Certification requirements waived as N/A" in
          [c[0] for c in sheets_engine.AUDIT_CHECKS],
          "sysAudit counts waived certification requirements")
    check("Certifications!$T" in wb["sysChecks"]["Q6"].value and
          '$Q6="Yes"' in wb["sysChecks"]["N6"].value and
          "Certs; " in wb["sysChecks"]["O6"].value,
          "certs gate graduation on sysChecks (matches GradChecklist note)")
    # sysFlags gained F:DismissReview at the END of the flag block, so the
    # block is E..R and Flag Count / Reasons are S / T (they were R / S)
    check("Certifications!$U" in wb["sysFlags"]["O6"].value and
          "cert copies outstanding" in wb["sysFlags"]["T6"].value,
          "cert warning flag in sysFlags with reason text")

    # Cert Expiration was a dead field: keyed, never read. It is now graded
    # against the LAST date that instructor appears on the Schedule, because
    # what TCOLE checks is that the instructor was licensed on the day taught
    # - not merely that the certificate is valid today.
    ins2 = wb["Instructors"]
    check(ins2["N5"].value == "Last Class Taught" and
          ins2["O5"].value == "Cert Status" and
          "nrSCH_Date" in str(ins2["N6"].value) and
          "MAX(" in str(ins2["N6"].value),
          "Instructors: last class taught computed from the schedule")
    ocell = str(ins2["O6"].value)
    check("EXPIRED BEFORE LAST CLASS TAUGHT" in ocell and
          "MISSING EXPIRATION" in ocell and
          "RENEW - expires during academy" in ocell and
          "$G6<$N6" in ocell,
          "Instructors: cert expiration graded against the date taught")
    check(ocell.count("(") == ocell.count(")"),
          "Instructors Cert Status parens balanced")
    check('LEFT($O6,7)<>"EXPIRED"' in str(ins2["K6"].value) and
          '$O6<>"MISSING EXPIRATION"' in str(ins2["K6"].value),
          "an expired/undocumented cert fails Audit Ready")
    check("nrInstrCertStat" in wb.defined_names and
          "nrInstrLastTaught" in wb.defined_names,
          "cert-status named ranges defined")
    # the InstructorBanks sheet asserted who may teach what and nothing
    # enforced it: the Schedule dropdown is warning-only and a multi-name
    # cell bypasses validation entirely.
    schP = str(wb["Schedule"]["P6"].value)
    check(wb["Schedule"]["P5"].value == "Bank Check" and
          '"NOT IN BANK"' in schP and "nrBankGrid" in schP and
          "nrBankTopics" in schP,
          "Schedule flags an instructor outside the topic's certified bank")
    # kept out of N's LET on purpose: LibreOffice cannot evaluate LET, and a
    # control that cannot be exercised on real data in the recalc sweep is a
    # control nobody has ever seen work
    check("_xlfn.LET" not in schP and "_xlpm." not in schP,
          "the bank check is LET-free so the recalc sweep can exercise it")
    check('COUNTA(INDEX(nrBankGrid' in schP,
          "bank enforcement is skipped for a topic with an empty bank")
    check("nrSCH_BankOK" in wb.defined_names,
          "nrSCH_BankOK defined")
    import sheets_engine as _SE2
    _audit_names = [c[0] for c in _SE2.AUDIT_CHECKS]
    for _want in ("Blocks taught outside the topic's certified bank",
                  "Instructor cert expired before a class they taught",
                  "Teaching instructor certs missing or expiring mid-academy"):
        check(_want in _audit_names, f"sysAudit check present: {_want}")
    wr = wb["Writing"]
    # "overdue missing" counts anything that is not an X, not merely
    # blanks: a stray mark used to erase the assignment from the counter,
    # from the red highlight and from "Writing Current?"
    check('COUNTIF(D6:AQ6,"X")' in wr["AR6"].value and
          '(UPPER(D6:AQ6)<>"X")' in wr["AS6"].value,
          "Writing grid counts X marks; anything else is not done")
    check(wr["D6"].alignment.horizontal == "center",
          "Writing X cells centered")
    check("InputGuide" in wb.sheetnames and
          "HYPERLINK" in str(wb["InputGuide"]["B7"].value),
          "InputGuide page with hyperlinks")
    check("HYPERLINK" in str(wb["Cadets"]["B1"].value) and
          "HYPERLINK" in str(wb["sysGrades"]["B1"].value),
          "home links back to Dashboard on sheets")
    dash_text = [str(c.value) for row in
                 wb["Dashboard"].iter_rows(min_row=5, max_row=12)
                 for c in row if isinstance(c.value, str)]
    check(any("Final Test" in t for t in dash_text) and
          any("State Test" in t for t in dash_text) and
          any("Graduation" in t for t in dash_text),
          "Dashboard date tiles (start/final/state/graduation)")
    check("nrSK_CoF" in wb["sysSkills"]["N6"].value,
          "firearms course-of-fire bests in sysSkills")
    se = wb["StateExam"]
    check("new BPOC required" in se["K6"].value,
          "state exam 3-attempt rule")
    # the category block on the official transcript / cadet profile must show
    # the FINAL EXAM average (sysGrades L), not the weighted composite (N)
    check("nrGRfinalExam" in wb.defined_names and
          str(wb.defined_names["nrGRfinalExam"].value).endswith("$L$6:$L$55"),
          "nrGRfinalExam points at the final-exam average column")
    for shname in ("Transcript", "CadetProfile"):
        sh = wb[shname]
        cells = [c for row in sh.iter_rows(min_row=5, max_row=40) for c in row
                 if isinstance(c.value, str) and "nrGRfinalExam" in c.value]
        check(len(cells) == 1,
              f"{shname} category block reads the final-exam average")
    check("nrCERTmissing" in str(wb["Dashboard"]["B18"].value or "") or
          any("nrCERTmissing" in str(c.value)
              for row in wb["Dashboard"].iter_rows(min_row=5, max_row=40)
              for c in row if isinstance(c.value, str)),
          "cert reminders on Dashboard")

    at = wb["Attendance"]
    check("nrMK_Link" in at["P6"].value and "CLEARED" in at["S6"].value,
          "per-event makeup reconciliation with CLEARED status")
    # the per-event banner and the sysAttendance owed balance must apply the
    # SAME criteria, or another cadet's (or another type's) makeup clears an
    # event the graduation engine still counts as owed
    check("nrMK_PID" in at["P6"].value and "nrMK_PID" in at["Q6"].value and
          "nrMK_PID" in at["S6"].value and
          '"Classroom"' in at["P6"].value and '"PT"' in at["Q6"].value,
          "Attendance makeup credit filtered by cadet PID + makeup type")
    # makeup credit is capped at the linked event's OWN duration, and the
    # sysAttendance roll-up is derived from those capped per-event figures —
    # otherwise a surplus credit paid off a different absence and the
    # graduation gate read "no makeup owed" over a still-OPEN event
    check(at["P6"].value.startswith('=IF(OR($B6="",N($I6)=0),"",MIN(N($I6),')
          and 'MIN(N($J6),' in at["Q6"].value,
          "Attendance caps per-event makeup credit at the event's own size")
    sat = wb["sysAttendance"]
    check("nrAT_MadeUpMin" in sat["F6"].value and
          "nrAT_MadeUpSess" in sat["N6"].value and
          "nrMK_Min" not in sat["F6"].value and
          "nrMK_Sess" not in sat["N6"].value,
          "sysAttendance made-up totals come from the capped per-event ledger")
    mk = wb["Makeup"]
    check(mk["N5"].value == "Row Check" and "WRONG CADET" in mk["N6"].value and
          "UNIT MISMATCH" in mk["N6"].value and
          "NO LINKED EVENT" in mk["N6"].value and
          "OK (no linked event)" not in mk["N6"].value and
          '$N6,2)="OK"' in mk["L6"].value,
          "Makeup row check gates Credit Applies (unlinked / wrong cadet / "
          "unit / type)")
    mk_dvs = [dv.formula1 for dv in wb["Makeup"].data_validations.dataValidation]
    # nrAT_IDlist, not nrAT_ID: the raw column is 800 formula cells, so the
    # picker listed one blank entry per unused row
    check(any((f or "").strip() == "=nrAT_IDlist" for f in mk_dvs)
          and "nrAT_IDlist" in wb.defined_names,
          "Makeup Linked Event dropdown = FILTERed attendance EventIDs")
    me = wb["Memos"]
    check("cfgMemoDueClassDays" in me["H6"].value and
          "OVERDUE" in me["L6"].value,
          "Memos: computed due dates + overdue status")
    me_dvs = [dv.formula1 for dv in wb["Memos"].data_validations.dataValidation]
    check(any("nrAllRefIDs" in (f or "") for f in me_dvs),
          "Memo Linked Ref dropdown of I/A/C IDs")
    check("nrME_Status" in wb["sysFlags"]["P6"].value and
          'nrAT_Cleared="OPEN"' in wb["sysFlags"]["Q6"].value,
          "overdue-memo and open-time warning flags")
    check('nrIN_Report="Yes"' in str(wb["EmailPreview"]["B65"].value or "") or
          any('nrIN_Report="Yes"' in str(c.value) for row in
              wb["EmailPreview"].iter_rows(min_row=55, max_row=75)
              for c in row if isinstance(c.value, str)),
          "email digest filters on Report-to-Agency marks")
    dl = wb["DailyLog"]
    check("nrCDdate" in dl["C6"].value and "nrME_Received" in dl["K6"].value,
          "DailyLog computes day #, class type and counters")

    check("nrAB_Date" in wb.defined_names and
          "cfgPolicyVersion" in wb.defined_names and
          "cfgBoardReviewed" in wb.defined_names and
          wb["AdvisoryBoard"]["D6"].value == "May 2026",
          "AdvisoryBoard governance alignment + meeting reference list")
    aud = wb["Audit"]
    # scan the whole sheet: the program-checks block grows whenever a check
    # is added to sysAudit, which pushes the enrollment-docs grid down
    found_ack = any("Rules Ack" == str(c.value) for row in
                    aud.iter_rows(min_row=5, max_row=aud.max_row)
                    for c in row)
    check(found_ack, "Rules Ack column in enrollment docs grid")

    st = wb["Settings"]
    tot_row = next(r for r in range(6, 60)
                   if st.cell(row=r, column=5).value == "cfgTotalScheduledMinutes")
    tot_chk = str(st.cell(row=tot_row, column=7).value or "")
    # SUMIFS(...,nrSCH_TimeCheck,"OK"), not SUM: a swapped start/end still
    # produces MOD-derived hours, and this is the figure the sheet tells the
    # coordinator to copy into cfgTotalScheduledMinutes (which scales the 5%
    # classroom attendance cap)
    tot_det = str(st.cell(row=tot_row, column=6).value or "")
    check('SUMIFS(nrSCH_Hrs,nrSCH_TimeCheck,"OK")' in tot_det
          and "No schedule entered yet" in tot_chk,
          "academy length cross-checked against the Schedule, "
          "impossible-time blocks excluded")
    # the silent window this check exists to close: New Academy Reset empties
    # the Schedule but leaves the previous academy's minutes in C, and an
    # empty Schedule used to take the blind "No schedule entered yet" branch
    check(f"AND($F${tot_row}=0,N($C${tot_row})=0)" in tot_chk and
          "Total Scheduled Minutes still reads" in tot_chk,
          "stale academy length is reported when the Schedule is empty")

    # every cadet grid keys off Cadets!B (PID) and every log resolves PID by
    # MATCH on the name, so a repeat of either silently merges two cadets
    cad_cf = " ".join(str(rule.formula[0]) for rng in
                      wb["Cadets"].conditional_formatting
                      for rule in rng.rules if rule.formula)
    check("COUNTIF(rngCadetPIDs,$B6)>1" in cad_cf and
          "COUNTIF(rngCadetNames,$F6)>1" in cad_cf,
          "duplicate PID / duplicate cadet name highlighted on Cadets")

    # ---- regression guards for the round-4 stress-test fixes -------------
    # a blank picker must not match the blank rows of the log it filters
    si = wb["SignIn"]
    check('(nrSCH_Date<>"")*(nrSCH_Date=cfgSignInDate)' in si["B10"].value,
          "SignIn schedule block ignores empty Schedule rows on a blank date")
    check('IF(cfgSignInDate=""' in si["B7"].value,
          "SignIn banner prints a blank date line, not January 0 1900")
    cprof = wb["CadetProfile"]
    for cell, nm in (("B27", "nrIN_PID"), ("B38", "nrCO_PID"),
                     ("B50", "nrAT_PID"), ("B61", "nrMK_PID")):
        v = cprof[cell].value
        check(v.count(f'({nm}<>"")') == 2,
              f"CadetProfile {cell} guards both FILTERs against blank PIDs")

    # chapter numbers are the workbook's join key: nrCHnum and EVERY cell of
    # EVERY dropdown sourced from it must share one data type, or an exact
    # XLOOKUP/MATCH silently returns #N/A across two auditor-facing pages
    cm = wb["ChapterMaster"]
    check(all(cm.cell(row=r, column=3).data_type == "s"
              for r in range(6, 50)), "nrCHnum stored as text")
    for sheet, rngs in (("ChapterPacket", ["C5"]), ("EvalSheet", ["C5"]),
                        ("ExamPlan", [f"H{r}" for r in range(6, 31)]),
                        ("WritingMaster", [f"D{r}" for r in range(6, 46)]),
                        ("ChapterMaster", [f"E{r}" for r in range(56, 71)])):
        wsx = wb[sheet]
        dvf = [dv.formula1 for dv in wsx.data_validations.dataValidation]
        check("=nrCHnum" in dvf, f"{sheet} chapter dropdown reads nrCHnum")
        bad = [c for c in rngs if wsx[c].number_format != "@"]
        check(not bad, f"{sheet} chapter cells are text-formatted {bad[:3]}")

    # an unscored attempt-2 row must not delete the failed exam it retests
    check(all('nrES_Att,2,nrES_Raw,">=0"' in es[c + "6"].value
              for c in "MQU") and
          'nrES_Att,">"&$K6,nrES_Att,"<=2",nrES_Raw,">=0"' in es["P6"].value and
          "RETEST ROW HAS NO SCORE" in es["X6"].value,
          "a retest only counts once it is scored (M/P/Q/U + Row Check)")
    # policy 300.5 = one retest per exam. Attempt 3+ used to be a legal
    # dropdown pick that recorded the FAILED attempt-1 score over the passed
    # retest with Row Check still reading OK.
    es_dvs = [dv.formula1 for dv in wb["ExamScores"].data_validations.dataValidation]
    check(any("lstExamAttemptNum" in (f or "") for f in es_dvs),
          "ExamScores Attempt # dropdown offers 1 and 2 only")
    check("ATTEMPT 3+ NOT ALLOWED" in es["X6"].value and
          'IF(N($K6)>2,""' in es["M6"].value and
          'IF(N($K6)>2,"No"' in es["P6"].value and
          'AND(N($K6)>=2' in es["R6"].value,
          "attempt 3+ records nothing, is flagged, and still opens a "
          "dismissal review")
    # deadlines may not roll onto days the calendar marks In Session? = No
    check("nrCDinsession" in es["T6"].value and
          "nrCDinsession" in wb["Memos"]["H6"].value,
          "retest/memo deadlines are gated on the in-session flag")
    # MAXIFS returns 0, not an error, so IFERROR never caught the empty case
    check("(not scheduled)" in wb["ExamSheet"]["B8"].value and
          "IFERROR(TEXT(_xlfn.MAXIFS" not in wb["ExamSheet"]["B8"].value,
          "ExamSheet 'Administered' cannot print a 1899 date")
    # a block may cross midnight; a swapped start/end may not go unnoticed
    sch = wb["Schedule"]
    check("MOD($E6-$D6,1)" in sch["F6"].value,
          "Schedule hours are midnight-safe (never negative)")
    check("CHECK TIMES" in str(sch["O6"].value or "") and
          "nrSCH_TimeCheck" in wb.defined_names and
          sch.auto_filter.ref == "B5:P5",
          "Schedule Time Check column, named and inside the filter "
          "(which now reaches P, the Bank Check)")
    sp = wb["Spelling"]
    check("SCORE OUT OF RANGE" in str(sp["S6"].value or "") and
          "nrSpellRowCheck" in wb.defined_names,
          "Spelling Row Check catches out-of-range scores")
    check(any(dv.type == "whole" for dv in
              sp.data_validations.dataValidation),
          "Spelling score grid carries 0-100 validation")
    # the computus helper block must keep the label the loop used to erase
    ctl = wb["Control"]
    check("do not edit" in str(ctl["P4"].value or "") and
          ctl["P5"].value == "Y" and ctl["AF5"].value == "GoodFri",
          "Control computus helper is labelled")

    # ---- regression guards for the round-5 stress-test fixes -------------
    ctl2 = wb["Control"]
    # A00: Christmas Eve and Christmas Day may never resolve to one date.
    # Eve is derived from the OBSERVED Christmas Day, not shifted separately.
    hol = {str(ctl2.cell(row=r, column=2).value): r for r in range(6, 20)}
    eve_f = str(ctl2.cell(row=hol["Christmas Eve"], column=3).value)
    day_f = str(ctl2.cell(row=hol["Christmas Day"], column=3).value)
    check("12,25" in eve_f and "12,24" not in eve_f and
          "IF(WEEKDAY(" in eve_f and ",2)=1,3,1)" in eve_f,
          "Christmas Eve is the weekday BEFORE the observed Christmas Day")
    check("12,25" in day_f, "Christmas Day keeps the weekend-shift rule")
    import datetime as _dt
    _bad = []
    for _y in range(2024, 2045):
        _x = _dt.date(_y, 12, 25)
        _w = _x.isoweekday()
        _obs = (_x - _dt.timedelta(1) if _w == 6 else
                _x + _dt.timedelta(1) if _w == 7 else _x)
        _eve = _obs - _dt.timedelta(3 if _obs.isoweekday() == 1 else 1)
        if _eve == _obs or _eve.isoweekday() > 5 or _obs.isoweekday() > 5:
            _bad.append(_y)
    check(not _bad, f"Christmas rule yields 2 distinct weekdays every year {_bad}")

    # A01: the WORKDAY.INTL START argument is guarded, not just the holidays
    check("ISNUMBER(cfgStartDate)" in ctl2["I6"].value and
          "IFERROR(WORKDAY.INTL(" in ctl2["I6"].value,
          "class-day calendar survives a blank / non-date Start Date")
    check("START DATE on Settings" in str(ctl2["H3"].value or ""),
          "Control says so when the Start Date is unusable")
    st2 = wb["Settings"]
    sd_row = next(r for r in range(6, 60)
                  if st2.cell(row=r, column=5).value == "cfgStartDate")
    check("FIX - Start Date" in str(st2.cell(row=sd_row, column=7).value or ""),
          "Settings flags a blank / non-date Start Date")

    # A29: a TEXT extra-closure date is no longer coerced to 0 and dropped
    n_h_v = sum(1 for r in range(6, 20) if ctl2.cell(row=r, column=2).value)
    m_extra = str(ctl2.cell(row=6 + 2 * n_h_v, column=13).value)
    check("ISNUMBER(F6)" in m_extra and "N(F6)" not in m_extra,
          "extra closure dates are ISNUMBER-guarded, never N()-coerced")
    check(ctl2["G5"].value == "Closure Check" and
          "IGNORED" in str(ctl2["G6"].value or "") and
          "nrExtraClosureCheck" in wb.defined_names,
          "Control has a Row Check for the extra-closure block")
    check(any(dv.type == "date" for dv in
              ctl2.data_validations.dataValidation),
          "extra closure cells carry date validation")

    # A09: the retest deadline survives the retest, and a late one is named
    check('$K6<>1' in es["T6"].value and 'IF($Q6<>"Yes"' not in es["T6"].value,
          "retest deadline keys off the failed attempt 1, not RetakeReq?")
    check("LATE RETEST" in es["U6"].value and "RETEST UNDATED" in es["U6"].value,
          "a retest taken after the deadline reads LATE RETEST")
    check('"LATE RETEST*"' in wb["sysFlags"]["L6"].value,
          "a late retest raises the sysFlags retest flag")

    # A11: nrCKfinalExamElig (sysChecks P) is referenced by something now
    check("nrCKfinalExamElig" in es["X6"].value and
          "500.1.H" in es["X6"].value,
          "final-PT-blocks-Final-Exam rule is wired into ExamScores Row Check")

    # A12: sysFlags carries a dismissal-review flag that reaches WatchList
    sf = wb["sysFlags"]
    check(sf["R5"].value == "F:DismissReview" and
          sf["S5"].value == "Flag Count" and sf["T5"].value == "Reasons" and
          "SUM($E6:$R6)" in sf["S6"].value and
          "DISMISSAL REVIEW OPEN" in sf["T6"].value,
          "sysFlags flags an open dismissal review (flag block E..R)")
    check(str(wb.defined_names["nrFLcount"].value).endswith("$S$6:$S$55") and
          str(wb.defined_names["nrFLreasons"].value).endswith("$T$6:$T$55"),
          "nrFLcount / nrFLreasons moved with the appended flag column")
    check("sysFlags!$T6" in wb["EmailPreview"]["J9"].value,
          "EmailPreview reads the relocated sysFlags Reasons column")

    # A10 / A23: Attendance finally has a Row Check
    at2 = wb["Attendance"]
    check(at2["T5"].value == "Row Check" and
          "MINUTES ON A PT EVENT" in at2["T6"].value and
          "SESSIONS ON A CLASSROOM EVENT" in at2["T6"].value and
          "COUNTED EVENT WITH NO MINUTES" in at2["T6"].value and
          "nrAT_RowCheck" in wb.defined_names and
          at2.auto_filter.ref == "B5:T5",
          "Attendance Row Check catches blank / mismatched units")

    # A15 / A22 / A28: the Makeup ledger cannot be moved by a bad row
    mk2 = wb["Makeup"]
    check("NO MAKEUP DATE" in mk2["N6"].value and
          "NEGATIVE CREDIT" in mk2["N6"].value and
          "NO MINUTES CREDITED" in mk2["N6"].value,
          "Makeup Row Check rejects a dateless, zero or negative credit")
    check("CLEARED (makeup date missing)" in at2["S6"].value and
          'IFERROR(_xlfn.MAXIFS' in at2["S6"].value,
          "a dateless makeup can no longer stamp CLEARED 12/30")
    import data_lists as _DL
    check(_DL.LISTS["Makeup Type"] == ["Classroom", "PT"],
          "Makeup Type offers only the two types the caps credit")

    # A37: an unscored skills placeholder must not supersede a real result
    sk2 = wb["Skills"]
    check('nrSK_Res,"Pass")' in sk2["O6"].value and
          'nrSK_Res,"Fail")' in sk2["O6"].value and
          'nrSK_Res,"Pass")' in sk2["N6"].value,
          "Skills status/attempts count only SCORED attempts")
    # ...and it must count ATTEMPT NUMBERS, not rows: one firearms attempt is
    # TWO rows (Course of Fire 1 and 2), which failed a cadet out an attempt
    # (or two) early and opened a 300.7 separation review on him
    check("COUNTIFS(" not in sk2["N6"].value and
          "_xlfn.MAXIFS(nrSK_Att" in sk2["N6"].value,
          "Skills 'Attempts Used' counts attempt numbers, not log rows")

    # A13 / A14 / A38: no more 12/30/1899 from an unguarded MAXIFS/MINIFS
    check("not yet taught" in str(wb["ChapterPacket"]["F8"].value or ""),
          "ChapterPacket 'Last taught' cannot print a 1899 date")
    check("nrELagency,$B6)=0" in wb["Agencies"]["I6"].value,
          "Agencies 'Last Email Sent' is blank, not 1899, when never emailed")
    check('cfgPreviewAgency),0)=0,"never"'
          in str(wb["EmailPreview"]["E5"].value or ""),
          "EmailPreview 'never' fallback is reachable")
    # A46: a retest badge is a claim about a retest that HAPPENED. An
    # attempt-2 row is logged when the retest is SCHEDULED, before it is
    # scored; and a FAILED retest records the raw attempt-1 score, not the cap
    check('nrES_Raw,">=0"' in str(wb["ExamSheet"]["I11"].value),
          "ExamSheet 'Retested' needs a SCORED retest")
    check('nrES_AttPass,"Yes"' in str(wb["EmailPreview"]["E9"].value) and
          'nrES_Raw,">=0"' in str(wb["EmailPreview"]["E9"].value) and
          "RETEST FAILED" in str(wb["EmailPreview"]["E9"].value),
          "EmailPreview 'cap 70' badge needs a retest that was scored AND "
          "passed")
    import io as _io
    _vba = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                 "src", "vba", "bpoc",
                                 "modAgencyEmail.bas"), encoding="utf-8").read()
    check('wsES.Range("L6:L1505"), ">=0"' in _vba and
          'wsES.Range("N6:N1505"), "Yes"' in _vba and
          "retest failed - first-attempt score shown" in _vba,
          "modAgencyEmail RetakeNote stays in lockstep with EmailPreview")
    check('IF(N(_xlpm.d)=0,"",_xlpm.d)' in wm["G6"].value,
          "WritingMaster 'Computed Assigned' is zero-guarded like its sibling")

    # A08: per-exam passing scores, not the global one
    check("rngEPpass" in str(wb["ExamSheet"]["H11"].value or ""),
          "ExamSheet Pass?/FAIL uses the exam's own passing score")
    sg_cf = " ".join(str(rule.formula[0]) for rng in
                     wb["ScoresGrid"].conditional_formatting
                     for rule in rng.rules if rule.formula)
    check("rngEPpass" in sg_cf,
          "ScoresGrid red cells use each exam's own passing score")

    # A31: a blank chapter picker must not match every blank row
    check('cfgPacketChapter=""' in str(wb["ChapterPacket"]["B20"].value or "") or
          any('cfgPacketChapter=""' in str(c.value)
              for row in wb["ChapterPacket"].iter_rows(min_row=18, max_row=40)
              for c in row if isinstance(c.value, str)),
          "ChapterPacket spills refuse to run on a blank chapter picker")

    # A36: chapters taught sort 1,2,10,20 - not 1,10,2,20
    check("SORTBY" in wb["Instructors"]["M6"].value,
          "'Chapters Taught' sorts numerically")

    # A42: Cl % / PT % carry a percent format under a % header
    check(wb["sysAttendance"]["I6"].number_format == "0.0%" and
          wb["sysAttendance"]["Q6"].number_format == "0.0%",
          "sysAttendance Cl % / PT % are percent-formatted")

    # A16 / A40 / A41: protection gaps
    # PrintCenter was the only green/OUTPUT tab shipping unprotected, even
    # though ClearButtons restores the protection state it finds and the
    # installer unprotects it without ever re-protecting
    for n in ("Audit", "Dashboard", "InputGuide", "PrintCenter"):
        check(wb[n].protection.sheet, f"{n} protected")
    _ps1x = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                  "tools", "Install-BPOC-VBA.ps1"),
                     encoding="utf-8").read()
    check("$pc.Protect($pw)" in _ps1x,
          "installer re-protects PrintCenter after placing its buttons")
    audx = wb["Audit"]
    docr = next(r for r in range(5, audx.max_row + 1)
                if str(audx.cell(row=r, column=3).value or "") == "Enroll App")
    check(not audx.cell(row=docr + 1, column=3).protection.locked,
          "Audit enrollment-docs inputs stay editable through protection")
    prgr = next(r for r in range(5, audx.max_row + 1)
                if str(audx.cell(row=r, column=6).value or "") == "Met?")
    check(not audx.cell(row=prgr + 1, column=6).protection.locked and
          audx.cell(row=prgr + 1, column=2).protection.locked,
          "Audit requirement answers unlocked, labels locked")
    unlocked_b1 = [ws.title for ws in wb.worksheets
                   if ws.protection.sheet and not ws["B1"].protection.locked]
    check(not unlocked_b1,
          f"B1 home links are LOCKED on protected sheets {unlocked_b1[:5]}")

    # A19: print titles must start at the first row of the print area
    for n in ("Addendum", "ExamSheet", "Audit", "Ranking", "WatchList",
              "GradChecklist", "Schedule"):
        wsx = wb[n]
        tr2 = wsx.print_title_rows
        if not tr2:
            check(True, f"{n} has no repeated print title row")
            continue
        pa = wsx.print_area
        pa = pa[0] if isinstance(pa, list) else str(pa)
        first_body = int(re.sub(r"[^0-9]", "",
                                pa.split("!")[-1].split(":")[0]))
        first_title = int(re.sub(r"[^0-9]", "", tr2.split(":")[0]))
        check(first_title == first_body,
              f"{n} print titles start at the top of the print area "
              f"({tr2} vs {pa})")

    # A03 / A02: capped handout spill, date-formatted digest
    check("_xlfn.TAKE(" in wb["WritingHandout"]["B11"].value,
          "WritingHandout digest is capped to the rows the page styles")
    check("cfgWritingDueTime" in str(wb["WritingHandout"]["E10"].value or ""),
          "cfgWritingDueTime drives the WritingHandout Due column header")
    epv = wb["EmailPreview"]
    since_r = next(r for r in range(55, 90)
                   if "nrIN_Report" in str(epv.cell(row=r, column=2).value or ""))
    # the whole spill, not just the first screenful: the digest is uncapped
    # by design (the VBA draft emits every matching row), so the date format
    # has to cover the worst case or item 17 onward prints as a serial
    fmt_rows = 0
    while epv.cell(row=since_r + fmt_rows, column=2).number_format != "General":
        fmt_rows += 1
    check(fmt_rows >= 300 and "_xlfn.TAKE(" not in
          str(epv.cell(row=since_r, column=2).value or ""),
          f"EmailPreview digest dates render as dates for the whole "
          f"uncapped spill ({fmt_rows} rows)")

    # A43: ExamPlan agrees with the date the exam was actually given
    check("nrES_Date" in wb["ExamPlan"]["I6"].value,
          "ExamPlan 'Exam Date' prefers the actual exam date")

    # A44 / A12 / A29: the new engine roll-ups reach the Audit sheet
    import sheets_engine as _SE
    names = [c[0] for c in _SE.AUDIT_CHECKS]
    for want in ("Makeup rows failing Row Check",
                 "Attendance rows failing Row Check",
                 "Open dismissal reviews (active cadets)",
                 "Extra closure dates that are not dates"):
        check(want in names, f"sysAudit check present: {want}")

    # A47: records management has CONFIRMED scans are the legal originals,
    # so no artifact may still describe that as unconfirmed / demand paper
    stale = [ws.title for ws in wb.worksheets
             if "unconfirmed" in str(ws["B4"].value or "").lower()
             or "keep the paper original" in str(ws["B4"].value or "").lower()]
    check(not stale,
          f"no sheet note calls the scans-are-originals question open {stale}")
    prg = [r for r in range(1, 200)
           if "CONFIRMED by TPD records management"
           in str(wb["Audit"].cell(row=r, column=2).value or "")]
    check(bool(prg),
          "Audit checklist records the confirmed scans-are-originals position")

    # ------------------------------------------------------------------
    # regressions for the stress-test round
    # ------------------------------------------------------------------
    # sysSkills "Failed Out Cats" must count CATEGORIES: firearms records one
    # attempt as TWO Course-of-Fire rows, so a bare row COUNTIFS made a
    # single board decision unable to close the review it raised.
    _ski = str(wb["sysSkills"]["I6"].value or "")
    check("rngSM_cat" in _ski and "SUMPRODUCT" in _ski,
          "sysSkills 'Failed Out Cats' counts categories, not log rows")

    # the emailed spelling figure and the previewed one must be the same
    # number: modAgencyEmail sends Spelling column D+cfgCurrentSpellingNum.
    _epf = str(wb["EmailPreview"]["F9"].value or "")
    _epg = str(wb["EmailPreview"]["G9"].value or "")
    check("INDEX(Spelling!$D6:$O6,cfgCurrentSpellingNum)" in _epf and
          "cfgCurrentSpellingNum>12" in _epf and
          "cfgCurrentSpellingNum>12" in _epg,
          "EmailPreview previews the spelling score the draft sends, and "
          "omits the flag whenever the draft omits the column")

    # a retest keyed against a first attempt that already passed silently
    # replaces the passing score with the cap
    check("RETEST AFTER A PASSING FIRST ATTEMPT" in
          str(wb["ExamScores"]["X6"].value or ""),
          "ExamScores Row Check flags a retest after a passing attempt 1")

    # one today-schedule panel on the Dashboard, capped like SignIn's
    _today = [c.value for row in wb["Dashboard"].iter_rows(min_row=1,
              max_row=40) for c in row
              if isinstance(c.value, str) and "nrSCH_Date=TODAY()" in c.value]
    _sign = [c.value for row in wb["SignIn"].iter_rows(min_row=1, max_row=30)
             for c in row if isinstance(c.value, str)
             and "cfgSignInDate" in c.value and "FILTER" in c.value]
    _cap = re.compile(r'\),(\d+)\),"')
    check(len(_today) == 1 and len(_sign) == 1 and
          _cap.search(_today[0]).group(1) == _cap.search(_sign[0]).group(1) and
          'nrSCH_Date<>""' in _today[0],
          "one Dashboard today-panel, same cap and blank-date guard as SignIn")

    # the SignIn strip inherits the ROSTER's column widths, so it may not be
    # a multi-column spill - nothing inside one can overflow
    check("HSTACK" not in _sign[0] and "IFERROR" in _sign[0],
          "SignIn schedule strip is a single overflowable column")

    # the instructor token check must count roster matches in the SAME
    # stripped text it counts tokens in, and must split on & and / too
    _sn = str(wb["Schedule"]["N6"].value or "")
    check("SEARCH(nrInstrNames,_xlpm.b)" in _sn and
          'SUBSTITUTE(_xlpm.b0,"&",",")' in _sn,
          "Schedule 'Instructor OK?' counts matches in the stripped, "
          "separator-normalised text")

    # cfgThresholdAfterExam ("Category avg enforced after this many exams")
    # used to live only on sysGrades Q/R, which nothing read: the setting was
    # inert and the academic gate enforced 70-in-each-category from exam #1.
    # V must consume Q/R, and the grace must lapse once every PLANNED exam of
    # that type is on file, or it becomes a permanent academic waiver.
    check('$Q6="No"' in str(wb["sysGrades"]["V6"].value or "") and
          "rngEPuse" in str(wb["sysChecks"]["R6"].value or ""),
          "the category-average grace period is live and cannot outlive the "
          "academy (graduation needs every PLANNED exam recorded)")

    # a memo with a cadet but no Assigned date must be reportable
    check('"(enter date)"' in str(wb["Memos"]["H6"].value or ""),
          "Memos with no Assigned date read CHECK DATE, not Pending forever")

    # per-academy inputs living on master/config sheets must be in the
    # New Academy Reset's clear list
    _reset = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                   "src", "vba", "bpoc",
                                   "modNewAcademy.bas"),
                      encoding="utf-8").read()
    for _sheet, _rng in (("WritingMaster", "I6:J45"), ("Control", "F6:F20")):
        check(f'ClearRange "{_sheet}", "{_rng}"' in _reset,
              f"New Academy Reset clears {_sheet}!{_rng}")

    # the button-width warning must not put -f in command-argument position:
    # PowerShell binds it to Write-Host -ForegroundColor and, under
    # $ErrorActionPreference='Stop', kills the install before SaveAs
    _ps1 = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                 "tools", "Install-BPOC-VBA.ps1"),
                    encoding="utf-8").read()
    check(not re.search(r"^\s*-f\s", _ps1, re.M) and
          '(anchor span {1:N0}pt)' in _ps1 and
          'right" -f $caption, $width)' in _ps1,
          "installer's width warning formats inside the parentheses")

    # ---- fixes from the V6 stress pass ------------------------------------
    import data_lists as _DL
    # every unlocked cfg* printable picker must carry a data validation:
    # on a protected sheet the picker is the only cell a user can type in
    sp = wb["SpellingPrint"]
    sp_dvs = [dv.formula1 for dv in sp.data_validations.dataValidation]
    check("=nrSpellTestNums" in sp_dvs and '"Test,Key"' in sp_dvs,
          "SpellingPrint test-number picker is validated 1-12")
    guard = "N(cfgSpellPrintNum)"
    words = [sp.cell(row=rr, column=cc).value
             for rr in range(12, 25) for cc in (3, 6)
             if isinstance(sp.cell(row=rr, column=cc).value, str)
             and "cfgSpellPrintNum" in sp.cell(row=rr, column=cc).value]
    check(len(words) == 25 and all(guard in w for w in words)
          and guard in str(sp["B8"].value or ""),
          "SpellingPrint words + heading refuse an out-of-range test number")
    # ...and the rule is enforced for EVERY unlocked cfg* picker, not just
    # SpellingPrint: ExamSheet, SignIn and WritingHandout each shipped an
    # unlocked, unvalidated cell as the only typeable cell on the sheet.
    _pickers = {}
    for _n, _dn in wb.defined_names.items():
        if not _n.startswith("cfg"):
            continue
        for _sh, _ref in _dn.destinations:
            if _sh in wb.sheetnames and wb[_sh].protection.sheet:
                _pickers.setdefault(_sh, set()).add(_ref.replace("$", ""))
    _unvalidated = []
    for _sh, _cells in _pickers.items():
        _ws = wb[_sh]
        _covered = set()
        for _dv in _ws.data_validations.dataValidation:
            for _rng in str(_dv.sqref).split():
                _covered.add(_rng.split(":")[0])
        for _c in _cells:
            if _ws[_c].protection.locked is False and _c not in _covered:
                _unvalidated.append(f"{_sh}!{_c}")
    check(not _unvalidated,
          f"every unlocked cfg* picker carries a validation: {_unvalidated}")

    # the exam Attempt # list must not offer 3-5 (Skills still may)
    check(_DL.LISTS["Exam Attempt #"] == ["1", "2"] and
          _DL.LISTS["Attempt #"] == ["1", "2", "3", "4", "5"],
          "exam attempts 1-2, skills attempts 1-5")
    # nothing in the workbook consumes an exam typed "Spelling"
    check("Spelling" not in _DL.LISTS["Exam Type"],
          "Exam Type dropdown offers no unwired 'Spelling' value")

    # a swapped start/end must not roll its bogus hours into the 736 number
    check('nrSCH_TimeCheck,"OK"' in wb["ChapterMaster"]["G6"].value,
          "ChapterMaster Delivered Hrs excludes impossible-time blocks")

    # final PT: a partially scored rubric is not a pass
    check('COUNT(T6:Z6)<7' in wb["PT"]["AB6"].value and
          '"Incomplete"' in wb["PT"]["AB6"].value,
          "final PT needs all seven events scored before it can read Yes")
    for cell in ("L6", "P6"):
        check('PT!$AB6="Incomplete"' in wb["sysChecks"][cell].value,
              f"sysChecks {cell[:1]} has an explicit PT Incomplete state")
    check('PT!$AB6="Incomplete"' in wb["sysFlags"]["M6"].value,
          "a partially scored final PT raises the PT flag")

    # ChapterPacket spills: cap, reservation and an overflow marker
    cp = wb["ChapterPacket"]
    check("),20)" in cp["B20"].value and
          cp.cell(row=40, column=2).value is not None and
          "more instructor(s) not shown" in cp.cell(row=40, column=2).value,
          "ChapterPacket instructor list holds 20 and reports overflow")
    check("more block(s) not shown" in
          str(cp.cell(row=74, column=2).value or ""),
          "ChapterPacket block list reports overflow")

    # cfgHomeAgency is the cell that drives the real emails
    st_dvs = [dv.formula1 for dv in wb["Settings"].data_validations.dataValidation]
    home_row = next(r for r in range(6, 60)
                    if wb["Settings"].cell(row=r, column=5).value == "cfgHomeAgency")
    check("=rngAgencyIDs" in st_dvs and
          "not an AgencyID" in str(wb["Settings"].cell(row=home_row,
                                                      column=7).value or ""),
          "Home Agency is validated and checked against the Agencies sheet")

    _vba = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                 "src", "vba", "bpoc", "modAgencyEmail.bas"),
                    encoding="utf-8").read()
    check("FlushPendingLog" in _vba and "LogRun wb, agID" not in _vba,
          "EmailLog rows are written only after a send is confirmed")
    check("matches no AgencyID" in _vba,
          "a bogus cfgHomeAgency is reported, not silently skipped")
    _pr = _io.open(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                "src", "vba", "bpoc", "modPrint.bas"),
                   encoding="utf-8").read()
    check('ws.PageSetup.PrintArea = "$B$5:$O$" & lastRow' in _pr,
          "the Schedule button prints the used rows, not all 900")

    # ---- fixes from the V6 stress pass, round 10 --------------------------
    # #0: ExamSheet is a filed TCOLE grade sheet. Every exam-score consumer
    # in the workbook must ask whether a score was RECORDED, never whether a
    # row merely exists: an excused first attempt is logged with a blank Raw
    # Score, and the bare existence test printed Raw 0 / Recorded 0 / FAIL.
    _ex = wb["ExamSheet"]
    check('nrES_Raw,">=0"' in str(_ex["F11"].value) and
          'nrES_Rec,">=0"' in str(_ex["G11"].value),
          "ExamSheet Raw/Recorded need a RECORDED score, not just a row")
    check(_ex["J10"].value == "Absence" and
          "nrES_Absence" in str(_ex["J11"].value) and
          "Excused - pending" in str(_ex["J11"].value),
          "ExamSheet says WHY a printed grade row is blank")
    # the COUNTIFS text criterion "<>" means "not an EMPTY CELL", so it
    # counts a formula cell returning "". It must appear nowhere.
    _badguard = [(ws.title, c.coordinate) for ws in wb.worksheets
                 for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and 'nrES_Rec,"<>"' in c.value]
    check(not _badguard,
          'no exam consumer guards on nrES_Rec,"<>" (it counts "")')

    # #4: the PREVIOUS scored attempt must carry the same filter as the
    # latest one, or an unscored attempt-1 row reads as a recorded 0 and
    # both fabricates a consecutive-fail flag and erases a real grade drop
    for _c in ("AA6", "AC6"):
        _v = str(wb["sysGrades"][_c].value)
        check(_v.count('nrES_Raw,">=0"') == 2,
              f"sysGrades {_c[:2]} filters BOTH the latest and previous "
              "attempt on a recorded score")

    # #6/#9: Skills was the only graded number with no Row Check and no
    # range guard - 68 keyed as 680 satisfied the ch.41 firearms gate
    _sk = wb["Skills"]
    check(_sk["S5"].value == "Row Check" and
          "SCORE OUT OF RANGE" in str(_sk["S6"].value) and
          "nrSK_RowCheck" in wb.defined_names,
          "Skills carries a Row Check on its Score column")
    for _c in ("L6", "M6", "N6", "O6"):
        _v = str(wb["sysSkills"][_c].value)
        check('nrSK_Score,">=0"' in _v and 'nrSK_Score,"<=100"' in _v,
              f"sysSkills {_c[:1]} excludes an out-of-range firearms score")

    # #7: the same guard on spelling, whose average feeds the weighted
    # grade, the rank and the printed valedictorian pick
    _spl = wb["Spelling"]
    check('">=0"' in str(_spl["P6"].value) and '"<=100"' in str(_spl["Q6"].value)
          and '">=0"' in str(_spl["D58"].value),
          "spelling average, count and class average exclude 0-100 outliers")

    # #5: the Audit enrollment-documents grid was consumed by NOTHING
    check("nrENRall" in wb.defined_names and
          any("nrENRall" in str(c.value) for row in wb["sysAudit"].iter_rows()
              for c in row if isinstance(c.value, str)) and
          "nrENRall" in str(wb["sysChecks"]["W6"].value) and
          '$W6="Yes"' in str(wb["sysChecks"]["N6"].value) and
          wb["GradChecklist"]["O5"].value == "Enroll Docs",
          "enrollment documents reach the audit engine and the grad gate")

    # #1: a dynamic-array panel must never answer "nothing to report"
    # because something inside it broke. The all-clear lives in FILTER's
    # own if_empty; the outer IFERROR keeps a diagnostic.
    def _iferror_masks(v):
        """True when an IFERROR wrapping a dynamic array falls back to a
        MESSAGE - i.e. an error inside the array is rendered as a claim
        about the data instead of as an error."""
        i = 0
        while True:
            j = v.find("IFERROR(", i)
            if j == -1:
                return False
            op = j + len("IFERROR")
            cp = postprocess._find_matching(v, op)
            if cp == -1:
                return False
            args = postprocess._split_top(v[op + 1:cp])
            if len(args) == 2 and "FILTER(" in args[0]:
                m = args[1].strip()
                if (len(m) >= 3 and m[0] == '"' and m[-1] == '"' and
                        m != postprocess.PANEL_DIAG):
                    return True
            i = j + len("IFERROR(")

    _masked = [(ws.title, c.coordinate) for ws in wb.worksheets
               for row in ws.iter_rows() for c in row
               if isinstance(c.value, str) and "FILTER(" in c.value
               and _iferror_masks(c.value)]
    check(not _masked,
          f"no panel hides an error behind an all-clear message: {_masked}")
    _diag = [(ws.title, c.coordinate) for ws in wb.worksheets
             for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)
             and postprocess.PANEL_DIAG.strip('"') in c.value]
    check(len(_diag) >= 18 and
          {t for t, _ in _diag} >= {"Dashboard", "Ranking", "WatchList",
                                    "CadetProfile", "Transcript",
                                    "ChapterPacket", "SignIn",
                                    "WritingHandout"},
          "every safety panel keeps a DIAGNOSTIC outer fallback")

    # #2: an assignment dated on a CLOSURE day resolved to week 0 and
    # appeared on no weekly handout at all
    _wh = str(wb["WritingHandout"]["B11"].value)
    check("nrCDweek,MAX(nrCDweek),1)" in _wh,
          "WritingHandout brackets a closure-day assigned date to the next "
          "class day")
    _whdv = [dv.formula1 for dv in
             wb["WritingHandout"].data_validations.dataValidation]
    check(_whdv, "WritingHandout week picker is validated")

    # #3: the agency-email "reported to nobody" net counted only a BLANK
    # AgencyID, and sat inside the home-agency block
    check("is not on the Agencies sheet" in _vba and
          "Application.Match(agOfCadet" in _vba and
          _vba.index("Application.Match(agOfCadet") >
          _vba.index("FlushPendingLog wb") - 4000,
          "agency emails report a cadet whose AgencyID resolves to nothing")

    check(wb.calculation.fullCalcOnLoad, "fullCalcOnLoad set")


def main():
    print("== postprocess unit checks ==")
    test_postprocess_units()
    print("== workbook structure ==")
    test_workbook()
    print()
    if failures:
        print(f"{len(failures)} FAILURE(S)")
        sys.exit(1)
    print("All checks passed.")


if __name__ == "__main__":
    main()
