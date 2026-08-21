"""Build the BPOC Academy Management Workbook (V6).

Usage:  python3 build/bpoc/build_bpoc.py
Output: workbooks/BPOC_Academy_Management_V6.xlsx
        (VBA installed separately via tools/Install-VBA.ps1 -> .xlsm)
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from openpyxl import Workbook

import sheets_config
import sheets_inputs
import sheets_engine
import sheets_outputs
import postprocess

TAB_INPUT = "2E75B6"     # blue tabs = data entry
TAB_CONFIG = "C8A24B"    # gold = setup/masters
TAB_OUTPUT = "1E7145"    # green = outputs/printables
TAB_SYS = "808080"       # gray = locked engine

TAB_COLORS = {
    TAB_CONFIG: ["StartHere", "InputGuide", "Settings", "Lists", "Agencies",
                 "Instructors", "InstructorBanks",
                 "ChapterMaster", "ExamMaster", "ExamPlan", "SkillsMaster",
                 "SpellingMaster", "WritingMaster", "Control", "Schedule"],
    TAB_INPUT: ["Cadets", "ExamScores", "Spelling", "Attendance", "Makeup",
                "Skills", "Writing", "Incidents", "Counseling", "PT",
                "Medical", "Certifications", "StateExam", "Memos", "DailyLog",
                "AdvisoryBoard", "DismissalLog", "EmailLog"],
    TAB_OUTPUT: ["Dashboard", "ScoresGrid", "Ranking", "WatchList",
                 "CadetProfile", "Transcript", "GradChecklist", "Audit",
                 "Addendum", "ChapterPacket", "ExamSheet", "SignIn",
                 "EvalSheet", "SpellingPrint", "WritingHandout",
                 "EmailPreview", "PrintCenter", "NamedRanges"],
    TAB_SYS: ["sysGrades", "sysAttendance", "sysSkills", "sysIncidents",
              "sysFlags", "sysChecks", "sysAwards", "sysAudit",
              "sysListsHelper"],
}

SHEET_ORDER = [
    "Dashboard", "StartHere", "InputGuide", "PrintCenter",
    "Cadets", "ExamScores", "Spelling", "Attendance", "Makeup", "Skills",
    "Writing", "Incidents", "Counseling", "Memos", "DailyLog", "PT",
    "Medical", "Certifications", "StateExam", "AdvisoryBoard", "DismissalLog",
    "ScoresGrid", "Ranking", "WatchList", "CadetProfile", "Transcript",
    "GradChecklist", "Audit", "Addendum", "ChapterPacket", "ExamSheet",
    "SignIn", "EvalSheet", "SpellingPrint", "WritingHandout",
    "EmailPreview", "EmailLog",
    "Settings", "Lists", "Agencies", "Instructors", "InstructorBanks",
    "ChapterMaster",
    "ExamMaster", "ExamPlan", "SkillsMaster", "SpellingMaster",
    "WritingMaster", "Control", "Schedule",
    "sysGrades", "sysAttendance", "sysSkills", "sysIncidents", "sysFlags",
    "sysChecks", "sysAwards", "sysAudit", "sysListsHelper", "NamedRanges",
]


def build():
    wb = Workbook()
    wb.remove(wb.active)

    sheets_config.build_all_config(wb)
    sheets_inputs.build_all_inputs(wb)
    sheets_engine.build_all_engine(wb)
    sheets_outputs.build_all_outputs(wb)

    for color, names in TAB_COLORS.items():
        for n in names:
            if n in wb.sheetnames:
                wb[n].sheet_properties.tabColor = color

    order = [n for n in SHEET_ORDER if n in wb.sheetnames]
    order += [n for n in wb.sheetnames if n not in order]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0

    # daily-use ergonomics: frozen headers + filterable logs
    FREEZE = {
        "Cadets": "D6", "ExamScores": "D6", "Spelling": "D6", "Writing": "D6",
        "PT": "D6", "Certifications": "D6", "ScoresGrid": "D6",
        "StateExam": "D6", "GradChecklist": "D6", "sysGrades": "D6",
        "sysAttendance": "D6", "sysSkills": "D6", "sysIncidents": "D6",
        "sysFlags": "D6", "sysChecks": "D6",
        "Attendance": "B6", "Makeup": "B6", "Skills": "B6", "Incidents": "B6",
        "Counseling": "B6", "Medical": "B6", "DismissalLog": "B6",
        "EmailLog": "B6", "Schedule": "B6", "Agencies": "B6",
        "Memos": "B6", "DailyLog": "B6", "AdvisoryBoard": "B12",
        "Instructors": "C6", "InstructorBanks": "C6", "ChapterMaster": "E6",
        "WritingMaster": "D6", "SpellingMaster": "C6", "ExamMaster": "B6",
        "ExamPlan": "B6",
    }
    for name, cell in FREEZE.items():
        if name in wb.sheetnames:
            wb[name].freeze_panes = cell
    FILTERS = {
        "ExamScores": "B5:X5", "Attendance": "B5:T5", "Makeup": "B5:N5",
        "Skills": "B5:R5", "Incidents": "B5:O5", "Counseling": "B5:N5",
        "Medical": "B5:N5", "Schedule": "B5:O5", "DismissalLog": "B5:O5",
        "EmailLog": "B5:I5", "Memos": "B5:N5", "DailyLog": "B5:N5",
    }
    for name, ref in FILTERS.items():
        if name in wb.sheetnames:
            wb[name].auto_filter.ref = ref

    # store modern functions with _xlfn/_xlws/_xlpm prefixes so Excel
    # resolves them instead of showing #NAME?
    fixed = postprocess.fix_workbook(wb)
    print(f"Prefixed {fixed} formulas for Excel compatibility")

    # force full recalculation on open so every formula evaluates
    wb.calculation.fullCalcOnLoad = True

    out_dir = os.path.join(os.path.dirname(os.path.dirname(HERE)), "workbooks")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, "BPOC_Academy_Management_V6.xlsx")
    wb.save(out)
    print(f"Saved {out}")
    print(f"Sheets: {len(wb.sheetnames)}  Named ranges: {len(wb.defined_names)}")
    return out


if __name__ == "__main__":
    build()
